import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import altair as alt
import re
import unicodedata
import math
import textwrap
import calendar
import base64
import json
from io import BytesIO
from datetime import datetime, date, timedelta, timezone
from zoneinfo import ZoneInfo
from urllib.parse import quote
from urllib import request, parse
from io import BytesIO
import matplotlib.pyplot as plt
from matplotlib import font_manager
from supabase import create_client
from PIL import Image, ImageDraw, ImageFont


# =========================================================
# CONFIGURACIÓN GENERAL
# =========================================================

APP_NAME = "GEN Control"
APP_SUBTITLE = "Cobranzas Inmobiliarias"

META_GESTIONES = 2400
META_COMPROMISOS = 550
META_DIARIA_GESTIONES = 98
META_DIARIA_COMPROMISOS = 25
JORNADA_INICIO_HORA = 8
JORNADA_FIN_HORA = 17

# =========================================================
# REGLA DEFINITIVA DE RECUPERACIÓN
# NO MODIFICAR SIN AUTORIZACIÓN
# =========================================================

META_RECUPERACION = 170400
CANTIDAD_OPERADORES = 8


OPERADORES = {
    "avargas": {
        "nombre": "Aracely Peña Vargas",
        "nombre_mensaje": "Aracely",
        "correo": "apena@gestionia.bo",
    },
    "cvaca": {
        "nombre": "Carla Fernanda Vaca Cespedes",
        "nombre_mensaje": "Carla",
        "correo": "cvaca@gestionia.bo",
    },
    "jborja": {
        "nombre": "James Abel Borja Chirinos",
        "nombre_mensaje": "James",
        "correo": "jborja@gestionia.bo",
    },
    "arodriguez": {
        "nombre": "Leen Alisson Rodriguez Espinoza",
        "nombre_mensaje": "Leen",
        "correo": "lrodriguez@gestionia.bo",
    },
    "malvarez": {
        "nombre": "Mirla Anahir Alvarez",
        "nombre_mensaje": "Anahir",
        "correo": "malvarez@gestionia.bo",
    },
    "projas": {
        "nombre": "Percy Daniel Rojas Ortega",
        "nombre_mensaje": "Percy",
        "correo": "projas@gestionia.bo",
    },
    "yarinez": {
        "nombre": "Yanine Ariñez Rivero",
        "nombre_mensaje": "Yanine",
        "correo": "yarinez@gestionia.bo",
    },
    "yrivas": {
        "nombre": "Yessica Rivas Blanco",
        "nombre_mensaje": "Yessica",
        "correo": "yrivas@gestionia.bo",
    },
}


# =========================================================
# HORARIOS OPERATIVOS POR OPERADOR
# Basados en el rol compartido por coordinación.
# La meta diaria se distribuye SOLO sobre tiempo efectivo de trabajo.
# El break de 30 min congela el esperado.
# =========================================================

HORARIOS_OPERADORES = {
    "arodriguez": {
        0: {"entrada": "08:30", "break_inicio": "13:00", "break_fin": "13:30", "salida": "15:30", "jornada_horas": 7},
        1: {"entrada": "09:00", "break_inicio": "13:00", "break_fin": "13:30", "salida": "16:00", "jornada_horas": 7},
        2: {"entrada": "08:30", "break_inicio": "13:00", "break_fin": "13:30", "salida": "15:30", "jornada_horas": 7},
        3: {"entrada": "09:00", "break_inicio": "13:00", "break_fin": "13:30", "salida": "16:00", "jornada_horas": 7},
        4: {"entrada": "08:30", "break_inicio": "13:00", "break_fin": "13:30", "salida": "15:30", "jornada_horas": 7},
        5: {"entrada": "08:00", "break_inicio": None, "break_fin": None, "salida": "13:00", "jornada_horas": 5},
    },
    "malvarez": {
        0: {"entrada": "08:30", "break_inicio": "12:30", "break_fin": "13:00", "salida": "15:30", "jornada_horas": 7},
        1: {"entrada": "08:30", "break_inicio": "12:30", "break_fin": "13:00", "salida": "15:30", "jornada_horas": 7},
        2: {"entrada": "08:30", "break_inicio": "12:30", "break_fin": "13:00", "salida": "15:30", "jornada_horas": 7},
        3: {"entrada": "08:30", "break_inicio": "12:30", "break_fin": "13:00", "salida": "15:30", "jornada_horas": 7},
        4: {"entrada": "08:30", "break_inicio": "12:30", "break_fin": "13:00", "salida": "15:30", "jornada_horas": 7},
        5: {"entrada": "08:30", "break_inicio": None, "break_fin": None, "salida": "13:30", "jornada_horas": 5},
    },
    "avargas": {
        0: {"entrada": "12:30", "break_inicio": "15:30", "break_fin": "16:00", "salida": "19:30", "jornada_horas": 7},
        1: {"entrada": "12:30", "break_inicio": "15:30", "break_fin": "16:00", "salida": "19:30", "jornada_horas": 7},
        2: {"entrada": "12:30", "break_inicio": "15:30", "break_fin": "16:00", "salida": "19:30", "jornada_horas": 7},
        3: {"entrada": "12:30", "break_inicio": "15:30", "break_fin": "16:00", "salida": "19:30", "jornada_horas": 7},
        4: {"entrada": "12:30", "break_inicio": "15:30", "break_fin": "16:00", "salida": "19:30", "jornada_horas": 7},
        5: {"entrada": "14:00", "break_inicio": None, "break_fin": None, "salida": "19:00", "jornada_horas": 5},
    },
    "yarinez": {
        0: {"entrada": "08:00", "break_inicio": "12:00", "break_fin": "12:30", "salida": "15:00", "jornada_horas": 7},
        1: {"entrada": "08:00", "break_inicio": "12:00", "break_fin": "12:30", "salida": "15:00", "jornada_horas": 7},
        2: {"entrada": "08:00", "break_inicio": "12:00", "break_fin": "12:30", "salida": "15:00", "jornada_horas": 7},
        3: {"entrada": "08:00", "break_inicio": "12:00", "break_fin": "12:30", "salida": "15:00", "jornada_horas": 7},
        4: {"entrada": "08:00", "break_inicio": "12:00", "break_fin": "12:30", "salida": "15:00", "jornada_horas": 7},
        5: {"entrada": "09:00", "break_inicio": None, "break_fin": None, "salida": "14:00", "jornada_horas": 5},
    },
    "jborja": {
        0: {"entrada": "12:00", "break_inicio": "13:00", "break_fin": "13:30", "salida": "20:00", "jornada_horas": 8},
        1: {"entrada": "12:00", "break_inicio": "13:00", "break_fin": "13:30", "salida": "20:00", "jornada_horas": 8},
        2: {"entrada": "12:00", "break_inicio": "13:00", "break_fin": "13:30", "salida": "20:00", "jornada_horas": 8},
        3: {"entrada": "12:00", "break_inicio": "13:00", "break_fin": "13:30", "salida": "20:00", "jornada_horas": 8},
        4: {"entrada": "12:00", "break_inicio": "13:00", "break_fin": "13:30", "salida": "20:00", "jornada_horas": 8},
        5: {"entrada": "12:00", "break_inicio": "13:00", "break_fin": "13:30", "salida": "20:00", "jornada_horas": 8},
    },
    "cvaca": {
        0: {"entrada": "13:00", "break_inicio": "16:00", "break_fin": "16:30", "salida": "20:00", "jornada_horas": 7},
        1: {"entrada": "13:00", "break_inicio": "16:00", "break_fin": "16:30", "salida": "20:00", "jornada_horas": 7},
        2: {"entrada": "13:00", "break_inicio": "16:00", "break_fin": "16:30", "salida": "20:00", "jornada_horas": 7},
        3: {"entrada": "13:00", "break_inicio": "16:00", "break_fin": "16:30", "salida": "20:00", "jornada_horas": 7},
        4: {"entrada": "13:00", "break_inicio": "16:00", "break_fin": "16:30", "salida": "20:00", "jornada_horas": 7},
        5: {"entrada": "14:30", "break_inicio": None, "break_fin": None, "salida": "19:30", "jornada_horas": 5},
    },
    "projas": {
        0: {"entrada": "10:00", "break_inicio": "12:30", "break_fin": "13:00", "salida": "18:00", "jornada_horas": 8},
        1: {"entrada": "10:00", "break_inicio": "12:30", "break_fin": "13:00", "salida": "18:00", "jornada_horas": 8},
        2: {"entrada": "10:00", "break_inicio": "12:30", "break_fin": "13:00", "salida": "18:00", "jornada_horas": 8},
        3: {"entrada": "10:00", "break_inicio": "12:30", "break_fin": "13:00", "salida": "18:00", "jornada_horas": 8},
        4: {"entrada": "10:00", "break_inicio": "12:30", "break_fin": "13:00", "salida": "18:00", "jornada_horas": 8},
        5: {"entrada": "09:30", "break_inicio": "12:30", "break_fin": "13:00", "salida": "17:30", "jornada_horas": 8},
    },
    "yrivas": {
        0: {"entrada": "08:00", "break_inicio": "12:00", "break_fin": "12:30", "salida": "15:00", "jornada_horas": 7},
        1: {"entrada": "08:00", "break_inicio": "12:00", "break_fin": "12:30", "salida": "15:00", "jornada_horas": 7},
        2: {"entrada": "08:00", "break_inicio": "12:00", "break_fin": "12:30", "salida": "15:00", "jornada_horas": 7},
        3: {"entrada": "08:00", "break_inicio": "12:00", "break_fin": "12:30", "salida": "15:00", "jornada_horas": 7},
        4: {"entrada": "08:00", "break_inicio": "12:00", "break_fin": "12:30", "salida": "15:00", "jornada_horas": 7},
        5: {"entrada": "08:00", "break_inicio": None, "break_fin": None, "salida": "13:00", "jornada_horas": 5},
    },
}


def _hora_en_fecha(fecha_base, hhmm):
    hora, minuto = [
        int(x)
        for x in str(hhmm).split(":")
    ]

    return datetime(
        fecha_base.year,
        fecha_base.month,
        fecha_base.day,
        hora,
        minuto,
        tzinfo=ZoneInfo("America/La_Paz"),
    )


def obtener_horario_operador(usuario, fecha=None):
    """
    Devuelve el horario real según operador + día de la semana.
    Lunes=0 ... Sábado=5. Domingo no tiene jornada.
    """
    if fecha is None:
        fecha = fecha_local_actual()

    if hasattr(fecha, "date") and not hasattr(fecha, "weekday"):
        fecha = fecha.date()

    dia_semana = fecha.weekday()

    if dia_semana == 6:
        return None

    return HORARIOS_OPERADORES.get(
        usuario,
        {},
    ).get(
        dia_semana
    )


def calcular_progreso_jornada_operador(
    usuario,
    corte,
):
    """
    Calcula el porcentaje REAL de jornada efectiva transcurrida.

    Reglas:
    - antes de entrada: 0%;
    - durante break: esperado congelado;
    - después de salida: 100%;
    - break no cuenta como tiempo productivo;
    - mujeres: jornada nominal de 7 h;
    - hombres: jornada nominal de 8 h;
    - break: 30 min.
    """
    horario = obtener_horario_operador(
        usuario,
        corte.date(),
    )

    if not horario:
        return {
            "horario_configurado": False,
            "estado_jornada": "Horario pendiente",
            "proporcion": 0.0,
            "minutos_efectivos_transcurridos": 0,
            "minutos_efectivos_totales": 0,
            "entrada": None,
            "break_inicio": None,
            "break_fin": None,
            "salida": None,
        }

    entrada = _hora_en_fecha(
        corte,
        horario["entrada"],
    )
    tiene_break = bool(
        horario.get("break_inicio")
        and horario.get("break_fin")
    )

    break_inicio = (
        _hora_en_fecha(
            corte,
            horario["break_inicio"],
        )
        if tiene_break
        else None
    )
    break_fin = (
        _hora_en_fecha(
            corte,
            horario["break_fin"],
        )
        if tiene_break
        else None
    )
    salida = _hora_en_fecha(
        corte,
        horario["salida"],
    )

    # Si corte viene sin tzinfo (por pandas), asignar Bolivia.
    if corte.tzinfo is None:
        corte = corte.replace(
            tzinfo=ZoneInfo(
                "America/La_Paz"
            )
        )

    if tiene_break:
        tramo_1 = max(
            (break_inicio - entrada).total_seconds() / 60,
            0,
        )
        tramo_2 = max(
            (salida - break_fin).total_seconds() / 60,
            0,
        )
        minutos_totales = tramo_1 + tramo_2

        if corte < entrada:
            minutos_transcurridos = 0
            estado = "Jornada aún no iniciada"
        elif corte < break_inicio:
            minutos_transcurridos = min(
                (corte - entrada).total_seconds() / 60,
                tramo_1,
            )
            estado = "En jornada"
        elif corte < break_fin:
            minutos_transcurridos = tramo_1
            estado = "En break"
        elif corte < salida:
            minutos_transcurridos = tramo_1 + min(
                (corte - break_fin).total_seconds() / 60,
                tramo_2,
            )
            estado = "En jornada"
        else:
            minutos_transcurridos = minutos_totales
            estado = "Jornada finalizada"
    else:
        minutos_totales = max(
            (salida - entrada).total_seconds() / 60,
            0,
        )

        if corte < entrada:
            minutos_transcurridos = 0
            estado = "Jornada aún no iniciada"
        elif corte < salida:
            minutos_transcurridos = min(
                (corte - entrada).total_seconds() / 60,
                minutos_totales,
            )
            estado = "En jornada"
        else:
            minutos_transcurridos = minutos_totales
            estado = "Jornada finalizada"

    proporcion = (
        minutos_transcurridos
        / minutos_totales
        if minutos_totales
        else 0
    )

    return {
        "horario_configurado": True,
        "estado_jornada": estado,
        "proporcion": min(
            max(
                proporcion,
                0,
            ),
            1,
        ),
        "minutos_efectivos_transcurridos": int(
            round(
                minutos_transcurridos
            )
        ),
        "minutos_efectivos_totales": int(
            round(
                minutos_totales
            )
        ),
        "entrada": horario["entrada"],
        "break_inicio": horario["break_inicio"],
        "break_fin": horario["break_fin"],
        "salida": horario["salida"],
        "jornada_horas": horario[
            "jornada_horas"
        ],
    }


st.set_page_config(
    page_title=APP_NAME,
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)


# =========================================================
# DISEÑO VISUAL EJECUTIVO — FINAL VISUAL
# =========================================================
st.markdown(
    """
    <style>
    :root{
        --gen-navy:#102A43;
        --gen-navy-2:#163A5F;
        --gen-blue:#2F80ED;
        --gen-cyan:#22B8CF;
        --gen-green:#12B76A;
        --gen-orange:#F79009;
        --gen-red:#F04438;
        --gen-purple:#7A5AF8;
        --gen-bg:#F4F7FB;
        --gen-card:#FFFFFF;
        --gen-border:#E6ECF3;
        --gen-text:#172B4D;
        --gen-muted:#6B7C93;
    }

    /* Fondo general */
    .stApp{
        background:
            radial-gradient(circle at 92% 4%, rgba(47,128,237,.07), transparent 24rem),
            radial-gradient(circle at 15% 92%, rgba(34,184,207,.05), transparent 28rem),
            var(--gen-bg);
    }

    .block-container{
        max-width: 1500px;
        padding-top: 1.35rem !important;
        padding-bottom: 3.5rem !important;
        padding-left: 2.25rem !important;
        padding-right: 2.25rem !important;
    }

    /* Tipografía y jerarquía */
    h1,h2,h3{
        letter-spacing:-.025em !important;
        color:var(--gen-text) !important;
    }
    h2{font-weight:800!important}
    h3{font-weight:780!important}

    /* Sidebar */
    [data-testid="stSidebar"]{
        background:
            linear-gradient(180deg,#0E2742 0%,#102A43 55%,#0B2037 100%) !important;
        border-right:1px solid rgba(255,255,255,.06)!important;
        box-shadow:16px 0 40px rgba(16,42,67,.08);
    }

    [data-testid="stSidebar"] div[role="radiogroup"] label{
        min-height:42px;
        border-radius:12px!important;
        margin:2px 3px!important;
        padding:9px 11px!important;
    }

    [data-testid="stSidebar"] div[role="radiogroup"] label:has(input:checked){
        background:
            linear-gradient(90deg,rgba(47,128,237,.30),rgba(34,184,207,.14))!important;
        border:1px solid rgba(112,189,255,.18)!important;
        box-shadow:
            inset 3px 0 0 #47D7D0,
            0 8px 18px rgba(0,0,0,.08)!important;
    }

    .sidebar-logo{
        background:linear-gradient(135deg,#2F80ED,#22B8CF)!important;
        box-shadow:0 10px 22px rgba(47,128,237,.24)!important;
    }

    .sidebar-status-card{
        background:
            linear-gradient(135deg,rgba(34,184,207,.18),rgba(47,128,237,.16))!important;
        border:1px solid rgba(106,203,213,.18)!important;
        box-shadow:0 12px 24px rgba(0,0,0,.08)!important;
    }

    /* Hero principal */
    .hero-card{
        position:relative;
        overflow:hidden;
        border:none!important;
        border-radius:20px!important;
        padding:24px 26px!important;
        margin-bottom:18px!important;
        background:
            linear-gradient(115deg,#102A43 0%,#163A5F 56%,#1D5D7C 100%)!important;
        box-shadow:0 18px 42px rgba(16,42,67,.16)!important;
    }

    .hero-card:after{
        content:"";
        position:absolute;
        width:220px;height:220px;
        right:-55px;top:-100px;
        border-radius:50%;
        background:rgba(71,215,208,.12);
        box-shadow:
            -115px 160px 0 15px rgba(47,128,237,.10);
        pointer-events:none;
    }

    .hero-title{
        position:relative;
        z-index:1;
        color:#fff!important;
        font-size:26px!important;
        font-weight:850!important;
        letter-spacing:-.035em;
    }

    .hero-subtitle{
        position:relative;
        z-index:1;
        color:#BCD0E5!important;
        font-size:12px!important;
        margin-top:5px!important;
    }

    /* Estado del equipo */
    .team-state-v79{
        border-radius:16px!important;
        padding:14px 18px!important;
        border:1px solid var(--gen-border)!important;
        box-shadow:0 8px 24px rgba(16,42,67,.05)!important;
        background:#fff!important;
        margin-bottom:14px!important;
    }

    .team-state-green-v79{
        border-left:5px solid var(--gen-green)!important;
        background:linear-gradient(90deg,#F0FDF6,#FFFFFF 34%)!important;
    }
    .team-state-orange-v79{
        border-left:5px solid var(--gen-orange)!important;
        background:linear-gradient(90deg,#FFF8EB,#FFFFFF 34%)!important;
    }
    .team-state-red-v79{
        border-left:5px solid var(--gen-red)!important;
        background:linear-gradient(90deg,#FFF3F2,#FFFFFF 34%)!important;
    }

    .team-state-title-v79{
        font-size:17px!important;
        font-weight:830!important;
        color:var(--gen-text)!important;
    }
    .team-state-sub-v79{
        font-size:10px!important;
        color:var(--gen-muted)!important;
        margin-top:2px!important;
    }
    .team-focus-v79{
        border-radius:999px!important;
        padding:7px 11px!important;
        background:#F2F6FA!important;
        color:#40566D!important;
        font-size:10px!important;
        font-weight:750!important;
    }

    /* KPIs */
    .kpi-card-v79{
        min-height:126px;
        border-radius:17px!important;
        padding:15px 17px!important;
        border:1px solid var(--gen-border)!important;
        background:#fff!important;
        box-shadow:0 10px 26px rgba(16,42,67,.055)!important;
        transition:transform .14s ease, box-shadow .14s ease;
    }
    .kpi-card-v79:hover{
        transform:translateY(-2px);
        box-shadow:0 16px 34px rgba(16,42,67,.09)!important;
    }
    .kpi-card-blue-v79{border-top:3px solid var(--gen-blue)!important}
    .kpi-card-orange-v79{border-top:3px solid var(--gen-orange)!important}
    .kpi-card-green-v79{border-top:3px solid var(--gen-green)!important}
    .kpi-card-purple-v79{border-top:3px solid var(--gen-purple)!important}

    .kpi-label-v79{
        font-size:10px!important;
        font-weight:800!important;
        color:#61758A!important;
        text-transform:uppercase;
        letter-spacing:.04em;
    }
    .kpi-value-v79{
        color:#102A43!important;
        font-size:27px!important;
        line-height:1.05!important;
        font-weight:860!important;
        margin:10px 0 7px!important;
        letter-spacing:-.04em;
    }
    .kpi-foot-v79{
        font-size:9px!important;
        color:#7A8DA1!important;
    }

    /* Avance vs esperado */
    .compare-card-v79{
        border-radius:16px!important;
        background:#fff!important;
        border:1px solid var(--gen-border)!important;
        padding:15px 16px!important;
        box-shadow:0 8px 22px rgba(16,42,67,.045)!important;
    }
    .compare-title-v79{
        font-size:10px!important;
        text-transform:uppercase;
        letter-spacing:.04em;
        color:#708399!important;
        font-weight:800!important;
    }
    .compare-value-v79{
        font-size:25px!important;
        color:#102A43!important;
        font-weight:850!important;
        margin:7px 0 5px!important;
    }

    /* Contenedores estándar */
    [data-testid="stVerticalBlockBorderWrapper"]{
        border:1px solid var(--gen-border)!important;
        border-radius:16px!important;
        background:#fff!important;
        box-shadow:0 7px 22px rgba(16,42,67,.035)!important;
    }

    /* Métricas Streamlit */
    [data-testid="stMetric"]{
        border:1px solid var(--gen-border)!important;
        background:#fff!important;
        border-radius:14px!important;
        padding:10px 12px!important;
        box-shadow:none!important;
    }
    [data-testid="stMetricValue"]{
        color:#102A43!important;
        font-weight:850!important;
        letter-spacing:-.035em!important;
    }
    [data-testid="stMetricLabel"] p{
        color:#71849A!important;
        font-weight:700!important;
    }

    /* Inputs */
    [data-baseweb="input"] > div,
    [data-baseweb="select"] > div,
    [data-testid="stTextArea"] textarea{
        border-radius:11px!important;
        border-color:#DDE6EF!important;
        background:#fff!important;
    }

    /* Tabs */
    button[data-baseweb="tab"]{
        border-radius:10px!important;
        padding:9px 14px!important;
        font-weight:750!important;
        color:#5C7085!important;
    }
    button[data-baseweb="tab"][aria-selected="true"]{
        color:#12395B!important;
        background:#EAF3FB!important;
    }

    /* Botones */
    [data-testid="stButton"] button{
        min-height:40px;
        border-radius:11px!important;
        font-weight:760!important;
        transition:transform .12s ease, box-shadow .12s ease!important;
    }
    [data-testid="stButton"] button:hover{
        transform:translateY(-1px);
        box-shadow:0 7px 16px rgba(16,42,67,.10)!important;
    }
    [data-testid="stButton"] button[kind="primary"]{
        background:linear-gradient(135deg,#2F80ED,#2570D7)!important;
        border:none!important;
        color:#fff!important;
    }

    /* Dataframes */
    [data-testid="stDataFrame"]{
        border:1px solid var(--gen-border)!important;
        border-radius:14px!important;
        overflow:hidden!important;
        box-shadow:0 8px 24px rgba(16,42,67,.04)!important;
        background:#fff!important;
    }

    /* Expanders */
    [data-testid="stExpander"]{
        border:1px solid var(--gen-border)!important;
        border-radius:13px!important;
        background:#fff!important;
        box-shadow:none!important;
    }

    /* Alertas */
    [data-testid="stAlert"]{
        border-radius:13px!important;
        border-width:1px!important;
    }

    /* Ranking / alertas */
    .leader-strip-v77{
        border-radius:13px!important;
        padding:11px 13px!important;
        box-shadow:none!important;
    }
    .alert-summary-v77{
        border-radius:16px!important;
        border:1px solid #F3D7D2!important;
        background:linear-gradient(90deg,#FFF7F5,#FFFFFF)!important;
        box-shadow:0 8px 20px rgba(240,68,56,.04)!important;
    }

    /* Mensajes diarios: tarjetas operadores */
    .op-head-v74{margin-bottom:9px!important}
    .op-name-v74{
        font-size:16px!important;
        color:#172B4D!important;
        font-weight:850!important;
        letter-spacing:-.02em;
    }
    .op-contact-v74{
        color:#7A8EA5!important;
        font-size:9px!important;
    }
    .schedule-v74{
        border-radius:8px!important;
        background:#F6F9FC!important;
        border:1px solid #E7EDF4!important;
        color:#60748A!important;
    }
    .action-v74{
        border-radius:9px!important;
        font-weight:760!important;
    }
    .monthly-v74{
        border-top:1px solid #EDF1F5!important;
    }

    /* Scrollbar sutil */
    ::-webkit-scrollbar{width:10px;height:10px}
    ::-webkit-scrollbar-track{background:transparent}
    ::-webkit-scrollbar-thumb{
        background:#C7D3DF;
        border-radius:999px;
        border:3px solid transparent;
        background-clip:padding-box;
    }

    /* Responsive */
    @media(max-width:1100px){
        .block-container{
            padding-left:1.35rem!important;
            padding-right:1.35rem!important;
        }
        .kpi-value-v79{font-size:23px!important}
    }
    
    /* =========================================================
       V20 · MENSAJES DIARIOS · UI EJECUTIVA
       ========================================================= */
    .msg-v20-title{
        font-size:30px;
        line-height:1.08;
        font-weight:850;
        color:#13233A;
        letter-spacing:-.035em;
        margin:0 0 3px 0;
    }
    .msg-v20-sub{
        color:#6C7B90;
        font-size:13px;
        margin-bottom:14px;
    }
    .msg-v20-top{
        display:flex;
        gap:10px;
        justify-content:flex-end;
        margin-bottom:6px;
    }
    .msg-v20-chip{
        min-width:130px;
        padding:10px 14px;
        border:1px solid #DCE6F2;
        border-radius:12px;
        background:#FFFFFF;
        box-shadow:0 4px 14px rgba(25,55,90,.04);
    }
    .msg-v20-chip .k{
        font-size:10px;
        color:#7B8BA0;
        font-weight:750;
        text-transform:uppercase;
        letter-spacing:.04em;
    }
    .msg-v20-chip .v{
        font-size:15px;
        color:#173B72;
        font-weight:850;
        margin-top:2px;
    }
    .msg-v20-protect{
        border:1px solid #D8E4F2;
        border-radius:14px;
        background:linear-gradient(135deg,#FFFFFF 0%,#F7FAFF 100%);
        padding:15px 17px;
        margin:8px 0 14px;
        box-shadow:0 5px 18px rgba(32,64,110,.04);
    }
    .msg-v20-protect-title{
        font-size:14px;
        font-weight:850;
        color:#1856B4;
    }
    .msg-v20-protect-sub{
        color:#667A91;
        font-size:11px;
        margin-top:4px;
    }
    .msg-v20-sendbar{
        border:1px solid #D9E7FB;
        background:linear-gradient(90deg,#EEF5FF 0%,#F8FBFF 100%);
        border-radius:13px;
        padding:11px 14px;
        margin:10px 0 12px;
        color:#35597E;
        font-size:11px;
    }
    .msg-v20-sendbar strong{
        color:#173B72;
        font-size:13px;
    }
    .msg-v20-headrow{
        display:grid;
        grid-template-columns:2.3fr .9fr 1.25fr 1.65fr 1.65fr 1.15fr;
        gap:12px;
        color:#728198;
        font-size:9px;
        font-weight:800;
        text-transform:uppercase;
        letter-spacing:.04em;
        padding:0 13px 7px;
    }
    .msg-v20-opname{
        font-size:13px;
        line-height:1.2;
        font-weight:850;
        color:#172B4D;
    }
    .msg-v20-user{
        font-size:9px;
        color:#7A8EA5;
        margin-top:2px;
    }
    .msg-v20-status{
        display:inline-block;
        border-radius:999px;
        padding:5px 8px;
        font-size:9px;
        font-weight:800;
        white-space:nowrap;
    }
    .msg-v20-green{background:#E7F8EF;color:#11834C;}
    .msg-v20-orange{background:#FFF2DF;color:#B46500;}
    .msg-v20-red{background:#FFE9EB;color:#C53B4B;}
    .msg-v20-gray{background:#F0F3F7;color:#6E7D90;}
    .msg-v20-small{
        font-size:10px;
        color:#43556B;
        line-height:1.45;
    }
    .msg-v20-strong{
        font-weight:850;
        color:#172B4D;
    }
    .msg-v20-progress{
        height:5px;
        border-radius:999px;
        background:#E9EEF5;
        overflow:hidden;
        margin-top:4px;
    }
    .msg-v20-fill-g{
        height:100%;
        border-radius:999px;
        background:#2E7BEF;
    }
    .msg-v20-fill-c{
        height:100%;
        border-radius:999px;
        background:#2DBD78;
    }
    .msg-v20-side-title{
        font-size:14px;
        font-weight:850;
        color:#172B4D;
        margin-bottom:2px;
    }
    .msg-v20-side-sub{
        font-size:10px;
        color:#7A8EA5;
        margin-bottom:10px;
    }
    .msg-v20-side-card{
        border:1px solid #E0E7F0;
        border-radius:12px;
        padding:12px;
        background:#FFFFFF;
        margin-bottom:10px;
    }
    .msg-v20-side-ok{
        border:1px solid #CDEEDC;
        border-radius:12px;
        padding:12px;
        background:#EFFBF4;
        color:#16834F;
        font-size:11px;
        line-height:1.55;
        margin-top:10px;
    }
    .msg-v20-side-info{
        border:1px solid #D9E7FB;
        border-radius:12px;
        padding:12px;
        background:#F1F6FF;
        color:#35597E;
        font-size:10px;
        line-height:1.55;
        margin-top:10px;
    }

    div[data-testid="stVerticalBlockBorderWrapper"]{
        border-color:#E1E8F0!important;
        box-shadow:0 4px 15px rgba(24,52,88,.035);
    }


    /* =========================================================
       V21 · MEJORAS VISUALES DE ALTO IMPACTO
       ========================================================= */

    .stApp{
        background:
            radial-gradient(circle at top right, rgba(50,115,255,.04), transparent 30%),
            linear-gradient(180deg,#F7F9FC 0%,#FBFCFE 100%);
    }

    .block-container{
        max-width:1480px!important;
        padding-top:1.15rem!important;
        padding-bottom:2rem!important;
    }

    section[data-testid="stSidebar"]{
        background:linear-gradient(180deg,#102E4F 0%,#0C223C 100%)!important;
        border-right:1px solid rgba(255,255,255,.05);
    }

    .stButton>button{
        min-height:42px;
        border-radius:11px!important;
        font-weight:780!important;
        border:1px solid #D8E2ED!important;
        box-shadow:0 3px 10px rgba(20,52,88,.045);
        transition:.15s ease;
    }
    .stButton>button:hover{
        transform:translateY(-1px);
        box-shadow:0 7px 18px rgba(20,52,88,.075);
    }
    .stButton>button[kind="primary"]{
        background:linear-gradient(135deg,#1C6FF2 0%,#1657D3 100%)!important;
        color:#fff!important;
        border:none!important;
    }

    div[data-baseweb="input"]>div,
    div[data-baseweb="select"]>div,
    div[data-baseweb="base-input"]{
        border-radius:11px!important;
        border-color:#DCE5EF!important;
        min-height:42px!important;
        background:#FFFFFF!important;
    }

    div[data-testid="stVerticalBlockBorderWrapper"]{
        border:1px solid #E0E7F0!important;
        border-radius:14px!important;
        background:#FFFFFF!important;
        box-shadow:0 6px 18px rgba(23,52,88,.035)!important;
    }

    .msg-v20-title{
        font-size:31px!important;
        color:#132A46!important;
        margin-bottom:4px!important;
    }
    .msg-v20-sub{
        font-size:12px!important;
        color:#748399!important;
        margin-bottom:16px!important;
    }
    .msg-v20-chip{
        border-radius:13px!important;
        box-shadow:0 5px 16px rgba(21,58,104,.035)!important;
    }
    .msg-v20-protect{
        border-radius:15px!important;
        padding:16px 18px!important;
        background:linear-gradient(135deg,#FFFFFF 0%,#F5F9FF 100%)!important;
        box-shadow:0 6px 18px rgba(28,59,102,.035)!important;
    }
    .msg-v20-sendbar{
        border-radius:14px!important;
        padding:12px 15px!important;
        background:linear-gradient(90deg,#EDF5FF 0%,#F7FBFF 100%)!important;
    }

    .v21-kpi-grid{
        display:grid;
        grid-template-columns:repeat(4,minmax(0,1fr));
        gap:12px;
        margin:4px 0 14px;
    }
    .v21-kpi{
        border:1px solid #E0E7F0;
        background:#FFFFFF;
        border-radius:14px;
        padding:14px 15px;
        box-shadow:0 5px 16px rgba(25,55,90,.03);
    }
    .v21-kpi .k{
        font-size:9px;
        color:#7E8B9C;
        font-weight:850;
        text-transform:uppercase;
        letter-spacing:.05em;
    }
    .v21-kpi .v{
        font-size:23px;
        font-weight:900;
        line-height:1.05;
        margin-top:5px;
        color:#17314E;
    }
    .v21-kpi .s{
        font-size:9px;
        color:#8592A3;
        margin-top:4px;
    }
    .v21-blue .v{color:#1E66D8;}
    .v21-green .v{color:#13804D;}
    .v21-purple .v{color:#7257CA;}
    .v21-orange .v{color:#B76A00;}

    .msg-v20-headrow{
        background:#F8FAFD;
        border:1px solid #E7EDF4;
        border-radius:10px;
        padding:8px 13px!important;
        margin-bottom:7px!important;
    }

    .msg-v20-opname{
        font-size:13px!important;
        font-weight:880!important;
        color:#172B4D!important;
    }
    .msg-v20-user{
        color:#8794A5!important;
        font-size:9px!important;
    }

    .msg-v20-status{
        border-radius:999px!important;
        padding:5px 8px!important;
        font-size:9px!important;
    }

    .msg-v20-progress{
        height:6px!important;
        background:#E9EEF4!important;
    }

    .msg-v20-side-title{
        font-size:15px!important;
        color:#172B4D!important;
        font-weight:880!important;
    }

    .msg-v20-side-card,
    .msg-v20-side-ok,
    .msg-v20-side-info{
        border-radius:13px!important;
        box-shadow:0 5px 16px rgba(22,49,85,.025)!important;
    }

    div[data-testid="stAlert"]{
        border-radius:11px!important;
        padding:.65rem .8rem!important;
    }

    details{
        border-radius:12px!important;
        border:1px solid #E2E8F0!important;
        background:#FFFFFF!important;
    }

    @media(max-width:1100px){
        .v21-kpi-grid{grid-template-columns:repeat(2,minmax(0,1fr));}
    }
    @media(max-width:720px){
        .v21-kpi-grid{grid-template-columns:1fr;}
    }

    /* =========================================================
       V22 · SISTEMA VISUAL GLOBAL GEN CONTROL
       ========================================================= */

    :root{
        --gen-navy:#0F2B46;
        --gen-blue:#1E6DEB;
        --gen-cyan:#2D9CDB;
        --gen-green:#15965C;
        --gen-orange:#C97900;
        --gen-red:#C63D4D;
        --gen-purple:#7257C9;
        --gen-text:#172B4D;
        --gen-muted:#738297;
        --gen-border:#E0E7F0;
        --gen-bg:#F7F9FC;
        --gen-card:#FFFFFF;
    }

    /* Oculta el encabezado genérico duplicado. Cada módulo tendrá su propio header. */
    .gen-header{display:none!important;}

    /* Estructura general */
    .stApp{
        background:
            radial-gradient(circle at 92% 2%, rgba(36,105,220,.045), transparent 28%),
            linear-gradient(180deg,#F7F9FC 0%,#FBFCFE 100%)!important;
    }
    .block-container{
        max-width:1460px!important;
        padding-top:1.0rem!important;
        padding-left:1.65rem!important;
        padding-right:1.65rem!important;
        padding-bottom:2.2rem!important;
    }

    /* Tipografía */
    h1,h2,h3,h4{
        letter-spacing:-.025em!important;
        color:var(--gen-text)!important;
    }
    p,span,label{
        -webkit-font-smoothing:antialiased;
    }

    /* Header universal de página */
    .page-head-v22{
        position:relative;
        overflow:hidden;
        border-radius:18px;
        padding:18px 20px;
        margin:0 0 14px 0;
        background:linear-gradient(120deg,#102A43 0%,#173B5F 64%,#1B5D78 100%);
        box-shadow:0 12px 30px rgba(16,42,67,.10);
        color:white;
    }
    .page-head-v22:after{
        content:"";
        position:absolute;
        width:170px;
        height:170px;
        border-radius:50%;
        right:-52px;
        top:-86px;
        background:rgba(79,190,220,.11);
    }
    .page-head-kicker-v22{
        display:inline-flex;
        align-items:center;
        gap:6px;
        padding:4px 8px;
        border-radius:999px;
        background:rgba(255,255,255,.10);
        color:#CFE2F0;
        font-size:8px;
        font-weight:850;
        letter-spacing:.06em;
        text-transform:uppercase;
        margin-bottom:7px;
    }
    .page-head-title-v22{
        position:relative;
        z-index:1;
        font-size:24px;
        line-height:1.08;
        font-weight:900;
        color:#FFFFFF!important;
        letter-spacing:-.03em;
    }
    .page-head-sub-v22{
        position:relative;
        z-index:1;
        margin-top:5px;
        color:#C8D8E5;
        font-size:10px;
        line-height:1.45;
    }

    /* Títulos de sección */
    .section-head-v22{
        margin:15px 0 8px;
        display:flex;
        align-items:end;
        justify-content:space-between;
        gap:10px;
    }
    .section-title-v22{
        font-size:16px;
        font-weight:880;
        color:var(--gen-text);
    }
    .section-sub-v22{
        font-size:9px;
        color:var(--gen-muted);
        margin-top:2px;
    }

    /* Cards */
    div[data-testid="stVerticalBlockBorderWrapper"]{
        border:1px solid var(--gen-border)!important;
        border-radius:14px!important;
        background:#FFFFFF!important;
        box-shadow:0 5px 16px rgba(23,52,88,.03)!important;
    }

    /* Métricas nativas */
    div[data-testid="stMetric"]{
        background:#FFFFFF!important;
        border:1px solid var(--gen-border)!important;
        border-radius:13px!important;
        padding:11px 12px!important;
        box-shadow:0 4px 13px rgba(23,52,88,.025)!important;
    }
    div[data-testid="stMetric"] label{
        color:#78879A!important;
        font-size:10px!important;
        font-weight:750!important;
    }
    div[data-testid="stMetricValue"]{
        color:#172B4D!important;
        font-size:22px!important;
        font-weight:880!important;
        letter-spacing:-.025em!important;
    }

    /* Inputs */
    div[data-baseweb="input"]>div,
    div[data-baseweb="select"]>div,
    div[data-baseweb="base-input"]{
        background:#FFFFFF!important;
        border-color:#DDE6EF!important;
        border-radius:10px!important;
        min-height:40px!important;
    }
    div[data-testid="stNumberInput"] input,
    div[data-testid="stTextInput"] input{
        font-size:13px!important;
    }

    /* Botones */
    .stButton>button,
    .stDownloadButton>button{
        min-height:40px;
        border-radius:10px!important;
        font-weight:780!important;
        border:1px solid #D8E2ED!important;
        box-shadow:0 3px 9px rgba(20,52,88,.035);
        transition:all .14s ease;
    }
    .stButton>button:hover,
    .stDownloadButton>button:hover{
        transform:translateY(-1px);
        border-color:#BFD0E2!important;
        box-shadow:0 6px 15px rgba(20,52,88,.07);
    }
    .stButton>button[kind="primary"]{
        background:linear-gradient(135deg,#1D6EF0,#1659D3)!important;
        color:#FFFFFF!important;
        border:none!important;
    }

    /* Expander */
    details{
        border:1px solid #E1E8F0!important;
        border-radius:12px!important;
        background:#FFFFFF!important;
        overflow:hidden;
        box-shadow:0 3px 10px rgba(23,52,88,.02);
    }
    details summary{
        font-weight:760!important;
        color:#2A415C!important;
        padding:.8rem .9rem!important;
    }

    /* Tabs */
    button[data-baseweb="tab"]{
        font-weight:760!important;
        color:#65768B!important;
        padding:.7rem .9rem!important;
    }
    button[data-baseweb="tab"][aria-selected="true"]{
        color:#1E63D6!important;
    }

    /* Tablas */
    div[data-testid="stDataFrame"]{
        border:1px solid #E2E8F0!important;
        border-radius:13px!important;
        overflow:hidden!important;
        box-shadow:0 4px 14px rgba(23,52,88,.025);
    }

    /* Uploader */
    section[data-testid="stFileUploaderDropzone"]{
        border:1.5px dashed #BDD0E4!important;
        border-radius:14px!important;
        background:linear-gradient(180deg,#FFFFFF,#F8FBFF)!important;
        padding:1.25rem!important;
    }

    /* Alertas */
    div[data-testid="stAlert"]{
        border-radius:11px!important;
        padding:.68rem .82rem!important;
        box-shadow:none!important;
    }

    /* Sidebar */
    section[data-testid="stSidebar"]{
        background:linear-gradient(180deg,#102E4F 0%,#0C223C 100%)!important;
        border-right:1px solid rgba(255,255,255,.05)!important;
    }
    .sidebar-brand{margin-bottom:15px!important;}
    .sidebar-section-label{opacity:.72!important;}
    .sidebar-status-card{
        border-radius:13px!important;
        background:rgba(31,129,174,.18)!important;
        border-color:rgba(111,204,238,.18)!important;
    }
    .sidebar-profile{padding-top:4px!important;}
    section[data-testid="stSidebar"] hr{
        border-color:rgba(255,255,255,.08)!important;
        margin:.85rem 0!important;
    }

    /* Reduce aire excesivo */
    .element-container{margin-bottom:.15rem;}
    [data-testid="stVerticalBlock"]{gap:.65rem;}

    @media(max-width:1100px){
        .block-container{
            padding-left:1.1rem!important;
            padding-right:1.1rem!important;
        }
    }

    /* V24 · Corrección del espacio vacío en Mensajes diarios */
    div[data-testid="stVerticalBlockBorderWrapper"]{
        min-height:0!important;
        height:auto!important;
    }
    .msg-v20-sendbar{
        margin-top:6px!important;
        margin-bottom:8px!important;
    }
    .msg-v20-headrow{
        margin-top:2px!important;
    }

    /* =========================================================
       V25 FINAL · PULIDO VISUAL Y UX GLOBAL
       ========================================================= */

    /* Mejor densidad visual */
    .block-container{
        max-width:1440px!important;
        padding-top:.85rem!important;
        padding-bottom:1.8rem!important;
    }

    /* Labels más claros */
    label[data-testid="stWidgetLabel"] p{
        font-size:10px!important;
        font-weight:760!important;
        color:#53657A!important;
    }

    /* Separación consistente entre secciones */
    .section-head-v22{
        margin-top:12px!important;
        margin-bottom:7px!important;
    }

    /* Hero compacto */
    .page-head-v22{
        padding:15px 18px!important;
        margin-bottom:12px!important;
        border-radius:16px!important;
    }
    .page-head-title-v22{
        font-size:22px!important;
    }

    /* Tarjetas KPI menos altas */
    .v21-kpi{
        min-height:82px!important;
        padding:12px 14px!important;
    }
    .v21-kpi .v{
        font-size:21px!important;
    }

    /* Mensajes diarios */
    .msg-v20-title{
        font-size:27px!important;
    }
    .msg-v20-protect{
        padding:12px 15px!important;
        margin:6px 0 10px!important;
    }
    .msg-v20-chip{
        min-width:112px!important;
        padding:8px 11px!important;
    }
    .msg-v20-chip .v{
        font-size:13px!important;
    }
    .msg-v20-headrow{
        font-size:8.5px!important;
    }

    /* Evitar tarjetas innecesariamente grandes */
    div[data-testid="stVerticalBlockBorderWrapper"]{
        min-height:unset!important;
    }

    /* Tablas: mejor lectura */
    div[data-testid="stDataFrame"]{
        font-size:12px!important;
    }

    /* Estado de sidebar más compacto */
    .sidebar-status-card{
        padding:11px 12px!important;
        margin-top:8px!important;
    }

    /* Popovers */
    div[data-baseweb="popover"]{
        border-radius:12px!important;
    }

    /* Toggle */
    div[data-testid="stToggle"]{
        margin-top:0!important;
        margin-bottom:0!important;
    }

    /* Ocultar decoración Streamlit inferior cuando no aporta */
    footer{visibility:hidden;}

    /* Responsive real para laptops */
    @media(max-width:1280px){
        .block-container{
            padding-left:1rem!important;
            padding-right:1rem!important;
        }
        .msg-v20-headrow{
            grid-template-columns:2.1fr .85fr 1.15fr 1.45fr 1.55fr 1fr!important;
        }
    }

    @media(max-width:900px){
        .page-head-v22{padding:13px 14px!important;}
        .page-head-title-v22{font-size:20px!important;}
        .msg-v20-title{font-size:24px!important;}
        .msg-v20-top{justify-content:flex-start!important;}
    }

    /* =========================================================
       V26 · RANKING EJECUTIVO
       ========================================================= */
    .rank-head-v26{
        display:flex;
        align-items:center;
        justify-content:space-between;
        gap:12px;
        margin:14px 0 8px;
    }
    .rank-title-v26{
        font-size:19px;
        font-weight:900;
        color:#172B4D;
        letter-spacing:-.025em;
    }
    .rank-sub-v26{
        font-size:9px;
        color:#7C899A;
        margin-top:2px;
    }

    .rank-kpis-v26{
        display:grid;
        grid-template-columns:repeat(6,minmax(0,1fr));
        gap:9px;
        margin:10px 0 12px;
    }
    .rank-kpi-v26{
        border:1px solid #E0E7F0;
        background:#FFFFFF;
        border-radius:13px;
        padding:11px 12px;
        box-shadow:0 4px 14px rgba(23,52,88,.025);
        min-height:78px;
    }
    .rank-kpi-v26 .k{
        color:#7C899A;
        font-size:8px;
        font-weight:850;
        text-transform:uppercase;
        letter-spacing:.04em;
    }
    .rank-kpi-v26 .v{
        color:#172B4D;
        font-size:19px;
        font-weight:900;
        margin-top:5px;
        line-height:1.05;
    }
    .rank-kpi-v26 .s{
        color:#8592A3;
        font-size:8px;
        margin-top:4px;
    }
    .rank-kpi-v26.green .v{color:#13804D;}
    .rank-kpi-v26.orange .v{color:#B66A00;}
    .rank-kpi-v26.blue .v{color:#1E66D8;}
    .rank-kpi-v26.purple .v{color:#7257C9;}

    .rank-wrap-v26{
        border:1px solid #DFE7F0;
        border-radius:14px;
        overflow:hidden;
        background:#FFFFFF;
        box-shadow:0 5px 16px rgba(23,52,88,.028);
        margin-top:8px;
    }
    .rank-row-v26{
        display:grid;
        grid-template-columns:44px 2.1fr 1.55fr 1.55fr 1.8fr 1fr;
        align-items:center;
        gap:10px;
        padding:9px 12px;
        border-bottom:1px solid #EDF1F5;
        font-size:10px;
    }
    .rank-row-v26:last-child{border-bottom:none;}
    .rank-row-v26:hover{background:#FAFCFF;}
    .rank-row-v26.header{
        background:#F8FAFD;
        color:#65768A;
        font-size:8px;
        font-weight:850;
        text-transform:uppercase;
        letter-spacing:.035em;
        padding-top:8px;
        padding-bottom:8px;
    }
    .rank-row-v26.first{
        background:linear-gradient(90deg,#F0FBF5 0%,#FFFFFF 42%);
    }
    .rank-pos-v26{
        font-size:12px;
        font-weight:900;
        color:#173B72;
    }
    .rank-name-v26{
        font-size:10px;
        font-weight:850;
        color:#172B4D;
        line-height:1.25;
    }
    .rank-metric-v26{
        color:#172B4D;
        font-weight:800;
        font-size:10px;
    }
    .rank-metric-sub-v26{
        color:#7C899A;
        font-size:8px;
        margin-top:2px;
    }
    .rank-bar-v26{
        height:5px;
        border-radius:999px;
        background:#E9EEF4;
        overflow:hidden;
        margin-top:4px;
    }
    .rank-bar-v26 span{
        display:block;
        height:100%;
        border-radius:999px;
    }
    .rank-bar-v26 .g{background:#2877E8;}
    .rank-bar-v26 .c{background:#7555D9;}
    .rank-bar-v26 .r{background:#1D9B62;}

    .rank-pill-v26{
        display:inline-flex;
        align-items:center;
        justify-content:center;
        border-radius:999px;
        padding:5px 8px;
        font-size:8px;
        font-weight:850;
        white-space:nowrap;
    }
    .rank-pill-v26.ok{background:#E8F8EF;color:#12814B;}
    .rank-pill-v26.warn{background:#FFF1E0;color:#B76800;}
    .rank-pill-v26.bad{background:#FFE9EC;color:#C53B4B;}

    .rank-strip-grid-v26{
        display:grid;
        grid-template-columns:1fr 1fr;
        gap:10px;
        margin:10px 0 12px;
    }
    .rank-strip-v26{
        border-radius:12px;
        padding:10px 12px;
        font-size:9px;
        line-height:1.45;
    }
    .rank-strip-v26 strong{font-size:10px;}
    .rank-strip-v26.good{
        background:#EEFBF4;
        border:1px solid #CDEDDC;
        color:#137C4A;
    }
    .rank-strip-v26.warn{
        background:#FFF7ED;
        border:1px solid #F4D7AF;
        color:#A75C00;
    }

    .alerts-grid-v26{
        display:grid;
        grid-template-columns:1.2fr .9fr;
        gap:12px;
        margin-top:8px;
    }
    .alerts-card-v26{
        border:1px solid #E2E8F0;
        border-radius:13px;
        background:#FFFFFF;
        padding:12px 13px;
    }
    .alerts-card-v26 h4{
        margin:0 0 4px;
        font-size:12px;
        color:#172B4D;
    }
    .alerts-card-v26 p{
        margin:0 0 8px;
        font-size:8px;
        color:#7C899A;
    }
    .alert-line-v26{
        display:flex;
        justify-content:space-between;
        gap:10px;
        padding:6px 0;
        border-bottom:1px solid #EEF2F6;
        font-size:9px;
    }
    .alert-line-v26:last-child{border-bottom:none;}
    .alert-line-v26 .name{font-weight:800;color:#263D5B;}
    .alert-line-v26 .val{font-weight:850;color:#B65F00;}

    @media(max-width:1100px){
        .rank-kpis-v26{grid-template-columns:repeat(3,minmax(0,1fr));}
        .rank-row-v26{grid-template-columns:40px 1.8fr 1.35fr 1.35fr 1.55fr .9fr;}
    }

    /* =========================================================
       V29 · RESUMEN MENSUAL
       ========================================================= */
    .month-head-v29{
        display:flex;
        align-items:center;
        justify-content:space-between;
        gap:14px;
        margin:2px 0 12px;
    }
    .month-title-v29{
        font-size:22px;
        font-weight:900;
        color:#102A43;
        letter-spacing:-.03em;
    }
    .month-sub-v29{
        color:#6F8195;
        font-size:9px;
        margin-top:3px;
    }
    .month-badge-v29{
        padding:7px 10px;
        border-radius:999px;
        background:#EEF5FF;
        border:1px solid #D8E6F8;
        color:#245A8D;
        font-size:9px;
        font-weight:850;
        white-space:nowrap;
    }
    .month-grid-v29{
        display:grid;
        grid-template-columns:repeat(3,minmax(0,1fr));
        gap:11px;
        margin:10px 0 14px;
    }
    .month-card-v29{
        border:1px solid #E1E8F0;
        border-radius:15px;
        padding:14px 15px;
        background:#FFFFFF;
        box-shadow:0 7px 18px rgba(16,42,67,.035);
    }
    .month-card-v29 .lbl{
        font-size:8px;
        font-weight:850;
        color:#71849A;
        text-transform:uppercase;
        letter-spacing:.045em;
    }
    .month-card-v29 .val{
        font-size:23px;
        font-weight:900;
        color:#102A43;
        margin:6px 0 3px;
        letter-spacing:-.035em;
    }
    .month-card-v29 .sub{
        font-size:8px;
        color:#8192A4;
    }
    .month-card-v29.blue{border-top:3px solid #2E77E5;}
    .month-card-v29.orange{border-top:3px solid #D88700;}
    .month-card-v29.green{border-top:3px solid #2A9D66;}
    .month-close-v29{
        border:1px solid #D5E6F8;
        background:linear-gradient(90deg,#F3F8FF,#FBFDFF);
        border-radius:13px;
        padding:10px 12px;
        margin:8px 0 12px;
        font-size:9px;
        color:#35597E;
    }
    .month-close-v29 strong{color:#173B72;}
    @media(max-width:950px){
        .month-grid-v29{grid-template-columns:1fr;}
    }
</style>
    """,
    unsafe_allow_html=True,
)


# =========================================================
# VERSIÓN FINAL VISUAL · UI ejecutiva
# =========================================================

# =========================================================
# FUNCIONES AUXILIARES
# =========================================================

def quitar_acentos(texto):
    texto = str(texto)
    return "".join(
        c for c in unicodedata.normalize("NFD", texto)
        if unicodedata.category(c) != "Mn"
    )


def normalizar_texto(texto):
    texto = quitar_acentos(texto).lower().strip()
    texto = re.sub(r"\s+", " ", texto)
    return texto


def normalizar_columna(columna):
    texto = normalizar_texto(columna)
    texto = texto.replace("$", "")
    texto = texto.replace("bs.", "")
    texto = texto.replace("bs", "")
    texto = re.sub(r"[^a-z0-9 ]", "", texto)
    texto = re.sub(r"\s+", " ", texto).strip()
    return texto


def convertir_numero(valor):
    """
    Convierte valores como:
    99.892,47
    170400
    Bs 99.892,47
    $ 99,892.47
    """
    if pd.isna(valor):
        return 0.0

    if isinstance(valor, (int, float)):
        return float(valor)

    texto = str(valor).strip()

    if texto == "":
        return 0.0

    texto = texto.replace("Bs.", "")
    texto = texto.replace("Bs", "")
    texto = texto.replace("$", "")
    texto = texto.replace(" ", "")

    # Formato latino: 99.892,47
    if "," in texto:
        texto = texto.replace(".", "")
        texto = texto.replace(",", ".")
    else:
        # Si únicamente tiene puntos, determinar si es decimal
        partes = texto.split(".")
        if len(partes) > 2:
            texto = "".join(partes)
        elif len(partes) == 2 and len(partes[1]) == 3:
            texto = "".join(partes)

    texto = re.sub(r"[^0-9.\-]", "", texto)

    try:
        return float(texto)
    except Exception:
        return 0.0


def formato_entero(valor):
    try:
        return f"{int(round(valor)):,}".replace(",", ".")
    except Exception:
        return "0"


def formato_bs(valor):
    try:
        return (
            f"Bs {valor:,.2f}"
            .replace(",", "X")
            .replace(".", ",")
            .replace("X", ".")
        )
    except Exception:
        return "Bs 0,00"


def formato_usd(valor):
    try:
        return (
            f"USD {valor:,.2f}"
            .replace(",", "X")
            .replace(".", ",")
            .replace("X", ".")
        )
    except Exception:
        return "USD 0,00"


def formato_porcentaje(valor):
    try:
        return f"{valor:.2f}%".replace(".", ",")
    except Exception:
        return "0,00%"


def buscar_columna(df, palabras):
    """
    Busca una columna usando partes de su nombre.
    """
    columnas_norm = {
        col: normalizar_columna(col)
        for col in df.columns
    }

    for col_original, col_norm in columnas_norm.items():
        if all(palabra in col_norm for palabra in palabras):
            return col_original

    return None


# =========================================================
# CALENDARIO Y MENSAJES
# =========================================================

def fecha_local_actual():
    return datetime.now(ZoneInfo("America/La_Paz")).date()



def nombre_mes_es(numero_mes):
    meses = [
        "", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
        "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
    ]
    try:
        return meses[int(numero_mes)]
    except Exception:
        return str(numero_mes)


def periodo_recuperacion_actual(fecha_ref=None):
    """
    Regla definitiva:
    - Gestiones y Compromisos: siempre mes actual.
    - Recuperación: días 1–5 sigue cerrando el mes anterior.
    - Recuperación: desde día 6 corresponde al mes actual.
    """
    fecha_ref = fecha_ref or fecha_local_actual()

    if fecha_ref.day <= 5:
        primer_dia = fecha_ref.replace(day=1)
        mes_anterior = primer_dia - timedelta(days=1)
        return {
            "cierre_anterior": True,
            "mes": mes_anterior.month,
            "anio": mes_anterior.year,
            "nombre_mes": nombre_mes_es(mes_anterior.month),
            "etiqueta": f"Cierre {nombre_mes_es(mes_anterior.month)}",
            "esperado_pct": 100.0,
            "plazo": fecha_ref.replace(day=5),
        }

    return {
        "cierre_anterior": False,
        "mes": fecha_ref.month,
        "anio": fecha_ref.year,
        "nombre_mes": nombre_mes_es(fecha_ref.month),
        "etiqueta": nombre_mes_es(fecha_ref.month),
        "esperado_pct": None,
        "plazo": None,
    }


def esperado_indicador(indicador, esperado_mes):
    """Esperado correcto según indicador y período."""
    periodo = periodo_recuperacion_actual()
    if indicador == "Recuperación" and periodo["cierre_anterior"]:
        return 100.0
    return float(esperado_mes)



def ahora_bolivia():
    """
    Hora Bolivia calculada desde UTC real.
    Evita depender de la zona horaria del servidor de Streamlit.
    """
    return datetime.now(
        timezone.utc
    ).astimezone(
        ZoneInfo("America/La_Paz")
    )


def datetime_bolivia(valor):
    """
    Convierte cualquier timestamp guardado a hora Bolivia.

    Regla V96:
    - si viene con zona/offset: se convierte a America/La_Paz;
    - si viene sin zona (registros antiguos): se interpreta como UTC,
      que es como Supabase/Postgres suele entregar timestamptz serializado.
    """
    if valor is None or valor == "":
        return None

    try:
        if isinstance(valor, datetime):
            dt = valor
        else:
            dt = pd.to_datetime(
                valor
            ).to_pydatetime()

        if dt.tzinfo is None:
            dt = dt.replace(
                tzinfo=timezone.utc
            )

        return dt.astimezone(
            ZoneInfo("America/La_Paz")
        )
    except Exception:
        return None

def es_jornada_laboral(fecha):
    # Regla actual: lunes a sábado. Domingo no cuenta.
    return fecha.weekday() != 6


def jornadas_mes(fecha_ref):
    primer_dia = fecha_ref.replace(day=1)
    if fecha_ref.month == 12:
        siguiente_mes = fecha_ref.replace(
            year=fecha_ref.year + 1,
            month=1,
            day=1,
        )
    else:
        siguiente_mes = fecha_ref.replace(
            month=fecha_ref.month + 1,
            day=1,
        )

    ultimo_dia = siguiente_mes - timedelta(days=1)

    dias = []
    actual = primer_dia

    while actual <= ultimo_dia:
        if es_jornada_laboral(actual):
            dias.append(actual)
        actual += timedelta(days=1)

    return dias


def calcular_jornadas(fecha_ref=None):
    fecha_ref = fecha_ref or fecha_local_actual()
    dias = jornadas_mes(fecha_ref)

    transcurridas = len(
        [d for d in dias if d <= fecha_ref]
    )

    disponibles = len(
        [d for d in dias if d >= fecha_ref]
    )

    return {
        "total": len(dias),
        "transcurridas": transcurridas,
        "disponibles": disponibles,
        "esperado_pct": (
            transcurridas / len(dias) * 100
            if dias
            else 0
        ),
    }


def inicializar_calendario_mes(fecha_ref):
    clave_mes = f"{fecha_ref.year:04d}-{fecha_ref.month:02d}"

    if clave_mes not in st.session_state.calendario_laboral:
        dias_mes = calendar.monthrange(
            fecha_ref.year,
            fecha_ref.month,
        )[1]

        st.session_state.calendario_laboral[clave_mes] = {
            dia: (
                date(
                    fecha_ref.year,
                    fecha_ref.month,
                    dia,
                ).weekday() != 6
            )
            for dia in range(1, dias_mes + 1)
        }

    return clave_mes


def meta_diaria_compromisos_calendario_v35(fecha_ref=None, acumulado_actual=0):
    """Meta diaria de compromisos basada en calendario laboral real."""
    fecha_ref = fecha_ref or fecha_local_actual()
    jornadas = jornadas_configuradas(fecha_ref)

    meta_mensual = int(
        st.session_state.get("meta_compromisos_cfg", META_COMPROMISOS)
    )

    total_jornadas = max(int(jornadas.get("total", 0)), 1)
    disponibles = max(int(jornadas.get("disponibles", 0)), 1)

    faltante = max(meta_mensual - int(acumulado_actual or 0), 0)

    if faltante <= 0:
        return {
            "meta_diaria": 0,
            "base_mes": 0,
            "necesaria": 0,
            "faltante": 0,
            "jornadas_total": total_jornadas,
            "jornadas_disponibles": disponibles,
        }

    base_mes = math.ceil(meta_mensual / total_jornadas)
    necesaria = math.ceil(faltante / disponibles)

    return {
        "meta_diaria": max(base_mes, necesaria),
        "base_mes": base_mes,
        "necesaria": necesaria,
        "faltante": faltante,
        "jornadas_total": total_jornadas,
        "jornadas_disponibles": disponibles,
    }



def jornadas_configuradas(fecha_ref=None):
    fecha_ref = fecha_ref or fecha_local_actual()
    clave_mes = inicializar_calendario_mes(fecha_ref)

    calendario_mes = st.session_state.calendario_laboral[
        clave_mes
    ]

    dias_laborales = [
        date(
            fecha_ref.year,
            fecha_ref.month,
            dia,
        )
        for dia, es_laboral in calendario_mes.items()
        if es_laboral
    ]

    transcurridas = len(
        [d for d in dias_laborales if d <= fecha_ref]
    )

    disponibles = len(
        [d for d in dias_laborales if d >= fecha_ref]
    )

    return {
        "total": len(dias_laborales),
        "transcurridas": transcurridas,
        "disponibles": disponibles,
        "esperado_pct": (
            transcurridas / len(dias_laborales) * 100
            if dias_laborales
            else 0
        ),
        "dias": dias_laborales,
        "clave_mes": clave_mes,
    }


def metas_actuales():
    return {
        "gestiones": st.session_state.meta_gestiones_cfg,
        "compromisos": st.session_state.meta_compromisos_cfg,
        "recuperacion": st.session_state.meta_recuperacion_cfg,
        "diaria_compromisos": st.session_state.meta_diaria_compromisos_cfg,
    }


def objetivo_hoy_gestiones(acumulado, jornadas_disponibles):
    meta = st.session_state.meta_gestiones_cfg
    faltante = max(meta - acumulado, 0)

    if faltante <= 0:
        return 0

    if jornadas_disponibles <= 0:
        return int(math.ceil(faltante))

    return int(
        math.ceil(faltante / jornadas_disponibles)
    )


def objetivo_hoy_compromisos(acumulado, jornadas_disponibles):
    meta = st.session_state.meta_compromisos_cfg
    minimo_diario = (
        st.session_state.meta_diaria_compromisos_cfg
    )

    faltante = max(meta - acumulado, 0)

    # Aunque la meta mensual ya esté cumplida,
    # se mantiene el mínimo diario definido.
    if faltante <= 0:
        return int(minimo_diario)

    if jornadas_disponibles <= 0:
        return max(
            int(math.ceil(faltante)),
            int(minimo_diario),
        )

    recuperacion_diaria = int(
        math.ceil(faltante / jornadas_disponibles)
    )

    return max(
        int(minimo_diario),
        recuperacion_diaria,
    )


def clasificar_avance(porcentaje, esperado):
    if porcentaje >= 100:
        return "Meta cumplida"
    if porcentaje >= esperado + 5:
        return "Excelente avance"
    if porcentaje >= esperado - 3:
        return "Buen avance"
    if porcentaje >= esperado - 10:
        return "En seguimiento"
    return "Reforzar"




def promesas_es_mes_actual_v32(fecha_ref=None):
    """
    Valida el período REAL de Promesas.
    Solo se considera mes actual cuando existe un marcador explícito
    de mes/año guardado al procesar el archivo.
    Esto evita reutilizar acumulados del mes anterior después de F5,
    redeploy o restauración desde Supabase.
    """
    fecha_ref = fecha_ref or fecha_local_actual()

    mes = st.session_state.get("promesas_mes_operativo_v32")
    anio = st.session_state.get("promesas_anio_operativo_v32")

    try:
        return (
            int(mes) == int(fecha_ref.month)
            and int(anio) == int(fecha_ref.year)
        )
    except Exception:
        return False


def sanear_resultado_para_mes_actual_v32(resultado_df):
    """
    Para Mensajes:
    - Si Promesas no está confirmado para el mes actual, G/C = 0.
    - Recuperación se conserva porque puede corresponder al cierre anterior.
    """
    if resultado_df is None or resultado_df.empty:
        return resultado_df

    if promesas_es_mes_actual_v32():
        return resultado_df.copy()

    return limpiar_gestiones_compromisos_mes_anterior_v28(
        resultado_df.copy()
    )



def cargar_cierre_mes_anterior_v33(fecha_ref=None):
    """
    Busca en resultados_diarios el último cierre disponible ANTES
    del primer día del mes actual.
    Devuelve una fila por operador con Gestiones/Compromisos de cierre.
    """
    fecha_ref = fecha_ref or fecha_local_actual()
    sb = get_supabase()

    if sb is None:
        return None

    primer_dia = fecha_ref.replace(day=1)

    try:
        resp = (
            sb.table("resultados_diarios")
            .select("fecha,usuario,operador,gestiones,compromisos")
            .lt("fecha", primer_dia.isoformat())
            .order("fecha", desc=True)
            .limit(200)
            .execute()
        )

        hist = pd.DataFrame(resp.data or [])
        if hist.empty:
            return None

        hist["fecha"] = pd.to_datetime(
            hist["fecha"],
            errors="coerce",
        ).dt.date
        hist = hist.dropna(subset=["fecha"])

        if hist.empty:
            return None

        # Para evitar mezclar días, tomamos el último registro disponible
        # de cada operador antes del cambio de mes.
        hist = (
            hist.sort_values(
                ["usuario", "fecha"],
                ascending=[True, False],
                kind="stable",
            )
            .drop_duplicates(
                subset=["usuario"],
                keep="first",
            )
        )

        return hist

    except Exception:
        return None


def aplicar_corte_gestiones_compromisos_v33(
    resultado_df,
    fecha_ref=None,
):
    """
    Corrige el problema real del reporte Promesas cuando el sistema origen
    sigue mostrando acumulados del mes anterior al iniciar un mes nuevo.

    Regla:
    - Gestiones/Compromisos deben representar SOLO el mes actual.
    - Si el reporte ya reinició, se usa el valor tal cual.
    - Si el reporte continúa acumulado, se resta el último cierre del mes anterior.
    - Recuperación NO se toca; del 1 al 5 puede seguir cerrando el mes anterior.
    """
    if resultado_df is None or resultado_df.empty:
        return resultado_df

    fecha_ref = fecha_ref or fecha_local_actual()
    limpio = resultado_df.copy()

    cierre = cargar_cierre_mes_anterior_v33(fecha_ref)

    # Fallback de seguridad: en los primeros días un acumulado cercano o superior
    # a la meta mensual es claramente arrastre del mes anterior.
    if cierre is None or cierre.empty:
        if fecha_ref.day <= 5:
            meta_g = float(
                st.session_state.get(
                    "meta_gestiones_cfg",
                    META_GESTIONES,
                ) or META_GESTIONES
            )
            meta_c = float(
                st.session_state.get(
                    "meta_compromisos_cfg",
                    META_COMPROMISOS,
                ) or META_COMPROMISOS
            )

            for idx, fila in limpio.iterrows():
                g_raw = float(fila.get("Gestiones", 0) or 0)
                c_raw = float(fila.get("Compromisos", 0) or 0)

                if g_raw >= meta_g * 0.50:
                    limpio.at[idx, "Gestiones"] = 0
                    limpio.at[idx, "% Gestiones"] = 0.0

                if c_raw >= meta_c * 0.50:
                    limpio.at[idx, "Compromisos"] = 0
                    limpio.at[idx, "% Compromisos"] = 0.0

        return limpio

    cierre_idx = cierre.set_index(
        cierre["usuario"].astype(str)
    )

    meta_g = float(
        st.session_state.get(
            "meta_gestiones_cfg",
            META_GESTIONES,
        ) or META_GESTIONES
    )
    meta_c = float(
        st.session_state.get(
            "meta_compromisos_cfg",
            META_COMPROMISOS,
        ) or META_COMPROMISOS
    )

    for idx, fila in limpio.iterrows():
        usuario = str(fila.get("Usuario", ""))

        if usuario not in cierre_idx.index:
            continue

        base = cierre_idx.loc[usuario]
        if isinstance(base, pd.DataFrame):
            base = base.iloc[0]

        g_raw = float(fila.get("Gestiones", 0) or 0)
        c_raw = float(fila.get("Compromisos", 0) or 0)

        g_base = float(base.get("gestiones", 0) or 0)
        c_base = float(base.get("compromisos", 0) or 0)

        # Si el sistema origen ya reinició, raw < base y se conserva raw.
        # Si no reinició, calculamos solo lo generado desde el cambio de mes.
        g_mes = (
            max(g_raw - g_base, 0)
            if g_raw >= g_base and g_base > 0
            else max(g_raw, 0)
        )
        c_mes = (
            max(c_raw - c_base, 0)
            if c_raw >= c_base and c_base > 0
            else max(c_raw, 0)
        )

        limpio.at[idx, "Gestiones"] = int(round(g_mes))
        limpio.at[idx, "Compromisos"] = int(round(c_mes))
        limpio.at[idx, "% Gestiones"] = (
            g_mes / meta_g * 100 if meta_g else 0.0
        )
        limpio.at[idx, "% Compromisos"] = (
            c_mes / meta_c * 100 if meta_c else 0.0
        )

    return limpio



def snapshot_es_mes_actual_v28(snapshot, fecha_ref=None):
    """Valida si el snapshot fue actualizado dentro del mes calendario actual."""
    fecha_ref = fecha_ref or fecha_local_actual()
    ts = (snapshot or {}).get("actualizado_en")

    if not ts:
        return False

    try:
        fecha_snap = datetime_bolivia(ts).date()
    except Exception:
        try:
            fecha_snap = pd.to_datetime(ts, errors="coerce").date()
        except Exception:
            return False

    return (
        fecha_snap.year == fecha_ref.year
        and fecha_snap.month == fecha_ref.month
    )


def limpiar_gestiones_compromisos_mes_anterior_v28(resultado_df):
    """
    Conserva recuperación del cierre anterior, pero evita arrastrar
    Gestiones y Compromisos al nuevo mes.
    """
    if resultado_df is None or resultado_df.empty:
        return resultado_df

    limpio = resultado_df.copy()

    for col in ["Gestiones", "Compromisos", "Compromisos cumplidos"]:
        if col in limpio.columns:
            limpio[col] = 0

    for col in ["% Gestiones", "% Compromisos"]:
        if col in limpio.columns:
            limpio[col] = 0.0

    return limpio


def combinar_inicio_mes_v28(resultado_nuevo, resultado_cierre):
    """
    Días 1–5:
    - Gestiones/Compromisos: toma el reporte NUEVO del mes actual.
    - Recuperación: conserva el último cierre del mes anterior,
      salvo que el reporte nuevo ya traiga una recuperación mayor/actualizada.
    """
    if resultado_nuevo is None or resultado_nuevo.empty:
        return resultado_nuevo
    if resultado_cierre is None or resultado_cierre.empty:
        return resultado_nuevo

    nuevo = resultado_nuevo.copy()
    cierre = resultado_cierre.copy()

    if "Usuario" not in nuevo.columns or "Usuario" not in cierre.columns:
        return nuevo

    cierre_idx = cierre.set_index("Usuario", drop=False)

    cols_rec = [
        "Recuperación individual",
        "Recuperación acumulada",
        "% Recuperación",
    ]

    for i, fila in nuevo.iterrows():
        usuario = fila.get("Usuario")
        if usuario not in cierre_idx.index:
            continue

        anterior = cierre_idx.loc[usuario]
        if isinstance(anterior, pd.DataFrame):
            anterior = anterior.iloc[0]

        rec_nueva = float(fila.get("Recuperación acumulada", 0) or 0)
        rec_anterior = float(anterior.get("Recuperación acumulada", 0) or 0)

        # Durante el cierre se conserva el valor más actualizado disponible.
        if rec_anterior > rec_nueva:
            for col in cols_rec:
                if col in nuevo.columns and col in cierre.columns:
                    nuevo.at[i, col] = anterior.get(col, nuevo.at[i, col])

    return nuevo




def resumen_mes_actual_v29():
    """
    Construye el resumen mensual visible:
    - Gestiones y Compromisos: SOLO mes actual.
    - Recuperación: si estamos del 1 al 5, se muestra aparte como cierre anterior.
    """
    resultado = st.session_state.get("resultado_operadores")
    resultado = aplicar_corte_gestiones_compromisos_v33(
        resultado,
        fecha_local_actual(),
    )
    hoy = fecha_local_actual()
    periodo_rec = periodo_recuperacion_actual(hoy)

    if resultado is None or resultado.empty:
        total_g = 0
        total_c = 0
        total_r = 0.0
        prom_g = 0.0
        prom_c = 0.0
        prom_r = 0.0
    else:
        total_g = float(resultado["Gestiones"].sum()) if "Gestiones" in resultado.columns else 0.0
        total_c = float(resultado["Compromisos"].sum()) if "Compromisos" in resultado.columns else 0.0
        total_r = float(resultado["Recuperación acumulada"].sum()) if "Recuperación acumulada" in resultado.columns else 0.0

        prom_g = float(resultado["% Gestiones"].mean()) if "% Gestiones" in resultado.columns else 0.0
        prom_c = float(resultado["% Compromisos"].mean()) if "% Compromisos" in resultado.columns else 0.0
        prom_r = float(resultado["% Recuperación"].mean()) if "% Recuperación" in resultado.columns else 0.0

    return {
        "mes_actual": nombre_mes_es(hoy.month),
        "anio": hoy.year,
        "total_gestiones": total_g,
        "total_compromisos": total_c,
        "total_recuperacion": total_r,
        "promedio_gestiones": prom_g,
        "promedio_compromisos": prom_c,
        "promedio_recuperacion": prom_r,
        "periodo_recuperacion": periodo_rec,
    }



def obtener_fila_operador_actual(usuario):
    """
    Devuelve SIEMPRE la fila más reciente del operador desde
    st.session_state.resultado_operadores.

    Así Ranking, tarjetas, vista previa y Telegram usan exactamente
    la misma fuente y no pueden enviar cifras de una versión anterior.
    """
    resultado_actual = st.session_state.get(
        "resultado_operadores"
    )

    if (
        resultado_actual is None
        or resultado_actual.empty
    ):
        return None

    coincidencia = resultado_actual[
        resultado_actual["Usuario"]
        .astype(str)
        == str(usuario)
    ]

    if coincidencia.empty:
        return None

    return coincidencia.iloc[0].copy()



def enviar_resumen_grupo_actualizado(callcenter_df):
    """Envía al grupo el resumen general actualizado del equipo."""
    try:
        if callcenter_df is None or callcenter_df.empty:
            return False, "No hay reporte CallCenter cargado."

        chat_grupo = obtener_telegram_group_chat_id()
        if not chat_grupo:
            return False, "Falta TELEGRAM_GROUP_CHAT_ID."

        mensaje_grupo = generar_mensaje_resumen_gestiones_grupo(callcenter_df)
        imagen_grupo = generar_imagen_resumen_gestiones_grupo(callcenter_df)

        ok_txt, det_txt = enviar_mensaje_telegram(
            chat_grupo,
            mensaje_grupo,
        )
        if not ok_txt:
            return False, det_txt

        ok_img, det_img = enviar_foto_telegram(
            chat_grupo,
            imagen_grupo,
            "📊 Gestiones y compromisos por operador",
        )
        if not ok_img:
            return False, det_img

        return True, "Resumen general enviado al grupo."
    except Exception as e:
        return False, str(e)



def generar_mensaje_operador_actual(
    usuario,
    jornadas_info,
):
    # V34: el mensaje privado usa CallCenter para G/C del mes actual.
    # Promesas queda reservado para Recuperación y no bloquea este mensaje.
    fila_actual = obtener_fila_operador_actual(
        usuario
    )

    if fila_actual is None:
        return None

    # Protección final V33: aunque exista una sesión antigua,
    # el mensaje se calcula sobre el corte mensual real.
    fila_df_v33 = pd.DataFrame([fila_actual])
    fila_df_v33 = aplicar_corte_gestiones_compromisos_v33(
        fila_df_v33,
        fecha_local_actual(),
    )
    fila_actual = fila_df_v33.iloc[0].copy()

    # Devuelve el mensaje individual tal como lo genera GEN Control,
    # sin enlace ni invitación adicional para escribir a coordinación.
    return generar_mensaje_diario(
        fila_actual,
        jornadas_info,
    )


def generar_mensaje_diario(fila, jornadas_info):
    usuario = fila["Usuario"]
    datos = OPERADORES.get(usuario, {})
    nombre = datos.get(
        "nombre_mensaje",
        fila["Operador"].split()[0],
    )

    # V34 · Gestiones y Compromisos del mensaje salen del
    # CallCenter filtrado estrictamente al mes actual.
    acumulado_mes_v34 = calcular_acumulado_mes_callcenter_v34(
        usuario,
        st.session_state.get("callcenter_df"),
        fecha_local_actual(),
    )

    if acumulado_mes_v34["disponible"]:
        gestiones = int(
            acumulado_mes_v34["gestiones_mes"]
        )
        compromisos = int(
            acumulado_mes_v34["compromisos_mes"] or 0
        )
    else:
        # Nunca reutilizar cifras del mes anterior si no hay CallCenter actual.
        gestiones = 0
        compromisos = 0

    recuperacion = float(fila["Recuperación acumulada"])

    meta_g = int(st.session_state.meta_gestiones_cfg)
    meta_c = int(st.session_state.meta_compromisos_cfg)
    meta_r = float(st.session_state.meta_recuperacion_cfg)

    pct_g = gestiones / meta_g * 100 if meta_g else 0
    pct_c = compromisos / meta_c * 100 if meta_c else 0
    pct_r = recuperacion / meta_r * 100 if meta_r else 0

    esperado = jornadas_info["esperado_pct"]

    minimo_g = int(
        st.session_state.get(
            "meta_diaria_gestiones_cfg",
            META_DIARIA_GESTIONES,
        )
    )
    calculo_comp_msg_v35 = meta_diaria_compromisos_calendario_v35(
        fecha_local_actual(),
        compromisos,
    )
    minimo_c = int(calculo_comp_msg_v35["meta_diaria"])

    faltante_r = max(meta_r - recuperacion, 0)

    check_g = " ✅" if gestiones >= meta_g else ""
    check_c = " ✅" if compromisos >= meta_c else ""
    check_r = " ✅" if recuperacion >= meta_r else ""

    linea_g = (
        f"🔹 Gestiones: {formato_entero(gestiones)} / "
        f"{formato_entero(meta_g)} — {formato_porcentaje(pct_g)}{check_g}"
    )
    linea_c = (
        f"🔹 Compromisos: {formato_entero(compromisos)} / "
        f"{formato_entero(meta_c)} — {formato_porcentaje(pct_c)}{check_c}"
    )
    linea_r = (
        f"🔹 Recuperación: {formato_usd(recuperacion)} / "
        f"{formato_usd(meta_r)} — {formato_porcentaje(pct_r)}{check_r}"
    )

    if gestiones >= meta_g and compromisos >= meta_c:
        cierre = (
            "Mantengamos los mínimos diarios y reforcemos la recuperación "
            "para sostener el resultado."
        )
    elif pct_r < esperado - 10:
        cierre = (
            "Mantengamos el ritmo y reforcemos la recuperación "
            "para acercarnos a la meta mensual."
        )
    else:
        cierre = (
            "Mantengamos el ritmo diario para seguir avanzando "
            "hacia las metas del mes."
        )

    saludo_individual, emoji_individual = saludo_segun_hora()

    avance_hora = calcular_avance_hora_operador(
        usuario,
        st.session_state.callcenter_df,
    )

    bloque_hoy = ""

    if avance_hora["disponible"]:
        horario_msg = avance_hora.get(
            "horario"
        ) or {}

        linea_horario_msg = ""

        if horario_msg.get(
            "horario_configurado"
        ):
            linea_horario_msg = (
                f"🕒 Jornada: "
                f"{horario_msg['entrada']}–{horario_msg['salida']} "
                f"· Break {horario_msg['break_inicio']}–{horario_msg['break_fin']}\n"
                f"📍 Estado: {avance_hora.get('estado_jornada', '')}\n"
            )

        bloque_hoy = (
            f"\n\n⏱️ Avance de hoy · corte {avance_hora['hora_corte']}\n"
            f"{linea_horario_msg}"
            f"📞 Gestiones: {formato_entero(avance_hora['gestiones_hoy'])} / "
            f"{formato_entero(minimo_g)}\n"
            f"{texto_estado_ritmo(avance_hora['delta_gestiones'], 'gestiones')}\n"
            f"🎯 Faltan {formato_entero(avance_hora['faltan_gestiones'])} "
            f"para el mínimo diario"
        )

        if avance_hora.get(
            "compromisos_disponibles"
        ):
            bloque_hoy += (
                f"\n\n🤝 Compromisos: "
                f"{formato_entero(avance_hora['compromisos_hoy'])} / "
                f"{formato_entero(minimo_c)}\n"
                f"{texto_estado_ritmo(avance_hora['delta_compromisos'], 'compromisos')}\n"
                f"🎯 Faltan "
                f"{formato_entero(avance_hora['faltan_compromisos'])} "
                f"para el mínimo diario"
            )

    # V87 · La recuperación no se repite en cada seguimiento.
    # Se muestra en el primer seguimiento individual del día y vuelve
    # a aparecer desde las 17:00 como referencia de cierre.
    ahora_mensaje_v87 = ahora_bolivia()
    ya_recibio_seguimiento_v87 = envio_ya_realizado_hoy(
        usuario,
        "seguimiento",
    )
    # V19 · En seguimiento privado solo se muestran Gestiones y Compromisos.
    # Recuperación permanece disponible en Resumen/reportes, pero no se envía
    # en el mensaje individual.
    mostrar_recuperacion_v87 = False
    bloque_mes_v87 = (
        f"📊 Acumulado de {nombre_mes_es(fecha_local_actual().month)}\n"
        f"{linea_g}\n"
        f"{linea_c}"
    )

    # En los seguimientos intermedios el cierre se enfoca en lo que el
    # operador puede corregir durante la jornada: gestiones y compromisos.
    if (
        ya_recibio_seguimiento_v87
        and ahora_mensaje_v87.hour < 17
    ):
        if avance_hora.get("disponible"):
            cierre_v87 = (
                "Sigamos enfocados en el avance de hoy y en recuperar "
                "cualquier brecha antes del siguiente corte."
            )
        else:
            cierre_v87 = (
                "Mantengamos el ritmo diario para continuar avanzando "
                "hacia las metas del mes."
            )
    elif ahora_mensaje_v87.hour >= 17:
        cierre_v87 = (
            "Aprovechemos las horas restantes para cerrar la jornada "
            "con el mejor cumplimiento posible."
        )
    else:
        cierre_v87 = cierre

    mensaje = (
        f"{saludo_individual}, {nombre}. {emoji_individual}\n\n"
        f"{bloque_mes_v87}"
        f"{bloque_hoy}\n\n"
        f"{cierre_v87} 💪"
    )

    return {
        "mensaje": mensaje,
        "objetivo_gestiones": minimo_g,
        "objetivo_compromisos": minimo_c,
        "faltante_recuperacion": faltante_r,
        "avance_hora": avance_hora,
        "estado_gestiones": clasificar_avance(pct_g, esperado),
        "estado_compromisos": clasificar_avance(pct_c, esperado),
        "estado_recuperacion": clasificar_avance(pct_r, esperado),
    }


# =========================================================
# ORDENAMIENTO DE TABLAS
# =========================================================

def controles_ordenamiento(
    df,
    columnas,
    key_prefix,
    columna_default=None,
    descendente_default=True,
    etiquetas=None,
):
    """
    Ordena la tabla ANTES de formatear los valores.
    etiquetas permite mostrar nombres más claros al usuario.
    """
    columnas_validas = [
        c for c in columnas
        if c in df.columns
    ]

    if not columnas_validas:
        return df

    if columna_default not in columnas_validas:
        columna_default = columnas_validas[0]

    etiquetas = etiquetas or {
        c: c for c in columnas_validas
    }

    inverso = {
        etiquetas.get(c, c): c
        for c in columnas_validas
    }

    opciones_visibles = [
        etiquetas.get(c, c)
        for c in columnas_validas
    ]

    default_visible = etiquetas.get(
        columna_default,
        columna_default,
    )

    c1, c2 = st.columns([2, 1])

    with c1:
        visible_sel = st.selectbox(
            "Ordenar resultados por",
            opciones_visibles,
            index=opciones_visibles.index(
                default_visible
            ),
            key=f"{key_prefix}_columna",
        )

    with c2:
        sentido = st.radio(
            "Sentido",
            ["Mayor → menor", "Menor → mayor"],
            index=0 if descendente_default else 1,
            horizontal=True,
            key=f"{key_prefix}_sentido",
        )

    columna_orden = inverso[visible_sel]

    ordenado = df.sort_values(
        by=columna_orden,
        ascending=(sentido == "Menor → mayor"),
        na_position="last",
        kind="stable",
    ).reset_index(drop=True)

    return ordenado


# =========================================================
# LECTOR FLEXIBLE DE ARCHIVOS
# =========================================================

def leer_archivo(archivo):
    nombre = archivo.name.lower()
    contenido = archivo.getvalue()

    # CSV
    if nombre.endswith(".csv"):
        for sep in [";", ",", "\t"]:
            try:
                df = pd.read_csv(
                    BytesIO(contenido),
                    sep=sep,
                    encoding="utf-8",
                )
                if len(df.columns) > 1:
                    return df
            except Exception:
                pass

        try:
            return pd.read_csv(
                BytesIO(contenido),
                sep=";",
                encoding="latin1",
            )
        except Exception as e:
            raise ValueError(f"No se pudo leer el CSV: {e}")

    # XLSX
    if nombre.endswith(".xlsx"):
        try:
            return pd.read_excel(
                BytesIO(contenido),
                engine="openpyxl",
            )
        except Exception as e:
            raise ValueError(f"No se pudo leer el XLSX: {e}")

    # XLS que realmente viene como tabla HTML
    if nombre.endswith(".xls"):
        try:
            tablas = pd.read_html(BytesIO(contenido))

            if not tablas:
                raise ValueError("No se encontraron tablas.")

            # Tomamos la tabla con mayor cantidad de registros
            df = max(tablas, key=lambda x: len(x))

            return df

        except Exception:
            # Intento alternativo por si fuera un XLS real
            try:
                return pd.read_excel(BytesIO(contenido))
            except Exception as e:
                raise ValueError(
                    "No se pudo interpretar el archivo .xls. "
                    f"Detalle: {e}"
                )

    raise ValueError("Formato de archivo no reconocido.")


# =========================================================
# DETECCIÓN DEL TIPO DE REPORTE
# =========================================================

def detectar_tipo_reporte(df):
    columnas = [
        normalizar_columna(c)
        for c in df.columns
    ]

    texto_columnas = " | ".join(columnas)

    criterios_promesas = [
        "nombre usuario",
        "total gestion",
        "total compromisos",
        "compromisos cumplidos",
    ]

    coincidencias = sum(
        criterio in texto_columnas
        for criterio in criterios_promesas
    )

    if coincidencias >= 3:
        return "PROMESAS"

    criterios_callcenter = [
        "contrato",
        "cliente",
        "usuario",
    ]

    coincidencias_call = sum(
        criterio in texto_columnas
        for criterio in criterios_callcenter
    )

    if coincidencias_call >= 2 and len(df) > 100:
        return "CALLCENTER"

    return "DESCONOCIDO"


# =========================================================
# PROCESAMIENTO DEFINITIVO DE PROMESAS / RECUPERACIÓN
# =========================================================

def procesar_promesas(df):
    """
    REGLA DEFINITIVA:

    1. Tomar Compromisos Cumplidos en $ del registro SIN USUARIO.
    2. Dividir ese monto entre 8 operadores.
    3. Sumar esa parte al monto individual de cada operador.
    4. Recuperación ajustada / Bs 170.400 * 100.
    5. Resultado principal = porcentaje.
    """

    df = df.copy()

    # Limpiar nombres de columnas multinivel si existieran
    if isinstance(df.columns, pd.MultiIndex):
        df.columns = [
            " ".join(
                str(x)
                for x in col
                if str(x) != "nan"
            ).strip()
            for col in df.columns
        ]

    col_usuario = buscar_columna(
        df,
        ["nombre", "usuario"],
    )

    col_gestiones = buscar_columna(
        df,
        ["total", "gestion"],
    )

    col_compromisos = buscar_columna(
        df,
        ["total", "compromisos"],
    )

    col_cumplidos = buscar_columna(
        df,
        ["compromisos", "cumplidos"],
    )

    # Buscar específicamente columna monetaria de cumplidos.
    columnas_cumplidos = []

    for col in df.columns:
        norm = normalizar_columna(col)

        if (
            "compromisos" in norm
            and "cumplidos" in norm
        ):
            columnas_cumplidos.append(col)

    col_cumplidos_monto = None

    # Normalmente la última columna coincidente es la monetaria.
    if columnas_cumplidos:
        col_cumplidos_monto = columnas_cumplidos[-1]

    if col_usuario is None:
        raise ValueError(
            "No se encontró la columna 'Nombre Usuario'."
        )

    if col_cumplidos_monto is None:
        raise ValueError(
            "No se encontró la columna de "
            "'Compromisos Cumplidos en $'."
        )

    # Normalizar usuario
    df["_usuario_norm"] = (
        df[col_usuario]
        .astype(str)
        .apply(normalizar_texto)
    )

    # -----------------------------------------------------
    # DETECTAR FILA SIN USUARIO
    # -----------------------------------------------------

    mascara_sin_usuario = df["_usuario_norm"].apply(
        lambda x:
        "sin usuario" in x
        or x in ["sinusuario", "sin user"]
    )

    filas_sin_usuario = df[mascara_sin_usuario]

    if filas_sin_usuario.empty:
        monto_sin_usuario = 0.0
    else:
        monto_sin_usuario = filas_sin_usuario[
            col_cumplidos_monto
        ].apply(convertir_numero).sum()

    # REGLA FIJA: DIVIDIR ENTRE 8
    distribucion_por_operador = (
        monto_sin_usuario / CANTIDAD_OPERADORES
    )

    resultados = []

    # -----------------------------------------------------
    # LOS 8 OPERADORES OFICIALES
    # -----------------------------------------------------

    for usuario, datos in OPERADORES.items():

        usuario_norm = normalizar_texto(usuario)

        mascara = df["_usuario_norm"] == usuario_norm

        fila_operador = df[mascara]

        if fila_operador.empty:

            gestiones = 0
            compromisos = 0
            cumplidos = 0
            monto_individual = 0.0

        else:

            fila = fila_operador.iloc[0]

            gestiones = (
                convertir_numero(fila[col_gestiones])
                if col_gestiones
                else 0
            )

            compromisos = (
                convertir_numero(fila[col_compromisos])
                if col_compromisos
                else 0
            )

            cumplidos = (
                convertir_numero(fila[col_cumplidos])
                if col_cumplidos
                else 0
            )

            monto_individual = convertir_numero(
                fila[col_cumplidos_monto]
            )

        # =================================================
        # CÁLCULO OFICIAL DE RECUPERACIÓN
        # =================================================

        recuperacion_ajustada = (
            monto_individual
            + distribucion_por_operador
        )

        porcentaje_recuperacion = (
            recuperacion_ajustada
            / st.session_state.meta_recuperacion_cfg
            * 100
        )

        porcentaje_gestiones = (
            gestiones
            / st.session_state.meta_gestiones_cfg
            * 100
            if st.session_state.meta_gestiones_cfg
            else 0
        )

        porcentaje_compromisos = (
            compromisos
            / st.session_state.meta_compromisos_cfg
            * 100
            if st.session_state.meta_compromisos_cfg
            else 0
        )

        resultados.append(
            {
                "Usuario": usuario,
                "Operador": datos["nombre"],
                "Correo": datos["correo"],

                "Gestiones": int(round(gestiones)),
                "% Gestiones": porcentaje_gestiones,

                "Compromisos": int(round(compromisos)),
                "% Compromisos": porcentaje_compromisos,

                "Compromisos cumplidos": int(
                    round(cumplidos)
                ),

                "Recuperación original": monto_individual,

                "Distribución sin usuario":
                    distribucion_por_operador,

                "Recuperación acumulada":
                    recuperacion_ajustada,

                "% Recuperación":
                    porcentaje_recuperacion,
            }
        )

    resultado_df = pd.DataFrame(resultados)

    # V33: Gestiones y Compromisos deben ser siempre del mes actual.
    # Si Promesas aún arrastra el acumulado anterior, se resta el cierre
    # histórico del mes previo. Recuperación permanece intacta.
    resultado_df = aplicar_corte_gestiones_compromisos_v33(
        resultado_df,
        fecha_local_actual(),
    )

    return (
        resultado_df,
        monto_sin_usuario,
        distribucion_por_operador,
    )




def mostrar_boton_copiar_imagen(imagen_bytes):
    """
    Renderiza un botón que intenta copiar el PNG directamente
    al portapapeles del navegador para pegarlo con Ctrl+V.
    """
    if hasattr(imagen_bytes, "getvalue"):
        raw = imagen_bytes.getvalue()
    else:
        raw = imagen_bytes

    imagen_b64 = base64.b64encode(raw).decode("utf-8")

    html = f"""
    <div style="font-family:Arial,sans-serif;">
      <button
        id="copy-image-btn"
        style="
          width:100%;
          padding:0.62rem 1rem;
          border:1px solid #d0d5dd;
          border-radius:0.5rem;
          background:#ffffff;
          color:#101828;
          font-size:14px;
          font-weight:600;
          cursor:pointer;
        "
      >
        📋 Copiar imagen
      </button>

      <div
        id="copy-status"
        style="
          margin-top:6px;
          font-size:12px;
          min-height:18px;
          color:#475467;
        "
      ></div>
    </div>

    <script>
      const btn = document.getElementById("copy-image-btn");
      const status = document.getElementById("copy-status");

      btn.addEventListener("click", async () => {{
        try {{
          status.textContent = "Copiando imagen...";

          const response = await fetch(
            "data:image/png;base64,{imagen_b64}"
          );

          const blob = await response.blob();

          if (!navigator.clipboard || !window.ClipboardItem) {{
            throw new Error(
              "Tu navegador no permite copiar imágenes directamente."
            );
          }}

          await navigator.clipboard.write([
            new ClipboardItem({{
              "image/png": blob
            }})
          ]);

          status.textContent =
            "✅ Imagen copiada. Ya puedes pegarla con Ctrl + V.";
          status.style.color = "#067647";

        }} catch (error) {{
          console.error(error);
          status.textContent =
            "⚠️ El navegador bloqueó el portapapeles. Usa Descargar imagen como alternativa.";
          status.style.color = "#b54708";
        }}
      }});
    </script>
    """

    components.html(
        html,
        height=92,
        scrolling=False,
    )


# =========================================================
# IMAGEN EJECUTIVA DE AVANCE DE RECUPERACIÓN
# =========================================================

def generar_imagen_avance_recuperacion(
    tabla_general,
    fecha_reporte,
    meta_individual,
):
    """
    Genera una imagen PNG compacta y legible para pegar directamente
    en Outlook o WhatsApp.
    """
    tabla = tabla_general.copy()

    columnas_requeridas = [
        "Operador",
        "Recuperación acumulada",
        "% Recuperación",
    ]

    faltantes = [
        c for c in columnas_requeridas
        if c not in tabla.columns
    ]

    if faltantes:
        raise ValueError(
            "No se puede generar la imagen. Faltan columnas: "
            + ", ".join(faltantes)
        )

    tabla = tabla.sort_values(
        "% Recuperación",
        ascending=False,
        kind="stable",
    ).reset_index(drop=True)

    total_equipo = float(
        tabla["Recuperación acumulada"].sum()
    )

    meta_equipo = float(meta_individual) * len(tabla)

    pct_equipo = (
        total_equipo / meta_equipo * 100
        if meta_equipo
        else 0
    )

    falta_equipo = max(
        meta_equipo - total_equipo,
        0,
    )

    # Imagen pensada para correo: aprox. 820 px de ancho.
    fig, ax = plt.subplots(
        figsize=(7.4, 5.6),
        dpi=110,
    )
    ax.axis("off")

    titulo = (
        f"AVANCE DE RECUPERACIÓN · "
        f"{fecha_reporte.strftime('%d/%m/%Y')}"
    )

    fig.text(
        0.04,
        0.955,
        titulo,
        fontsize=15,
        fontweight="bold",
        ha="left",
        va="top",
    )

    fig.text(
        0.04,
        0.915,
        (
            f"Meta mensual por operador: "
            f"{formato_usd(meta_individual)}"
        ),
        fontsize=9.5,
        ha="left",
        va="top",
    )

    fig.text(
        0.04,
        0.875,
        (
            f"Equipo: {formato_usd(total_equipo)}   ·   "
            f"Cumplimiento: {formato_porcentaje(pct_equipo)}   ·   "
            f"Falta: {formato_usd(falta_equipo)}"
        ),
        fontsize=9.5,
        fontweight="bold",
        ha="left",
        va="top",
    )

    # Nombres más compactos para que nunca se corten.
    def nombre_corto(nombre):
        partes = str(nombre).split()

        if len(partes) <= 2:
            return str(nombre)

        # Nombre + último apellido.
        return f"{partes[0]} {partes[-1]}"

    filas = []

    for i, fila in tabla.iterrows():

        recuperacion = float(
            fila["Recuperación acumulada"]
        )

        porcentaje = float(
            fila["% Recuperación"]
        )

        falta = max(
            float(meta_individual) - recuperacion,
            0,
        )

        filas.append(
            [
                str(i + 1),
                nombre_corto(
                    fila["Operador"]
                ),
                formato_usd(
                    recuperacion
                ),
                formato_porcentaje(
                    porcentaje
                ),
                formato_usd(
                    falta
                ),
            ]
        )

    tabla_plot = ax.table(
        cellText=filas,
        colLabels=[
            "#",
            "Operador",
            "Recuperación",
            "Cumpl.",
            "Falta",
        ],
        cellLoc="left",
        colLoc="left",
        bbox=[
            0.035,
            0.08,
            0.93,
            0.70,
        ],
        colWidths=[
            0.055,
            0.24,
            0.245,
            0.16,
            0.245,
        ],
    )

    tabla_plot.auto_set_font_size(
        False
    )

    tabla_plot.set_fontsize(
        8.8
    )

    for (
        fila_idx,
        col_idx,
    ), celda in tabla_plot.get_celld().items():

        celda.set_linewidth(
            0.6
        )

        if fila_idx == 0:
            celda.get_text().set_fontweight(
                "bold"
            )

        if (
            fila_idx > 0
            and col_idx == 1
        ):
            celda.get_text().set_fontweight(
                "bold"
            )

    fig.text(
        0.04,
        0.025,
        (
            "GEN Control · Recuperación acumulada en USD · "
            "Distribución Sin usuario ÷ 8"
        ),
        fontsize=7.8,
        ha="left",
        va="bottom",
    )

    buffer = BytesIO()

    fig.savefig(
        buffer,
        format="png",
        dpi=110,
        bbox_inches="tight",
        pad_inches=0.12,
    )

    plt.close(
        fig
    )

    buffer.seek(
        0
    )

    return buffer




# =========================================================
# SUPABASE / PERSISTENCIA
# =========================================================

@st.cache_resource
def get_supabase():
    """
    Conexión robusta a Supabase.
    Acepta distintas formas de guardar los Secrets en Streamlit.
    """
    try:
        url = None
        key = None

        # Opción 1: claves en la raíz
        try:
            url = st.secrets.get("SUPABASE_URL")
            key = st.secrets.get("SUPABASE_KEY")
        except Exception:
            pass

        # Opción 2: sección [supabase]
        if not url or not key:
            try:
                supa = st.secrets.get("supabase", {})
                url = (
                    url
                    or supa.get("url")
                    or supa.get("SUPABASE_URL")
                )
                key = (
                    key
                    or supa.get("key")
                    or supa.get("SUPABASE_KEY")
                )
            except Exception:
                pass

        if not url:
            raise ValueError(
                "No se encontró SUPABASE_URL en Streamlit Secrets."
            )

        if not key:
            raise ValueError(
                "No se encontró SUPABASE_KEY en Streamlit Secrets."
            )

        cliente = create_client(
            str(url).strip(),
            str(key).strip(),
        )

        # Prueba real de conexión
        cliente.table("configuracion").select(
            "id"
        ).limit(1).execute()

        st.session_state["_supabase_error"] = ""
        return cliente

    except Exception as e:
        st.session_state["_supabase_error"] = str(e)
        return None


def supabase_disponible():
    return get_supabase() is not None


def diagnostico_supabase():
    """
    Devuelve un mensaje seguro; nunca muestra URL ni key.
    """
    if get_supabase() is not None:
        return True, "Supabase conectado correctamente."

    error = st.session_state.get(
        "_supabase_error",
        "No se pudo establecer conexión.",
    )

    return False, error


def cargar_configuracion_supabase():
    sb = get_supabase()

    if sb is None:
        return

    try:
        resp = (
            sb.table("configuracion")
            .select("*")
            .eq("id", 1)
            .execute()
        )

        if resp.data:
            fila = resp.data[0]

            st.session_state.meta_gestiones_cfg = int(
                fila.get(
                    "meta_gestiones",
                    st.session_state.meta_gestiones_cfg,
                )
            )

            st.session_state.meta_compromisos_cfg = int(
                fila.get(
                    "meta_compromisos",
                    st.session_state.meta_compromisos_cfg,
                )
            )

            st.session_state.meta_recuperacion_cfg = float(
                fila.get(
                    "meta_recuperacion",
                    st.session_state.meta_recuperacion_cfg,
                )
            )

            st.session_state.meta_diaria_compromisos_cfg = int(
                fila.get(
                    "meta_diaria_compromisos",
                    st.session_state.meta_diaria_compromisos_cfg,
                )
            )
    except Exception:
        pass


def guardar_configuracion_supabase():
    sb = get_supabase()

    if sb is None:
        return False, "Supabase no está conectado."

    payload = {
        "id": 1,
        "meta_gestiones": int(
            st.session_state.meta_gestiones_cfg
        ),
        "meta_compromisos": int(
            st.session_state.meta_compromisos_cfg
        ),
        "meta_recuperacion": float(
            st.session_state.meta_recuperacion_cfg
        ),
        "meta_diaria_compromisos": int(
            st.session_state.meta_diaria_compromisos_cfg
        ),
        "updated_at": ahora_bolivia().isoformat(),
    }

    try:
        (
            sb.table("configuracion")
            .upsert(payload)
            .execute()
        )
        return True, "Configuración guardada en Supabase."
    except Exception as e:
        return False, str(e)


def guardar_calendario_supabase(
    anio,
    mes,
    calendario_mes,
):
    sb = get_supabase()

    if sb is None:
        return False, "Supabase no está conectado."

    registros = []

    for dia, laboral in calendario_mes.items():
        fecha_dia = date(
            int(anio),
            int(mes),
            int(dia),
        )

        registros.append(
            {
                "fecha": fecha_dia.isoformat(),
                "es_laboral": bool(laboral),
            }
        )

    try:
        (
            sb.table("calendario_laboral")
            .upsert(
                registros,
                on_conflict="fecha",
            )
            .execute()
        )
        return True, "Calendario guardado."
    except Exception as e:
        return False, str(e)


def cargar_calendario_supabase(
    anio,
    mes,
):
    sb = get_supabase()

    if sb is None:
        return None

    inicio = date(
        int(anio),
        int(mes),
        1,
    )

    dias_mes = calendar.monthrange(
        int(anio),
        int(mes),
    )[1]

    fin = date(
        int(anio),
        int(mes),
        dias_mes,
    )

    try:
        resp = (
            sb.table("calendario_laboral")
            .select("fecha,es_laboral")
            .gte("fecha", inicio.isoformat())
            .lte("fecha", fin.isoformat())
            .execute()
        )

        if not resp.data:
            return None

        return {
            int(
                pd.to_datetime(
                    fila["fecha"]
                ).day
            ): bool(fila["es_laboral"])
            for fila in resp.data
        }
    except Exception:
        return None


def guardar_resultados_supabase(
    resultado_df,
    fecha_reporte,
    nombre_archivo,
):
    sb = get_supabase()

    if sb is None:
        return False, "Supabase no está conectado."

    registros = []

    for _, fila in resultado_df.iterrows():
        registros.append(
            {
                "fecha": fecha_reporte.isoformat(),
                "usuario": fila["Usuario"],
                "operador": fila["Operador"],
                "gestiones": int(
                    fila["Gestiones"]
                ),
                "compromisos": int(
                    fila["Compromisos"]
                ),
                "compromisos_cumplidos": int(
                    fila["Compromisos cumplidos"]
                ),
                "recuperacion_original": float(
                    fila["Recuperación original"]
                ),
                "distribucion_sin_usuario": float(
                    fila["Distribución sin usuario"]
                ),
                "recuperacion_acumulada": float(
                    fila["Recuperación acumulada"]
                ),
                "porcentaje_recuperacion": float(
                    fila["% Recuperación"]
                ),
                "archivo_origen": nombre_archivo,
            }
        )

    try:
        (
            sb.table("resultados_diarios")
            .upsert(
                registros,
                on_conflict="fecha,usuario",
            )
            .execute()
        )
        return True, "Resultados guardados en histórico."
    except Exception as e:
        return False, str(e)


def cargar_operadores_supabase():
    sb = get_supabase()

    if sb is None:
        return None

    try:
        resp = (
            sb.table("operadores")
            .select("*")
            .eq("activo", True)
            .order("nombre")
            .execute()
        )

        if not resp.data:
            return None

        return pd.DataFrame(resp.data)

    except Exception:
        return None


def guardar_operador_supabase(payload):
    sb = get_supabase()

    if sb is None:
        return False, "Supabase no está conectado."

    try:
        (
            sb.table("operadores")
            .upsert(
                payload,
                on_conflict="usuario",
            )
            .execute()
        )
        return True, "Operador guardado."
    except Exception as e:
        return False, str(e)


def sincronizar_operadores_base():
    """
    Carga inicial de los 8 operadores SIN sobrescribir datos ya guardados.
    Importante: no debe borrar teléfonos, correos editados ni nombres
    personalizados existentes en Supabase.
    """
    sb = get_supabase()

    if sb is None:
        return

    try:
        existentes_resp = (
            sb.table("operadores")
            .select("usuario")
            .execute()
        )

        usuarios_existentes = {
            str(fila["usuario"])
            for fila in (existentes_resp.data or [])
        }

        nuevos = []

        for usuario, datos in OPERADORES.items():
            if usuario in usuarios_existentes:
                continue

            nuevos.append(
                {
                    "usuario": usuario,
                    "nombre": datos["nombre"],
                    "nombre_mensaje": datos.get(
                        "nombre_mensaje",
                        datos["nombre"].split()[0],
                    ),
                    "correo": datos.get("correo", ""),
                    "telefono": datos.get("telefono", ""),
                    "telegram_chat_id": "",
                    "activo": True,
                }
            )

        if nuevos:
            (
                sb.table("operadores")
                .insert(nuevos)
                .execute()
            )

    except Exception:
        pass


def guardar_carga_supabase(
    nombre_archivo,
    tipo,
    registros,
):
    sb = get_supabase()

    if sb is None:
        return

    try:
        sb.table("cargas").insert(
            {
                "fecha_carga": datetime.now(
                    ZoneInfo("America/La_Paz")
                ).isoformat(),
                "nombre_archivo": nombre_archivo,
                "tipo_reporte": tipo,
                "registros": int(registros),
            }
        ).execute()
    except Exception:
        pass



# =========================================================
# PERSISTENCIA DEL ESTADO OPERATIVO — V93
# =========================================================

def guardar_estado_operativo_supabase(clave, nombre_archivo="", datos=None):
    sb = get_supabase()
    if sb is None:
        return False, "Supabase no está conectado."

    payload = {
        "clave": str(clave),
        "actualizado_en": ahora_bolivia().isoformat(),
        "nombre_archivo": str(nombre_archivo or ""),
        "datos": datos or {},
    }

    try:
        (
            sb.table("estado_operativo")
            .upsert(payload, on_conflict="clave")
            .execute()
        )
        return True, "Estado guardado."
    except Exception as e:
        return False, str(e)


def cargar_estado_operativo_supabase(clave):
    sb = get_supabase()
    if sb is None:
        return None

    try:
        resp = (
            sb.table("estado_operativo")
            .select("*")
            .eq("clave", str(clave))
            .limit(1)
            .execute()
        )
        if resp.data:
            return resp.data[0]
    except Exception:
        pass

    return None


def dataframe_a_json_v93(df):
    if df is None or df.empty:
        return []

    return json.loads(
        df.to_json(
            orient="records",
            date_format="iso",
            force_ascii=False,
        )
    )


def guardar_snapshot_promesas_v93(
    resultado_df,
    monto_sin_usuario,
    distribucion,
    nombre_archivo,
):
    return guardar_estado_operativo_supabase(
        "promesas_actual",
        nombre_archivo,
        {
            "resultado_operadores": dataframe_a_json_v93(
                resultado_df
            ),
            "monto_sin_usuario": float(
                monto_sin_usuario or 0.0
            ),
            "distribucion_sin_usuario": float(
                distribucion or 0.0
            ),
            "mes_operativo": int(fecha_local_actual().month),
            "anio_operativo": int(fecha_local_actual().year),
        },
    )


def guardar_snapshot_callcenter_v93(df, nombre_archivo):
    if df is None or df.empty:
        return False, "CallCenter vacío."

    # Solo se guardan columnas útiles para los cálculos de GEN Control.
    cols = []
    grupos = [
        ["fecha"],
        ["usuario"],
        ["compromiso"],
        ["tipo gestion", "tipo gestión"],
        ["contacto"],
        ["resultado"],
        ["monto($us)", "monto"],
    ]

    for candidatos in grupos:
        col = buscar_columna(df, candidatos)
        if col is not None and col not in cols:
            cols.append(col)

    compacto = df[cols].copy() if cols else df.copy()

    return guardar_estado_operativo_supabase(
        "callcenter_actual",
        nombre_archivo,
        {
            "filas": dataframe_a_json_v93(compacto),
            "registros": int(len(compacto)),
        },
    )


def restaurar_estado_operativo_v93():
    """
    Recupera automáticamente los últimos reportes procesados
    después de F5, cierre de pestaña, reinicio o nueva sesión.
    """
    if get_supabase() is None:
        return

    if st.session_state.get("resultado_operadores") is None:
        snap = cargar_estado_operativo_supabase(
            "promesas_actual"
        )
        if snap:
            datos = snap.get("datos") or {}
            filas = datos.get("resultado_operadores") or []

            if filas:
                resultado_restaurado_v28 = pd.DataFrame(
                    filas
                )

                hoy_v28 = fecha_local_actual()
                snapshot_mes_actual_v28 = snapshot_es_mes_actual_v28(
                    snap,
                    hoy_v28,
                )

                # Si cambió el mes, jamás se arrastran Gestiones ni
                # Compromisos del mes anterior.
                if not snapshot_mes_actual_v28:
                    if hoy_v28.day <= 5:
                        # Únicamente Recuperación permanece visible
                        # como cierre del mes anterior.
                        resultado_restaurado_v28 = (
                            limpiar_gestiones_compromisos_mes_anterior_v28(
                                resultado_restaurado_v28
                            )
                        )
                        st.session_state["cierre_recuperacion_anterior_v28"] = (
                            pd.DataFrame(filas)
                        )
                    else:
                        # Desde el día 6 tampoco debe permanecer el cierre anterior.
                        resultado_restaurado_v28 = (
                            limpiar_gestiones_compromisos_mes_anterior_v28(
                                resultado_restaurado_v28
                            )
                        )
                        for col_rec in [
                            "Recuperación individual",
                            "Recuperación acumulada",
                            "% Recuperación",
                        ]:
                            if col_rec in resultado_restaurado_v28.columns:
                                resultado_restaurado_v28[col_rec] = 0.0

                st.session_state.resultado_operadores = (
                    resultado_restaurado_v28
                )
                st.session_state.monto_sin_usuario = float(
                    datos.get("monto_sin_usuario", 0.0) or 0.0
                )
                st.session_state.distribucion_sin_usuario = float(
                    datos.get("distribucion_sin_usuario", 0.0) or 0.0
                )
                st.session_state.promesas_nombre_archivo = str(
                    snap.get("nombre_archivo") or ""
                )
                st.session_state.promesas_mes_operativo_v32 = (
                    datos.get("mes_operativo")
                )
                st.session_state.promesas_anio_operativo_v32 = (
                    datos.get("anio_operativo")
                )

                ts = snap.get("actualizado_en")
                if ts:
                    try:
                        st.session_state.promesas_cargado_en = (
                            datetime_bolivia(ts)
                        )
                    except Exception:
                        pass

    call = st.session_state.get("callcenter_df")
    if call is None or (
        hasattr(call, "empty") and call.empty
    ):
        snap = cargar_estado_operativo_supabase(
            "callcenter_actual"
        )
        if snap:
            datos = snap.get("datos") or {}
            filas = datos.get("filas") or []

            if filas:
                st.session_state.callcenter_df = pd.DataFrame(
                    filas
                )
                st.session_state.callcenter_nombre_archivo = str(
                    snap.get("nombre_archivo") or ""
                )

                ts = snap.get("actualizado_en")
                if ts:
                    try:
                        st.session_state.callcenter_cargado_en = (
                            datetime_bolivia(ts)
                        )
                    except Exception:
                        pass


# =========================================================
# SESSION STATE
# =========================================================

if "resultado_operadores" not in st.session_state:
    st.session_state.resultado_operadores = None

if "monto_sin_usuario" not in st.session_state:
    st.session_state.monto_sin_usuario = 0.0

if "distribucion_sin_usuario" not in st.session_state:
    st.session_state.distribucion_sin_usuario = 0.0

if "archivos_cargados" not in st.session_state:
    st.session_state.archivos_cargados = []

if "callcenter_df" not in st.session_state:
    st.session_state.callcenter_df = None

if "callcenter_cargado_en" not in st.session_state:
    st.session_state.callcenter_cargado_en = None

if "callcenter_nombre_archivo" not in st.session_state:
    st.session_state.callcenter_nombre_archivo = ""

if "promesas_cargado_en" not in st.session_state:
    st.session_state.promesas_cargado_en = None

if "promesas_nombre_archivo" not in st.session_state:
    st.session_state.promesas_nombre_archivo = ""

if "promesas_mes_operativo_v32" not in st.session_state:
    st.session_state.promesas_mes_operativo_v32 = None

if "promesas_anio_operativo_v32" not in st.session_state:
    st.session_state.promesas_anio_operativo_v32 = None

if "envios_diarios_cache" not in st.session_state:
    st.session_state.envios_diarios_cache = {}

if "permitir_envio_fuera_turno" not in st.session_state:
    st.session_state.permitir_envio_fuera_turno = False


if "config_desbloqueada" not in st.session_state:
    st.session_state.config_desbloqueada = False

if "meta_gestiones_cfg" not in st.session_state:
    st.session_state.meta_gestiones_cfg = META_GESTIONES

if "meta_compromisos_cfg" not in st.session_state:
    st.session_state.meta_compromisos_cfg = META_COMPROMISOS

if "meta_recuperacion_cfg" not in st.session_state:
    st.session_state.meta_recuperacion_cfg = META_RECUPERACION

if "meta_diaria_gestiones_cfg" not in st.session_state:
    st.session_state.meta_diaria_gestiones_cfg = META_DIARIA_GESTIONES

if "meta_diaria_compromisos_cfg" not in st.session_state:
    st.session_state.meta_diaria_compromisos_cfg = META_DIARIA_COMPROMISOS

if "calendario_laboral" not in st.session_state:
    st.session_state.calendario_laboral = {}

if "config_supabase_cargada" not in st.session_state:
    st.session_state.config_supabase_cargada = False

if "operadores_supabase_cargados" not in st.session_state:
    st.session_state.operadores_supabase_cargados = False

if "operador_guardado_ok" not in st.session_state:
    st.session_state.operador_guardado_ok = False

if not st.session_state.config_supabase_cargada:
    cargar_configuracion_supabase()
    st.session_state.config_supabase_cargada = True

if (
    supabase_disponible()
    and not st.session_state.operadores_supabase_cargados
):
    sincronizar_operadores_base()
    st.session_state.operadores_supabase_cargados = True

if "estado_operativo_restaurado_v93" not in st.session_state:
    st.session_state.estado_operativo_restaurado_v93 = False

if not st.session_state.estado_operativo_restaurado_v93:
    restaurar_estado_operativo_v93()

    # V33: sanear inmediatamente cualquier acumulado heredado,
    # incluso si el snapshot fue guardado nuevamente en septiembre.
    if (
        st.session_state.get("resultado_operadores") is not None
        and not st.session_state.resultado_operadores.empty
    ):
        st.session_state.resultado_operadores = (
            aplicar_corte_gestiones_compromisos_v33(
                st.session_state.resultado_operadores,
                fecha_local_actual(),
            )
        )

    st.session_state.estado_operativo_restaurado_v93 = True




# =========================================================
# TELEGRAM
# =========================================================

def obtener_telegram_group_chat_id():
    try:
        chat_id = st.secrets.get("TELEGRAM_GROUP_CHAT_ID")
        if chat_id:
            return str(chat_id).strip()

        telegram_cfg = st.secrets.get("telegram", {})
        chat_id = (
            telegram_cfg.get("group_chat_id")
            or telegram_cfg.get("TELEGRAM_GROUP_CHAT_ID")
        )
        return str(chat_id).strip() if chat_id else ""
    except Exception:
        return ""



def obtener_telegram_coordinador_chat_id():
    try:
        chat_id = st.secrets.get("TELEGRAM_COORDINADOR_CHAT_ID")
        if chat_id:
            return normalizar_telegram_chat_id(chat_id)

        telegram_cfg = st.secrets.get("telegram", {})
        chat_id = (
            telegram_cfg.get("coordinador_chat_id")
            or telegram_cfg.get("TELEGRAM_COORDINADOR_CHAT_ID")
        )
        return normalizar_telegram_chat_id(chat_id)
    except Exception:
        return ""


def obtener_telegram_bot_token():
    try:
        token = st.secrets.get("TELEGRAM_BOT_TOKEN")
        if token:
            return str(token).strip()

        telegram_cfg = st.secrets.get("telegram", {})
        token = (
            telegram_cfg.get("bot_token")
            or telegram_cfg.get("TELEGRAM_BOT_TOKEN")
        )
        return str(token).strip() if token else ""
    except Exception:
        return ""



def normalizar_telegram_chat_id(valor):
    """
    Devuelve un Chat ID válido o cadena vacía.

    Evita contar como configurados valores provenientes de Supabase
    como None, nan, null o cadenas vacías.
    Acepta IDs privados positivos y grupos negativos.
    """
    if valor is None:
        return ""

    try:
        if pd.isna(valor):
            return ""
    except Exception:
        pass

    texto = str(valor).strip()

    if texto.lower() in {
        "",
        "none",
        "nan",
        "null",
        "<na>",
    }:
        return ""

    # Telegram Chat ID debe ser numérico; grupos pueden llevar signo negativo.
    if not re.fullmatch(r"-?\d+", texto):
        return ""

    return texto


def enviar_mensaje_telegram(chat_id, texto):
    token = obtener_telegram_bot_token()

    if not token:
        return False, "Falta TELEGRAM_BOT_TOKEN en Streamlit Secrets."

    chat_id = str(chat_id or "").strip()

    if not chat_id:
        return False, "El operador no tiene Telegram Chat ID."

    url = f"https://api.telegram.org/bot{token}/sendMessage"

    payload = parse.urlencode(
        {
            "chat_id": chat_id,
            "text": str(texto),
            "disable_web_page_preview": "true",
        }
    ).encode("utf-8")

    try:
        req = request.Request(
            url,
            data=payload,
            method="POST",
            headers={"Content-Type": "application/x-www-form-urlencoded"},
        )

        with request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))

        if data.get("ok") is True:
            return True, "Mensaje enviado."

        return False, str(
            data.get("description", "Telegram rechazó el mensaje.")
        )
    except Exception as e:
        return False, str(e)



def _fuente_reporte(size, bold=False):
    """
    Fuente escalable REAL para las imágenes de Telegram.

    Usa matplotlib.font_manager para encontrar una fuente TrueType
    instalada en el entorno de Streamlit. Esto evita que PIL caiga en
    ImageFont.load_default(), que era la causa de que todo se viera
    diminuto aunque el código pidiera 40, 50 o 60 px.
    """
    try:
        propiedad = font_manager.FontProperties(
            family="DejaVu Sans",
            weight="bold" if bold else "normal",
        )

        ruta_fuente = font_manager.findfont(
            propiedad,
            fallback_to_default=True,
        )

        return ImageFont.truetype(
            ruta_fuente,
            size=int(size),
        )

    except Exception:
        # Segundo intento con la fuente incluida por matplotlib.
        try:
            ruta_fuente = font_manager.findfont(
                "DejaVu Sans",
                fallback_to_default=True,
            )
            return ImageFont.truetype(
                ruta_fuente,
                size=int(size),
            )
        except Exception:
            # Solo como último recurso.
            return ImageFont.load_default()



def preparar_resumen_gestiones_grupo(callcenter_df):
    """
    Resume el CallCenter del día por los 8 operadores:
    Contactado, Sin contacto, Total gestión, Compromisos y Monto comprometido.
    """
    if callcenter_df is None or callcenter_df.empty:
        return pd.DataFrame()

    df = callcenter_df.copy()
    col_fecha = buscar_columna(df, ["fecha"])
    col_usuario = buscar_columna(df, ["usuario"])
    col_contacto = buscar_columna(df, ["contacto"])
    col_compromiso = buscar_columna(df, ["compromiso"])
    col_monto = buscar_columna(df, ["monto($us)", "monto", "monto us", "monto($)"])

    if col_fecha is None or col_usuario is None:
        return pd.DataFrame()

    df["_fecha_dt"] = pd.to_datetime(df[col_fecha], dayfirst=True, errors="coerce")
    df = df.dropna(subset=["_fecha_dt"]).copy()

    hoy = fecha_local_actual()
    df = df[df["_fecha_dt"].dt.date == hoy].copy()

    df["_usuario_norm"] = df[col_usuario].astype(str).apply(normalizar_texto)
    df = df[df["_usuario_norm"].isin(OPERADORES.keys())].copy()

    if df.empty:
        return pd.DataFrame()

    if col_contacto:
        contacto = df[col_contacto].fillna("").astype(str).str.strip().str.lower()
        # Se consideran contacto los registros que no estén vacíos ni marcados explícitamente como no contacto.
        no_contacto_tokens = (
            contacto.eq("")
            | contacto.eq("nan")
            | contacto.str.contains("sin contacto", regex=False)
            | contacto.str.contains("no contact", regex=False)
            | contacto.str.contains("no contesta", regex=False)
        )
        df["_contactado"] = ~no_contacto_tokens
    else:
        df["_contactado"] = False

    df["_sin_contacto"] = ~df["_contactado"]

    if col_compromiso:
        compromiso = df[col_compromiso].fillna("").astype(str).str.strip()
        df["_compromiso"] = (
            compromiso.ne("")
            & compromiso.str.lower().ne("nan")
        )
    else:
        df["_compromiso"] = False

    if col_monto:
        monto_txt = (
            df[col_monto]
            .fillna(0)
            .astype(str)
            .str.replace(",", "", regex=False)
            .str.replace("$", "", regex=False)
            .str.replace("us", "", regex=False)
            .str.strip()
        )
        df["_monto"] = pd.to_numeric(monto_txt, errors="coerce").fillna(0.0)
    else:
        df["_monto"] = 0.0

    filas = []
    for usuario, datos in OPERADORES.items():
        op = df[df["_usuario_norm"] == usuario]
        filas.append({
            "Usuario": usuario,
            "Operador": datos.get("nombre_mensaje", datos.get("nombre", usuario)),
            "Contactado": int(op["_contactado"].sum()) if not op.empty else 0,
            "Sin contacto": int(op["_sin_contacto"].sum()) if not op.empty else 0,
            "Total gestión": int(len(op)),
            "Compromisos": int(op["_compromiso"].sum()) if not op.empty else 0,
            "Monto comprometido": float(op.loc[op["_compromiso"], "_monto"].sum()) if not op.empty else 0.0,
        })

    return pd.DataFrame(filas)


def generar_imagen_resumen_gestiones_grupo(callcenter_df):
    """Resumen visual del equipo sin monto comprometido."""
    import io
    import matplotlib.pyplot as plt
    import numpy as np
    from matplotlib.patches import FancyBboxPatch

    tabla = preparar_resumen_gestiones_grupo(callcenter_df)
    if tabla.empty:
        raise ValueError("No hay datos del día para generar el resumen.")

    nombres = tabla["Operador"].astype(str).tolist()
    x = np.arange(len(nombres))
    total_g = int(tabla["Total gestión"].sum())
    total_c = int(tabla["Compromisos"].sum())
    activos = int((tabla["Total gestión"] > 0).sum())

    corte = obtener_corte_callcenter(callcenter_df)
    corte_txt = corte.strftime("%H:%M") if corte is not None else ahora_bolivia().strftime("%H:%M")
    fecha_txt = fecha_local_actual().strftime("%d/%m/%Y")

    fig = plt.figure(figsize=(18, 10.5), dpi=150, facecolor="white")
    gs = fig.add_gridspec(4, 2, height_ratios=[0.85, 1.15, 4.8, 0.72], hspace=0.32, wspace=0.18)

    ax_header = fig.add_subplot(gs[0, :]); ax_header.axis("off")
    ax_header.text(0.5, 0.70, "GEN Control · Avance del Equipo", ha="center", va="center", fontsize=24, fontweight="bold")
    ax_header.text(0.5, 0.20, f"Corte: {fecha_txt} · {corte_txt}", ha="center", va="center", fontsize=13)

    ax_kpi = fig.add_subplot(gs[1, :]); ax_kpi.axis("off")
    kpis = [("GESTIONES TOTALES", formato_entero(total_g)), ("COMPROMISOS", formato_entero(total_c)), ("OPERADORES CON ACTIVIDAD", f"{activos} / {CANTIDAD_OPERADORES}")]
    card_w, gap = 0.295, 0.025
    x_start = (1 - (3*card_w + 2*gap))/2
    for i,(titulo,valor) in enumerate(kpis):
        x0=x_start+i*(card_w+gap)
        card=FancyBboxPatch((x0,0.10),card_w,0.78,boxstyle="round,pad=0.012,rounding_size=0.022",linewidth=1.0,edgecolor="#D7DEE8",facecolor="#FFFFFF",transform=ax_kpi.transAxes)
        ax_kpi.add_patch(card)
        ax_kpi.text(x0+0.025,0.64,titulo,fontsize=10.5,fontweight="bold",va="center",ha="left",transform=ax_kpi.transAxes)
        ax_kpi.text(x0+0.025,0.35,valor,fontsize=20,fontweight="bold",va="center",ha="left",transform=ax_kpi.transAxes)

    ax1=fig.add_subplot(gs[2,0]); width=.25
    b1=ax1.bar(x-width,tabla["Contactado"],width,label="Contactado")
    b2=ax1.bar(x,tabla["Sin contacto"],width,label="Sin contacto")
    b3=ax1.bar(x+width,tabla["Total gestión"],width,label="Total gestión")
    ax1.set_title("Total gestión por agente",fontsize=15,fontweight="bold",loc="left",pad=14)
    ax1.set_xticks(x); ax1.set_xticklabels(nombres,fontsize=9); ax1.grid(axis="y",alpha=.20); ax1.set_axisbelow(True)
    for bars in (b1,b2,b3): ax1.bar_label(bars,padding=3,fontsize=8.5)
    ax1.legend(loc="upper center",bbox_to_anchor=(.5,-.10),ncol=3,frameon=False,fontsize=9)

    ax2=fig.add_subplot(gs[2,1])
    c1=ax2.bar(x,tabla["Compromisos"],width=.52,label="Compromisos de pago")
    ax2.set_title("Compromisos de pago por agente",fontsize=15,fontweight="bold",loc="left",pad=14)
    ax2.set_xticks(x); ax2.set_xticklabels(nombres,fontsize=9); ax2.grid(axis="y",alpha=.20); ax2.set_axisbelow(True)
    ax2.bar_label(c1,padding=3,fontsize=9)
    ax2.legend(loc="upper center",bbox_to_anchor=(.5,-.10),ncol=1,frameon=False,fontsize=9)

    ax_footer=fig.add_subplot(gs[3,:]); ax_footer.axis("off")
    footer=FancyBboxPatch((.015,.14),.97,.68,boxstyle="round,pad=0.012,rounding_size=0.025",linewidth=0,facecolor="#EEF5FF",transform=ax_footer.transAxes)
    ax_footer.add_patch(footer)
    ax_footer.text(.035,.48,"Recordatorio:",fontsize=11.5,fontweight="bold",va="center",ha="left",transform=ax_footer.transAxes)
    ax_footer.text(.145,.48,"Mantengamos el ritmo de gestiones y la generación de compromisos para alcanzar las metas del día y del mes.",fontsize=11.5,va="center",ha="left",transform=ax_footer.transAxes)

    fig.tight_layout(rect=[.02,.02,.98,.98])
    out=io.BytesIO(); fig.savefig(out,format="png",bbox_inches="tight",facecolor="white"); plt.close(fig); out.seek(0)
    return out


def generar_mensaje_resumen_gestiones_grupo(callcenter_df):
    tabla = preparar_resumen_gestiones_grupo(callcenter_df)
    if tabla.empty:
        return "📊 No hay datos del día disponibles para generar el resumen."
    total_g = int(tabla["Total gestión"].sum())
    total_c = int(tabla["Compromisos"].sum())
    return (
        f"📊 AVANCE DE GESTIONES Y COMPROMISOS | {fecha_local_actual().strftime('%d/%m/%Y')}\n\n"
        f"📞 Gestiones del equipo: {formato_entero(total_g)}\n"
        f"🤝 Compromisos: {formato_entero(total_c)}\n\n"
        "Adjunto el detalle actualizado por operador.\n"
        "Sigamos avanzando con enfoque para cumplir las metas del día. 💪"
    )

def generar_imagen_recuperacion_telegram(
    tabla_general,
    meta_individual,
):
    """
    UNA sola imagen vertical para Telegram,
    inspirada en la referencia aprobada:
    - nombres grandes;
    - porcentaje grande a la derecha;
    - monto debajo del nombre;
    - barra de avance visible;
    - 8 operadores en una sola imagen;
    - diseño limpio y proporcional para celular.
    """
    import io

    tabla = tabla_general.copy().sort_values(
        "% Recuperación",
        ascending=False,
        kind="stable",
    ).reset_index(drop=True)

    # Formato vertical similar a 1080 x 1536.
    W = 1080
    H = 1540
    margin = 22

    navy = (7, 29, 52)
    green = (67, 166, 68)
    green_dark = (30, 137, 46)
    white = (255, 255, 255)
    bg = (247, 249, 251)
    dark = (18, 29, 42)
    border = (218, 224, 230)
    bar_bg = (224, 229, 234)
    gold = (242, 182, 24)
    silver = (172, 181, 191)
    bronze = (195, 108, 48)

    img = Image.new("RGB", (W, H), bg)
    d = ImageDraw.Draw(img)

    # Tipografías grandes, como en la referencia.
    f_title = _fuente_reporte(62, True)
    f_date = _fuente_reporte(34, True)

    f_name = _fuente_reporte(58, True)
    f_name_long = _fuente_reporte(52, True)
    f_name_xlong = _fuente_reporte(46, True)

    f_amount = _fuente_reporte(36, True)
    f_pct = _fuente_reporte(52, True)
    f_pos = _fuente_reporte(40, True)
    f_footer = _fuente_reporte(28, True)

    # -----------------------------------------------------
    # HEADER
    # -----------------------------------------------------
    header_y1 = margin
    header_y2 = 155

    d.rounded_rectangle(
        (
            margin,
            header_y1,
            W - margin,
            header_y2,
        ),
        radius=28,
        fill=navy,
    )

    d.text(
        (
            margin + 34,
            header_y1 + 27,
        ),
        "RANKING DE RECUPERACIÓN",
        font=f_title,
        fill=white,
    )

    d.text(
        (
            margin + 34,
            header_y1 + 91,
        ),
        fecha_local_actual().strftime("%d/%m/%Y"),
        font=f_date,
        fill=green,
    )

    # -----------------------------------------------------
    # TARJETAS
    # -----------------------------------------------------
    cards_top = 177
    card_h = 153
    gap = 10

    for i, fila in tabla.iterrows():
        y1 = cards_top + i * (card_h + gap)
        y2 = y1 + card_h

        d.rounded_rectangle(
            (
                margin,
                y1,
                W - margin,
                y2,
            ),
            radius=24,
            fill=white,
            outline=border,
            width=2,
        )

        # Puesto
        pos_color = (
            gold if i == 0
            else silver if i == 1
            else bronze if i == 2
            else navy
        )

        badge_x1 = margin + 18
        badge_y1 = y1 + 28
        badge_x2 = badge_x1 + 78
        badge_y2 = badge_y1 + 78

        d.rounded_rectangle(
            (
                badge_x1,
                badge_y1,
                badge_x2,
                badge_y2,
            ),
            radius=16,
            fill=pos_color,
        )

        pos_txt = str(i + 1)
        pos_box = d.textbbox(
            (0, 0),
            pos_txt,
            font=f_pos,
        )

        d.text(
            (
                (badge_x1 + badge_x2) / 2
                - (pos_box[2] - pos_box[0]) / 2,
                badge_y1 + 18,
            ),
            pos_txt,
            font=f_pos,
            fill=white,
        )

        nombre = str(
            fila["Operador"]
        ).strip()

        recuperacion = float(
            fila["Recuperación acumulada"]
        )

        porcentaje = float(
            fila["% Recuperación"]
        )

        # El nombre es el elemento principal.
        name_x = badge_x2 + 34

        # Porcentaje grande a la derecha.
        pct_txt = formato_porcentaje(
            porcentaje
        )
        pct_box = d.textbbox(
            (0, 0),
            pct_txt,
            font=f_pct,
        )
        pct_w = pct_box[2] - pct_box[0]

        # Reservar espacio al porcentaje.
        available_name_w = (
            W
            - margin
            - 30
            - pct_w
            - name_x
        )

        fuente_nombre = f_name

        name_box = d.textbbox(
            (0, 0),
            nombre,
            font=fuente_nombre,
        )

        if (
            name_box[2] - name_box[0]
            > available_name_w
        ):
            fuente_nombre = f_name_long
            name_box = d.textbbox(
                (0, 0),
                nombre,
                font=fuente_nombre,
            )

        if (
            name_box[2] - name_box[0]
            > available_name_w
        ):
            fuente_nombre = f_name_xlong

        d.text(
            (
                name_x,
                y1 + 23,
            ),
            nombre,
            font=fuente_nombre,
            fill=dark,
        )

        # Monto debajo del nombre.
        d.text(
            (
                name_x,
                y1 + 87,
            ),
            formato_usd(
                recuperacion
            ),
            font=f_amount,
            fill=green_dark,
        )

        # % grande, alineado a la derecha.
        d.text(
            (
                W
                - margin
                - 28
                - pct_w,
                y1 + 45,
            ),
            pct_txt,
            font=f_pct,
            fill=dark,
        )

        # Barra de progreso, gruesa y limpia.
        bar_x1 = name_x
        bar_x2 = W - margin - 28
        bar_y = y2 - 23

        d.rounded_rectangle(
            (
                bar_x1,
                bar_y,
                bar_x2,
                bar_y + 14,
            ),
            radius=7,
            fill=bar_bg,
        )

        fill_x = bar_x1 + int(
            (
                bar_x2 - bar_x1
            )
            * min(
                max(
                    porcentaje / 100,
                    0,
                ),
                1,
            )
        )

        if fill_x > bar_x1:
            d.rounded_rectangle(
                (
                    bar_x1,
                    bar_y,
                    fill_x,
                    bar_y + 14,
                ),
                radius=7,
                fill=green,
            )

    # -----------------------------------------------------
    # FOOTER
    # -----------------------------------------------------
    footer_y1 = 1490
    footer_y2 = H - margin

    d.rounded_rectangle(
        (
            margin,
            footer_y1,
            W - margin,
            footer_y2,
        ),
        radius=22,
        fill=navy,
    )

    footer_txt = "🏆 Ranking actualizado de recuperación"

    footer_box = d.textbbox(
        (0, 0),
        footer_txt,
        font=f_footer,
    )

    d.text(
        (
            (W - (footer_box[2] - footer_box[0])) / 2,
            footer_y1 + 16,
        ),
        footer_txt,
        font=f_footer,
        fill=white,
    )

    buffer = io.BytesIO()

    img.save(
        buffer,
        format="PNG",
        optimize=True,
    )

    buffer.seek(0)

    return buffer


def enviar_foto_telegram(chat_id, imagen_bytes, caption=""):
    """
    Envía una imagen PNG directamente a Telegram mediante sendPhoto.
    """
    token = obtener_telegram_bot_token()
    chat_id = str(chat_id or "").strip()

    if not token:
        return False, "Falta TELEGRAM_BOT_TOKEN en Streamlit Secrets."
    if not chat_id:
        return False, "Falta el Chat ID de Telegram."

    import uuid

    boundary = "----GENControl" + uuid.uuid4().hex
    partes = []

    def campo(nombre, valor):
        partes.append(f"--{boundary}\r\n".encode())
        partes.append(
            f'Content-Disposition: form-data; name="{nombre}"\r\n\r\n'.encode()
        )
        partes.append(str(valor).encode("utf-8"))
        partes.append(b"\r\n")

    campo("chat_id", chat_id)
    if caption:
        campo("caption", caption)

    partes.append(f"--{boundary}\r\n".encode())
    partes.append(
        b'Content-Disposition: form-data; name="photo"; filename="avance_recuperacion.png"\r\n'
    )
    partes.append(b"Content-Type: image/png\r\n\r\n")
    partes.append(imagen_bytes.getvalue())
    partes.append(b"\r\n")
    partes.append(f"--{boundary}--\r\n".encode())

    body = b"".join(partes)
    url = f"https://api.telegram.org/bot{token}/sendPhoto"

    try:
        req = request.Request(
            url,
            data=body,
            method="POST",
            headers={
                "Content-Type": f"multipart/form-data; boundary={boundary}",
                "Content-Length": str(len(body)),
            },
        )
        with request.urlopen(req, timeout=30) as resp:
            data = json.loads(resp.read().decode("utf-8"))

        if data.get("ok") is True:
            return True, "Imagen enviada."
        return False, str(data.get("description", "Telegram rechazó la imagen."))
    except Exception as e:
        return False, str(e)

def saludo_segun_hora():
    hora = ahora_bolivia().hour

    if 5 <= hora < 12:
        return "Buenos días", "☀️"
    elif 12 <= hora < 19:
        return "Buenas tardes", "🌤️"
    else:
        return "Buenas noches", "🌙"


def generar_mensaje_grupo_recuperacion(
    tabla_general,
    meta_individual,
):
    """
    Texto breve que acompaña la imagen del ranking en Telegram.
    El saludo cambia automáticamente según la hora de Bolivia.
    """
    tabla = tabla_general.copy()

    total_rec = float(
        tabla["Recuperación acumulada"].sum()
    )

    meta_equipo = (
        float(meta_individual)
        * len(tabla)
    )

    pct_equipo = (
        total_rec / meta_equipo * 100
        if meta_equipo
        else 0
    )

    falta_equipo = max(
        meta_equipo - total_rec,
        0,
    )

    saludo, emoji_saludo = saludo_segun_hora()

    return (
        f"{emoji_saludo} {saludo}, equipo. 💪\n\n"
        f"📊 AVANCE DE RECUPERACIÓN | "
        f"{fecha_local_actual().strftime('%d/%m/%Y')}\n\n"
        f"💰 Recuperado: {formato_usd(total_rec)}\n"
        f"🎯 Meta equipo: {formato_usd(meta_equipo)}\n"
        f"📈 Cumplimiento: {formato_porcentaje(pct_equipo)}\n"
        f"📉 Brecha pendiente: {formato_usd(falta_equipo)}\n\n"
        f"Adjunto el ranking actualizado de recuperación por operador.\n\n"
        f"Sigamos con enfoque y constancia para alcanzar la meta del mes. 🚀"
    )



def obtener_usuarios_telegram_bot():
    """
    Lee getUpdates del bot y devuelve los chats privados que le escribieron.
    No envía mensajes.
    """
    token = obtener_telegram_bot_token()

    if not token:
        return [], "Falta TELEGRAM_BOT_TOKEN en Secrets."

    try:
        url = f"https://api.telegram.org/bot{token}/getUpdates"

        with request.urlopen(url, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))

        if data.get("ok") is not True:
            return [], str(
                data.get(
                    "description",
                    "Telegram no devolvió los usuarios.",
                )
            )

        encontrados = {}

        for update in data.get("result", []):
            mensaje = (
                update.get("message")
                or update.get("edited_message")
                or update.get("channel_post")
                or {}
            )

            chat = mensaje.get("chat", {}) or {}
            remitente = mensaje.get("from", {}) or {}

            chat_id = chat.get("id")

            # Solo chats privados para operadores.
            if not chat_id or chat.get("type") != "private":
                continue

            nombre = " ".join(
                [
                    str(remitente.get("first_name") or "").strip(),
                    str(remitente.get("last_name") or "").strip(),
                ]
            ).strip()

            username = str(
                remitente.get("username") or ""
            ).strip()

            encontrados[str(chat_id)] = {
                "chat_id": str(chat_id),
                "nombre": nombre or "Sin nombre",
                "username": (
                    f"@{username}"
                    if username
                    else "Sin username"
                ),
            }

        usuarios = list(encontrados.values())

        usuarios.sort(
            key=lambda x: x["nombre"].lower()
        )

        if not usuarios:
            return [], (
                "Todavía no hay chats privados disponibles. "
                "Cada operador debe abrir el bot y enviar /start."
            )

        return usuarios, ""

    except Exception as e:
        return [], str(e)


def probar_conexion_telegram():
    token = obtener_telegram_bot_token()

    if not token:
        return False, "Falta TELEGRAM_BOT_TOKEN en Secrets."

    try:
        url = f"https://api.telegram.org/bot{token}/getMe"

        with request.urlopen(url, timeout=10) as resp:
            data = json.loads(resp.read().decode("utf-8"))

        if data.get("ok") is True:
            username = data.get("result", {}).get("username", "bot")
            return True, f"Conectado con @{username}"

        return False, str(
            data.get("description", "No se pudo validar el bot.")
        )
    except Exception as e:
        return False, str(e)



# =========================================================
# HORA DE CORTE DEL ARCHIVO CALLCENTER
# =========================================================

def obtener_corte_callcenter(callcenter_df):
    """
    Hora de corte operativa del reporte.

    Prioridad:
    1) momento en que GEN CallCenter fue cargado a la app;
    2) última fecha/hora contenida en el archivo, como respaldo.

    Esto evita que el esperado siga avanzando después de cargar
    un archivo que ya quedó congelado.
    """
    cargado_en = st.session_state.get(
        "callcenter_cargado_en"
    )

    if cargado_en is not None:
        return (
            datetime_bolivia(cargado_en)
            or cargado_en
        )

    if callcenter_df is None or callcenter_df.empty:
        return None

    df = callcenter_df.copy()
    col_fecha = buscar_columna(df, ["fecha"])

    if col_fecha is None:
        return None

    serie = pd.to_datetime(
        df[col_fecha],
        dayfirst=True,
        errors="coerce",
    ).dropna()

    if serie.empty:
        return None

    return serie.max()


# =========================================================
# AVANCE DE HOY / A LA HORA
# =========================================================

def calcular_acumulado_mes_callcenter_v34(
    usuario,
    callcenter_df=None,
    fecha_ref=None,
):
    """
    Fuente definitiva para Gestiones y Compromisos de los mensajes:
    GEN CallCenter filtrado por MES/AÑO ACTUAL.

    Esto evita por completo que Promesas arrastre acumulados del mes anterior.
    """
    fecha_ref = fecha_ref or fecha_local_actual()

    base = {
        "disponible": False,
        "gestiones_mes": 0,
        "compromisos_mes": None,
        "compromisos_disponibles": False,
        "mes": fecha_ref.month,
        "anio": fecha_ref.year,
    }

    if callcenter_df is None or callcenter_df.empty:
        return base

    df = callcenter_df.copy()

    col_fecha = buscar_columna(df, ["fecha"])
    col_usuario = buscar_columna(df, ["usuario"])
    col_compromiso = buscar_columna(df, ["compromiso"])

    if col_fecha is None or col_usuario is None:
        return base

    df["_fecha_hora_v34"] = pd.to_datetime(
        df[col_fecha],
        dayfirst=True,
        errors="coerce",
    )
    df = df.dropna(subset=["_fecha_hora_v34"])

    # SOLO mes y año actuales.
    df = df[
        (df["_fecha_hora_v34"].dt.year == int(fecha_ref.year))
        & (df["_fecha_hora_v34"].dt.month == int(fecha_ref.month))
    ].copy()

    if df.empty:
        return base

    datos_op = OPERADORES.get(usuario, {})

    aliases = {
        normalizar_texto(usuario),
        normalizar_texto(datos_op.get("nombre", "")),
        normalizar_texto(datos_op.get("nombre_mensaje", "")),
        normalizar_texto(datos_op.get("correo", "")),
    }

    correo_op = str(datos_op.get("correo", "")).strip()
    if "@" in correo_op:
        aliases.add(
            normalizar_texto(
                correo_op.split("@")[0]
            )
        )

    aliases = {x for x in aliases if x}

    df["_usuario_norm_v34"] = (
        df[col_usuario]
        .astype(str)
        .apply(normalizar_texto)
    )

    mascara_usuario = df["_usuario_norm_v34"].isin(aliases)

    if not mascara_usuario.any():
        aliases_largos = [x for x in aliases if len(x) >= 4]
        mascara_usuario = df["_usuario_norm_v34"].apply(
            lambda valor: any(
                alias in valor or valor in alias
                for alias in aliases_largos
            )
        )

    df_op = df[mascara_usuario].copy()

    if df_op.empty:
        return {
            **base,
            "disponible": True,
            "gestiones_mes": 0,
            "compromisos_mes": 0 if col_compromiso is not None else None,
            "compromisos_disponibles": col_compromiso is not None,
        }

    gestiones_mes = int(len(df_op))

    compromisos_disponibles = col_compromiso is not None
    compromisos_mes = None

    if compromisos_disponibles:
        compromiso_txt = (
            df_op[col_compromiso]
            .astype(str)
            .str.strip()
        )

        compromisos_mes = int(
            (
                df_op[col_compromiso].notna()
                & (compromiso_txt != "")
                & (compromiso_txt.str.lower() != "nan")
            ).sum()
        )

    return {
        "disponible": True,
        "gestiones_mes": gestiones_mes,
        "compromisos_mes": compromisos_mes,
        "compromisos_disponibles": compromisos_disponibles,
        "mes": fecha_ref.month,
        "anio": fecha_ref.year,
    }



def calcular_avance_hora_operador(
    usuario,
    callcenter_df=None,
    ahora=None,
):
    meta_g_dia = int(
        st.session_state.get(
            "meta_diaria_gestiones_cfg",
            META_DIARIA_GESTIONES,
        )
    )
    meta_c_dia = int(
        st.session_state.meta_diaria_compromisos_cfg
    )

    corte_operativo = obtener_corte_callcenter(
        callcenter_df
    )

    base = {
        "disponible": False,
        "gestiones_hoy": 0,
        "compromisos_hoy": None,
        "compromisos_disponibles": False,
        "esperado_gestiones": 0,
        "esperado_compromisos": 0,
        "delta_gestiones": 0,
        "delta_compromisos": None,
        "faltan_gestiones": meta_g_dia,
        "faltan_compromisos": None,
        "hora_corte": "--:--",
        "fecha_corte": None,
        "progreso_jornada_pct": 0.0,
        "estado_jornada": "Sin corte",
        "horario": None,
    }

    if (
        corte_operativo is None
        or callcenter_df is None
        or callcenter_df.empty
    ):
        return base

    if hasattr(
        corte_operativo,
        "to_pydatetime",
    ):
        corte_operativo = (
            corte_operativo.to_pydatetime()
        )

    # Asegurar tz Bolivia.
    if corte_operativo.tzinfo is None:
        corte_operativo = (
            corte_operativo.replace(
                tzinfo=ZoneInfo(
                    "America/La_Paz"
                )
            )
        )

    df = callcenter_df.copy()

    col_fecha = buscar_columna(
        df,
        ["fecha"],
    )
    col_usuario = buscar_columna(
        df,
        ["usuario"],
    )
    col_compromiso = buscar_columna(
        df,
        ["compromiso"],
    )

    if col_fecha is None or col_usuario is None:
        return base

    df["_fecha_hora"] = pd.to_datetime(
        df[col_fecha],
        dayfirst=True,
        errors="coerce",
    )

    df = df.dropna(
        subset=["_fecha_hora"]
    )

    fechas_validas = (
        df["_fecha_hora"]
        .dt.date
        .dropna()
    )

    if fechas_validas.empty:
        return base

    fecha_corte = max(
        fechas_validas
    )

    datos_op = OPERADORES.get(
        usuario,
        {},
    )

    aliases = {
        normalizar_texto(usuario),
        normalizar_texto(
            datos_op.get("nombre", "")
        ),
        normalizar_texto(
            datos_op.get(
                "nombre_mensaje",
                "",
            )
        ),
        normalizar_texto(
            datos_op.get("correo", "")
        ),
    }

    correo_op = str(
        datos_op.get("correo", "")
    ).strip()

    if "@" in correo_op:
        aliases.add(
            normalizar_texto(
                correo_op.split("@")[0]
            )
        )

    aliases = {
        x
        for x in aliases
        if x
    }

    df["_usuario_norm"] = (
        df[col_usuario]
        .astype(str)
        .apply(normalizar_texto)
    )

    mascara_usuario = (
        df["_usuario_norm"].isin(
            aliases
        )
    )

    if not mascara_usuario.any():
        aliases_largos = [
            x
            for x in aliases
            if len(x) >= 4
        ]

        mascara_usuario = (
            df["_usuario_norm"]
            .apply(
                lambda valor: any(
                    alias in valor
                    or valor in alias
                    for alias in aliases_largos
                )
            )
        )

    df_hoy = df[
        mascara_usuario
        & (
            df["_fecha_hora"].dt.date
            == fecha_corte
        )
    ].copy()

    gestiones_hoy = int(
        len(df_hoy)
    )

    compromisos_disponibles = (
        col_compromiso is not None
    )

    compromisos_hoy = None

    if compromisos_disponibles:
        compromiso_txt = (
            df_hoy[col_compromiso]
            .astype(str)
            .str.strip()
        )

        compromisos_hoy = int(
            (
                df_hoy[
                    col_compromiso
                ].notna()
                & (
                    compromiso_txt
                    != ""
                )
                & (
                    compromiso_txt
                    .str.lower()
                    != "nan"
                )
            ).sum()
        )

    # La fecha viene del archivo y la hora del momento de carga/corte.
    corte_para_ritmo = (
        corte_operativo.replace(
            year=fecha_corte.year,
            month=fecha_corte.month,
            day=fecha_corte.day,
        )
    )

    jornada = (
        calcular_progreso_jornada_operador(
            usuario,
            corte_para_ritmo,
        )
    )

    proporcion = float(
        jornada.get(
            "proporcion",
            0,
        )
    )

    esperado_g = int(
        math.ceil(
            meta_g_dia
            * proporcion
        )
    )

    esperado_c = int(
        math.ceil(
            meta_c_dia
            * proporcion
        )
    )

    delta_c = (
        compromisos_hoy
        - esperado_c
        if compromisos_disponibles
        else None
    )

    faltan_c = (
        max(
            meta_c_dia
            - compromisos_hoy,
            0,
        )
        if compromisos_disponibles
        else None
    )

    return {
        "disponible": True,
        "gestiones_hoy": gestiones_hoy,
        "compromisos_hoy": compromisos_hoy,
        "compromisos_disponibles": compromisos_disponibles,
        "esperado_gestiones": esperado_g,
        "esperado_compromisos": esperado_c,
        "delta_gestiones": (
            gestiones_hoy
            - esperado_g
        ),
        "delta_compromisos": delta_c,
        "faltan_gestiones": max(
            meta_g_dia
            - gestiones_hoy,
            0,
        ),
        "faltan_compromisos": faltan_c,
        "hora_corte": (
            corte_para_ritmo
            .strftime("%H:%M")
        ),
        "fecha_corte": fecha_corte,
        "progreso_jornada_pct": (
            proporcion * 100
        ),
        "estado_jornada": jornada.get(
            "estado_jornada",
            "",
        ),
        "horario": jornada,
    }


def texto_estado_ritmo(delta, indicador):
    if delta > 0:
        return (
            f"🟢 Vas {formato_entero(delta)} {indicador} "
            f"por encima del ritmo esperado"
        )
    if delta == 0:
        return "🟢 Vas exactamente en el ritmo esperado"
    return (
        f"🔴 Vas {formato_entero(abs(delta))} {indicador} "
        f"por debajo del ritmo esperado"
    )



# =========================================================
# CONTROL DE ENVÍOS DIARIOS
# =========================================================

def _clave_envio_local(usuario, fecha=None, tipo="seguimiento"):
    fecha = fecha or fecha_local_actual()
    return f"{fecha.isoformat()}|{usuario}|{tipo}"


def obtener_envio_diario(usuario, fecha=None, tipo="seguimiento"):
    """
    V96: Supabase es la fuente principal del último envío.
    Session State queda únicamente como respaldo si Supabase no responde.
    """
    fecha = fecha or fecha_local_actual()
    clave = _clave_envio_local(
        usuario,
        fecha,
        tipo,
    )

    sb = get_supabase()

    if sb is not None:
        try:
            resp = (
                sb.table("envios_mensajes")
                .select("*")
                .eq("fecha", fecha.isoformat())
                .eq("usuario", str(usuario))
                .eq("tipo", str(tipo))
                .eq("estado", "enviado")
                .order(
                    "fecha_hora",
                    desc=True,
                )
                .limit(1)
                .execute()
            )

            if resp.data:
                registro = resp.data[0]
                st.session_state.envios_diarios_cache[
                    clave
                ] = registro
                return registro
        except Exception:
            pass

    cache = st.session_state.get(
        "envios_diarios_cache",
        {},
    )
    return cache.get(clave)


def registrar_envio_diario(
    usuario,
    operador,
    canal="telegram",
    tipo="seguimiento",
    detalle="",
):
    """
    Registra un envío exitoso. El cambio de fecha reinicia
    automáticamente el control porque la clave incluye la fecha.
    """
    ahora_local = ahora_bolivia()
    ahora_utc = datetime.now(
        timezone.utc
    )

    registro = {
        "fecha": ahora_local.date().isoformat(),
        "fecha_hora": ahora_utc.isoformat(),
        "usuario": str(usuario),
        "operador": str(operador),
        "canal": str(canal),
        "tipo": str(tipo),
        "estado": "enviado",
        "detalle": str(detalle or ""),
    }

    clave = _clave_envio_local(
        usuario,
        ahora_local.date(),
        tipo,
    )

    st.session_state.envios_diarios_cache[
        clave
    ] = registro

    sb = get_supabase()

    if sb is not None:
        try:
            (
                sb.table("envios_mensajes")
                .insert(registro)
                .execute()
            )
        except Exception:
            # La app sigue funcionando con Session State.
            pass

    return registro


def envio_ya_realizado_hoy(
    usuario,
    tipo="seguimiento",
):
    return obtener_envio_diario(
        usuario,
        fecha_local_actual(),
        tipo,
    ) is not None


def hora_envio_hoy(
    usuario,
    tipo="seguimiento",
):
    registro = obtener_envio_diario(
        usuario,
        fecha_local_actual(),
        tipo,
    )

    if not registro:
        return ""

    fecha_hora = str(
        registro.get(
            "fecha_hora",
            "",
        )
    )

    try:
        dt = datetime_bolivia(
            fecha_hora
        )
        return (
            dt.strftime("%H:%M")
            if dt is not None
            else ""
        )
    except Exception:
        return ""




CORTES_SEGUIMIENTO = ["10:30", "13:30", "15:30", "17:30"]
MINUTOS_RECOMENDADOS_ENTRE_SEGUIMIENTOS = 60


def ultimo_envio_datetime(
    usuario,
    tipo="seguimiento",
):
    registro = obtener_envio_diario(
        usuario,
        fecha_local_actual(),
        tipo,
    )

    if not registro:
        return None

    valor = str(
        registro.get(
            "fecha_hora",
            "",
        )
    ).strip()

    if not valor:
        return None

    try:
        return datetime_bolivia(
            valor
        )
    except Exception:
        return None


def minutos_desde_ultimo_envio(
    usuario,
    momento=None,
    tipo="seguimiento",
):
    momento = momento or ahora_bolivia()

    ultimo = ultimo_envio_datetime(
        usuario,
        tipo,
    )

    if ultimo is None:
        return None

    minutos = int(
        (momento - ultimo).total_seconds()
        // 60
    )

    return max(
        minutos,
        0,
    )


def seguimiento_muy_reciente(
    usuario,
    momento=None,
    minimo_minutos=MINUTOS_RECOMENDADOS_ENTRE_SEGUIMIENTOS,
):
    minutos = minutos_desde_ultimo_envio(
        usuario,
        momento,
    )

    return (
        minutos is not None
        and minutos < minimo_minutos
    )


def informacion_corte_recomendado(
    momento=None,
):
    """
    Devuelve el corte recomendado más útil para la hora actual.
    Los cortes son orientativos; nunca bloquean un envío manual.
    """
    momento = momento or ahora_bolivia()

    cortes_dt = []

    for hora_txt in CORTES_SEGUIMIENTO:
        cortes_dt.append(
            (
                hora_txt,
                _hora_en_fecha(
                    momento,
                    hora_txt,
                ),
            )
        )

    for hora_txt, corte_dt in cortes_dt:
        if momento <= corte_dt:
            minutos = int(
                (corte_dt - momento).total_seconds()
                // 60
            )
            return {
                "hora": hora_txt,
                "estado": "proximo",
                "minutos": max(minutos, 0),
                "texto": (
                    f"Próximo corte recomendado: {hora_txt}"
                    if minutos > 0
                    else f"Corte recomendado ahora: {hora_txt}"
                ),
            }

    return {
        "hora": CORTES_SEGUIMIENTO[-1],
        "estado": "finalizado",
        "minutos": None,
        "texto": "Los cortes recomendados del día ya finalizaron.",
    }


def resumen_frecuencia_seguimiento(
    usuarios,
    momento=None,
):
    momento = momento or ahora_bolivia()

    recientes = []
    disponibles = []

    for usuario in usuarios:
        minutos = minutos_desde_ultimo_envio(
            usuario,
            momento,
        )

        if minutos is None:
            disponibles.append(
                usuario
            )
        elif minutos < MINUTOS_RECOMENDADOS_ENTRE_SEGUIMIENTOS:
            recientes.append(
                (
                    usuario,
                    minutos,
                )
            )
        else:
            disponibles.append(
                usuario
            )

    return recientes, disponibles


def puede_enviar_seguimiento(usuario, momento=None):
    """
    Permite envío normal dentro de turno.
    Si coordinación activa el modo excepcional, también permite
    enviar fuera de turno.
    """
    return (
        operador_en_turno(
            usuario,
            momento,
        )
        or bool(
            st.session_state.get(
                "permitir_envio_fuera_turno",
                False,
            )
        )
    )


def operador_en_turno(
    usuario,
    momento=None,
):
    """
    True únicamente si el operador está dentro de su jornada actual:
    entrada <= hora actual < salida.

    El break sigue siendo parte del turno, aunque el cálculo productivo
    permanezca congelado durante esos 30 minutos.
    """
    momento = momento or ahora_bolivia()

    horario = obtener_horario_operador(
        usuario,
        momento.date(),
    )

    if not horario:
        return False

    entrada = _hora_en_fecha(
        momento,
        horario["entrada"],
    )
    salida = _hora_en_fecha(
        momento,
        horario["salida"],
    )

    return entrada <= momento < salida


def texto_estado_turno(
    usuario,
    momento=None,
):
    momento = momento or ahora_bolivia()

    horario = obtener_horario_operador(
        usuario,
        momento.date(),
    )

    if not horario:
        return "Horario no configurado"

    entrada = _hora_en_fecha(
        momento,
        horario["entrada"],
    )
    salida = _hora_en_fecha(
        momento,
        horario["salida"],
    )

    if momento < entrada:
        return f"Inicia a las {horario['entrada']}"

    if momento >= salida:
        return f"Fuera de turno · salió {horario['salida']}"

    return f"En turno · hasta {horario['salida']}"


def operador_fuera_de_horario(
    usuario,
    momento=None,
):
    """
    Devuelve True si el momento actual ya es posterior
    a la hora de salida configurada del operador.
    """
    momento = momento or ahora_bolivia()

    horario = obtener_horario_operador(
        usuario,
        momento.date(),
    )

    if not horario:
        return False

    salida = _hora_en_fecha(
        momento,
        horario["salida"],
    )

    return momento >= salida



def _clave_aviso_grupal(fecha=None):
    fecha = fecha or fecha_local_actual()
    return f"aviso_grupal_seguimiento|{fecha.isoformat()}"


def contar_avisos_grupales_hoy():
    """
    Cuenta cuántos avisos grupales de seguimiento se enviaron hoy.
    Usa Supabase si está disponible y Session State como respaldo.
    """
    fecha = fecha_local_actual()
    clave = _clave_aviso_grupal(fecha)

    cache = st.session_state.setdefault(
        "avisos_grupales_cache",
        {},
    )

    # Supabase es la fuente preferida porque conserva el conteo
    # aunque Streamlit reinicie la sesión.
    sb = get_supabase()
    if sb is not None:
        try:
            resp = (
                sb.table("envios_mensajes")
                .select("fecha_hora")
                .eq("fecha", fecha.isoformat())
                .eq("tipo", "aviso_grupal_seguimiento")
                .eq("estado", "enviado")
                .execute()
            )
            if resp.data is not None:
                total = len(resp.data)
                cache[clave] = total
                return total
        except Exception:
            pass

    return int(cache.get(clave, 0) or 0)


def registrar_aviso_grupal_enviado(detalle=""):
    ahora_local = ahora_bolivia()
    ahora_utc = datetime.now(
        timezone.utc
    )
    clave = _clave_aviso_grupal(
        ahora_local.date()
    )

    cache = st.session_state.setdefault(
        "avisos_grupales_cache",
        {},
    )
    cache[clave] = int(
        cache.get(clave, 0) or 0
    ) + 1

    registro = {
        "fecha": ahora_local.date().isoformat(),
        "fecha_hora": ahora_utc.isoformat(),
        "usuario": "grupo",
        "operador": "Grupo Inmobiliaria",
        "canal": "telegram",
        "tipo": "aviso_grupal_seguimiento",
        "estado": "enviado",
        "detalle": str(detalle or ""),
    }

    sb = get_supabase()
    if sb is not None:
        try:
            (
                sb.table("envios_mensajes")
                .insert(registro)
                .execute()
            )
        except Exception:
            pass

    return registro


def generar_aviso_grupo_envios(
    operadores_enviados,
    operadores_fuera_turno=None,
    numero_seguimiento=None,
):
    """
    Genera un aviso grupal distinto según el momento de la jornada y
    la cantidad de seguimientos grupales realizados hoy.

    No expone nombres ni resultados individuales.
    """
    operadores_enviados = list(
        operadores_enviados or []
    )

    if not operadores_enviados:
        return ""

    ahora = ahora_bolivia()

    if numero_seguimiento is None:
        numero_seguimiento = (
            contar_avisos_grupales_hoy()
            + 1
        )

    hora = ahora.strftime("%H:%M")

    # V89 · El texto no depende de decir "primer/segundo seguimiento".
    # Así sigue siendo correcto aunque Streamlit reinicie la sesión o el
    # historial anterior no se encuentre en Supabase.
    if ahora.hour < 12:
        cuerpo = (
            "Se realizó un seguimiento individual de avance a los operadores "
            "que se encuentran actualmente de turno.\n\n"
            "Mantengamos el ritmo para continuar avanzando con las metas del día. 💪"
        )

    elif ahora.hour < 17:
        cuerpo = (
            "Se realizó una nueva actualización y seguimiento individual "
            "de los avances a los operadores que se encuentran de turno.\n\n"
            "Continuemos enfocados en las metas del día y en recuperar "
            "cualquier brecha pendiente. 💪"
        )

    else:
        cuerpo = (
            "Se realizó un nuevo corte y seguimiento individual a los operadores "
            "que continúan de turno.\n\n"
            "Aprovechemos las horas restantes para cerrar la jornada "
            "con el mejor cumplimiento posible. 💪"
        )

    return (
        f"📊 Seguimiento de avance · {hora}\n\n"
        + cuerpo
    )


def enviar_aviso_grupo_post_envio(
    operadores_enviados,
    operadores_fuera_turno=None,
):
    chat_grupo = obtener_telegram_group_chat_id()

    if not chat_grupo:
        return False, "No está configurado TELEGRAM_GROUP_CHAT_ID."

    numero_seguimiento = (
        contar_avisos_grupales_hoy()
        + 1
    )

    mensaje = generar_aviso_grupo_envios(
        operadores_enviados,
        operadores_fuera_turno,
        numero_seguimiento=numero_seguimiento,
    )

    if not mensaje:
        return False, "No hubo envíos individuales para informar."

    ok, detalle = enviar_mensaje_telegram(
        chat_grupo,
        mensaje,
    )

    if ok:
        registrar_aviso_grupal_enviado(
            detalle=(
                f"Seguimiento grupal #{numero_seguimiento}. "
                f"Operadores con envío individual: "
                f"{len(list(operadores_enviados or []))}."
            )
        )

    return ok, detalle

def enviar_copia_coordinador(operador, mensaje_original, detalle_envio):
    chat_coord = obtener_telegram_coordinador_chat_id()

    if not chat_coord:
        return False, "Falta TELEGRAM_COORDINADOR_CHAT_ID."

    copia = (
        f"✅ MENSAJE ENVIADO\n\n"
        f"👤 Operador: {operador}\n"
        f"📬 Estado: {detalle_envio}\n\n"
        f"--- Copia del mensaje ---\n"
        f"{mensaje_original}"
    )

    return enviar_mensaje_telegram(chat_coord, copia)


# =========================================================
# TELÉFONOS
# =========================================================

def normalizar_telefono_whatsapp(valor):
    numero = re.sub(
        r"\D",
        "",
        str(valor or ""),
    )

    if not numero:
        return ""

    # Bolivia: si se ingresan 8 dígitos, agregar automáticamente 591.
    if len(numero) == 8:
        numero = "591" + numero

    return numero


# =========================================================
# ESTILO
# =========================================================

st.markdown(
    """
    <style>

        .block-container {
            padding-top: 1.5rem;
            padding-bottom: 3rem;
        }

        [data-testid="stSidebar"] {
            background:
                linear-gradient(180deg, #10233f 0%, #0d1d35 100%);
            border-right: 1px solid rgba(255,255,255,.06);
        }

        [data-testid="stSidebar"] > div:first-child {
            padding-top: 1rem;
        }

        [data-testid="stSidebar"] * {
            color: white;
        }

        /* Marca / encabezado */
        .sidebar-brand {
            display:flex;
            align-items:center;
            gap:12px;
            padding:10px 8px 16px 8px;
        }

        .sidebar-logo {
            width:42px;
            height:42px;
            border-radius:12px;
            display:flex;
            align-items:center;
            justify-content:center;
            background:linear-gradient(135deg,#2f80ed 0%,#37c9a5 100%);
            font-size:20px;
            font-weight:800;
            box-shadow:0 8px 20px rgba(0,0,0,.18);
        }

        .sidebar-brand-title {
            font-size:17px;
            font-weight:800;
            line-height:1.15;
        }

        .sidebar-brand-sub {
            font-size:11px;
            color:#a8b6ca !important;
            margin-top:3px;
        }

        .sidebar-section-label {
            color:#7183a0 !important;
            font-size:10px;
            font-weight:800;
            letter-spacing:.10em;
            text-transform:uppercase;
            margin:8px 0 6px 8px;
        }

        /* Navegación */
        [data-testid="stSidebar"] div[role="radiogroup"] {
            gap:4px;
        }

        [data-testid="stSidebar"] div[role="radiogroup"] label {
            border-radius:11px;
            padding:8px 10px;
            transition:background .15s ease, transform .15s ease;
            border:1px solid transparent;
        }

        [data-testid="stSidebar"] div[role="radiogroup"] label:hover {
            background:rgba(255,255,255,.06);
        }

        [data-testid="stSidebar"] div[role="radiogroup"] label:has(input:checked) {
            background:linear-gradient(90deg, rgba(47,128,237,.32), rgba(55,201,165,.14));
            border:1px solid rgba(93,163,255,.22);
            box-shadow:inset 3px 0 0 #53d4b3;
        }

        [data-testid="stSidebar"] div[role="radiogroup"] label p {
            font-size:13px;
            font-weight:650;
        }

        /* Tarjeta de actualización */
        .sidebar-status-card {
            margin-top:10px;
            padding:14px;
            border-radius:14px;
            background:linear-gradient(135deg, rgba(26,106,112,.38), rgba(14,65,88,.52));
            border:1px solid rgba(85,214,173,.14);
        }

        .sidebar-status-top {
            display:flex;
            align-items:center;
            gap:7px;
            font-size:12px;
            font-weight:700;
        }

        .sidebar-status-dot {
            width:8px;
            height:8px;
            border-radius:50%;
            background:#45d19a;
            box-shadow:0 0 0 4px rgba(69,209,154,.10);
        }

        .sidebar-status-date {
            font-size:13px;
            font-weight:800;
            margin-top:10px;
        }

        .sidebar-status-time {
            color:#8fa1ba !important;
            font-size:10px;
            margin-top:2px;
        }

        /* Perfil */
        .sidebar-profile {
            display:flex;
            align-items:center;
            gap:10px;
            padding:12px 8px;
            margin-top:6px;
        }

        .sidebar-avatar {
            width:38px;
            height:38px;
            border-radius:50%;
            display:flex;
            align-items:center;
            justify-content:center;
            background:linear-gradient(135deg,#5a67d8,#805ad5);
            font-weight:800;
            font-size:13px;
        }

        .sidebar-profile-name {
            font-size:13px;
            font-weight:800;
        }

        .sidebar-profile-role {
            color:#8fa1ba !important;
            font-size:10px;
            margin-top:2px;
        }

        .sidebar-version {
            margin:12px 0 4px 0;
            padding:9px 11px;
            border-radius:10px;
            background:rgba(255,255,255,.04);
            border:1px solid rgba(255,255,255,.06);
            display:flex;
            justify-content:space-between;
            align-items:center;
            font-size:10px;
            color:#8fa1ba !important;
        }

        .sidebar-version-badge {
            background:rgba(69,209,154,.14);
            color:#6ee7b7 !important;
            border-radius:999px;
            padding:3px 7px;
            font-weight:700;
        }

        [data-testid="stSidebar"] hr {
            border-color:rgba(255,255,255,.08);
            margin:.75rem 0;
        }

        .gen-header {
            padding: 8px 0 18px 0;
        }

        .gen-title {
            font-size: 30px;
            font-weight: 800;
            margin: 0;
        }

        .gen-subtitle {
            color: #667085;
            font-size: 15px;
            margin-top: 2px;
        }

        .metric-card {
            background: white;
            border: 1px solid #e6eaf0;
            border-radius: 14px;
            padding: 20px;
            min-height: 145px;
            box-shadow: 0 2px 8px rgba(16,24,40,.05);
        }

        .metric-label {
            color: #667085;
            font-size: 13px;
            margin-bottom: 8px;
        }

        .metric-value {
            color: #101828;
            font-size: 28px;
            font-weight: 800;
        }

        .metric-detail {
            color: #667085;
            font-size: 12px;
            margin-top: 8px;
        }

        .status-box {
            background: #f8fafc;
            border: 1px solid #e6eaf0;
            border-radius: 14px;
            padding: 20px;
        }

        .hero-card {
            background: linear-gradient(135deg, #172a4a 0%, #24456f 100%);
            color: white;
            border-radius: 18px;
            padding: 22px 24px;
            margin-bottom: 18px;
            box-shadow: 0 6px 18px rgba(16,24,40,.10);
        }

        .hero-title {
            font-size: 24px;
            font-weight: 800;
            margin-bottom: 4px;
        }

        .hero-subtitle {
            font-size: 13px;
            opacity: .86;
        }

        .kpi-card {
            background: white;
            border: 1px solid #e8edf3;
            border-radius: 16px;
            padding: 18px;
            min-height: 132px;
            box-shadow: 0 3px 10px rgba(16,24,40,.05);
        }

        .kpi-title {
            color: #667085;
            font-size: 12px;
            font-weight: 700;
            text-transform: uppercase;
            letter-spacing: .04em;
        }

        .kpi-value {
            color: #101828;
            font-size: 28px;
            font-weight: 800;
            margin-top: 6px;
        }

        .kpi-foot {
            color: #667085;
            font-size: 12px;
            margin-top: 8px;
        }

        .section-title {
            font-size: 20px;
            font-weight: 800;
            margin-top: 12px;
            margin-bottom: 8px;
        }

        .soft-card {
            background: #ffffff;
            border: 1px solid #e8edf3;
            border-radius: 14px;
            padding: 16px;
        }


        /* ===== MENSAJES DIARIOS V32 ===== */

        .daily-panel {
            background: linear-gradient(135deg, #102846 0%, #214e7d 100%);
            color: white;
            border-radius: 18px;
            padding: 20px 24px;
            margin-bottom: 16px;
            box-shadow: 0 8px 22px rgba(16,24,40,.10);
        }

        .daily-panel-title {
            font-size: 23px;
            font-weight: 800;
            margin-bottom: 4px;
        }

        .daily-panel-sub {
            font-size: 13px;
            opacity: .88;
        }

        .section-chip {
            display: inline-block;
            background: #eef4ff;
            color: #24456f;
            border-radius: 999px;
            padding: 5px 10px;
            font-size: 12px;
            font-weight: 700;
            margin-bottom: 8px;
        }

        div[data-testid="stMetric"] {
            background: #ffffff;
            border: 1px solid #e7ebf0;
            padding: 12px 14px;
            border-radius: 12px;
        }

        div[data-testid="stExpander"] {
            border-radius: 12px;
        }


        /* ===== TARJETAS DE OPERADORES V53 ===== */

        .operator-card-head {
            display:flex;
            justify-content:space-between;
            align-items:flex-start;
            gap:12px;
            margin-bottom:10px;
        }

        .operator-card-name {
            font-size:21px;
            font-weight:800;
            color:#101828;
            line-height:1.2;
        }

        .operator-status-pill {
            display:inline-flex;
            align-items:center;
            border-radius:999px;
            padding:5px 10px;
            font-size:12px;
            font-weight:700;
            white-space:nowrap;
        }

        .status-red {
            background:#fff1f3;
            color:#c01048;
        }

        .status-orange {
            background:#fff6ed;
            color:#b54708;
        }

        .status-yellow {
            background:#fffaeb;
            color:#b54708;
        }

        .status-green {
            background:#ecfdf3;
            color:#067647;
        }

        .operator-meta-line {
            font-size:12px;
            color:#667085;
            margin-top:-2px;
            margin-bottom:10px;
        }

        .operator-progress-label {
            display:flex;
            justify-content:space-between;
            align-items:center;
            font-size:11px;
            color:#667085;
            margin-top:7px;
            margin-bottom:2px;
        }

        .operator-divider {
            height:1px;
            background:#eef1f5;
            margin:10px 0 8px 0;
        }


        .messages-summary {
            background:#ffffff;
            border:1px solid #e6eaf0;
            border-radius:16px;
            padding:14px 18px;
            margin:8px 0 14px 0;
        }

        .operator-daily-box {
            background:#f4f8ff;
            border:1px solid #cfe0ff;
            border-radius:10px;
            padding:9px 12px;
            margin:8px 0 10px 0;
            color:#24456f;
            font-size:12px;
            font-weight:700;
        }

        .operator-small-line {
            color:#667085;
            font-size:11px;
            margin-top:4px;
        }

        .operator-channel {
            color:#667085;
            font-size:11px;
            margin-top:3px;
        }


        /* ===== MENSAJES INDIVIDUALES V59 ===== */

        .messages-title-wrap {
            display:flex;
            align-items:center;
            gap:12px;
            margin-bottom:4px;
        }

        .messages-title-icon {
            width:42px;
            height:42px;
            border-radius:13px;
            display:flex;
            align-items:center;
            justify-content:center;
            background:#f1edff;
            color:#6c4ff8;
            font-size:22px;
            font-weight:800;
        }

        .messages-title-text {
            font-size:28px;
            font-weight:800;
            color:#101828;
            line-height:1.1;
        }

        .messages-subtitle-text {
            color:#667085;
            font-size:13px;
            margin-bottom:16px;
        }

        .summary-tile {
            background:#ffffff;
            border:1px solid #e6eaf0;
            border-radius:14px;
            padding:14px 15px;
            min-height:94px;
            box-shadow:0 2px 8px rgba(16,24,40,.04);
        }

        .summary-tile-label {
            color:#667085;
            font-size:11px;
            font-weight:700;
            text-transform:uppercase;
            letter-spacing:.04em;
        }

        .summary-tile-value {
            color:#101828;
            font-size:24px;
            font-weight:800;
            margin-top:5px;
        }

        .summary-tile-foot {
            color:#667085;
            font-size:11px;
            margin-top:3px;
        }

        .operator-head-v59 {
            display:flex;
            justify-content:space-between;
            align-items:flex-start;
            gap:12px;
            margin-bottom:10px;
        }

        .operator-id-wrap-v59 {
            display:flex;
            align-items:center;
            gap:10px;
            min-width:0;
        }

        .operator-avatar-v59 {
            min-width:44px;
            width:44px;
            height:44px;
            border-radius:50%;
            display:flex;
            align-items:center;
            justify-content:center;
            background:#f4ecff;
            color:#6c4ff8;
            font-size:14px;
            font-weight:800;
        }

        .operator-name-v59 {
            color:#101828;
            font-size:18px;
            font-weight:800;
            line-height:1.15;
        }

        .operator-mail-v59 {
            color:#667085;
            font-size:11px;
            margin-top:4px;
        }

        .operator-status-wrap-v59 {
            text-align:right;
            flex-shrink:0;
        }

        .operator-status-note-v59 {
            color:#667085;
            font-size:10px;
            margin-top:5px;
        }

        .daily-target-v59 {
            display:flex;
            align-items:center;
            justify-content:center;
            gap:10px;
            background:#f4f8ff;
            border:1px solid #cfe0ff;
            color:#24456f;
            border-radius:10px;
            padding:10px 12px;
            margin:8px 0 10px 0;
            font-size:12px;
            font-weight:800;
        }

        .toolbar-note-v59 {
            color:#667085;
            font-size:11px;
            margin-top:4px;
        }

        .operator-metric-caption-v59 {
            color:#667085;
            font-size:11px;
            margin-top:4px;
        }


        /* ===== TARJETAS COMPACTAS V61 ===== */
        .operator-head-v59 {
            margin-bottom:4px !important;
        }

        .operator-avatar-v59 {
            min-width:36px !important;
            width:36px !important;
            height:36px !important;
            font-size:12px !important;
        }

        .operator-name-v59 {
            font-size:16px !important;
        }

        .operator-mail-v59,
        .operator-status-note-v59 {
            margin-top:2px !important;
        }

        .daily-target-v59 {
            padding:6px 10px !important;
            margin:4px 0 6px 0 !important;
            font-size:11px !important;
            min-height:0 !important;
        }

        /* Reduce padding de contenedores con borde dentro del área principal */
        [data-testid="stMainBlockContainer"] div[data-testid="stVerticalBlockBorderWrapper"] > div {
            padding-top:8px;
            padding-bottom:8px;
        }

        /* Métricas más compactas manteniendo números legibles */
        [data-testid="stMetric"] {
            padding:4px 8px !important;
        }

        [data-testid="stMetricLabel"] {
            margin-bottom:0 !important;
        }

        [data-testid="stMetricLabel"] p {
            font-size:10px !important;
        }

        [data-testid="stMetricValue"] {
            font-size:22px !important;
            line-height:1.05 !important;
        }

        [data-testid="stMetricDelta"] {
            font-size:10px !important;
        }

        /* Barras y captions más juntos */
        [data-testid="stProgress"] {
            margin-top:-5px !important;
            margin-bottom:-7px !important;
        }

        [data-testid="stCaptionContainer"] {
            margin-top:-2px !important;
            margin-bottom:-4px !important;
        }

        [data-testid="stCaptionContainer"] p {
            font-size:10px !important;
        }

        /* Botones inferiores menos altos */
        [data-testid="stBaseButton-secondary"],
        [data-testid="stLinkButton"] a {
            min-height:32px !important;
            padding-top:4px !important;
            padding-bottom:4px !important;
        }


        /* ===== PRIORIZACIÓN VISUAL V62 ===== */

        .priority-chip-v62 {
            display:inline-flex;
            align-items:center;
            gap:5px;
            border-radius:999px;
            padding:4px 8px;
            font-size:10px;
            font-weight:800;
            margin-top:5px;
        }

        .priority-red-v62 {
            background:#fff1f3;
            color:#c01048;
        }

        .priority-orange-v62 {
            background:#fff6ed;
            color:#b54708;
        }

        .priority-green-v62 {
            background:#ecfdf3;
            color:#067647;
        }

        .channel-mini-v62 {
            display:flex;
            align-items:center;
            gap:10px;
            color:#667085;
            font-size:10px;
            margin-top:4px;
            flex-wrap:wrap;
        }

        .metric-box-v62 {
            background:#ffffff;
            border:1px solid #e7ebf0;
            border-radius:12px;
            padding:8px 9px 7px 9px;
            min-height:112px;
        }

        .metric-label-v62 {
            color:#667085;
            font-size:9px;
            font-weight:800;
            letter-spacing:.04em;
            text-transform:uppercase;
        }

        .metric-value-v62 {
            color:#101828;
            font-size:20px;
            font-weight:800;
            line-height:1.05;
            margin-top:5px;
            white-space:nowrap;
            overflow:hidden;
            text-overflow:ellipsis;
        }

        .metric-sub-v62 {
            color:#667085;
            font-size:10px;
            margin-top:3px;
            white-space:nowrap;
            overflow:hidden;
            text-overflow:ellipsis;
        }

        .metric-bar-bg-v62 {
            width:100%;
            height:7px;
            background:#eef1f4;
            border-radius:999px;
            margin-top:8px;
            overflow:hidden;
        }

        .metric-bar-fill-v62 {
            height:100%;
            border-radius:999px;
        }

        .bar-red-v62 {
            background:#e5484d;
        }

        .bar-orange-v62 {
            background:#f59e0b;
        }

        .bar-green-v62 {
            background:#22a447;
        }

        .metric-foot-v62 {
            color:#667085;
            font-size:9px;
            margin-top:6px;
        }

        .today-mini-v62 {
            color:#667085;
            font-size:10px;
            font-weight:700;
            margin:6px 0 7px 2px;
        }

        /* ===== CONTROL OPERATIVO V63 ===== */

        .status-summary-v63 {
            background:#ffffff;
            border:1px solid #e6eaf0;
            border-radius:14px;
            padding:11px 13px;
            min-height:74px;
        }

        .status-summary-label-v63 {
            color:#667085;
            font-size:10px;
            font-weight:800;
            text-transform:uppercase;
            letter-spacing:.04em;
        }

        .status-summary-value-v63 {
            color:#101828;
            font-size:22px;
            font-weight:800;
            margin-top:3px;
        }

        .status-summary-foot-v63 {
            color:#667085;
            font-size:10px;
            margin-top:2px;
        }

        .expected-note-v63 {
            display:flex;
            align-items:center;
            justify-content:space-between;
            gap:10px;
            background:#f8fafc;
            border:1px solid #e7ebf0;
            border-radius:10px;
            padding:8px 10px;
            margin:6px 0 10px 0;
            font-size:10px;
            color:#475467;
        }

        .expected-note-v63 strong {
            color:#101828;
        }

        .operator-alert-v63 {
            margin-top:4px;
            font-size:10px;
            font-weight:700;
        }

        .operator-alert-red-v63 {
            color:#c01048;
        }

        .operator-alert-orange-v63 {
            color:#b54708;
        }

        .operator-alert-green-v63 {
            color:#067647;
        }

        /* ===== FASE VISUAL V64 ===== */

        .top-action-row-v64 {
            display:flex;
            justify-content:flex-end;
            gap:10px;
            margin:2px 0 10px 0;
        }

        .color-kpi-v64 {
            border-radius:14px;
            padding:13px 14px;
            min-height:86px;
            border:1px solid;
            box-shadow:0 2px 8px rgba(16,24,40,.04);
        }

        .kpi-purple-v64 {
            background:#f6f3ff;
            border-color:#dfd6ff;
        }

        .kpi-blue-v64 {
            background:#eef7ff;
            border-color:#cfe6ff;
        }

        .kpi-green-v64 {
            background:#effcf4;
            border-color:#ccefd8;
        }

        .kpi-orange-v64 {
            background:#fff7ed;
            border-color:#fed7aa;
        }

        .kpi-red-v64 {
            background:#fff1f3;
            border-color:#fecdd3;
        }

        .kpi-value-color-v64 {
            font-size:24px;
            font-weight:800;
            color:#101828;
            margin-top:4px;
        }

        .kpi-label-color-v64 {
            font-size:10px;
            font-weight:800;
            letter-spacing:.04em;
            text-transform:uppercase;
        }

        .kpi-foot-color-v64 {
            font-size:10px;
            color:#667085;
            margin-top:3px;
        }

        .kpi-icon-v64 {
            font-size:20px;
            margin-bottom:2px;
        }

        .operator-card-accent-v64 {
            position:relative;
        }

        .operator-card-accent-v64::before {
            content:"";
            position:absolute;
            left:-1px;
            top:12px;
            bottom:12px;
            width:4px;
            border-radius:999px;
            background:#d0d5dd;
        }

        .channel-ok-v64 {
            color:#067647;
            font-weight:700;
        }

        .channel-pending-v64 {
            color:#b54708;
            font-weight:700;
        }

        .gap-badge-v64 {
            display:inline-flex;
            align-items:center;
            gap:4px;
            border-radius:999px;
            padding:4px 8px;
            font-size:10px;
            font-weight:800;
        }

        .gap-red-v64 {
            background:#fff1f3;
            color:#c01048;
        }

        .gap-orange-v64 {
            background:#fff6ed;
            color:#b54708;
        }

        .gap-green-v64 {
            background:#ecfdf3;
            color:#067647;
        }

        .recovery-focus-v64 {
            color:#b54708;
            font-weight:800;
        }

        .toolbar-shell-v64 {
            background:#ffffff;
            border:1px solid #e6eaf0;
            border-radius:14px;
            padding:10px 12px 2px 12px;
            margin:8px 0 10px 0;
        }

        .mini-chip-v64 {
            display:inline-flex;
            align-items:center;
            gap:5px;
            border-radius:999px;
            padding:4px 8px;
            background:#f2f4f7;
            color:#475467;
            font-size:10px;
            font-weight:700;
            margin-right:6px;
        }

        /* ===== FASE V66 · CONTROL DEL DÍA ===== */

        .hello-v66 {
            display:flex;
            justify-content:space-between;
            align-items:flex-start;
            gap:16px;
            margin-bottom:12px;
        }

        .hello-title-v66 {
            font-size:27px;
            font-weight:850;
            color:#101828;
            line-height:1.05;
        }

        .hello-sub-v66 {
            color:#667085;
            font-size:12px;
            margin-top:5px;
        }

        .kpi-v66 {
            border-radius:15px;
            padding:13px 14px;
            min-height:92px;
            border:1px solid #e5e9ef;
            box-shadow:0 2px 8px rgba(16,24,40,.04);
            background:#fff;
        }

        .kpi-v66-purple {background:linear-gradient(135deg,#f7f3ff,#fff);border-color:#e4d8ff;}
        .kpi-v66-blue   {background:linear-gradient(135deg,#edf7ff,#fff);border-color:#cde6ff;}
        .kpi-v66-green  {background:linear-gradient(135deg,#eefcf4,#fff);border-color:#ccefd8;}
        .kpi-v66-orange {background:linear-gradient(135deg,#fff7ed,#fff);border-color:#fed7aa;}
        .kpi-v66-red    {background:linear-gradient(135deg,#fff1f3,#fff);border-color:#fecdd3;}

        .kpi-label-v66 {
            color:#667085;
            font-size:9px;
            font-weight:850;
            letter-spacing:.05em;
            text-transform:uppercase;
        }

        .kpi-value-v66 {
            color:#101828;
            font-size:23px;
            font-weight:850;
            margin-top:5px;
        }

        .kpi-foot-v66 {
            color:#667085;
            font-size:10px;
            margin-top:4px;
        }

        .day-shell-v66 {
            background:#fff;
            border:1px solid #e5e9ef;
            border-radius:16px;
            padding:14px;
            margin:10px 0 12px 0;
            box-shadow:0 2px 10px rgba(16,24,40,.04);
        }

        .day-title-v66 {
            color:#101828;
            font-size:15px;
            font-weight:850;
            margin-bottom:10px;
        }

        .day-grid-v66 {
            display:grid;
            grid-template-columns:1fr 1.6fr 1.6fr;
            gap:12px;
        }

        .day-card-v66 {
            border:1px solid #e7ebf0;
            border-radius:13px;
            padding:12px;
            background:#fbfcfe;
        }

        .day-card-blue-v66 {
            background:#f2f7ff;
            border-color:#d4e4ff;
        }

        .day-card-green-v66 {
            background:#f1fbf5;
            border-color:#d2efdc;
        }

        .day-card-label-v66 {
            color:#667085;
            font-size:9px;
            font-weight:850;
            text-transform:uppercase;
            letter-spacing:.05em;
        }

        .day-main-v66 {
            color:#101828;
            font-size:24px;
            font-weight:850;
            margin-top:5px;
        }

        .day-detail-v66 {
            color:#667085;
            font-size:10px;
            margin-top:4px;
        }

        .day-delta-red-v66 {
            color:#c01048;
            font-weight:850;
        }

        .day-delta-green-v66 {
            color:#067647;
            font-weight:850;
        }

        .day-progress-bg-v66 {
            height:8px;
            border-radius:999px;
            overflow:hidden;
            background:#e9edf2;
            margin-top:9px;
        }

        .day-progress-fill-v66 {
            height:100%;
            border-radius:999px;
        }

        .operator-today-shell-v66 {
            border-radius:11px;
            padding:9px 10px;
            margin:6px 0 7px 0;
            border:1px solid #e7ebf0;
            background:#f8fafc;
        }

        .operator-today-title-v66 {
            font-size:9px;
            font-weight:850;
            color:#667085;
            text-transform:uppercase;
            letter-spacing:.04em;
            margin-bottom:7px;
        }

        .operator-today-grid-v66 {
            display:grid;
            grid-template-columns:1fr 1fr;
            gap:8px;
        }

        .operator-today-item-v66 {
            background:#fff;
            border:1px solid #e8ecf1;
            border-radius:9px;
            padding:7px 8px;
        }

        .operator-today-value-v66 {
            font-size:16px;
            font-weight:850;
            color:#101828;
        }

        .operator-today-sub-v66 {
            font-size:9px;
            color:#667085;
            margin-top:2px;
        }

        .monthly-strip-v66 {
            display:grid;
            grid-template-columns:1fr 1fr 1.2fr;
            gap:6px;
            margin-top:7px;
        }

        .monthly-item-v66 {
            border-top:1px solid #eef1f4;
            padding-top:6px;
        }

        .monthly-label-v66 {
            color:#98a2b3;
            font-size:8px;
            font-weight:800;
            text-transform:uppercase;
        }

        .monthly-value-v66 {
            color:#344054;
            font-size:11px;
            font-weight:750;
            margin-top:2px;
        }

        .status-critical-v66 {
            background:#fff1f3;
            border-color:#fecdd3 !important;
        }

        .status-attention-v66 {
            background:#fffaf2;
            border-color:#fed7aa !important;
        }

        .status-good-v66 {
            background:#f3fbf6;
            border-color:#ccefd8 !important;
        }

        .legend-v66 {
            background:#f8fafc;
            border:1px solid #e7ebf0;
            border-radius:11px;
            padding:8px 10px;
            margin-top:10px;
            color:#667085;
            font-size:10px;
        }

        /* ===== V71 · TARJETAS OPERATIVAS REFINADAS ===== */

        .op-head-v71 {
            display:flex;
            justify-content:space-between;
            gap:10px;
            align-items:flex-start;
            margin-bottom:6px;
        }

        .op-name-v71 {
            color:#101828;
            font-size:17px;
            font-weight:850;
            line-height:1.15;
        }

        .op-contact-v71 {
            color:#667085;
            font-size:10px;
            margin-top:3px;
        }

        .op-state-v71 {
            display:inline-flex;
            align-items:center;
            border-radius:999px;
            padding:4px 8px;
            font-size:10px;
            font-weight:850;
        }

        .op-red-v71 {
            color:#c01048;
            background:#fff1f3;
        }

        .op-orange-v71 {
            color:#b54708;
            background:#fff6ed;
        }

        .op-green-v71 {
            color:#067647;
            background:#ecfdf3;
        }

        .op-gray-v71 {
            color:#475467;
            background:#f2f4f7;
        }

        .schedule-v71 {
            color:#667085;
            font-size:10px;
            padding:6px 8px;
            background:#f8fafc;
            border:1px solid #e7ebf0;
            border-radius:9px;
            margin-bottom:7px;
        }

        .today-label-v71 {
            color:#667085;
            font-size:9px;
            font-weight:850;
            text-transform:uppercase;
            letter-spacing:.04em;
            margin:3px 0 4px 0;
        }

        .month-bar-v71 {
            border-top:1px solid #eef1f4;
            margin-top:7px;
            padding-top:6px;
        }

        .legend-v71 {
            background:#f8fafc;
            border:1px solid #e7ebf0;
            border-radius:11px;
            padding:8px 10px;
            color:#667085;
            font-size:10px;
            margin-top:8px;
        }

        /* Menos aire vertical dentro de las tarjetas */
        [data-testid="stVerticalBlockBorderWrapper"] [data-testid="stMetric"] {
            padding:5px 8px !important;
        }

        [data-testid="stVerticalBlockBorderWrapper"] [data-testid="stMetricValue"] {
            font-size:22px !important;
        }

        [data-testid="stVerticalBlockBorderWrapper"] [data-testid="stMetricLabel"] p {
            font-size:10px !important;
        }

        [data-testid="stVerticalBlockBorderWrapper"] [data-testid="stMetricDelta"] {
            font-size:10px !important;
        }

        /* ===== V72 · GRID 4 COLUMNAS ===== */
        .op-title-v72{
            font-size:15px;
            font-weight:850;
            color:#101828;
            line-height:1.15;
        }
        .op-meta-v72{
            font-size:9px;
            color:#667085;
            margin-top:3px;
        }
        .op-status-v72{
            display:inline-flex;
            padding:3px 7px;
            border-radius:999px;
            font-size:9px;
            font-weight:850;
        }
        .op-red-v72{background:#fff1f3;color:#c01048;}
        .op-orange-v72{background:#fff6ed;color:#b54708;}
        .op-green-v72{background:#ecfdf3;color:#067647;}
        .op-gray-v72{background:#f2f4f7;color:#475467;}
        .schedule-v72{
            font-size:9px;
            color:#667085;
            margin:4px 0 6px 0;
        }
        .month-line-v72{
            font-size:9px;
            color:#667085;
            margin-top:5px;
            padding-top:5px;
            border-top:1px solid #eef1f4;
        }
        .month-strong-v72{
            color:#344054;
            font-weight:800;
        }

        /* compactar métricas dentro de tarjetas */
        [data-testid="stVerticalBlockBorderWrapper"] [data-testid="stMetric"]{
            padding:3px 5px !important;
        }
        [data-testid="stVerticalBlockBorderWrapper"] [data-testid="stMetricValue"]{
            font-size:18px !important;
            line-height:1.05 !important;
        }
        [data-testid="stVerticalBlockBorderWrapper"] [data-testid="stMetricLabel"] p{
            font-size:9px !important;
        }
        [data-testid="stVerticalBlockBorderWrapper"] [data-testid="stMetricDelta"]{
            font-size:9px !important;
        }
        [data-testid="stVerticalBlockBorderWrapper"] [data-testid="stCaptionContainer"] p{
            font-size:9px !important;
        }

        /* ===== V73 · VISUAL OPERATIVA MEJORADA ===== */

        .op-card-v73 {
            border-radius:14px;
            padding:0;
        }

        .op-head-wrap-v73 {
            display:flex;
            justify-content:space-between;
            align-items:flex-start;
            gap:8px;
            margin-bottom:5px;
        }

        .op-title-v73 {
            font-size:15px;
            font-weight:850;
            color:#101828;
            line-height:1.15;
        }

        .op-sub-v73 {
            font-size:9px;
            color:#667085;
            margin-top:3px;
            line-height:1.25;
        }

        .status-pill-v73 {
            display:inline-flex;
            align-items:center;
            gap:4px;
            padding:3px 7px;
            border-radius:999px;
            font-size:9px;
            font-weight:850;
            white-space:nowrap;
        }

        .status-red-v73 {
            background:#fff1f3;
            color:#c01048;
        }

        .status-orange-v73 {
            background:#fff6ed;
            color:#b54708;
        }

        .status-green-v73 {
            background:#ecfdf3;
            color:#067647;
        }

        .status-gray-v73 {
            background:#f2f4f7;
            color:#475467;
        }

        .schedule-row-v73 {
            background:#f8fafc;
            border:1px solid #eef1f4;
            border-radius:8px;
            padding:5px 7px;
            font-size:9px;
            color:#667085;
            margin:4px 0 6px 0;
        }

        .today-kicker-v73 {
            font-size:8px;
            color:#98a2b3;
            text-transform:uppercase;
            letter-spacing:.06em;
            font-weight:850;
            margin:2px 0 4px 0;
        }

        .progress-mini-v73 {
            height:5px;
            background:#eef1f4;
            border-radius:999px;
            overflow:hidden;
            margin-top:5px;
        }

        .progress-mini-fill-v73 {
            height:100%;
            border-radius:999px;
        }

        .fill-red-v73 {background:#e5484d;}
        .fill-orange-v73 {background:#f59e0b;}
        .fill-green-v73 {background:#22a447;}

        .month-summary-v73 {
            display:flex;
            justify-content:space-between;
            align-items:center;
            gap:6px;
            margin-top:6px;
            padding-top:6px;
            border-top:1px solid #eef1f4;
            font-size:9px;
            color:#667085;
        }

        .month-summary-v73 strong {
            color:#344054;
        }

        .focus-note-v73 {
            border-radius:8px;
            padding:5px 7px;
            margin-top:5px;
            font-size:9px;
            font-weight:800;
        }

        .focus-red-v73 {
            background:#fff1f3;
            color:#c01048;
        }

        .focus-orange-v73 {
            background:#fff6ed;
            color:#b54708;
        }

        .focus-green-v73 {
            background:#ecfdf3;
            color:#067647;
        }

        /* tarjetas algo más altas pero mejor balanceadas */
        [data-testid="stVerticalBlockBorderWrapper"] {
            border-radius:14px !important;
        }

        [data-testid="stVerticalBlockBorderWrapper"] [data-testid="stMetric"]{
            padding:3px 4px !important;
        }

        [data-testid="stVerticalBlockBorderWrapper"] [data-testid="stMetricValue"]{
            font-size:19px !important;
        }

        [data-testid="stVerticalBlockBorderWrapper"] [data-testid="stMetricLabel"] p{
            font-size:9px !important;
            font-weight:700 !important;
        }

        [data-testid="stVerticalBlockBorderWrapper"] [data-testid="stMetricDelta"]{
            font-size:9px !important;
        }

        [data-testid="stVerticalBlockBorderWrapper"] [data-testid="stCaptionContainer"] p{
            font-size:8px !important;
        }

        /* Botones más discretos */
        [data-testid="stVerticalBlockBorderWrapper"] button,
        [data-testid="stVerticalBlockBorderWrapper"] a[data-testid="stBaseButton-secondary"]{
            min-height:30px !important;
            font-size:10px !important;
        }

        /* ===== V74 · TARJETAS 3 COLUMNAS / MÁS LEGIBLES ===== */

        .op-head-v74{
            display:flex;
            justify-content:space-between;
            gap:10px;
            align-items:flex-start;
            margin-bottom:6px;
        }

        .op-name-v74{
            font-size:16px;
            line-height:1.15;
            font-weight:850;
            color:#101828;
        }

        .op-contact-v74{
            margin-top:3px;
            font-size:10px;
            color:#667085;
        }

        .status-v74{
            display:inline-flex;
            align-items:center;
            padding:4px 8px;
            border-radius:999px;
            font-size:10px;
            font-weight:850;
            white-space:nowrap;
        }

        .v74-red{background:#fff1f3;color:#c01048;}
        .v74-orange{background:#fff6ed;color:#b54708;}
        .v74-green{background:#ecfdf3;color:#067647;}
        .v74-gray{background:#f2f4f7;color:#475467;}

        .schedule-v74{
            background:#f8fafc;
            border:1px solid #e7ebf0;
            border-radius:9px;
            padding:6px 8px;
            font-size:10px;
            color:#667085;
            margin-bottom:7px;
        }

        .kicker-v74{
            color:#98a2b3;
            font-size:9px;
            font-weight:850;
            letter-spacing:.05em;
            text-transform:uppercase;
            margin:2px 0 5px 0;
        }

        .mini-progress-v74{
            height:6px;
            border-radius:999px;
            background:#edf0f4;
            overflow:hidden;
            margin-top:5px;
        }

        .mini-fill-v74{
            height:100%;
            border-radius:999px;
        }

        .mini-red-v74{background:#e5484d;}
        .mini-orange-v74{background:#f59e0b;}
        .mini-green-v74{background:#22a447;}

        .action-v74{
            border-radius:9px;
            padding:7px 9px;
            margin:7px 0 6px 0;
            font-size:10px;
            font-weight:800;
        }

        .action-red-v74{background:#fff1f3;color:#c01048;}
        .action-orange-v74{background:#fff6ed;color:#b54708;}
        .action-green-v74{background:#ecfdf3;color:#067647;}
        .action-gray-v74{background:#f2f4f7;color:#475467;}

        .monthly-v74{
            border-top:1px solid #eef1f4;
            margin-top:7px;
            padding-top:6px;
            display:flex;
            justify-content:space-between;
            gap:8px;
            font-size:10px;
            color:#667085;
        }

        .monthly-v74 strong{
            color:#344054;
        }

        /* Métricas más cómodas que V73 */
        [data-testid="stVerticalBlockBorderWrapper"] [data-testid="stMetric"]{
            padding:5px 6px !important;
        }

        [data-testid="stVerticalBlockBorderWrapper"] [data-testid="stMetricValue"]{
            font-size:21px !important;
            line-height:1.05 !important;
        }

        [data-testid="stVerticalBlockBorderWrapper"] [data-testid="stMetricLabel"] p{
            font-size:10px !important;
            font-weight:700 !important;
        }

        [data-testid="stVerticalBlockBorderWrapper"] [data-testid="stMetricDelta"]{
            font-size:10px !important;
        }

        [data-testid="stVerticalBlockBorderWrapper"] [data-testid="stCaptionContainer"] p{
            font-size:9px !important;
        }

        [data-testid="stVerticalBlockBorderWrapper"] button{
            min-height:32px !important;
            font-size:10px !important;
        }

        /* ===== V77 · ALERTAS COMPACTAS ===== */

        .alert-summary-v77 {
            border:1px solid #f4d7a2;
            background:linear-gradient(135deg,#fffaf0,#fffef8);
            border-radius:14px;
            padding:12px 14px;
            margin-top:8px;
        }

        .alert-summary-title-v77 {
            font-size:13px;
            font-weight:850;
            color:#7a4b00;
        }

        .alert-summary-sub-v77 {
            font-size:10px;
            color:#8a6b2d;
            margin-top:3px;
        }

        .alert-row-v77 {
            display:flex;
            justify-content:space-between;
            gap:10px;
            align-items:center;
            padding:6px 0;
            border-bottom:1px solid rgba(122,75,0,.08);
            font-size:10px;
        }

        .alert-row-v77:last-child {
            border-bottom:none;
        }

        .alert-name-v77 {
            color:#344054;
            font-weight:750;
        }

        .alert-value-v77 {
            color:#b54708;
            font-weight:850;
            white-space:nowrap;
        }

        .leader-strip-v77 {
            border-radius:11px;
            padding:8px 10px;
            font-size:10px;
            font-weight:750;
            min-height:42px;
            display:flex;
            align-items:center;
        }

        .leader-green-v77 {
            background:#ecfdf3;
            color:#067647;
            border:1px solid #ccefd8;
        }

        .leader-orange-v77 {
            background:#fff7ed;
            color:#b54708;
            border:1px solid #fed7aa;
        }

        /* ===== V79 · RESUMEN EJECUTIVO ===== */

        .team-state-v79{
            border-radius:14px;
            padding:12px 14px;
            margin:8px 0 12px 0;
            border:1px solid;
            display:flex;
            justify-content:space-between;
            align-items:center;
            gap:14px;
        }

        .team-state-red-v79{
            background:#fff1f3;
            border-color:#fecdd3;
            color:#9f1239;
        }

        .team-state-orange-v79{
            background:#fff7ed;
            border-color:#fed7aa;
            color:#9a3412;
        }

        .team-state-green-v79{
            background:#ecfdf3;
            border-color:#bbf7d0;
            color:#166534;
        }

        .team-state-title-v79{
            font-size:15px;
            font-weight:850;
        }

        .team-state-sub-v79{
            font-size:10px;
            margin-top:3px;
            opacity:.82;
        }

        .team-focus-v79{
            font-size:11px;
            font-weight:800;
            white-space:nowrap;
        }

        .kpi-card-v79{
            border:1px solid #e7ebf0;
            border-radius:14px;
            padding:12px 13px;
            min-height:112px;
            background:#fff;
            box-shadow:0 2px 8px rgba(16,24,40,.04);
        }

        .kpi-card-blue-v79{
            background:linear-gradient(135deg,#eef7ff,#ffffff);
            border-color:#cfe6ff;
        }

        .kpi-card-orange-v79{
            background:linear-gradient(135deg,#fff7ed,#ffffff);
            border-color:#fed7aa;
        }

        .kpi-card-green-v79{
            background:linear-gradient(135deg,#ecfdf3,#ffffff);
            border-color:#bbf7d0;
        }

        .kpi-card-purple-v79{
            background:linear-gradient(135deg,#f6f3ff,#ffffff);
            border-color:#ded5ff;
        }

        .kpi-label-v79{
            font-size:9px;
            font-weight:850;
            letter-spacing:.05em;
            text-transform:uppercase;
            color:#667085;
        }

        .kpi-value-v79{
            font-size:24px;
            font-weight:850;
            color:#101828;
            margin-top:5px;
        }

        .kpi-foot-v79{
            font-size:10px;
            color:#667085;
            margin-top:4px;
        }

        .compare-card-v79{
            border:1px solid #e7ebf0;
            border-radius:13px;
            padding:11px 12px;
            background:#fff;
            min-height:118px;
        }

        .compare-title-v79{
            font-size:11px;
            font-weight:850;
            color:#344054;
        }

        .compare-value-v79{
            font-size:21px;
            font-weight:850;
            color:#101828;
            margin-top:4px;
        }

        .compare-sub-v79{
            font-size:10px;
            color:#667085;
            margin-top:3px;
        }

        .compare-gap-red-v79{
            color:#c01048;
            font-weight:850;
        }

        .compare-gap-orange-v79{
            color:#b54708;
            font-weight:850;
        }

        .compare-gap-green-v79{
            color:#067647;
            font-weight:850;
        }
    </style>
    """,
    unsafe_allow_html=True,
)


# =========================================================
# MENÚ LATERAL
# =========================================================

with st.sidebar:

    st.markdown(
        """
        <div class="sidebar-brand">
            <div class="sidebar-logo">G</div>
            <div>
                <div class="sidebar-brand-title">GEN Control</div>
                <div class="sidebar-brand-sub">Cobranzas Inmobiliarias</div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.markdown(
        '<div class="sidebar-section-label">Navegación</div>',
        unsafe_allow_html=True,
    )

    menu = st.radio(
        "Navegación",
        [
            "🏠 Resumen",
            "📈 Comportamiento diario",
            "✉️ Mensajes diarios",
            "📥 Cargar reportes",
            "🗂️ Histórico",
            "💰 Bonos",
            "👥 Equipo",
            "⚙️ Configuración",
        ],
        label_visibility="collapsed",
    )

    st.markdown("---")

    ahora_sidebar = (
        st.session_state.get(
            "callcenter_cargado_en"
        )
        or st.session_state.get(
            "promesas_cargado_en"
        )
        or ahora_bolivia()
    )

    ahora_sidebar = (
        datetime_bolivia(ahora_sidebar)
        or ahora_sidebar
    )

    st.markdown(
        f"""
        <div class="sidebar-status-card">
            <div class="sidebar-status-top">
                <span class="sidebar-status-dot"></span>
                Datos actualizados
            </div>
            <div class="sidebar-status-date">
                {ahora_sidebar.strftime('%d/%m/%Y')}
            </div>
            <div class="sidebar-status-time">
                {ahora_sidebar.strftime('%H:%M')} · Hora Bolivia
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.markdown("---")

    st.markdown(
        """
        <div class="sidebar-profile">
            <div class="sidebar-avatar">JC</div>
            <div>
                <div class="sidebar-profile-name">José Carlos</div>
                <div class="sidebar-profile-role">Coordinador</div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.markdown(
        """
        <div class="sidebar-version">
            <span>GEN Control</span>
            <span class="sidebar-version-badge">Actualizado</span>
        </div>
        """,
        unsafe_allow_html=True,
    )


# =========================================================
# ENCABEZADO
# =========================================================

st.markdown(
    f"""
    <div class="gen-header">
        <div class="gen-title">{APP_NAME}</div>
        <div class="gen-subtitle">{APP_SUBTITLE}</div>
    </div>
    """,
    unsafe_allow_html=True,
)


# =========================================================
# RESUMEN
# =========================================================


def estado_datos_para_seguimiento(resultado=None):
    ahora = datetime.now(ZoneInfo("America/La_Paz"))
    hoy = ahora.date()
    promesas_ok = resultado is not None and not resultado.empty
    promesas_cargado = st.session_state.get("promesas_cargado_en")
    call_df = st.session_state.get("callcenter_df")
    call_ok = call_df is not None and not call_df.empty
    call_cargado = st.session_state.get("callcenter_cargado_en")
    corte = obtener_corte_callcenter(call_df) if call_ok else None
    if hasattr(corte, "to_pydatetime"):
        corte = corte.to_pydatetime()

    minutos_call = None
    if call_cargado is not None:
        try:
            minutos_call = max(int((ahora-call_cargado).total_seconds()//60),0)
        except Exception:
            pass

    fecha_operativa_call = None
    if call_ok:
        col_fecha = buscar_columna(call_df, ["fecha"])
        if col_fecha is not None:
            serie = pd.to_datetime(call_df[col_fecha], dayfirst=True, errors="coerce").dropna()
            if not serie.empty:
                fecha_operativa_call = serie.max().date()

    usuarios_detectados = 0
    if promesas_ok and "Usuario" in resultado.columns:
        usuarios_detectados = int(resultado["Usuario"].astype(str).nunique())

    recuperacion_ok = (
        promesas_ok
        and "Recuperación acumulada" in resultado.columns
        and resultado["Recuperación acumulada"].notna().any()
    )

    bloqueos, avisos = [], []
    if not promesas_ok:
        bloqueos.append("Falta Promesas de Pago")
    if not call_ok:
        bloqueos.append("Falta GEN CallCenter")
    elif fecha_operativa_call is not None and fecha_operativa_call != hoy:
        bloqueos.append(f"CallCenter corresponde al {fecha_operativa_call.strftime('%d/%m/%Y')}")
    if usuarios_detectados and usuarios_detectados < CANTIDAD_OPERADORES:
        avisos.append(f"Solo se identificaron {usuarios_detectados}/{CANTIDAD_OPERADORES} operadores")
    if call_ok and minutos_call is not None and minutos_call > 180:
        avisos.append(f"CallCenter fue cargado hace {minutos_call//60} h {minutos_call%60} min")

    if bloqueos:
        nivel,titulo="rojo","Datos no listos para enviar"
    elif avisos:
        nivel,titulo="naranja","Revisar antes de enviar"
    else:
        nivel,titulo="verde","Datos listos para seguimiento"

    return {
        "nivel":nivel,"titulo":titulo,"promesas_ok":promesas_ok,
        "promesas_cargado":promesas_cargado,"call_ok":call_ok,
        "call_cargado":call_cargado,"corte":corte,"minutos_call":minutos_call,
        "fecha_operativa_call":fecha_operativa_call,
        "usuarios_detectados":usuarios_detectados,"recuperacion_ok":recuperacion_ok,
        "razones_bloqueo":bloqueos,"advertencias":avisos,
        "bloquear_envio":bool(bloqueos),
    }


def operador_habilitado_para_envio(usuario, momento=None):
    """En turno + habilitación individual + habilitación global."""
    return bool(
        operador_en_turno(usuario, momento)
        or st.session_state.get(
            f"override_fuera_turno_{usuario}",
            False,
        )
        or st.session_state.get(
            "permitir_envio_fuera_turno",
            False,
        )
    )


if menu == "🏠 Resumen":

    resultado = st.session_state.resultado_operadores
    jornadas_info = jornadas_configuradas()

    st.markdown(
        f"""
        <div class="hero-card page-head-v22">
            <div style="
                position:relative;z-index:1;
                display:inline-flex;align-items:center;gap:6px;
                padding:5px 9px;border-radius:999px;
                background:rgba(255,255,255,.10);
                color:#CDE7F6;font-size:9px;font-weight:800;
                letter-spacing:.06em;text-transform:uppercase;
                margin-bottom:10px;
            ">
                ● GEN CONTROL · COBRANZAS INMOBILIARIAS
            </div>
            <div class="hero-title">Centro de control operativo</div>
            <div class="hero-subtitle">
                Seguimiento ejecutivo de metas, brechas y recuperación ·
                {fecha_local_actual().strftime("%d/%m/%Y")}
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    if resultado is None:
        st.info(
            "Carga el reporte de Promesas de Pago para visualizar "
            "el tablero operativo."
        )

    else:
        esperado = jornadas_info["esperado_pct"]

        promedio_gestiones = float(
            resultado["% Gestiones"].mean()
        )
        promedio_compromisos = float(
            resultado["% Compromisos"].mean()
        )
        promedio_recuperacion = float(
            resultado["% Recuperación"].mean()
        )

        total_gestiones = int(
            resultado["Gestiones"].sum()
        )
        total_compromisos = int(
            resultado["Compromisos"].sum()
        )
        total_recuperacion = float(
            resultado["Recuperación acumulada"].sum()
        )

        meta_equipo_gestiones = (
            st.session_state.meta_gestiones_cfg
            * CANTIDAD_OPERADORES
        )
        meta_equipo_compromisos = (
            st.session_state.meta_compromisos_cfg
            * CANTIDAD_OPERADORES
        )
        meta_equipo_recuperacion = (
            st.session_state.meta_recuperacion_cfg
            * CANTIDAD_OPERADORES
        )

        # Estado del equipo según la MAYOR brecha real.
        # No se promedian indicadores distintos para evitar ocultar rezagos.
        resumen_mes_v29 = resumen_mes_actual_v29()

        html_resumen_mes_v31 = (
            f'<div class="month-head-v29">'
            f'<div><div class="month-title-v29">Resumen de {resumen_mes_v29["mes_actual"]} {resumen_mes_v29["anio"]}</div>'
            f'<div class="month-sub-v29">Gestiones y Compromisos corresponden únicamente al mes actual.</div></div>'
            f'<div class="month-badge-v29">📅 Mes actual · {resumen_mes_v29["mes_actual"]}</div>'
            f'</div>'
            f'<div class="month-grid-v29">'
            f'<div class="month-card-v29 blue">'
            f'<div class="lbl">Gestiones del mes</div>'
            f'<div class="val">{formato_entero(resumen_mes_v29["total_gestiones"])}</div>'
            f'<div class="sub">{formato_porcentaje(resumen_mes_v29["promedio_gestiones"])} de cumplimiento promedio</div>'
            f'</div>'
            f'<div class="month-card-v29 orange">'
            f'<div class="lbl">Compromisos del mes</div>'
            f'<div class="val">{formato_entero(resumen_mes_v29["total_compromisos"])}</div>'
            f'<div class="sub">{formato_porcentaje(resumen_mes_v29["promedio_compromisos"])} de cumplimiento promedio</div>'
            f'</div>'
            f'<div class="month-card-v29 green">'
            f'<div class="lbl">Recuperación</div>'
            f'<div class="val">{formato_usd(resumen_mes_v29["total_recuperacion"])}</div>'
            f'<div class="sub">{resumen_mes_v29["periodo_recuperacion"]["etiqueta"]}</div>'
            f'</div>'
            f'</div>'
        )
        st.markdown(
            html_resumen_mes_v31,
            unsafe_allow_html=True,
        )

        if resumen_mes_v29["periodo_recuperacion"]["cierre_anterior"]:
            html_cierre_mes_v31 = (
                f'<div class="month-close-v29">'
                f'💰 <strong>Recuperación en cierre:</strong> del 1 al 5 se mantiene el cierre de '
                f'<strong>{resumen_mes_v29["periodo_recuperacion"]["nombre_mes"]}</strong>. '
                f'Gestiones y Compromisos ya pertenecen a '
                f'<strong>{resumen_mes_v29["mes_actual"]}</strong>.'
                f'</div>'
            )
            st.markdown(
                html_cierre_mes_v31,
                unsafe_allow_html=True,
            )

        periodo_rec_v27 = periodo_recuperacion_actual()
        esperado_recuperacion_v27 = esperado_indicador(
            "Recuperación",
            esperado,
        )

        brechas_equipo = {
            "Gestiones": promedio_gestiones - esperado,
            "Compromisos": promedio_compromisos - esperado,
            "Recuperación": (
                promedio_recuperacion
                - esperado_recuperacion_v27
            ),
        }

        indicador_prioritario = min(
            brechas_equipo,
            key=brechas_equipo.get,
        )

        mayor_brecha_equipo = float(
            brechas_equipo[
                indicador_prioritario
            ]
        )

        if mayor_brecha_equipo >= -3:
            estado_general = "🟢 Equipo en ritmo"
            estado_clase_v79 = "team-state-green-v79"
        elif mayor_brecha_equipo >= -10:
            estado_general = "🟠 Equipo en seguimiento"
            estado_clase_v79 = "team-state-orange-v79"
        else:
            estado_general = "🔴 Reforzar ritmo"
            estado_clase_v79 = "team-state-red-v79"

        if periodo_rec_v27["cierre_anterior"]:
            st.markdown(
                f"""
                <div style="
                    border:1px solid #D6E5F8;
                    background:linear-gradient(90deg,#EEF5FF,#F9FBFF);
                    border-radius:12px;
                    padding:10px 13px;
                    margin:4px 0 10px;
                    color:#35597E;
                    font-size:10px;
                ">
                    📅 <strong>Inicio de {nombre_mes_es(fecha_local_actual().month)}:</strong>
                    Gestiones y Compromisos ya se calculan para
                    <strong>{nombre_mes_es(fecha_local_actual().month)}</strong>.
                    Recuperación continúa como
                    <strong>{periodo_rec_v27["etiqueta"]}</strong>
                    y tiene plazo hasta el <strong>05/{fecha_local_actual().month:02d}</strong>.
                </div>
                """,
                unsafe_allow_html=True,
            )

        st.markdown(
            f"""
            <div class="team-state-v79 {estado_clase_v79}">
                <div>
                    <div class="team-state-title-v79">
                        {estado_general}
                    </div>
                    <div class="team-state-sub-v79">
                        Comparación contra el esperado a la fecha:
                        {formato_porcentaje(esperado)}
                    </div>
                </div>
                <div class="team-focus-v79">
                    Prioridad: {indicador_prioritario} ·
                    {formato_porcentaje(abs(mayor_brecha_equipo))} de brecha
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

        # -------------------------------------------------
        # KPI PRINCIPALES — V79
        # -------------------------------------------------

        c1, c2, c3, c4 = st.columns(4)

        with c1:
            st.markdown(
                f"""
                <div class="kpi-card-v79 kpi-card-blue-v79">
                    <div class="kpi-label-v79">📞 Cumplimiento Gestiones · {resumen_mes_v29["mes_actual"]}</div>
                    <div class="kpi-value-v79">{formato_entero(total_gestiones)}</div>
                    <div class="kpi-foot-v79">
                        {formato_porcentaje(promedio_gestiones)} ·
                        Meta equipo {formato_entero(meta_equipo_gestiones)}
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
            )

        with c2:
            st.markdown(
                f"""
                <div class="kpi-card-v79 kpi-card-orange-v79">
                    <div class="kpi-label-v79">🤝 Cumplimiento Compromisos · {resumen_mes_v29["mes_actual"]}</div>
                    <div class="kpi-value-v79">{formato_entero(total_compromisos)}</div>
                    <div class="kpi-foot-v79">
                        {formato_porcentaje(promedio_compromisos)} ·
                        Meta equipo {formato_entero(meta_equipo_compromisos)}
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
            )

        with c3:
            st.markdown(
                f"""
                <div class="kpi-card-v79 kpi-card-green-v79">
                    <div class="kpi-label-v79">💰 Recuperación · {periodo_rec_v27["etiqueta"]}</div>
                    <div class="kpi-value-v79">{formato_usd(total_recuperacion)}</div>
                    <div class="kpi-foot-v79">
                        {formato_porcentaje(promedio_recuperacion)} ·
                        Meta equipo {formato_usd(meta_equipo_recuperacion)}
                        {" · Plazo 05/" + fecha_local_actual().strftime("%m") if periodo_rec_v27["cierre_anterior"] else ""}
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
            )

        with c4:
            st.markdown(
                f"""
                <div class="kpi-card-v79 kpi-card-purple-v79">
                    <div class="kpi-label-v79">📈 Esperado a la fecha</div>
                    <div class="kpi-value-v79">{formato_porcentaje(esperado)}</div>
                    <div class="kpi-foot-v79">
                        {jornadas_info['disponibles']} jornadas disponibles contando hoy
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
            )

        st.write("")

        # -------------------------------------------------
        # AVANCE VS ESPERADO — V79
        # -------------------------------------------------

        st.markdown(f"### Avance de {resumen_mes_v29['mes_actual']} vs esperado")

        comparativos_v79 = [
            ("Gestiones", promedio_gestiones, esperado),
            ("Compromisos", promedio_compromisos, esperado),
            (
                f"Recuperación · {periodo_rec_v27['etiqueta']}",
                promedio_recuperacion,
                esperado_recuperacion_v27,
            ),
        ]

        cc1, cc2, cc3 = st.columns(3)

        for columna_cmp, (nombre_cmp, valor_cmp, esperado_cmp) in zip(
            [cc1, cc2, cc3],
            comparativos_v79,
        ):
            brecha_cmp = float(
                valor_cmp - esperado_cmp
            )

            if brecha_cmp >= -3:
                clase_gap = "compare-gap-green-v79"
                estado_gap = "En ritmo"
            elif brecha_cmp >= -10:
                clase_gap = "compare-gap-orange-v79"
                estado_gap = "Seguimiento"
            else:
                clase_gap = "compare-gap-red-v79"
                estado_gap = "Prioridad"

            with columna_cmp:
                st.markdown(
                    f"""
                    <div class="compare-card-v79">
                        <div class="compare-title-v79">{nombre_cmp}</div>
                        <div class="compare-value-v79">
                            {formato_porcentaje(valor_cmp)}
                        </div>
                        <div class="compare-sub-v79">
                            Esperado: {formato_porcentaje(esperado_cmp)}
                        </div>
                        <div class="compare-sub-v79 {clase_gap}">
                            {estado_gap} ·
                            Brecha {formato_porcentaje(brecha_cmp)}
                        </div>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )

        st.write("")

        # -------------------------------------------------
        # RANKING EJECUTIVO · V26
        # -------------------------------------------------

        st.markdown(
            """
            <div class="rank-head-v26">
                <div>
                    <div class="rank-title-v26">Ranking de operadores</div>
                    <div class="rank-sub-v26">
                        Gestiones y Compromisos: {resumen_mes_v29["mes_actual"]} ·
                        Recuperación: {resumen_mes_v29["periodo_recuperacion"]["etiqueta"]}.
                    </div>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

        ranking = resultado.copy()

        rf1, rf2 = st.columns([4.2, 1.1], vertical_alignment="bottom")

        with rf1:
            criterio = st.selectbox(
                "Ranking por",
                [
                    "Recuperación",
                    "Gestiones",
                    "Compromisos",
                ],
                key="ranking_simple_v76",
            )

        with rf2:
            menor_primero = st.checkbox(
                "Menor primero",
                value=False,
                key="ranking_menor_v76",
            )

        mapa_criterio = {
            "Recuperación": "% Recuperación",
            "Gestiones": "% Gestiones",
            "Compromisos": "% Compromisos",
        }
        columna_orden = mapa_criterio[criterio]

        esperado_ranking_v27 = (
            esperado_recuperacion_v27
            if criterio == "Recuperación"
            else esperado
        )

        ranking["Estado"] = ranking[columna_orden].apply(
            lambda x: clasificar_avance(
                float(x),
                esperado_ranking_v27,
            )
        )

        ranking = ranking.sort_values(
            columna_orden,
            ascending=menor_primero,
            kind="stable",
        ).reset_index(drop=True)

        ranking["Posición"] = ranking.index + 1

        # KPIs específicos del ranking
        total_ops_rank_v26 = len(ranking)
        lider_rank_v26 = ranking.iloc[0] if not ranking.empty else None
        ultimo_rank_v26 = ranking.iloc[-1] if not ranking.empty else None
        promedio_rank_v26 = float(ranking[columna_orden].mean()) if not ranking.empty else 0.0
        reforzar_rank_v26 = int(
            (ranking[columna_orden].astype(float) < esperado_ranking_v27 - 10).sum()
        ) if not ranking.empty else 0

        st.markdown(
            f"""
            <div class="rank-kpis-v26">
                <div class="rank-kpi-v26 blue">
                    <div class="k">Operadores</div>
                    <div class="v">{total_ops_rank_v26}</div>
                    <div class="s">Activos en el ranking</div>
                </div>
                <div class="rank-kpi-v26">
                    <div class="k">Gestiones equipo</div>
                    <div class="v">{formato_entero(total_gestiones)}</div>
                    <div class="s">Acumulado actual</div>
                </div>
                <div class="rank-kpi-v26 purple">
                    <div class="k">Compromisos equipo</div>
                    <div class="v">{formato_entero(total_compromisos)}</div>
                    <div class="s">Acumulado actual</div>
                </div>
                <div class="rank-kpi-v26 green">
                    <div class="k">Recuperación · {periodo_rec_v27["etiqueta"]}</div>
                    <div class="v">{formato_usd(total_recuperacion)}</div>
                    <div class="s">Acumulado actual</div>
                </div>
                <div class="rank-kpi-v26 blue">
                    <div class="k">Promedio {criterio.lower()}</div>
                    <div class="v">{formato_porcentaje(promedio_rank_v26)}</div>
                    <div class="s">Promedio del equipo</div>
                </div>
                <div class="rank-kpi-v26 orange">
                    <div class="k">A reforzar</div>
                    <div class="v">{reforzar_rank_v26}</div>
                    <div class="s">Brecha mayor a 10 p.p.</div>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

        # Tabla visual
        filas_rank_v26 = []

        for _, rr in ranking.iterrows():
            pos = int(rr["Posición"])
            pct_g = float(rr["% Gestiones"])
            pct_c = float(rr["% Compromisos"])
            pct_r = float(rr["% Recuperación"])
            pct_estado = float(rr[columna_orden])

            # Barras visuales: topadas para lectura, no para cálculo.
            bar_g = min(max(pct_g, 0), 100)
            bar_c = min(max(pct_c, 0), 100)
            bar_r = min(max(pct_r, 0), 100)

            if pct_estado >= esperado_ranking_v27 - 3:
                estado_cls = "ok"
            elif pct_estado >= esperado_ranking_v27 - 10:
                estado_cls = "warn"
            else:
                estado_cls = "bad"

            medalla = "🥇" if pos == 1 else ("🥈" if pos == 2 else ("🥉" if pos == 3 else ""))

            filas_rank_v26.append(
                f"""
                <div class="rank-row-v26 {'first' if pos == 1 else ''}">
                    <div class="rank-pos-v26">{pos} {medalla}</div>
                    <div class="rank-name-v26">{rr['Operador']}</div>

                    <div>
                        <div class="rank-metric-v26">{formato_entero(rr['Gestiones'])}</div>
                        <div class="rank-metric-sub-v26">{formato_porcentaje(pct_g)}</div>
                        <div class="rank-bar-v26"><span class="g" style="width:{bar_g:.1f}%"></span></div>
                    </div>

                    <div>
                        <div class="rank-metric-v26">{formato_entero(rr['Compromisos'])}</div>
                        <div class="rank-metric-sub-v26">{formato_porcentaje(pct_c)}</div>
                        <div class="rank-bar-v26"><span class="c" style="width:{bar_c:.1f}%"></span></div>
                    </div>

                    <div>
                        <div class="rank-metric-v26">{formato_usd(rr['Recuperación acumulada'])}</div>
                        <div class="rank-metric-sub-v26">{formato_porcentaje(pct_r)}</div>
                        <div class="rank-bar-v26"><span class="r" style="width:{bar_r:.1f}%"></span></div>
                    </div>

                    <div>
                        <span class="rank-pill-v26 {estado_cls}">{rr['Estado']}</span>
                    </div>
                </div>
                """
            )

        html_rank_v26 = (
            """
            <div class="rank-wrap-v26">
                <div class="rank-row-v26 header">
                    <div>#</div>
                    <div>Operador</div>
                    <div>Gestiones</div>
                    <div>Compromisos</div>
                    <div>Recuperación</div>
                    <div>Estado</div>
                </div>
            """
            + "".join(
                parte.strip().replace("\n", " ")
                for parte in filas_rank_v26
            )
            + "</div>"
        )

        st.markdown(
            html_rank_v26,
            unsafe_allow_html=True,
        )

        if not ranking.empty:
            lider = ranking.iloc[0]
            seguimiento = ranking.iloc[-1]

            st.markdown(
                f"""
                <div class="rank-strip-grid-v26">
                    <div class="rank-strip-v26 good">
                        🏆 <strong>Líder en {criterio.lower()}</strong><br>
                        {lider['Operador']} · {formato_porcentaje(lider[columna_orden])}
                    </div>
                    <div class="rank-strip-v26 warn">
                        🎯 <strong>Mayor oportunidad de mejora</strong><br>
                        {seguimiento['Operador']} · {formato_porcentaje(seguimiento[columna_orden])}
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
            )

        # -------------------------------------------------
        # ALERTAS Y RECOMENDACIONES · V26
        # -------------------------------------------------
        alertas_df = ranking[
            ranking[columna_orden].astype(float) < esperado_ranking_v27 - 10
        ].copy()

        st.markdown(
            """
            <div class="rank-head-v26" style="margin-top:8px;">
                <div>
                    <div class="rank-title-v26">Alertas y recomendaciones</div>
                    <div class="rank-sub-v26">
                        Prioridades del equipo según el indicador seleccionado.
                    </div>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

        if alertas_df.empty:
            st.success(
                f"✅ No hay brechas críticas en {criterio.lower()}."
            )
        else:
            alertas_df = alertas_df.sort_values(
                columna_orden,
                ascending=True,
                kind="stable",
            )

            lineas_alerta_v26 = []
            for _, ar in alertas_df.iterrows():
                valor_ar = float(ar[columna_orden])
                brecha_ar = max(esperado_ranking_v27 - valor_ar, 0)

                lineas_alerta_v26.append(
                    f"""
                    <div class="alert-line-v26">
                        <span class="name">{ar['Operador']}</span>
                        <span class="val">
                            {formato_porcentaje(valor_ar)} ·
                            brecha {formato_porcentaje(brecha_ar)}
                        </span>
                    </div>
                    """
                )

            recomendaciones_v26 = [
                "Priorizar seguimiento a quienes tengan la mayor brecha.",
                "Mantener el ritmo de gestiones aunque una meta mensual ya esté cumplida.",
                "Revisar compromisos y recuperación antes del cierre del día.",
            ]

            st.markdown(
                f"""
                <div class="alerts-grid-v26">
                    <div class="alerts-card-v26">
                        <h4>⚠️ {len(alertas_df)} operador(es) requieren atención</h4>
                        <p>Ordenados desde la mayor brecha hasta la menor.</p>
                        {''.join(x.strip().replace(chr(10),' ') for x in lineas_alerta_v26)}
                    </div>
                    <div class="alerts-card-v26" style="background:#F3F7FF;border-color:#D8E5F7;">
                        <h4>💡 Recomendaciones</h4>
                        <p>Acciones rápidas para el seguimiento de hoy.</p>
                        <div class="alert-line-v26"><span class="name">✓ {recomendaciones_v26[0]}</span></div>
                        <div class="alert-line-v26"><span class="name">✓ {recomendaciones_v26[1]}</span></div>
                        <div class="alert-line-v26"><span class="name">✓ {recomendaciones_v26[2]}</span></div>
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
            )


# =========================================================
# COMPORTAMIENTO DIARIO
# =========================================================

elif menu == "📈 Comportamiento diario":

    # =====================================================
    # COMPORTAMIENTO DIARIO · DISEÑO EJECUTIVO
    # =====================================================
    st.markdown(
        """
        <style>
        .beh-hero{
            background:linear-gradient(120deg,#102A43 0%,#163A5F 62%,#1A6080 100%);
            border-radius:19px;
            padding:20px 22px;
            margin:0 0 15px 0;
            box-shadow:0 14px 32px rgba(16,42,67,.12);
            position:relative;
            overflow:hidden;
        }
        .beh-hero:after{
            content:"";
            position:absolute;
            width:180px;height:180px;border-radius:50%;
            right:-50px;top:-90px;
            background:rgba(70,214,208,.11);
        }
        .beh-kicker{
            display:inline-block;
            padding:4px 8px;
            border-radius:999px;
            background:rgba(255,255,255,.10);
            color:#C8DCEB;
            font-size:8px;
            font-weight:800;
            letter-spacing:.06em;
            text-transform:uppercase;
            margin-bottom:7px;
        }
        .beh-title{
            color:white;
            font-size:24px;
            line-height:1.1;
            font-weight:850;
            letter-spacing:-.03em;
        }
        .beh-sub{
            color:#BDD0E0;
            font-size:10px;
            margin-top:5px;
        }
        .beh-kpi{
            min-height:124px;
            border:1px solid #E4EBF3;
            border-radius:16px;
            background:#FFFFFF;
            padding:14px 15px 13px;
            box-shadow:0 8px 22px rgba(16,42,67,.045);
        }
        .beh-kpi-label{
            font-size:9px;
            color:#708399;
            font-weight:800;
            text-transform:uppercase;
            letter-spacing:.04em;
        }
        .beh-kpi-value{
            font-size:25px;
            line-height:1.05;
            color:#102A43;
            font-weight:850;
            letter-spacing:-.035em;
            margin:8px 0 5px;
        }
        .beh-kpi-sub{
            color:#71849A;
            font-size:9px;
            line-height:1.35;
        }
        .beh-section{
            margin:18px 0 9px;
            font-size:18px;
            font-weight:830;
            color:#102A43;
            letter-spacing:-.025em;
        }
        .beh-section-sub{
            color:#71849A;
            font-size:9px;
            margin-top:-6px;
            margin-bottom:9px;
        }
        .beh-summary{
            border:1px solid #E4EBF3;
            border-radius:15px;
            background:#FFFFFF;
            padding:13px 15px;
            min-height:136px;
            box-shadow:0 7px 20px rgba(16,42,67,.035);
        }
        .beh-summary-title{
            font-size:11px;
            font-weight:820;
            color:#102A43;
            margin-bottom:9px;
        }
        .beh-row{
            display:flex;
            justify-content:space-between;
            gap:14px;
            padding:5px 0;
            border-bottom:1px solid #F0F3F7;
            font-size:9px;
            color:#65798E;
        }
        .beh-row:last-child{border-bottom:none}
        .beh-row strong{color:#183B5B}
        .beh-good{color:#067647!important}
        .beh-warn{color:#B54708!important}
        .beh-bad{color:#B42318!important}
        .beh-pill{
            display:inline-block;
            border-radius:999px;
            padding:4px 8px;
            background:#EEF6FF;
            color:#245A8D;
            font-size:8px;
            font-weight:800;
        }
        .chart-mini-grid{
            display:grid;
            grid-template-columns:repeat(4,minmax(0,1fr));
            gap:7px;
            margin:8px 0 4px;
        }
        .chart-mini{
            border:1px solid #E7EDF4;
            background:#F8FAFC;
            border-radius:10px;
            padding:8px 9px;
        }
        .chart-mini-label{
            color:#7A8DA1;
            font-size:8px;
            font-weight:750;
            text-transform:uppercase;
            letter-spacing:.03em;
        }
        .chart-mini-value{
            color:#16324F;
            font-size:13px;
            font-weight:850;
            margin-top:2px;
        }
        @media(max-width:950px){
            .chart-mini-grid{grid-template-columns:repeat(2,minmax(0,1fr));}
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    st.markdown(
        """
        <div class="beh-hero">
            <div class="beh-kicker">GEN Control · Analítica operativa</div>
            <div class="beh-title">Comportamiento diario</div>
            <div class="beh-sub">
                Evolución real de gestiones y compromisos, cumplimiento diario y días que requieren atención.
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    callcenter = st.session_state.callcenter_df

    if callcenter is None or callcenter.empty:
        st.warning(
            "Carga primero el reporte GEN CallCenter en Cargar reportes."
        )

    else:
        df_cc = callcenter.copy()

        col_fecha = buscar_columna(
            df_cc,
            ["fecha"],
        )
        col_usuario = buscar_columna(
            df_cc,
            ["usuario"],
        )
        col_compromiso = buscar_columna(
            df_cc,
            ["compromiso"],
        )

        if col_fecha is None or col_usuario is None:
            st.error(
                "No se encontraron las columnas Fecha y Usuario "
                "necesarias para el análisis diario."
            )

        else:
            df_cc["Fecha_dt"] = pd.to_datetime(
                df_cc[col_fecha],
                dayfirst=True,
                errors="coerce",
            )
            df_cc = df_cc.dropna(
                subset=["Fecha_dt"]
            )
            df_cc["Fecha_dia"] = (
                df_cc["Fecha_dt"].dt.date
            )

            df_cc["_usuario_norm"] = (
                df_cc[col_usuario]
                .astype(str)
                .apply(normalizar_texto)
            )

            df_cc = df_cc[
                df_cc["_usuario_norm"].isin(
                    list(OPERADORES.keys())
                )
            ].copy()

            if col_compromiso:
                compromiso_txt = (
                    df_cc[col_compromiso]
                    .astype(str)
                    .str.strip()
                )
                df_cc["_tiene_compromiso"] = (
                    df_cc[col_compromiso].notna()
                    & (compromiso_txt != "")
                    & (
                        compromiso_txt.str.lower()
                        != "nan"
                    )
                )
            else:
                df_cc["_tiene_compromiso"] = False

            if df_cc.empty:
                st.warning(
                    "El reporte cargado no contiene registros válidos "
                    "para los operadores configurados."
                )
            else:
                fecha_min = df_cc["Fecha_dia"].min()
                fecha_max = df_cc["Fecha_dia"].max()

                # V8 · Calendario operativo real:
                # los días con datos no equivalen necesariamente a días laborables.
                # La evaluación usa el calendario configurado (feriados incluidos).
                def es_dia_laboral_configurado(fecha_obj):
                    clave = f"{fecha_obj.year:04d}-{fecha_obj.month:02d}"

                    # Inicializar mes localmente si aún no existe.
                    if clave not in st.session_state.calendario_laboral:
                        inicializar_calendario_mes(
                            date(fecha_obj.year, fecha_obj.month, 1)
                        )

                        # Recuperar configuración guardada en Supabase.
                        if supabase_disponible():
                            calendario_guardado_v8 = cargar_calendario_supabase(
                                fecha_obj.year,
                                fecha_obj.month,
                            )
                            if calendario_guardado_v8:
                                st.session_state.calendario_laboral[
                                    clave
                                ].update(calendario_guardado_v8)

                    return bool(
                        st.session_state.calendario_laboral[
                            clave
                        ].get(
                            fecha_obj.day,
                            fecha_obj.weekday() != 6,
                        )
                    )

                f1, f2 = st.columns([2, 1])

                with f1:
                    rango = st.date_input(
                        "Periodo de análisis",
                        value=(
                            fecha_min,
                            fecha_max,
                        ),
                        min_value=fecha_min,
                        max_value=fecha_max,
                        key="periodo_comportamiento_final",
                    )

                with f2:
                    operador_sel = st.selectbox(
                        "Operador",
                        ["Todos"] + [
                            datos["nombre"]
                            for datos in OPERADORES.values()
                        ],
                        key="operador_comportamiento_final",
                    )

                if (
                    isinstance(rango, tuple)
                    and len(rango) == 2
                ):
                    inicio, fin = rango
                else:
                    inicio = fin = rango

                filtrado = df_cc[
                    (df_cc["Fecha_dia"] >= inicio)
                    & (df_cc["Fecha_dia"] <= fin)
                ].copy()

                usuario_sel = None

                if operador_sel != "Todos":
                    usuario_sel = next(
                        usuario
                        for usuario, datos in OPERADORES.items()
                        if datos["nombre"] == operador_sel
                    )

                    filtrado = filtrado[
                        filtrado["_usuario_norm"]
                        == usuario_sel
                    ].copy()

                diario = (
                    filtrado
                    .groupby("Fecha_dia")
                    .agg(
                        Gestiones=("Fecha_dia", "size"),
                        Compromisos=(
                            "_tiene_compromiso",
                            "sum",
                        ),
                        Operadores_activos=(
                            "_usuario_norm",
                            "nunique",
                        ),
                    )
                    .reset_index()
                    .sort_values("Fecha_dia")
                )

                diario["Compromisos"] = (
                    diario["Compromisos"].astype(int)
                )

                # V8 · La fuente de verdad es el calendario laboral configurado.
                # Un domingo puede tener actividad por cobertura especial, pero no
                # aumenta el número de jornadas laborables del equipo si sustituye
                # un descanso compensatorio entre semana.
                hoy_bolivia_comp = fecha_local_actual()
                diario["_fecha_ts"] = pd.to_datetime(diario["Fecha_dia"])
                diario["_weekday"] = diario["_fecha_ts"].dt.weekday
                diario["_es_domingo"] = diario["_weekday"].eq(6)
                diario["_es_hoy"] = diario["Fecha_dia"].eq(hoy_bolivia_comp)
                diario["_es_laboral_cfg"] = diario["Fecha_dia"].apply(
                    es_dia_laboral_configurado
                )
                diario["_cobertura_especial"] = (
                    diario["_es_domingo"]
                    & (diario["Operadores_activos"] > 0)
                )

                total_gestiones = int(
                    diario["Gestiones"].sum()
                ) if not diario.empty else 0

                total_compromisos = int(
                    diario["Compromisos"].sum()
                ) if not diario.empty else 0

                dias = int(len(diario))

                prom_g = (
                    total_gestiones / dias
                    if dias
                    else 0
                )
                prom_c = (
                    total_compromisos / dias
                    if dias
                    else 0
                )

                conversion = (
                    total_compromisos
                    / total_gestiones
                    * 100
                    if total_gestiones
                    else 0
                )

                # Meta diaria correcta:
                # lunes-sábado tienen meta; domingo no se evalúa.
                # En vista general se usa la cantidad de operadores con registros,
                # pero nunca se asigna meta a domingo.
                if operador_sel == "Todos":
                    diario["Meta_gestiones"] = (
                        diario["Operadores_activos"]
                        * META_DIARIA_GESTIONES
                    )
                    diario["Meta_compromisos"] = (
                        diario["Operadores_activos"]
                        * META_DIARIA_COMPROMISOS
                    )
                else:
                    diario["Meta_gestiones"] = META_DIARIA_GESTIONES
                    diario["Meta_compromisos"] = META_DIARIA_COMPROMISOS

                # Los días no laborables (domingos/feriados configurados)
                # pueden mostrar actividad real, pero no se usan para cumplimiento
                # del equipo ni para ampliar el denominador mensual.
                diario.loc[
                    ~diario["_es_laboral_cfg"],
                    ["Meta_gestiones", "Meta_compromisos"],
                ] = 0

                diario["_jornada_completa"] = (
                    diario["_es_laboral_cfg"]
                    & ~diario["_es_hoy"]
                )

                diario["Cumplimiento_gestiones"] = (
                    diario["Gestiones"]
                    / diario["Meta_gestiones"].replace(0, pd.NA)
                    * 100
                )
                diario["Cumplimiento_compromisos"] = (
                    diario["Compromisos"]
                    / diario["Meta_compromisos"].replace(0, pd.NA)
                    * 100
                )

                diario["Conversion"] = (
                    diario["Compromisos"]
                    / diario["Gestiones"].replace(0, pd.NA)
                    * 100
                ).fillna(0)

                # Total de jornadas laborables programadas en el periodo,
                # aunque no exista registro en el CallCenter ese día.
                fechas_periodo_v8 = pd.date_range(
                    start=inicio,
                    end=fin,
                    freq="D",
                )
                jornadas_programadas_v8 = [
                    d.date()
                    for d in fechas_periodo_v8
                    if es_dia_laboral_configurado(d.date())
                ]
                total_jornadas_programadas_v8 = len(
                    jornadas_programadas_v8
                )

                feriados_periodo_v8 = [
                    d.date()
                    for d in fechas_periodo_v8
                    if (
                        d.weekday() != 6
                        and not es_dia_laboral_configurado(d.date())
                    )
                ]
                total_feriados_v8 = len(feriados_periodo_v8)

                coberturas_domingo_v8 = int(
                    diario["_cobertura_especial"].sum()
                )

                if not diario.empty:
                    idx_mejor_g = diario["Gestiones"].idxmax()
                    idx_mejor_c = diario["Compromisos"].idxmax()
                    mejor_g = diario.loc[idx_mejor_g]
                    mejor_c = diario.loc[idx_mejor_c]
                else:
                    mejor_g = None
                    mejor_c = None

                # ------------------------------
                # KPIs
                # ------------------------------
                k1, k2, k3, k4, k5 = st.columns(5)

                with k1:
                    st.markdown(
                        f"""
                        <div class="beh-kpi">
                            <div class="beh-kpi-label">📞 Gestiones</div>
                            <div class="beh-kpi-value">{formato_entero(total_gestiones)}</div>
                            <div class="beh-kpi-sub">
                                Promedio diario · <b>{formato_entero(prom_g)}</b>
                            </div>
                        </div>
                        """,
                        unsafe_allow_html=True,
                    )

                with k2:
                    st.markdown(
                        f"""
                        <div class="beh-kpi">
                            <div class="beh-kpi-label">🎯 Compromisos</div>
                            <div class="beh-kpi-value">{formato_entero(total_compromisos)}</div>
                            <div class="beh-kpi-sub">
                                Promedio diario · <b>{formato_entero(prom_c)}</b>
                            </div>
                        </div>
                        """,
                        unsafe_allow_html=True,
                    )

                with k3:
                    fecha_mejor_g = (
                        pd.Timestamp(mejor_g["Fecha_dia"]).strftime("%d/%m/%Y")
                        if mejor_g is not None
                        else "—"
                    )
                    valor_mejor_g = (
                        formato_entero(mejor_g["Gestiones"])
                        if mejor_g is not None
                        else "0"
                    )
                    st.markdown(
                        f"""
                        <div class="beh-kpi">
                            <div class="beh-kpi-label">🏆 Mejor día · Gestiones</div>
                            <div class="beh-kpi-value">{valor_mejor_g}</div>
                            <div class="beh-kpi-sub">{fecha_mejor_g}</div>
                        </div>
                        """,
                        unsafe_allow_html=True,
                    )

                with k4:
                    fecha_mejor_c = (
                        pd.Timestamp(mejor_c["Fecha_dia"]).strftime("%d/%m/%Y")
                        if mejor_c is not None
                        else "—"
                    )
                    valor_mejor_c = (
                        formato_entero(mejor_c["Compromisos"])
                        if mejor_c is not None
                        else "0"
                    )
                    st.markdown(
                        f"""
                        <div class="beh-kpi">
                            <div class="beh-kpi-label">⭐ Mejor día · Compromisos</div>
                            <div class="beh-kpi-value">{valor_mejor_c}</div>
                            <div class="beh-kpi-sub">{fecha_mejor_c}</div>
                        </div>
                        """,
                        unsafe_allow_html=True,
                    )

                with k5:
                    st.markdown(
                        f"""
                        <div class="beh-kpi">
                            <div class="beh-kpi-label">🔄 Conversión</div>
                            <div class="beh-kpi-value">{conversion:.1f}%</div>
                            <div class="beh-kpi-sub">
                                Compromisos / Gestiones
                            </div>
                        </div>
                        """,
                        unsafe_allow_html=True,
                    )

                if diario.empty:
                    st.info(
                        "No hay información dentro del periodo seleccionado."
                    )

                else:
                    # Cálculos generales usados por gráficos y resumen.
                    diario_eval = diario[
                        diario["_jornada_completa"]
                    ].copy()

                    cumplimiento_g_prom = (
                        diario_eval["Gestiones"].sum()
                        / diario_eval["Meta_gestiones"].sum()
                        * 100
                        if diario_eval["Meta_gestiones"].sum()
                        else 0
                    )

                    cumplimiento_c_prom = (
                        diario_eval["Compromisos"].sum()
                        / diario_eval["Meta_compromisos"].sum()
                        * 100
                        if diario_eval["Meta_compromisos"].sum()
                        else 0
                    )

                    dias_meta_g = int(
                        (
                            diario_eval["Gestiones"]
                            >= diario_eval["Meta_gestiones"]
                        ).sum()
                    )
                    dias_meta_c = int(
                        (
                            diario_eval["Compromisos"]
                            >= diario_eval["Meta_compromisos"]
                        ).sum()
                    )

                    mejores = diario_eval.sort_values(
                        ["Cumplimiento_gestiones", "Gestiones"],
                        ascending=[False, False],
                    ).head(3)

                    peores = diario_eval.sort_values(
                        ["Cumplimiento_gestiones", "Gestiones"],
                        ascending=[True, True],
                    ).head(3)

                    # ------------------------------
                    # GRÁFICOS SEPARADOS
                    # ------------------------------
                    if bool(diario["_es_hoy"].any()):
                        st.caption(
                            "● Hoy: avance parcial · no afecta cumplimiento, brechas ni tendencias hasta el cierre."
                        )

                    # ==================================================
                    # EVOLUCIÓN DIARIA · REDISEÑO V7
                    # ==================================================
                    dias_evaluados = int(len(diario_eval))
                    dias_sobre_meta_g = int(
                        (diario_eval["Gestiones"] >= diario_eval["Meta_gestiones"]).sum()
                    )
                    dias_sobre_meta_c = int(
                        (diario_eval["Compromisos"] >= diario_eval["Meta_compromisos"]).sum()
                    )
                    pct_dias_g = (
                        dias_sobre_meta_g / dias_evaluados * 100
                        if dias_evaluados else 0
                    )
                    pct_dias_c = (
                        dias_sobre_meta_c / dias_evaluados * 100
                        if dias_evaluados else 0
                    )

                    chart_base = diario.copy()
                    chart_base["Fecha_plot"] = pd.to_datetime(
                        chart_base["Fecha_dia"]
                    )
                    chart_base["Meta_gestiones_plot"] = (
                        chart_base["Meta_gestiones"]
                        .where(chart_base["_jornada_completa"])
                    )
                    chart_base["Meta_compromisos_plot"] = (
                        chart_base["Meta_compromisos"]
                        .where(chart_base["_jornada_completa"])
                    )

                    def grafico_cumplimiento_limpio(
                        df_plot,
                        real_col,
                        meta_col,
                        titulo_real,
                    ):
                        base = alt.Chart(df_plot).encode(
                            x=alt.X(
                                "Fecha_plot:T",
                                title=None,
                                axis=alt.Axis(
                                    format="%d %b",
                                    labelAngle=0,
                                    labelOverlap=True,
                                    grid=False,
                                    tickCount="day",
                                ),
                            )
                        )

                        area = base.mark_area(
                            opacity=0.055,
                            interpolate="monotone",
                        ).encode(
                            y=alt.Y(
                                f"{real_col}:Q",
                                title=None,
                                scale=alt.Scale(zero=True),
                            )
                        )

                        real = base.mark_line(
                            strokeWidth=2.8,
                            interpolate="monotone",
                            point=alt.OverlayMarkDef(
                                filled=True,
                                size=42,
                            ),
                        ).encode(
                            y=alt.Y(
                                f"{real_col}:Q",
                                title=None,
                                scale=alt.Scale(zero=True),
                            ),
                            tooltip=[
                                alt.Tooltip(
                                    "Fecha_plot:T",
                                    title="Fecha",
                                    format="%d/%m/%Y",
                                ),
                                alt.Tooltip(
                                    f"{real_col}:Q",
                                    title=titulo_real,
                                    format=",",
                                ),
                                alt.Tooltip(
                                    f"{meta_col}:Q",
                                    title="Meta",
                                    format=",",
                                ),
                            ],
                        )

                        meta = base.mark_line(
                            strokeDash=[6, 5],
                            strokeWidth=1.6,
                            opacity=.75,
                        ).encode(
                            y=alt.Y(
                                f"{meta_col}:Q",
                                title=None,
                            )
                        )

                        return (
                            area + real + meta
                        ).properties(
                            height=250
                        ).configure_axis(
                            labelFontSize=10,
                            labelColor="#71849A",
                            gridColor="#EEF2F6",
                            domain=False,
                            tickColor="#D8E1EA",
                        ).configure_view(
                            strokeOpacity=0
                        )

                    st.markdown(
                        '<div class="beh-section">Evolución y cumplimiento</div>',
                        unsafe_allow_html=True,
                    )
                    st.caption(
                        "Resultado diario frente a la meta. Las jornadas en curso o no laborables no afectan el cumplimiento."
                    )

                    # Resumen compacto y entendible antes del gráfico
                    s1, s2, s3, s4 = st.columns(4)
                    with s1:
                        st.metric(
                            "Gestiones · promedio diario",
                            formato_entero(prom_g),
                            f"{cumplimiento_g_prom:.0f}% de cumplimiento",
                        )
                    with s2:
                        st.metric(
                            "Gestiones · días cumplidos",
                            f"{dias_sobre_meta_g} de {dias_evaluados}",
                            f"{pct_dias_g:.0f}% de las jornadas",
                        )
                    with s3:
                        st.metric(
                            "Compromisos · promedio diario",
                            formato_entero(prom_c),
                            f"{cumplimiento_c_prom:.0f}% de cumplimiento",
                        )
                    with s4:
                        st.metric(
                            "Compromisos · días cumplidos",
                            f"{dias_sobre_meta_c} de {dias_evaluados}",
                            f"{pct_dias_c:.0f}% de las jornadas",
                        )

                    tab_g, tab_c = st.tabs(
                        ["📞 Gestiones", "🎯 Compromisos"]
                    )

                    with tab_g:
                        g_left, g_right = st.columns([3.2, 1])

                        with g_left:
                            chart_g_final = grafico_cumplimiento_limpio(
                                chart_base,
                                "Gestiones",
                                "Meta_gestiones_plot",
                                "Gestiones",
                            )
                            st.altair_chart(
                                chart_g_final,
                                use_container_width=True,
                            )
                            st.caption(
                                "Línea continua: resultado real · línea segmentada: meta diaria."
                            )

                        with g_right:
                            st.markdown("##### Lectura")
                            st.metric(
                                "Cumplimiento",
                                f"{cumplimiento_g_prom:.0f}%",
                            )
                            st.metric(
                                "Meta cumplida",
                                f"{dias_sobre_meta_g} de {dias_evaluados} días",
                            )
                            brecha_graf_g = int(
                                diario_eval["Gestiones"].sum()
                                - diario_eval["Meta_gestiones"].sum()
                            )
                            st.metric(
                                "Brecha acumulada",
                                f"{brecha_graf_g:+,}".replace(",", "."),
                            )

                    with tab_c:
                        c_left, c_right = st.columns([3.2, 1])

                        with c_left:
                            chart_c_final = grafico_cumplimiento_limpio(
                                chart_base,
                                "Compromisos",
                                "Meta_compromisos_plot",
                                "Compromisos",
                            )
                            st.altair_chart(
                                chart_c_final,
                                use_container_width=True,
                            )
                            st.caption(
                                "Línea continua: resultado real · línea segmentada: meta diaria."
                            )

                        with c_right:
                            st.markdown("##### Lectura")
                            st.metric(
                                "Cumplimiento",
                                f"{cumplimiento_c_prom:.0f}%",
                            )
                            st.metric(
                                "Meta cumplida",
                                f"{dias_sobre_meta_c} de {dias_evaluados} días",
                            )
                            brecha_graf_c = int(
                                diario_eval["Compromisos"].sum()
                                - diario_eval["Meta_compromisos"].sum()
                            )
                            st.metric(
                                "Brecha acumulada",
                                f"{brecha_graf_c:+,}".replace(",", "."),
                            )

                    # ------------------------------
                    # RESUMEN EJECUTIVO DEL PERIODO
                    # ------------------------------
                    meta_g_total = int(
                        diario_eval["Meta_gestiones"].sum()
                    )
                    meta_c_total = int(
                        diario_eval["Meta_compromisos"].sum()
                    )

                    gestiones_eval = int(diario_eval["Gestiones"].sum())
                    compromisos_eval = int(diario_eval["Compromisos"].sum())

                    brecha_g_total = (
                        gestiones_eval - meta_g_total
                    )
                    brecha_c_total = (
                        compromisos_eval - meta_c_total
                    )

                    tendencia_g = 0.0
                    tendencia_c = 0.0
                    if len(diario_eval) >= 4:
                        mitad = max(len(diario_eval) // 2, 1)
                        prom_g_1 = diario_eval.iloc[:mitad]["Gestiones"].mean()
                        prom_g_2 = diario_eval.iloc[mitad:]["Gestiones"].mean()
                        prom_c_1 = diario_eval.iloc[:mitad]["Compromisos"].mean()
                        prom_c_2 = diario_eval.iloc[mitad:]["Compromisos"].mean()

                        if prom_g_1:
                            tendencia_g = (
                                (prom_g_2 - prom_g_1)
                                / prom_g_1
                                * 100
                            )
                        if prom_c_1:
                            tendencia_c = (
                                (prom_c_2 - prom_c_1)
                                / prom_c_1
                                * 100
                            )

                    promedio_meta_g = (
                        diario_eval["Meta_gestiones"].mean()
                        if dias
                        else 0
                    )
                    promedio_meta_c = (
                        diario_eval["Meta_compromisos"].mean()
                        if dias
                        else 0
                    )

                    st.markdown(
                        '<div class="beh-section">Lectura ejecutiva del periodo</div>',
                        unsafe_allow_html=True,
                    )
                    st.caption(
                        "Meta acumulada, brecha, tendencia y consistencia diaria."
                    )

                    e1, e2, e3, e4 = st.columns(4)

                    with e1:
                        st.metric(
                            "Meta acumulada · Gestiones",
                            formato_entero(meta_g_total),
                            f"{brecha_g_total:+,} de brecha".replace(",", "."),
                        )

                    with e2:
                        st.metric(
                            "Meta acumulada · Compromisos",
                            formato_entero(meta_c_total),
                            f"{brecha_c_total:+,} de brecha".replace(",", "."),
                        )

                    with e3:
                        st.metric(
                            "Tendencia · Gestiones",
                            f"{tendencia_g:+.1f}%",
                            "2ª mitad vs 1ª mitad",
                        )

                    with e4:
                        st.metric(
                            "Tendencia · Compromisos",
                            f"{tendencia_c:+.1f}%",
                            "2ª mitad vs 1ª mitad",
                        )

                    r1, r2, r3 = st.columns([1.05, 1.35, 1.35])

                    with r1:
                        with st.container(border=True):
                            st.markdown("#### 📌 Resumen del periodo")
                            st.write(f"**Jornadas laborables programadas:** {total_jornadas_programadas_v8}")
                            st.write(
                                f"**Cumplimiento gestiones:** {cumplimiento_g_prom:.0f}%"
                            )
                            st.write(
                                f"**Cumplimiento compromisos:** {cumplimiento_c_prom:.0f}%"
                            )
                            st.write(
                                f"**Feriados/no laborables entre semana:** {total_feriados_v8}"
                            )
                            if coberturas_domingo_v8:
                                st.write(
                                    f"**Coberturas especiales en domingo:** {coberturas_domingo_v8}"
                                )
                            st.write(
                                f"**Jornadas con meta de gestiones cumplida:** {dias_meta_g} / {len(diario_eval)}"
                            )
                            st.write(
                                f"**Jornadas con meta de compromisos cumplida:** {dias_meta_c} / {len(diario_eval)}"
                            )
                            st.caption(
                                f"Meta diaria promedio observada: "
                                f"{formato_entero(promedio_meta_g)} gestiones · "
                                f"{formato_entero(promedio_meta_c)} compromisos."
                            )

                    with r2:
                        with st.container(border=True):
                            st.markdown("#### 🏆 Días con mayor desempeño")
                            for pos, (_, fila_r) in enumerate(
                                mejores.iterrows(),
                                start=1,
                            ):
                                fecha_txt = pd.Timestamp(
                                    fila_r["Fecha_dia"]
                                ).strftime("%d/%m/%Y")
                                st.markdown(
                                    f"**{pos}. {fecha_txt}**  \n"
                                    f"📞 {formato_entero(fila_r['Gestiones'])} gestiones · "
                                    f"🎯 {formato_entero(fila_r['Compromisos'])} compromisos · "
                                    f"**{fila_r['Cumplimiento_gestiones']:.0f}%** de meta"
                                )

                    with r3:
                        with st.container(border=True):
                            st.markdown("#### ⚠️ Días con menor desempeño")
                            for pos, (_, fila_r) in enumerate(
                                peores.iterrows(),
                                start=1,
                            ):
                                fecha_txt = pd.Timestamp(
                                    fila_r["Fecha_dia"]
                                ).strftime("%d/%m/%Y")
                                brecha_dia = int(
                                    fila_r["Gestiones"]
                                    - fila_r["Meta_gestiones"]
                                )
                                st.markdown(
                                    f"**{pos}. {fecha_txt}**  \n"
                                    f"📞 {formato_entero(fila_r['Gestiones'])} gestiones · "
                                    f"🎯 {formato_entero(fila_r['Compromisos'])} compromisos · "
                                    f"brecha **{brecha_dia:+d}**"
                                )

                    # ------------------------------
                    # INSIGHTS AUTOMÁTICOS
                    # ------------------------------
                    st.markdown(
                        '<div class="beh-section">Qué está pasando</div>',
                        unsafe_allow_html=True,
                    )

                    dias_bajos_g = int(
                        (
                            diario_eval["Cumplimiento_gestiones"] < 100
                        ).sum()
                    )
                    dias_bajos_c = int(
                        (
                            diario_eval["Cumplimiento_compromisos"] < 100
                        ).sum()
                    )

                    mejor_fecha_txt = pd.Timestamp(
                        mejor_g["Fecha_dia"]
                    ).strftime("%d/%m/%Y")

                    peor_fila = peores.iloc[0]
                    peor_fecha_txt = pd.Timestamp(
                        peor_fila["Fecha_dia"]
                    ).strftime("%d/%m/%Y")

                    i1, i2, i3 = st.columns(3)

                    with i1:
                        if brecha_g_total >= 0:
                            st.success(
                                f"Gestiones están {formato_entero(abs(brecha_g_total))} "
                                "por encima de la meta acumulada del periodo."
                            )
                        else:
                            st.warning(
                                f"Gestiones están {formato_entero(abs(brecha_g_total))} "
                                "por debajo de la meta acumulada del periodo."
                            )

                    with i2:
                        st.info(
                            f"{dias_bajos_g} de {len(diario_eval)} jornadas cerradas quedaron por debajo de "
                            f"la meta de gestiones y {dias_bajos_c} de {len(diario_eval)} "
                            "por debajo de compromisos."
                        )

                    with i3:
                        st.info(
                            f"Mejor día: {mejor_fecha_txt}. "
                            f"Día que más atención requiere: {peor_fecha_txt}."
                        )

                    # ------------------------------
                    # DETALLE POR FECHA
                    # ------------------------------
                    st.markdown(
                        '<div class="beh-section">Detalle por fecha</div>',
                        unsafe_allow_html=True,
                    )
                    st.markdown(
                        '<div class="beh-section-sub">Lectura diaria con meta, cumplimiento y conversión.</div>',
                        unsafe_allow_html=True,
                    )

                    nombres_dia = {
                        0: "Lunes",
                        1: "Martes",
                        2: "Miércoles",
                        3: "Jueves",
                        4: "Viernes",
                        5: "Sábado",
                        6: "Domingo",
                    }

                    detalle = diario.copy()
                    detalle["Fecha"] = pd.to_datetime(
                        detalle["Fecha_dia"]
                    )
                    detalle["Día"] = (
                        detalle["Fecha"]
                        .dt.weekday
                        .map(nombres_dia)
                    )
                    detalle["Estado día"] = detalle.apply(
                        lambda r: (
                            "En curso"
                            if bool(r["_es_hoy"])
                            else (
                                "Cobertura especial"
                                if bool(r["_cobertura_especial"])
                                else (
                                    "Feriado / no laborable"
                                    if not bool(r["_es_laboral_cfg"])
                                    else "Cerrado"
                                )
                            )
                        ),
                        axis=1,
                    )
                    detalle["Fecha"] = (
                        detalle["Fecha"]
                        .dt.strftime("%d/%m/%Y")
                    )
                    detalle["% Gestiones"] = (
                        detalle["Cumplimiento_gestiones"]
                        / 100
                    )
                    detalle["% Compromisos"] = (
                        detalle["Cumplimiento_compromisos"]
                        / 100
                    )
                    detalle["Conversión"] = (
                        detalle["Conversion"]
                        / 100
                    )

                    detalle["Brecha gestiones"] = (
                        detalle["Gestiones"]
                        - detalle["Meta_gestiones"]
                    )
                    detalle["Brecha compromisos"] = (
                        detalle["Compromisos"]
                        - detalle["Meta_compromisos"]
                    )

                    detalle_tabla = detalle[
                        [
                            "Fecha",
                            "Día",
                            "Estado día",
                            "Gestiones",
                            "Meta_gestiones",
                            "% Gestiones",
                            "Brecha gestiones",
                            "Compromisos",
                            "Meta_compromisos",
                            "% Compromisos",
                            "Brecha compromisos",
                            "Conversión",
                        ]
                    ].rename(
                        columns={
                            "Meta_gestiones": "Meta gestiones",
                            "Meta_compromisos": "Meta compromisos",
                        }
                    )

                    st.dataframe(
                        detalle_tabla.sort_values(
                            "Fecha",
                            ascending=False,
                        ),
                        use_container_width=True,
                        hide_index=True,
                        column_config={
                            "Fecha": st.column_config.TextColumn(
                                "Fecha",
                                width="small",
                            ),
                            "Día": st.column_config.TextColumn(
                                "Día",
                                width="small",
                            ),
                            "Gestiones": st.column_config.NumberColumn(
                                "Gestiones",
                                format="%d",
                            ),
                            "Meta gestiones": st.column_config.NumberColumn(
                                "Meta gestiones",
                                format="%d",
                            ),
                            "% Gestiones": st.column_config.ProgressColumn(
                                "% Gestiones",
                                min_value=0,
                                max_value=1,
                                format="%.0f%%",
                            ),
                            "Brecha gestiones": st.column_config.NumberColumn(
                                "Brecha gestiones",
                                format="%+d",
                            ),
                            "Compromisos": st.column_config.NumberColumn(
                                "Compromisos",
                                format="%d",
                            ),
                            "Meta compromisos": st.column_config.NumberColumn(
                                "Meta compromisos",
                                format="%d",
                            ),
                            "% Compromisos": st.column_config.ProgressColumn(
                                "% Compromisos",
                                min_value=0,
                                max_value=1,
                                format="%.0f%%",
                            ),
                            "Brecha compromisos": st.column_config.NumberColumn(
                                "Brecha compromisos",
                                format="%+d",
                            ),
                            "Conversión": st.column_config.NumberColumn(
                                "Conversión",
                                format="%.1f%%",
                            ),
                        },
                    )

                    # ------------------------------
                    # COMPARACIÓN ENTRE OPERADORES
                    # ------------------------------
                    if operador_sel == "Todos":
                        with st.expander(
                            "👥 Ver comparación acumulada entre operadores",
                            expanded=False,
                        ):
                            comp_op = (
                                filtrado
                                .groupby("_usuario_norm")
                                .agg(
                                    Gestiones=("Fecha_dia", "size"),
                                    Compromisos=(
                                        "_tiene_compromiso",
                                        "sum",
                                    ),
                                    Dias=("Fecha_dia", "nunique"),
                                )
                                .reset_index()
                            )

                            comp_op["Operador"] = comp_op[
                                "_usuario_norm"
                            ].map(
                                {
                                    u: d["nombre"]
                                    for u, d in OPERADORES.items()
                                }
                            )

                            comp_op["Prom. gestiones"] = (
                                comp_op["Gestiones"]
                                / comp_op["Dias"].replace(0, pd.NA)
                            ).fillna(0).round(1)

                            comp_op["Prom. compromisos"] = (
                                comp_op["Compromisos"]
                                / comp_op["Dias"].replace(0, pd.NA)
                            ).fillna(0).round(1)

                            comp_op["Conversión"] = (
                                comp_op["Compromisos"]
                                / comp_op["Gestiones"].replace(0, pd.NA)
                            ).fillna(0)

                            comp_op = controles_ordenamiento(
                                comp_op,
                                [
                                    "Gestiones",
                                    "Compromisos",
                                    "Prom. gestiones",
                                    "Operador",
                                ],
                                key_prefix="comparacion_diaria_final",
                                columna_default="Gestiones",
                                descendente_default=True,
                                etiquetas={
                                    "Gestiones": "Gestiones",
                                    "Compromisos": "Compromisos",
                                    "Prom. gestiones": "Promedio diario",
                                    "Operador": "Operador (A-Z / Z-A)",
                                },
                            )

                            st.dataframe(
                                comp_op[
                                    [
                                        "Operador",
                                        "Gestiones",
                                        "Compromisos",
                                        "Prom. gestiones",
                                        "Prom. compromisos",
                                        "Conversión",
                                    ]
                                ],
                                use_container_width=True,
                                hide_index=True,
                                column_config={
                                    "Conversión": st.column_config.NumberColumn(
                                        "Conversión",
                                        format="%.1f%%",
                                    ),
                                },
                            )


# =========================================================
# MENSAJES DIARIOS
# =========================================================

elif menu == "✉️ Mensajes diarios":

    # V32: el mensaje solo puede usar G/C del mes actual confirmado.
    resultado = sanear_resultado_para_mes_actual_v32(
        st.session_state.resultado_operadores
    )
    resultado = aplicar_corte_gestiones_compromisos_v33(
        resultado,
        fecha_local_actual(),
    )
    st.session_state.resultado_operadores = (
        resultado.copy()
        if resultado is not None
        else None
    )

    jornadas_info = jornadas_configuradas()

    periodo_rec_msg_v27 = periodo_recuperacion_actual()
    carga_promesas_mes_actual_v28 = promesas_es_mes_actual_v32()

    callcenter_mes_disponible_v34 = False
    if (
        st.session_state.get("callcenter_df") is not None
        and not st.session_state.callcenter_df.empty
    ):
        col_fecha_v34 = buscar_columna(
            st.session_state.callcenter_df,
            ["fecha"],
        )
        if col_fecha_v34 is not None:
            fechas_v34 = pd.to_datetime(
                st.session_state.callcenter_df[col_fecha_v34],
                dayfirst=True,
                errors="coerce",
            )
            hoy_v34 = fecha_local_actual()
            callcenter_mes_disponible_v34 = bool(
                (
                    (fechas_v34.dt.year == hoy_v34.year)
                    & (fechas_v34.dt.month == hoy_v34.month)
                ).any()
            )

    if not callcenter_mes_disponible_v34:
        st.warning(
            f"Carga GEN CallCenter de {nombre_mes_es(fecha_local_actual().month)} "
            "para calcular Gestiones y Compromisos acumulados del mes.",
            icon="🔄",
        )

    if periodo_rec_msg_v27["cierre_anterior"]:
        st.info(
            f"Gestiones y Compromisos: {nombre_mes_es(fecha_local_actual().month)} · "
            f"Recuperación: {periodo_rec_msg_v27['etiqueta']} hasta el 05/{fecha_local_actual().month:02d}.",
            icon="📅",
        )

    st.markdown("""
    <style>
    .daily-panel{padding:16px 18px!important;border-radius:14px!important;margin-bottom:12px!important;background:linear-gradient(135deg,#f8fbff,#eef5ff)!important;border:1px solid #dbe7f5!important}
    .daily-panel-title{font-size:23px!important;font-weight:800!important;color:#102a43!important}
    .daily-panel-sub{font-size:12px!important;color:#627d98!important;margin-top:2px!important}
    .data-ready-v86{border:1px solid #dbe5ef;border-radius:13px;padding:12px 14px;margin:6px 0 14px;background:#fff}
    .data-ready-head-v86{font-size:14px;font-weight:800;color:#102a43;margin-bottom:9px}
    .data-grid-v86{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:7px}
    .data-item-v86{background:#f7f9fc;border:1px solid #e7edf4;border-radius:9px;padding:8px 9px}
    .data-label-v86{font-size:9px;font-weight:800;color:#829ab1;text-transform:uppercase;letter-spacing:.35px}
    .data-value-v86{font-size:12px;font-weight:700;color:#243b53;margin-top:2px}
    div[data-testid="stMetric"]{background:#fff;border:1px solid #e4eaf1;padding:8px 10px;border-radius:11px}
    div[data-testid="stMetric"] label{font-size:10px!important;color:#627d98!important}
    div[data-testid="stMetricValue"]{font-size:19px!important;font-weight:800!important}
    div[data-testid="stVerticalBlockBorderWrapper"]{border-color:#e1e8f0!important;border-radius:13px!important}
    div[data-testid="stButton"] button{border-radius:9px!important;font-weight:700!important}
    @media(max-width:900px){.data-grid-v86{grid-template-columns:repeat(2,minmax(0,1fr))}}
    </style>
    """, unsafe_allow_html=True)
    st.markdown("""
    <style>
    /* V88 · Tarjetas con semáforo y mejor jerarquía */
    .op-head-v74{align-items:center!important;margin-bottom:8px!important}
    .op-name-v74{font-size:17px!important;line-height:1.12!important;font-weight:800!important}
    .op-contact-v74{font-size:10px!important;color:#7b8794!important;margin-top:4px!important}
    .status-v74{font-size:10px!important;padding:5px 9px!important;border:1px solid transparent!important}
    .v74-red{background:#fff0f1!important;color:#b42318!important;border-color:#fecaca!important}
    .v74-orange{background:#fff7ed!important;color:#b54708!important;border-color:#fed7aa!important}
    .v74-green{background:#ecfdf3!important;color:#067647!important;border-color:#abefc6!important}
    .v74-gray{background:#f3f4f6!important;color:#667085!important;border-color:#e5e7eb!important}
    .schedule-v74{font-size:10px!important;padding:6px 8px!important;background:#f9fbfd!important}
    .kicker-v74{font-size:10px!important;color:#667085!important;margin:7px 0 6px!important}
    .action-v74{font-size:11px!important;padding:8px 10px!important;margin:8px 0!important}
    .monthly-v74{font-size:10px!important;padding-top:7px!important;margin-top:8px!important}
    .monthly-v74 span:first-child{font-weight:800!important;color:#667085!important}
    [data-testid="stVerticalBlockBorderWrapper"] [data-testid="stMetricValue"]{
        font-size:24px!important;line-height:1.05!important
    }
    [data-testid="stVerticalBlockBorderWrapper"] [data-testid="stMetricLabel"] p{
        font-size:11px!important;font-weight:700!important
    }
    [data-testid="stVerticalBlockBorderWrapper"] [data-testid="stMetricDelta"]{
        font-size:11px!important
    }
    [data-testid="stVerticalBlockBorderWrapper"] [data-testid="stCaptionContainer"] p{
        font-size:10px!important;color:#667085!important
    }
    </style>
    """, unsafe_allow_html=True)



    st.markdown("""
    <style>
    /* V92 · Rediseño de Mensajes diarios */
    .daily-panel{
        padding:14px 18px!important;
        border-radius:14px!important;
        margin:0 0 12px 0!important;
        background:#f7faff!important;
        border:1px solid #dce7f3!important;
        box-shadow:none!important;
    }
    .daily-panel-title{font-size:22px!important}
    .daily-panel-sub{font-size:11px!important}

    .data-ready-v86{
        padding:10px 12px!important;
        margin:0 0 10px 0!important;
        border-radius:12px!important;
        background:#ffffff!important;
    }
    .data-ready-head-v86{
        font-size:12px!important;
        margin-bottom:7px!important;
    }
    .data-item-v86{
        padding:7px 9px!important;
        min-height:48px!important;
    }
    .data-label-v86{font-size:8px!important}
    .data-value-v86{font-size:11px!important}

    .control-v92{
        margin:8px 0 12px;
        border:1px solid #dfe7ef;
        border-radius:13px;
        padding:11px 13px;
        background:#fff;
    }
    .control-head-v92{
        display:flex;
        align-items:center;
        justify-content:space-between;
        gap:12px;
        margin-bottom:9px;
    }
    .control-head-v92 b{
        display:block;
        font-size:14px;
        color:#16324f;
    }
    .control-head-v92 span{
        display:block;
        margin-top:1px;
        font-size:9px;
        color:#829ab1;
    }
    .control-time-v92{
        font-size:18px;
        line-height:1;
        font-weight:800;
        color:#16324f;
    }
    .control-grid-v92{
        display:grid;
        grid-template-columns:repeat(5,minmax(0,1fr));
        gap:7px;
    }
    .control-grid-v92>div{
        background:#f8fafc;
        border:1px solid #edf1f5;
        border-radius:9px;
        padding:7px 9px;
    }
    .control-grid-v92 span{
        display:block;
        font-size:8px;
        color:#829ab1;
        text-transform:uppercase;
        letter-spacing:.25px;
    }
    .control-grid-v92 b{
        display:block;
        font-size:13px;
        margin-top:2px;
        color:#243b53;
    }

    /* Expander usados como herramientas secundarias */
    div[data-testid="stExpander"]{
        border:1px solid #e4eaf1!important;
        border-radius:11px!important;
        background:#fff!important;
    }
    div[data-testid="stExpander"] summary{
        font-weight:700!important;
        font-size:12px!important;
    }

    /* Quitar sensación de formulario gigante */
    div[data-testid="stTextInput"] input,
    div[data-testid="stSelectbox"]>div>div{
        min-height:38px!important;
    }

    @media(max-width:1000px){
        .control-grid-v92{grid-template-columns:repeat(3,minmax(0,1fr))}
    }
    </style>
    """, unsafe_allow_html=True)

    st.markdown("""
    <style>
    /* V93: la información técnica se conserva, pero no ocupa media pantalla */
    .data-ready-v86{display:none!important}
    .control-v92{display:none!important}
    .daily-panel{display:none!important}
    .legend-v71{margin-top:8px!important}
    </style>
    """, unsafe_allow_html=True)


    st.markdown(
        """
        <div style="
            background:linear-gradient(120deg,#102A43,#164A6A);
            border-radius:18px;padding:18px 20px;margin:2px 0 14px;
            box-shadow:0 14px 32px rgba(16,42,67,.13);
        ">
            <div style="
                display:inline-block;padding:4px 8px;border-radius:999px;
                background:rgba(255,255,255,.10);color:#BFD7E8;
                font-size:8px;font-weight:800;letter-spacing:.06em;
                text-transform:uppercase;margin-bottom:8px;
            ">Telegram · Seguimiento operativo</div>
            <div style="font-size:23px;font-weight:850;color:#fff;letter-spacing:-.03em;">
                Mensajes diarios
            </div>
            <div style="font-size:10px;color:#BED0E0;margin-top:3px;">
                Seguimiento individual, mensajes libres y comunicación al grupo.
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    if resultado is None:
        st.warning("Primero carga el reporte de Promesas de Pago.")
    else:
        operadores_db = cargar_operadores_supabase()
        datos_contacto = {}
        if operadores_db is not None and not operadores_db.empty:
            for _, op in operadores_db.iterrows():
                datos_contacto[str(op["usuario"])] = {
                    "correo": str(op.get("correo") or ""),
                    "telefono": str(op.get("telefono") or ""),
                    "telegram_chat_id": str(op.get("telegram_chat_id") or ""),
                    "nombre_mensaje": str(op.get("nombre_mensaje") or op.get("nombre") or ""),
                }

        validacion_v86 = estado_datos_para_seguimiento(resultado)
        icono_v86 = {"verde":"🟢","naranja":"🟠","rojo":"🔴"}[validacion_v86["nivel"]]

        promesas_txt_v86 = "Disponible" if validacion_v86["promesas_ok"] else "No cargado"
        if validacion_v86["promesas_cargado"] is not None:
            promesas_txt_v86 += " · " + validacion_v86["promesas_cargado"].strftime("%H:%M")

        if validacion_v86["call_ok"]:
            corte_v86 = validacion_v86["corte"]
            call_txt_v86 = f"Corte {corte_v86.strftime('%H:%M')}" if corte_v86 else "Cargado"
            if validacion_v86["minutos_call"] is not None:
                call_txt_v86 += f" · hace {validacion_v86['minutos_call']} min"
        else:
            call_txt_v86 = "No cargado"

        operadores_txt_v86 = f"{validacion_v86['usuarios_detectados']}/{CANTIDAD_OPERADORES} identificados"
        rec_txt_v86 = "Disponible" if validacion_v86["recuperacion_ok"] else "No disponible"

        html_v86 = (
            '<div class="data-ready-v86">'
            f'<div class="data-ready-head-v86">{icono_v86} {validacion_v86["titulo"]}</div>'
            '<div class="data-grid-v86">'
            f'<div class="data-item-v86"><div class="data-label-v86">Promesas</div><div class="data-value-v86">{promesas_txt_v86}</div></div>'
            f'<div class="data-item-v86"><div class="data-label-v86">CallCenter</div><div class="data-value-v86">{call_txt_v86}</div></div>'
            f'<div class="data-item-v86"><div class="data-label-v86">Operadores</div><div class="data-value-v86">{operadores_txt_v86}</div></div>'
            f'<div class="data-item-v86"><div class="data-label-v86">Recuperación</div><div class="data-value-v86">{rec_txt_v86}</div></div>'
            '</div></div>'
        )
        st.markdown(html_v86, unsafe_allow_html=True)

        if validacion_v86["razones_bloqueo"]:
            st.error("No se habilitará el envío masivo hasta corregir: " + " · ".join(validacion_v86["razones_bloqueo"]))
        elif validacion_v86["advertencias"]:
            st.warning("Revisión recomendada: " + " · ".join(validacion_v86["advertencias"]))

        # -------------------------------------------------
        # PANEL COMPACTO DE CONTROL — V92
        # -------------------------------------------------
        ahora_seguimiento = ahora_bolivia()
        info_corte = informacion_corte_recomendado(
            ahora_seguimiento
        )

        usuarios_resultado = (
            resultado["Usuario"]
            .astype(str)
            .tolist()
        )

        usuarios_turno_actual = [
            usuario_ctrl
            for usuario_ctrl in usuarios_resultado
            if (
                operador_en_turno(
                    usuario_ctrl,
                    ahora_seguimiento,
                )
                or st.session_state.get(
                    "permitir_envio_fuera_turno",
                    False,
                )
            )
        ]

        recientes_ctrl, listos_ctrl = (
            resumen_frecuencia_seguimiento(
                usuarios_turno_actual,
                ahora_seguimiento,
            )
        )

        proximo_corte_v92 = (
            info_corte["hora"]
            if info_corte["estado"] != "finalizado"
            else "Finalizado"
        )

        completadas_v92 = len(
            [
                d
                for d in jornadas_info["dias"]
                if d < fecha_local_actual()
            ]
        )

        html_control_v92 = (
            '<div class="control-v92">'
            '<div class="control-head-v92">'
            '<div><b>📊 Control del seguimiento</b>'
            '<span>Estado operativo del momento</span></div>'
            f'<div class="control-time-v92">{ahora_seguimiento.strftime("%H:%M")}</div>'
            '</div>'
            '<div class="control-grid-v92">'
            f'<div><span>Próximo corte</span><b>{proximo_corte_v92}</b></div>'
            f'<div><span>En turno</span><b>{len(usuarios_turno_actual)}/{len(resultado)}</b></div>'
            f'<div><span>Seguimiento &lt;60 min</span><b>{len(recientes_ctrl)}</b></div>'
            f'<div><span>Esperado mes</span><b>{formato_porcentaje(jornadas_info["esperado_pct"])}</b></div>'
            f'<div><span>Jornadas</span><b>{completadas_v92}/{jornadas_info["total"]}</b></div>'
            '</div>'
            '</div>'
        )
        st.markdown(
            html_control_v92,
            unsafe_allow_html=True,
        )

        if recientes_ctrl:
            nombres_recientes_v92 = []
            for usuario_rec, minutos_rec in recientes_ctrl:
                nombre_rec = OPERADORES.get(
                    usuario_rec,
                    {},
                ).get(
                    "nombre_mensaje",
                    usuario_rec,
                )
                nombres_recientes_v92.append(
                    f"{nombre_rec} ({minutos_rec} min)"
                )

            with st.expander(
                f"⚠️ {len(recientes_ctrl)} seguimiento(s) reciente(s)",
                expanded=False,
            ):
                st.caption(
                    " · ".join(nombres_recientes_v92)
                    + ". Puedes reenviar si es necesario; GEN Control solo avisa para evitar mensajes demasiado seguidos."
                )


        ahora_v93 = ahora_bolivia()
        corte_v93 = obtener_corte_callcenter(
            st.session_state.get("callcenter_df")
        )

        if corte_v93 is not None and hasattr(
            corte_v93,
            "to_pydatetime",
        ):
            corte_v93 = corte_v93.to_pydatetime()

        corte_txt_v93 = (
            corte_v93.strftime("%H:%M")
            if corte_v93 is not None
            else "--:--"
        )

        usuarios_v93 = resultado["Usuario"].astype(str).tolist()
        en_turno_v93 = sum(
            1
            for usuario_v93 in usuarios_v93
            if operador_en_turno(
                usuario_v93,
                ahora_v93,
            )
        )

        enviados_v93 = sum(
            1
            for usuario_v93 in usuarios_v93
            if envio_ya_realizado_hoy(
                usuario_v93,
                "seguimiento",
            )
        )

        proximo_v93 = informacion_corte_recomendado(
            ahora_v93
        )
        proximo_txt_v93 = (
            proximo_v93.get("hora", "--:--")
            if proximo_v93.get("estado") != "finalizado"
            else "Cierre"
        )

        st.markdown(
            '<div style="display:grid;grid-template-columns:repeat(4,1fr);gap:8px;margin:6px 0 12px;">'
            f'<div style="border:1px solid #e4e9ef;border-radius:10px;padding:9px 11px;"><span style="font-size:8px;color:#8291a3;text-transform:uppercase;">Corte cargado</span><b style="display:block;font-size:16px;color:#1d354d;">{corte_txt_v93}</b></div>'
            f'<div style="border:1px solid #e4e9ef;border-radius:10px;padding:9px 11px;"><span style="font-size:8px;color:#8291a3;text-transform:uppercase;">En turno</span><b style="display:block;font-size:16px;color:#1d354d;">{en_turno_v93}/{len(usuarios_v93)}</b></div>'
            f'<div style="border:1px solid #e4e9ef;border-radius:10px;padding:9px 11px;"><span style="font-size:8px;color:#8291a3;text-transform:uppercase;">Seguimiento hoy</span><b style="display:block;font-size:16px;color:#1d354d;">{enviados_v93}/{len(usuarios_v93)}</b></div>'
            f'<div style="border:1px solid #e4e9ef;border-radius:10px;padding:9px 11px;"><span style="font-size:8px;color:#8291a3;text-transform:uppercase;">Próximo corte</span><b style="display:block;font-size:16px;color:#1d354d;">{proximo_txt_v93}</b></div>'
            '</div>',
            unsafe_allow_html=True,
        )

        st.caption(
            "Los reportes y la hora del último seguimiento quedan guardados en Supabase y se restauran al actualizar la app."
        )

        with st.expander("📣 Grupo y recuperación", expanded=False):
            # -------------------------------------------------
            # V12 · RESUMEN GRUPAL DE GESTIONES Y COMPROMISOS
            # -------------------------------------------------
            st.markdown(
                '<span class="section-chip">RESUMEN OPERATIVO DEL EQUIPO</span>',
                unsafe_allow_html=True,
            )
            st.markdown("### Gestiones y compromisos")
            st.caption(
                "Infografía del corte actual para el grupo, independiente del seguimiento individual."
            )

            call_resumen_v12 = st.session_state.get("callcenter_df")
            tabla_resumen_v12 = preparar_resumen_gestiones_grupo(call_resumen_v12)

            if tabla_resumen_v12.empty:
                st.info("Carga el GEN CallCenter del día para generar el resumen grupal.")
            else:
                total_g_v12 = int(tabla_resumen_v12["Total gestión"].sum())
                total_c_v12 = int(tabla_resumen_v12["Compromisos"].sum())
                total_m_v12 = float(tabla_resumen_v12["Monto comprometido"].sum())

                rg1, rg2, rg3 = st.columns(3)
                rg1.metric("Gestiones del equipo", formato_entero(total_g_v12))
                rg2.metric("Compromisos", formato_entero(total_c_v12))
                rg3.metric("Monto comprometido", formato_usd(total_m_v12))

                mensaje_resumen_v12 = generar_mensaje_resumen_gestiones_grupo(call_resumen_v12)

                try:
                    imagen_resumen_v12 = generar_imagen_resumen_gestiones_grupo(call_resumen_v12)

                    # V13 · La imagen ya no ocupa espacio permanentemente.
                    # El usuario decide cuándo abrir/cerrar la vista previa.
                    if "mostrar_preview_resumen_grupal_v13" not in st.session_state:
                        st.session_state["mostrar_preview_resumen_grupal_v13"] = False

                    pv1, pv2 = st.columns([1, 1])
                    with pv1:
                        texto_preview_v13 = (
                            "🙈 Ocultar vista previa"
                            if st.session_state["mostrar_preview_resumen_grupal_v13"]
                            else "👁️ Visualizar resumen"
                        )
                        if st.button(
                            texto_preview_v13,
                            use_container_width=True,
                            key="toggle_preview_resumen_grupal_v13",
                        ):
                            st.session_state["mostrar_preview_resumen_grupal_v13"] = (
                                not st.session_state["mostrar_preview_resumen_grupal_v13"]
                            )
                            st.rerun()

                    with pv2:
                        st.download_button(
                            "🖼️ Descargar imagen",
                            data=imagen_resumen_v12.getvalue(),
                            file_name=f"resumen_gestiones_compromisos_{fecha_local_actual().isoformat()}.png",
                            mime="image/png",
                            use_container_width=True,
                            key="descargar_resumen_gestiones_v13",
                        )

                    if st.session_state["mostrar_preview_resumen_grupal_v13"]:
                        st.markdown("#### Vista previa antes de enviar")
                        st.caption(
                            "Esta es exactamente la imagen que se enviará al grupo con el corte actual."
                        )
                        st.image(
                            imagen_resumen_v12,
                            use_container_width=True,
                            caption="Resumen de gestiones y compromisos del equipo",
                        )

                        with st.expander(
                            "💬 Ver mensaje que acompañará la imagen",
                            expanded=True,
                        ):
                            st.text_area(
                                "Mensaje resumen",
                                value=mensaje_resumen_v12,
                                height=180,
                                disabled=True,
                                label_visibility="collapsed",
                                key="preview_resumen_gestiones_grupo_v13",
                            )

                    st.markdown("##### Envío")
                    chat_grupo_v12 = obtener_telegram_group_chat_id()
                    if st.button(
                        "📊 Enviar resumen grupal",
                        type="primary",
                        use_container_width=True,
                        disabled=not bool(chat_grupo_v12),
                        key="enviar_resumen_gestiones_compromisos_v13",
                    ):
                        ok_txt_v12, det_txt_v12 = enviar_mensaje_telegram(
                            chat_grupo_v12,
                            mensaje_resumen_v12,
                        )
                        if ok_txt_v12:
                            ok_img_v12, det_img_v12 = enviar_foto_telegram(
                                chat_grupo_v12,
                                imagen_resumen_v12,
                                "📊 Gestiones y compromisos por operador",
                            )
                            if ok_img_v12:
                                st.success("Resumen de gestiones y compromisos enviado al grupo.")
                            else:
                                st.error(f"El texto se envió, pero la imagen falló: {det_img_v12}")
                        else:
                            st.error(f"No se pudo enviar el resumen: {det_txt_v12}")

                    if not obtener_telegram_group_chat_id():
                        st.info("Configura TELEGRAM_GROUP_CHAT_ID para habilitar el envío al grupo.")

                except Exception as e:
                    st.error(f"No se pudo generar el resumen visual: {e}")

            st.divider()

            # -------------------------------------------------
            # AVANCE GENERAL DE RECUPERACIÓN
            # -------------------------------------------------
            meta_individual = float(st.session_state.meta_recuperacion_cfg)
            tabla_general = resultado[["Operador", "Recuperación acumulada", "% Recuperación"]].copy()
            tabla_general["Meta"] = meta_individual
            tabla_general["Falta"] = (meta_individual - tabla_general["Recuperación acumulada"]).clip(lower=0)
            tabla_general = tabla_general.sort_values("% Recuperación", ascending=False, kind="stable").reset_index(drop=True)

            total_recuperacion_equipo = float(tabla_general["Recuperación acumulada"].sum())
            meta_equipo = meta_individual * CANTIDAD_OPERADORES
            pct_equipo = total_recuperacion_equipo / meta_equipo * 100 if meta_equipo else 0
            falta_equipo = max(meta_equipo - total_recuperacion_equipo, 0)

            st.markdown(
                '<span class="section-chip">RECUPERACIÓN DEL EQUIPO</span>',
                unsafe_allow_html=True,
            )
            st.markdown("### Avance general")
            g1, g2, g3 = st.columns(3)
            g1.metric("Recuperación del equipo", formato_usd(total_recuperacion_equipo))
            g2.metric("Cumplimiento", formato_porcentaje(pct_equipo))
            g3.metric("Brecha total", formato_usd(falta_equipo))

            mensaje_general = (
                f"📊 AVANCE DE RECUPERACIÓN – {fecha_local_actual().strftime('%d/%m/%Y')}\n\n"
                f"Buenos días, equipo. Comparto el avance acumulado de recuperación a la fecha, "
                f"considerando una meta mensual de {formato_usd(meta_individual)} por operador.\n\n"
                "Revisemos nuestro porcentaje de cumplimiento y la brecha pendiente. "
                "Mantengamos el enfoque en recuperación para continuar avanzando hacia la meta mensual. 💪"
            )

            with st.expander("Ver mensaje general y tabla de recuperación", expanded=False):
                st.text_area("Mensaje general", value=mensaje_general, height=135, key="mensaje_general_recuperacion_v22", label_visibility="collapsed")
                tabla_compartir = pd.DataFrame({
                    "Operador": tabla_general["Operador"],
                    "Recuperación": tabla_general["Recuperación acumulada"].apply(formato_usd),
                    "Cumplimiento": tabla_general["% Recuperación"].apply(formato_porcentaje),
                    "Falta": tabla_general["Falta"].apply(formato_usd),
                })
                st.dataframe(tabla_compartir, use_container_width=True, hide_index=True)

                imagen_recuperacion = generar_imagen_avance_recuperacion(tabla_general, fecha_local_actual(), meta_individual)
                st.image(imagen_recuperacion, caption="Imagen lista para correo o Telegram", width=760)
                bi1, bi2, bi3 = st.columns(3)
                with bi1:
                    mostrar_boton_copiar_imagen(imagen_recuperacion)
                with bi2:
                    st.download_button("🖼️ Descargar imagen", data=imagen_recuperacion.getvalue(), file_name=f"avance_recuperacion_{fecha_local_actual().isoformat()}.png", mime="image/png", use_container_width=True)
                with bi3:
                    mailto_general = f"mailto:cobranza@gestiona.bo?subject={quote('Avance de recuperación')}&body={quote(mensaje_general)}"
                    st.link_button("✉️ Correo general", mailto_general, use_container_width=True)

            st.divider()

            # -------------------------------------------------
            # ENCABEZADO + CONTROL DEL DÍA — V66
            # -------------------------------------------------
            corte_callcenter_v68 = obtener_corte_callcenter(
                st.session_state.callcenter_df
            )

            ahora_v66 = (
                corte_callcenter_v68.to_pydatetime()
                if hasattr(
                    corte_callcenter_v68,
                    "to_pydatetime",
                )
                else corte_callcenter_v68
            )

            if ahora_v66 is None:
                ahora_v66 = datetime.now(
                    ZoneInfo("America/La_Paz")
                )

            saludo_v66, emoji_v66 = saludo_segun_hora()

            st.markdown(
                textwrap.dedent(f"""
                <div class="hello-v66">
                    <div>
                        <div class="hello-title-v66">
                            {saludo_v66}, José Carlos. {emoji_v66}
                        </div>
                        <div class="hello-sub-v66">
                            {ahora_v66.strftime('%d/%m/%Y')} ·
                            {ahora_v66.strftime('%H:%M')} hrs ·
                            Seguimiento de metas y avance a la hora
                        </div>
                    </div>
                </div>
                """),
                unsafe_allow_html=True,
            )

            top_a1, top_a2, top_a3 = st.columns([5, 1.25, 1.25])

            with top_a2:
                with st.popover(
                    "📄 Mensaje general",
                    use_container_width=True,
                ):
                    st.text_area(
                        "Mensaje general",
                        value=mensaje_general,
                        height=220,
                        disabled=True,
                        label_visibility="collapsed",
                        key="mensaje_general_popover_v66",
                    )

            with top_a3:
                with st.popover(
                    "📊 Tabla recuperación",
                    use_container_width=True,
                ):
                    st.dataframe(
                        tabla_compartir,
                        use_container_width=True,
                        hide_index=True,
                    )

            # -------------------------------------------------
            # KPIs COMPACTOS
            # -------------------------------------------------
            operadores_con_correo = 0
            operadores_con_telegram = 0

            for _, fila_tmp in resultado.iterrows():
                usuario_tmp = fila_tmp["Usuario"]
                contacto_tmp = datos_contacto.get(
                    usuario_tmp,
                    {},
                )

                correo_tmp = (
                    contacto_tmp.get("correo")
                    or str(
                        fila_tmp.get("Correo", "")
                    ).strip()
                )

                telegram_tmp = (
                    normalizar_telegram_chat_id(
                        contacto_tmp.get(
                            "telegram_chat_id",
                            "",
                        )
                    )
                )

                if correo_tmp:
                    operadores_con_correo += 1

                if telegram_tmp:
                    operadores_con_telegram += 1

            promedio_rec_v66 = float(
                resultado["% Recuperación"].mean()
            )

            esperado_resumen = float(
                jornadas_info["esperado_pct"]
            )

            k1, k2, k3, k4, k5, k6 = st.columns(6)

            kpis_v66 = [
                (
                    k1,
                    "kpi-v66-purple",
                    "👥 OPERADORES",
                    str(len(resultado)),
                    "Activos en el reporte",
                ),
                (
                    k2,
                    "kpi-v66-blue",
                    "✉️ CORREOS",
                    str(operadores_con_correo),
                    "Configurados",
                ),
                (
                    k3,
                    "kpi-v66-green",
                    "✈️ TELEGRAM",
                    str(operadores_con_telegram),
                    "Configurados",
                ),
                (
                    k4,
                    "kpi-v66-orange",
                    "🎯 MÍNIMOS",
                    "98 / 25",
                    "Gestiones / compromisos",
                ),
                (
                    k5,
                    "kpi-v66-blue",
                    "📈 ESPERADO MES",
                    formato_porcentaje(esperado_resumen),
                    "Según jornadas transcurridas",
                ),
                (
                    k6,
                    "kpi-v66-purple",
                    "💰 RECUPERACIÓN",
                    formato_porcentaje(promedio_rec_v66),
                    "Promedio mensual",
                ),
            ]

            for columna_kpi, clase_kpi, etiqueta_kpi, valor_kpi, pie_kpi in kpis_v66:
                with columna_kpi:
                    st.markdown(
                        f"""
                        <div class="kpi-v66 {clase_kpi}">
                            <div class="kpi-label-v66">{etiqueta_kpi}</div>
                            <div class="kpi-value-v66">{valor_kpi}</div>
                            <div class="kpi-foot-v66">{pie_kpi}</div>
                        </div>
                        """,
                        unsafe_allow_html=True,
                    )

        # -------------------------------------------------
        # MENSAJE LIBRE POR TELEGRAM — V90
        # -------------------------------------------------
        with st.expander(
            "💬 Escribir mensaje libre",
            expanded=False,
        ):
            st.caption(
                "Escribe un aviso libre y elige a quién enviarlo. No cuenta como seguimiento."
            )

            opciones_manual_v90 = []
            mapa_manual_v90 = {}

            for _, fila_manual_v90 in resultado.iterrows():
                usuario_manual_v90 = str(
                    fila_manual_v90["Usuario"]
                )
                nombre_manual_v90 = str(
                    fila_manual_v90["Operador"]
                )

                chat_manual_v90 = normalizar_telegram_chat_id(
                    datos_contacto.get(
                        usuario_manual_v90,
                        {},
                    ).get(
                        "telegram_chat_id",
                        "",
                    )
                )

                if not chat_manual_v90:
                    continue

                etiqueta_manual_v90 = (
                    f"{nombre_manual_v90} · @{usuario_manual_v90}"
                )

                opciones_manual_v90.append(
                    etiqueta_manual_v90
                )
                mapa_manual_v90[
                    etiqueta_manual_v90
                ] = {
                    "usuario": usuario_manual_v90,
                    "nombre": nombre_manual_v90,
                    "chat_id": chat_manual_v90,
                }

            tipo_destino_manual_v90 = st.radio(
                "Destinatarios",
                [
                    "Un operador",
                    "Varios operadores",
                    "Todos los operadores de turno",
                ],
                horizontal=True,
                key="tipo_destino_manual_v90",
            )

            destinatarios_manual_v90 = []

            if tipo_destino_manual_v90 == "Un operador":
                seleccion_manual_v90 = st.selectbox(
                    "Seleccionar operador",
                    opciones_manual_v90,
                    index=None,
                    placeholder="Elige un operador...",
                    key="operador_manual_v90",
                )

                if seleccion_manual_v90:
                    destinatarios_manual_v90 = [
                        mapa_manual_v90[
                            seleccion_manual_v90
                        ]
                    ]

            elif tipo_destino_manual_v90 == "Varios operadores":
                seleccion_multiple_manual_v90 = st.multiselect(
                    "Seleccionar operadores",
                    opciones_manual_v90,
                    key="operadores_manual_v90",
                )

                destinatarios_manual_v90 = [
                    mapa_manual_v90[x]
                    for x in seleccion_multiple_manual_v90
                ]

            else:
                ahora_manual_v90 = datetime.now(
                    ZoneInfo("America/La_Paz")
                )

                destinatarios_manual_v90 = [
                    datos_manual_v90
                    for datos_manual_v90 in mapa_manual_v90.values()
                    if operador_en_turno(
                        datos_manual_v90["usuario"],
                        ahora_manual_v90,
                    )
                ]

                st.info(
                    f"Se enviará únicamente a quienes estén de turno ahora: "
                    f"{len(destinatarios_manual_v90)} operador(es)."
                )

            mensaje_manual_v90 = st.text_area(
                "Mensaje",
                placeholder=(
                    "Ejemplo: Por favor, prioricemos los compromisos pendientes "
                    "durante este corte. Gracias."
                ),
                height=130,
                max_chars=3500,
                key="texto_manual_telegram_v90",
            )

            if destinatarios_manual_v90:
                nombres_destino_manual_v90 = ", ".join(
                    x["nombre"]
                    for x in destinatarios_manual_v90
                )

                st.caption(
                    f"Destinatarios: {nombres_destino_manual_v90}"
                )

            confirmar_manual_v90 = st.checkbox(
                "Confirmo que revisé el mensaje y los destinatarios.",
                key="confirmar_manual_v90",
            )

            enviar_manual_v90 = st.button(
                f"✈️ Enviar mensaje personalizado ({len(destinatarios_manual_v90)})",
                use_container_width=True,
                type="primary",
                disabled=(
                    not destinatarios_manual_v90
                    or not mensaje_manual_v90.strip()
                    or not confirmar_manual_v90
                ),
                key="enviar_manual_telegram_v90",
            )

            if enviar_manual_v90:
                enviados_manual_v90 = []
                errores_manual_v90 = []

                for destino_manual_v90 in destinatarios_manual_v90:
                    ok_manual_v90, detalle_manual_v90 = enviar_mensaje_telegram(
                        destino_manual_v90["chat_id"],
                        mensaje_manual_v90.strip(),
                    )

                    if ok_manual_v90:
                        enviados_manual_v90.append(
                            destino_manual_v90["nombre"]
                        )

                        registrar_envio_diario(
                            destino_manual_v90["usuario"],
                            destino_manual_v90["nombre"],
                            canal="telegram",
                            tipo="mensaje_manual",
                            detalle=detalle_manual_v90,
                        )

                    else:
                        errores_manual_v90.append(
                            f"{destino_manual_v90['nombre']}: {detalle_manual_v90}"
                        )

                if enviados_manual_v90:
                    st.success(
                        f"Mensaje enviado correctamente a "
                        f"{len(enviados_manual_v90)} operador(es)."
                    )

                if errores_manual_v90:
                    st.warning(
                        "No se pudo enviar a:\n\n- "
                        + "\n- ".join(
                            errores_manual_v90
                        )
                    )

        # -------------------------------------------------
        # FILTROS + ENVÍO MASIVO
        # -------------------------------------------------
        ahora_ui_v20 = ahora_bolivia()
        modo_fuera_final = bool(
            st.session_state.get(
                "permitir_envio_fuera_turno",
                False,
            )
        )

        ui_h1, ui_h2 = st.columns([3.8, 2.2], vertical_alignment="top")
        with ui_h1:
            st.markdown(
                """
                <div class="msg-v20-title">✈️ Mensajes diarios</div>
                <div class="msg-v20-sub">
                    Envía seguimientos diarios a los operadores de cobranzas.
                </div>
                """,
                unsafe_allow_html=True,
            )
        with ui_h2:
            st.markdown(
                f"""
                <div class="msg-v20-top">
                    <div class="msg-v20-chip">
                        <div class="k">Hora actual</div>
                        <div class="v">🕒 {ahora_ui_v20.strftime("%H:%M")}</div>
                    </div>
                    <div class="msg-v20-chip">
                        <div class="k">Fecha</div>
                        <div class="v">📅 {ahora_ui_v20.strftime("%d/%m/%Y")}</div>
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
            )


        # V21 · Resumen rápido de Mensajes diarios
        ahora_kpi_v21 = ahora_bolivia()
        total_ops_v21 = len(OPERADORES)
        en_turno_v21 = sum(
            1 for usuario_v21 in OPERADORES
            if operador_en_turno(usuario_v21, ahora_kpi_v21)
        )
        telegram_ok_v21 = sum(
            1 for usuario_v21 in OPERADORES
            if normalizar_telegram_chat_id(
                datos_contacto.get(usuario_v21, {}).get("telegram_chat_id", "")
            )
        )
        ultimo_corte_v21 = "--:--"
        try:
            corte_v21 = obtener_corte_callcenter(st.session_state.callcenter_df)
            if corte_v21 is not None:
                ultimo_corte_v21 = corte_v21.strftime("%H:%M")
        except Exception:
            pass

        st.markdown(
            f"""
            <div class="v21-kpi-grid">
                <div class="v21-kpi v21-blue">
                    <div class="k">Operadores en turno</div>
                    <div class="v">{en_turno_v21} / {total_ops_v21}</div>
                    <div class="s">Según horario actual</div>
                </div>
                <div class="v21-kpi v21-green">
                    <div class="k">Telegram configurado</div>
                    <div class="v">{telegram_ok_v21} / {total_ops_v21}</div>
                    <div class="s">Listos para recibir seguimiento</div>
                </div>
                <div class="v21-kpi v21-purple">
                    <div class="k">Resumen grupal</div>
                    <div class="v">1</div>
                    <div class="s">Se envía una sola vez al grupo</div>
                </div>
                <div class="v21-kpi v21-orange">
                    <div class="k">Último corte CallCenter</div>
                    <div class="v">{ultimo_corte_v21}</div>
                    <div class="s">Base usada para el avance</div>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

        st.markdown(
            """
            <div class="msg-v20-protect">
                <div class="msg-v20-protect-title">🛡️ Protección de horario activa</div>
                <div class="msg-v20-protect-sub">
                    Los operadores fuera de turno están protegidos. Los mensajes se envían
                    dentro del horario laboral, salvo que habilites una excepción.
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

        pcol1, pcol2 = st.columns([3.8, 1.2], vertical_alignment="center")
        with pcol1:
            permitir_fuera_v98 = st.toggle(
                "Permitir envío fuera de turno",
                value=bool(
                    st.session_state.get(
                        "permitir_envio_fuera_turno",
                        False,
                    )
                ),
                key="permitir_fuera_turno_v98",
                help="Úsalo solo cuando necesites enviar un mensaje después de la jornada.",
            )
            st.session_state.permitir_envio_fuera_turno = permitir_fuera_v98
        with pcol2:
            if permitir_fuera_v98:
                st.warning("Excepción activa", icon="⚠️")
            else:
                st.success("Protección activa", icon="🔒")

        # -------------------------------------------------
        # V24 · FILTROS + ENVÍO MASIVO COMPACTO
        # -------------------------------------------------
        with st.container(border=True):
            st.markdown(
                """
                <div style="margin-bottom:4px;">
                    <div style="font-size:13px;font-weight:850;color:#172B4D;">Filtros y selección</div>
                    <div style="font-size:9px;color:#7A8EA5;">Filtra operadores y define a quién dar seguimiento.</div>
                </div>
                """,
                unsafe_allow_html=True,
            )

            f1, f2, f3, f4 = st.columns([2.0, 1.0, 1.1, 1.1])

            with f1:
                buscar_operador = st.text_input(
                    "Buscar operador",
                    placeholder="Nombre o usuario...",
                    key="buscar_operador_mensajes_v63",
                ).strip()

            with f2:
                filtro_estado = st.selectbox(
                    "Estado",
                    [
                        "Todos",
                        "⚠️ Prioridad",
                        "Reforzar",
                        "Seguimiento",
                        "Buen avance",
                        "Excelente",
                    ],
                    key="filtro_estado_mensajes_v63",
                )

            with f3:
                filtro_canal = st.selectbox(
                    "Telegram",
                    [
                        "Todos",
                        "Configurado",
                        "Pendiente",
                    ],
                    key="filtro_canal_mensajes_v63",
                )

            with f4:
                ordenar_mensajes = st.selectbox(
                    "Ordenar por",
                    [
                        "Prioridad de hoy",
                        "Mayor prioridad",
                        "Nombre A-Z",
                        "Recuperación mayor",
                        "Recuperación menor",
                        "Gestiones mayor",
                        "Compromisos mayor",
                    ],
                    key="orden_mensajes_v63",
                )

        # Preparar destinatarios fuera del panel de filtros para evitar
        # que los avisos y resultados estiren toda la tarjeta.
        telegram_configurados_top = [
            usuario
            for usuario in resultado["Usuario"].tolist()
            if normalizar_telegram_chat_id(
                datos_contacto.get(usuario, {}).get("telegram_chat_id", "")
            )
        ]

        momento_envio_top = datetime.now(
            ZoneInfo("America/La_Paz")
        )

        telegram_en_turno_top = [
            usuario
            for usuario in telegram_configurados_top
            if operador_en_turno(usuario, momento_envio_top)
        ]

        telegram_pendientes_top = [
            usuario
            for usuario in telegram_configurados_top
            if operador_habilitado_para_envio(
                usuario,
                momento_envio_top,
            )
        ]

        telegram_fuera_turno_top = [
            usuario
            for usuario in telegram_configurados_top
            if not operador_en_turno(
                usuario,
                momento_envio_top,
            )
        ]

        recientes_masivo, _ = resumen_frecuencia_seguimiento(
            telegram_pendientes_top,
            momento_envio_top,
        )

        info_envio_col, boton_envio_col = st.columns(
            [4.2, 1.45],
            vertical_alignment="center",
        )

        with info_envio_col:
            aviso_reciente_v24 = (
                f" · ⚠️ {len(recientes_masivo)} con seguimiento hace menos de 60 min"
                if recientes_masivo
                else ""
            )
            st.markdown(
                f"""
                <div style="
                    border:1px solid #D9E7FB;
                    background:linear-gradient(90deg,#F1F6FF,#F8FBFF);
                    border-radius:12px;
                    padding:10px 13px;
                    min-height:42px;
                ">
                    <div style="font-size:12px;font-weight:850;color:#23466F;">
                        👥 {len(telegram_pendientes_top)} operadores listos para seguimiento
                    </div>
                    <div style="font-size:9px;color:#71839A;margin-top:2px;">
                        En turno: {len(telegram_en_turno_top)} ·
                        Fuera de turno: {len(telegram_fuera_turno_top)}
                        {aviso_reciente_v24}
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
            )

        with boton_envio_col:
            enviar_masivo_v24 = st.button(
                f"✈️ Enviar seleccionados ({len(telegram_pendientes_top)})",
                use_container_width=True,
                type="primary",
                disabled=(
                    len(telegram_pendientes_top) == 0
                    or validacion_v86["bloquear_envio"]
                ),
                key="enviar_todos_top_v86",
            )

        if enviar_masivo_v24:
            enviados_top = []
            errores_top = []

            for _, fila_envio in resultado.iterrows():
                usuario_envio = fila_envio["Usuario"]

                chat_id_envio = normalizar_telegram_chat_id(
                    datos_contacto.get(
                        usuario_envio, {}
                    ).get("telegram_chat_id", "")
                )

                if not chat_id_envio:
                    continue

                if not operador_habilitado_para_envio(
                    usuario_envio,
                    momento_envio_top,
                ):
                    continue

                calculo_envio = generar_mensaje_operador_actual(
                    usuario_envio,
                    jornadas_info,
                )

                if calculo_envio is None:
                    errores_top.append(
                        f"{fila_envio['Operador']}: sin datos actuales."
                    )
                    continue

                ok_envio, detalle_envio = enviar_mensaje_telegram(
                    chat_id_envio,
                    calculo_envio["mensaje"],
                )

                if ok_envio:
                    enviados_top.append(
                        fila_envio["Operador"]
                    )

                    registrar_envio_diario(
                        usuario_envio,
                        fila_envio["Operador"],
                        canal="telegram",
                        tipo="seguimiento",
                        detalle=detalle_envio,
                    )

                    enviar_copia_coordinador(
                        fila_envio["Operador"],
                        calculo_envio["mensaje"],
                        detalle_envio,
                    )
                else:
                    errores_top.append(
                        f"{fila_envio['Operador']}: {detalle_envio}"
                    )

            if enviados_top:
                ok_aviso_grupo, detalle_aviso_grupo = (
                    enviar_resumen_grupo_actualizado(
                        st.session_state.callcenter_df
                    )
                )

                r1, r2 = st.columns(2)
                with r1:
                    st.success(
                        f"✅ {len(enviados_top)} mensajes individuales enviados."
                    )
                with r2:
                    if ok_aviso_grupo:
                        st.info(
                            "📊 Resumen enviado al grupo."
                        )
                    else:
                        st.warning(
                            "Resumen grupal pendiente."
                        )

            if errores_top:
                with st.expander(
                    f"⚠️ Ver {len(errores_top)} envío(s) con problema",
                    expanded=False,
                ):
                    st.write("\n".join(f"• {e}" for e in errores_top))

        # -------------------------------------------------
        # FILTRAR Y ORDENAR
        # -------------------------------------------------
        resultado_vista = resultado.copy()

        if buscar_operador:
            termino = normalizar_texto(buscar_operador)

            mascara_busqueda = (
                resultado_vista["Operador"]
                .astype(str)
                .apply(normalizar_texto)
                .str.contains(
                    termino,
                    regex=False,
                )
                |
                resultado_vista["Usuario"]
                .astype(str)
                .apply(normalizar_texto)
                .str.contains(
                    termino,
                    regex=False,
                )
            )

            resultado_vista = resultado_vista[
                mascara_busqueda
            ].copy()

        # Brecha contra el avance esperado a la fecha.
        esperado_v63 = float(jornadas_info["esperado_pct"])

        resultado_vista["_brecha_prioridad"] = resultado_vista.apply(
            lambda r: max(
                esperado_v63 - float(r["% Gestiones"]),
                esperado_v63 - float(r["% Compromisos"]),
                esperado_v63 - float(r["% Recuperación"]),
            ),
            axis=1,
        )

        if ordenar_mensajes == "Prioridad de hoy":
            def prioridad_hoy_sort_v66(row):
                av = calcular_avance_hora_operador(
                    row["Usuario"],
                    st.session_state.callcenter_df,
                )
                if not av.get("disponible"):
                    return 0
                return min(
                    int(av["delta_gestiones"]),
                    int(av["delta_compromisos"]),
                )

            resultado_vista["_prioridad_hoy_v66"] = (
                resultado_vista.apply(
                    prioridad_hoy_sort_v66,
                    axis=1,
                )
            )

            resultado_vista = resultado_vista.sort_values(
                "_prioridad_hoy_v66",
                ascending=True,
                kind="stable",
            )

        elif ordenar_mensajes == "Mayor prioridad":
            resultado_vista = resultado_vista.sort_values(
                "_brecha_prioridad",
                ascending=False,
                kind="stable",
            )
        elif ordenar_mensajes == "Nombre A-Z":
            resultado_vista = resultado_vista.sort_values(
                "Operador",
                ascending=True,
                kind="stable",
            )
        elif ordenar_mensajes == "Recuperación mayor":
            resultado_vista = resultado_vista.sort_values(
                "% Recuperación",
                ascending=False,
                kind="stable",
            )
        elif ordenar_mensajes == "Recuperación menor":
            resultado_vista = resultado_vista.sort_values(
                "% Recuperación",
                ascending=True,
                kind="stable",
            )
        elif ordenar_mensajes == "Gestiones mayor":
            resultado_vista = resultado_vista.sort_values(
                "Gestiones",
                ascending=False,
                kind="stable",
            )
        else:
            resultado_vista = resultado_vista.sort_values(
                "Compromisos",
                ascending=False,
                kind="stable",
            )

        # -------------------------------------------------
        # PREPARAR ESTADO
        # -------------------------------------------------
        filas_preparadas = []

        for _, fila_pre in resultado_vista.iterrows():
            fila_pre = fila_pre.copy()
            usuario_pre = fila_pre["Usuario"]

            contacto_pre = datos_contacto.get(
                usuario_pre,
                {},
            )

            nombre_guardado_pre = str(
                contacto_pre.get(
                    "nombre_mensaje",
                    "",
                )
            ).strip()

            nombre_original_pre = OPERADORES.get(
                usuario_pre,
                {},
            ).get(
                "nombre_mensaje",
                fila_pre["Operador"].split()[0],
            )

            if nombre_guardado_pre:
                OPERADORES[usuario_pre][
                    "nombre_mensaje"
                ] = nombre_guardado_pre

            chat_id_pre = normalizar_telegram_chat_id(
                contacto_pre.get(
                    "telegram_chat_id",
                    "",
                )
            )

            if (
                filtro_canal == "Configurado"
                and not chat_id_pre
            ):
                continue

            if (
                filtro_canal == "Pendiente"
                and chat_id_pre
            ):
                continue

            fila_actual_pre = obtener_fila_operador_actual(
                usuario_pre
            )

            if fila_actual_pre is not None:
                fila_pre = fila_actual_pre

            acumulado_pre_v34 = calcular_acumulado_mes_callcenter_v34(
                usuario_pre,
                st.session_state.get("callcenter_df"),
                fecha_local_actual(),
            )

            if acumulado_pre_v34["disponible"]:
                calculo_pre = generar_mensaje_diario(
                    fila_pre,
                    jornadas_info,
                )
            else:
                saludo_pre, emoji_pre = saludo_segun_hora()
                nombre_pre = OPERADORES.get(
                    usuario_pre,
                    {},
                ).get(
                    "nombre_mensaje",
                    fila_pre["Operador"].split()[0],
                )
                calculo_pre = {
                    "mensaje": (
                        f"{saludo_pre}, {nombre_pre}. {emoji_pre}\n\n"
                        f"📊 Acumulado de {nombre_mes_es(fecha_local_actual().month)}\n"
                        "🔹 Gestiones: pendiente de cargar CallCenter del mes actual\n"
                        "🔹 Compromisos: pendiente de cargar CallCenter del mes actual\n\n"
                        "Carga el GEN CallCenter actualizado para generar el seguimiento."
                    ),
                    "estado_gestiones": "Sin datos",
                    "estado_compromisos": "Sin datos",
                    "estado_recuperacion": "Sin datos",
                    "avance_hora": calcular_avance_hora_operador(
                        usuario_pre,
                        st.session_state.callcenter_df,
                    ),
                }

            OPERADORES[usuario_pre][
                "nombre_mensaje"
            ] = nombre_original_pre

            estados_pre = [
                calculo_pre["estado_gestiones"],
                calculo_pre["estado_compromisos"],
            ]

            if not acumulado_pre_v34["disponible"]:
                estado_pre = "Sin datos"
                clase_pre = "status-gray"
            elif "Reforzar" in estados_pre:
                estado_pre = "Reforzar"
                clase_pre = "status-red"
            elif "En seguimiento" in estados_pre:
                estado_pre = "Seguimiento"
                clase_pre = "status-orange"
            elif "Buen avance" in estados_pre:
                estado_pre = "Buen avance"
                clase_pre = "status-yellow"
            else:
                estado_pre = "Excelente"
                clase_pre = "status-green"

            if filtro_estado == "⚠️ Prioridad":
                esperado_pre = float(
                    jornadas_info["esperado_pct"]
                )
                brecha_max_pre = max(
                    esperado_pre - float(fila_pre["% Gestiones"]),
                    esperado_pre - float(fila_pre["% Compromisos"]),
                )
                if brecha_max_pre < 15:
                    continue
            elif (
                filtro_estado != "Todos"
                and estado_pre != filtro_estado
            ):
                continue

            filas_preparadas.append(
                (
                    fila_pre,
                    calculo_pre,
                    estado_pre,
                    clase_pre,
                )
            )

        if not filas_preparadas:
            st.info(
                "No hay operadores que coincidan con los filtros seleccionados."
            )

        st.caption(
            "Semáforo: 🔴 requiere seguimiento · 🟠 bajo el ritmo esperado · "
            "🟢 en ritmo o adelantado · ⚪ sin datos o fuera del momento de seguimiento."
        )

        # -------------------------------------------------
        # -------------------------------------------------
        # OPERADORES · V20 · TABLA EJECUTIVA + PREVIEW
        # -------------------------------------------------
        elegibles_v20 = [
            item for item in filas_preparadas
            if calcular_acumulado_mes_callcenter_v34(
                item[0]["Usuario"],
                st.session_state.get("callcenter_df"),
                fecha_local_actual(),
            )["disponible"]
            and normalizar_telegram_chat_id(
                datos_contacto.get(item[0]["Usuario"], {}).get("telegram_chat_id", "")
            )
            and operador_habilitado_para_envio(item[0]["Usuario"])
        ]

        st.markdown(
            f"""
            <div class="msg-v20-sendbar">
                <strong>👥 {len(elegibles_v20)} operadores listos para seguimiento</strong><br>
                Se enviará un mensaje individual a cada operador habilitado y una sola actualización general al grupo.
            </div>
            """,
            unsafe_allow_html=True,
        )

        area_ops_v20, area_side_v20 = st.columns([4.7, 1.45], gap="medium")

        with area_ops_v20:
            st.markdown(
                """
                <div class="msg-v20-headrow">
                    <div>Operador</div>
                    <div>Estado</div>
                    <div>Horario laboral</div>
                    <div>Prioridad de hoy</div>
                    <div>Avance de hoy</div>
                    <div>Acción</div>
                </div>
                """,
                unsafe_allow_html=True,
            )

            for fila, calculo, _, _ in filas_preparadas:
                usuario = fila["Usuario"]
                contacto = datos_contacto.get(usuario, {})
                telegram_chat_id = normalizar_telegram_chat_id(
                    contacto.get("telegram_chat_id", "")
                )
                avance = calculo.get("avance_hora", {}) or {}
                horario = avance.get("horario") or {}

                dg = int(avance.get("delta_gestiones", 0)) if avance.get("disponible") else 0
                dc_raw = avance.get("delta_compromisos")
                dc = int(dc_raw) if dc_raw is not None else 0
                estado_jornada = avance.get("estado_jornada", "")

                if not avance.get("disponible"):
                    estado_txt_v20, estado_cls_v20 = "Sin datos", "msg-v20-gray"
                elif estado_jornada == "Jornada aún no iniciada":
                    estado_txt_v20, estado_cls_v20 = "Aún no inicia", "msg-v20-gray"
                elif min(dg, dc) <= -10:
                    estado_txt_v20, estado_cls_v20 = "Seguimiento", "msg-v20-red"
                elif min(dg, dc) < 0:
                    estado_txt_v20, estado_cls_v20 = "Atención", "msg-v20-orange"
                else:
                    estado_txt_v20, estado_cls_v20 = "En turno", "msg-v20-green"

                esperado_g_v20 = max(int(avance.get("esperado_gestiones", 0)), 1)
                esperado_c_v20 = max(int(avance.get("esperado_compromisos", 0)), 1)
                gest_hoy_v20 = int(avance.get("gestiones_hoy", 0))
                comp_hoy_v20 = int(avance.get("compromisos_hoy", 0)) if avance.get("compromisos_disponibles") else 0

                pct_g_v20 = min(100.0, max(0.0, gest_hoy_v20 / esperado_g_v20 * 100))
                pct_c_v20 = min(
                    100.0,
                    max(
                        0.0,
                        (comp_hoy_v20 / esperado_c_v20 * 100)
                        if avance.get("compromisos_disponibles")
                        else 0.0,
                    ),
                )

                turno_txt_v20 = (
                    f"{horario.get('entrada','--:--')}–{horario.get('salida','--:--')}"
                    if horario.get("horario_configurado")
                    else "Sin horario"
                )

                prioridad_g_v20 = (
                    f"{abs(dg)} gestiones por recuperar"
                    if dg < 0 else f"{dg:+d} gestiones vs esperado"
                )
                prioridad_c_v20 = (
                    f"{abs(dc)} compromisos por recuperar"
                    if dc < 0 else f"{dc:+d} compromisos vs esperado"
                )

                en_turno_v20 = operador_en_turno(usuario)
                habilitado_v20 = bool(
                    acumulado_pre_v34["disponible"]
                    and telegram_chat_id
                    and operador_habilitado_para_envio(usuario)
                )

                with st.container(border=True):
                    c1, c2, c3, c4, c5, c6 = st.columns(
                        [2.3, .9, 1.25, 1.65, 1.65, 1.15],
                        vertical_alignment="center",
                    )

                    with c1:
                        iniciales_v20 = "".join(
                            [p[0] for p in str(fila["Operador"]).split()[:2]]
                        ).upper()
                        st.markdown(
                            f"""
                            <div class="msg-v20-opname">◉ {fila['Operador']}</div>
                            <div class="msg-v20-user">@{usuario} · {'Telegram ✓' if telegram_chat_id else 'Telegram pendiente'}</div>
                            """,
                            unsafe_allow_html=True,
                        )

                    with c2:
                        st.markdown(
                            f'<span class="msg-v20-status {estado_cls_v20}">{estado_txt_v20}</span>',
                            unsafe_allow_html=True,
                        )

                    with c3:
                        st.markdown(
                            f"""
                            <div class="msg-v20-small">
                                <span class="msg-v20-strong">{turno_txt_v20}</span><br>
                                {estado_jornada or '—'}
                            </div>
                            """,
                            unsafe_allow_html=True,
                        )

                    with c4:
                        st.markdown(
                            f"""
                            <div class="msg-v20-small">
                                📞 <span class="msg-v20-strong">{prioridad_g_v20}</span><br>
                                🤝 {prioridad_c_v20}
                            </div>
                            """,
                            unsafe_allow_html=True,
                        )

                    with c5:
                        st.markdown(
                            f"""
                            <div class="msg-v20-small">
                                📞 Gestiones: <span class="msg-v20-strong">{gest_hoy_v20} / {esperado_g_v20}</span>
                                <div class="msg-v20-progress"><div class="msg-v20-fill-g" style="width:{pct_g_v20:.0f}%"></div></div>
                                🤝 Compromisos: <span class="msg-v20-strong">{comp_hoy_v20 if avance.get('compromisos_disponibles') else 'Sin dato'} / {esperado_c_v20 if avance.get('compromisos_disponibles') else '—'}</span>
                                <div class="msg-v20-progress"><div class="msg-v20-fill-c" style="width:{pct_c_v20:.0f}%"></div></div>
                            </div>
                            """,
                            unsafe_allow_html=True,
                        )

                    with c6:
                        if habilitado_v20:
                            if st.button(
                                "✈️ Enviar",
                                use_container_width=True,
                                key=f"telegram_v20_{usuario}",
                            ):
                                calculo_actual = generar_mensaje_operador_actual(
                                    usuario,
                                    jornadas_info,
                                )
                                if calculo_actual is None:
                                    st.error("Sin datos actuales.")
                                else:
                                    ok_tg, detalle_tg = enviar_mensaje_telegram(
                                        telegram_chat_id,
                                        calculo_actual["mensaje"],
                                    )
                                    if ok_tg:
                                        registrar_envio_diario(
                                            usuario,
                                            fila["Operador"],
                                            canal="telegram",
                                            tipo="seguimiento",
                                            detalle=detalle_tg,
                                        )
                                        enviar_copia_coordinador(
                                            fila["Operador"],
                                            calculo_actual["mensaje"],
                                            detalle_tg,
                                        )
                                        ok_resumen_auto, det_resumen_auto = (
                                            enviar_resumen_grupo_actualizado(
                                                st.session_state.callcenter_df
                                            )
                                        )
                                        st.success("Enviado")
                                        if not ok_resumen_auto:
                                            st.warning("Resumen grupal pendiente")
                                    else:
                                        st.error("Error de envío")
                        elif not en_turno_v20:
                            st.button(
                                "🔒 Fuera turno",
                                disabled=True,
                                use_container_width=True,
                                key=f"bloqueado_v20_{usuario}",
                            )
                        else:
                            st.button(
                                "Telegram pendiente",
                                disabled=True,
                                use_container_width=True,
                                key=f"pendiente_v20_{usuario}",
                            )

                        with st.popover("👁️ Ver", use_container_width=True):
                            calculo_preview = generar_mensaje_operador_actual(
                                usuario,
                                jornadas_info,
                            )
                            if calculo_preview is not None:
                                mensaje_preview = calculo_preview["mensaje"]
                            else:
                                # `calculo` ya fue saneado arriba en V32.
                                mensaje_preview = calculo.get(
                                    "mensaje",
                                    "Carga el reporte del mes actual para generar el mensaje.",
                                )
                            st.text_area(
                                "Vista previa",
                                value=mensaje_preview,
                                height=240,
                                disabled=True,
                                label_visibility="collapsed",
                                key=f"msg_v20_{usuario}",
                            )

        with area_side_v20:
            st.markdown(
                """
                <div class="msg-v20-side-title">Resumen general del equipo</div>
                <div class="msg-v20-side-sub">
                    Vista previa de la imagen que se enviará al grupo.
                </div>
                """,
                unsafe_allow_html=True,
            )

            try:
                preview_v20 = generar_imagen_resumen_gestiones_grupo(
                    st.session_state.callcenter_df
                )
                st.image(
                    preview_v20.getvalue(),
                    use_container_width=True,
                )
            except Exception as e:
                st.info("Carga CallCenter para generar la vista previa.")

            st.markdown(
                f"""
                <div class="msg-v20-side-ok">
                    <strong>¿Qué se enviará?</strong><br><br>
                    👤 {len(elegibles_v20)} mensajes individuales<br>
                    👥 1 resumen general al grupo
                </div>
                """,
                unsafe_allow_html=True,
            )

            st.markdown(
                """
                <div class="msg-v20-side-info">
                    <strong>Consejo</strong><br><br>
                    Revisa la vista previa antes del seguimiento.
                    El resumen grupal se actualiza con el último CallCenter cargado.
                </div>
                """,
                unsafe_allow_html=True,
            )

            if st.button(
                "📊 Ver resumen completo",
                use_container_width=True,
                key="ver_resumen_completo_v20",
            ):
                st.session_state["mostrar_resumen_completo_v20"] = True

            if st.session_state.get("mostrar_resumen_completo_v20", False):
                with st.expander("Resumen completo", expanded=True):
                    try:
                        preview_full_v20 = generar_imagen_resumen_gestiones_grupo(
                            st.session_state.callcenter_df
                        )
                        st.image(
                            preview_full_v20.getvalue(),
                            use_container_width=True,
                        )
                    except Exception:
                        st.info("No hay vista previa disponible.")

        st.markdown(
            """
            <div class="legend-v71">
                <strong>Estados del día:</strong>
                🔴 Crítico = brecha importante ·
                🟠 Atención = debajo del esperado ·
                🟢 En ritmo = dentro del esperado ·
                🟢 Adelantado = por encima del esperado ·
                ⚪ Aún no inicia = todavía no comenzó su jornada.
            </div>
            """,
            unsafe_allow_html=True,
        )

        with st.expander("📊 Recuperación y herramientas adicionales", expanded=False):
            coordinador_chat_id = obtener_telegram_coordinador_chat_id()

            if coordinador_chat_id:
                st.success(
                    "✅ Copia al coordinador activa. Recibirás una confirmación "
                    "con el mensaje exacto después de cada envío individual exitoso."
                )
            else:
                st.info(
                    "Agrega TELEGRAM_COORDINADOR_CHAT_ID en Streamlit Secrets "
                    "para recibir copia y confirmación de cada envío."
                )


            telegram_group_chat_id = (
                obtener_telegram_group_chat_id()
            )

            mensaje_grupo = (
                generar_mensaje_grupo_recuperacion(
                    tabla_general,
                    meta_individual,
                )
            )


            # -------------------------------------------------
            # VISTA PREVIA TELEGRAM — NO ENVÍA NADA
            # -------------------------------------------------
            st.markdown("#### 👁️ Vista previa antes de enviar")

            if st.button(
                "👁️ Ver cómo llegará al grupo",
                use_container_width=True,
                key="preview_telegram_grupo",
            ):
                st.session_state["mostrar_preview_telegram"] = True

            if st.session_state.get(
                "mostrar_preview_telegram",
                False,
            ):
                with st.container(border=True):
                    st.caption(
                        "Vista previa local. No se ha enviado nada a Telegram."
                    )

                    st.markdown("**Mensaje de texto**")
                    st.text_area(
                        "Vista previa del mensaje",
                        value=mensaje_grupo,
                        height=235,
                        disabled=True,
                        label_visibility="collapsed",
                        key="preview_texto_telegram",
                    )

                    st.markdown("**Imagen del ranking**")

                    try:
                        imagen_preview = (
                            generar_imagen_recuperacion_telegram(
                                tabla_general,
                                meta_individual,
                            )
                        )

                        st.image(
                            imagen_preview,
                            use_container_width=True,
                        )

                    except Exception as e:
                        st.error(
                            f"No se pudo generar la vista previa: {e}"
                        )

                    if st.button(
                        "✖️ Cerrar vista previa",
                        use_container_width=True,
                        key="cerrar_preview_telegram",
                    ):
                        st.session_state[
                            "mostrar_preview_telegram"
                        ] = False
                        st.rerun()

            st.divider()

            tg_col1, tg_col2 = st.columns(2)

            with tg_col1:
                if st.button(
                    "📤 Enviar ahora al grupo",
                    type="primary",
                    use_container_width=True,
                    disabled=(
                        not bool(
                            telegram_group_chat_id
                        )
                    ),
                    key="enviar_resumen_grupo_telegram",
                ):
                    # 1. Enviar texto general.
                    ok_texto, detalle_texto = (
                        enviar_mensaje_telegram(
                            telegram_group_chat_id,
                            mensaje_grupo,
                        )
                    )

                    if not ok_texto:
                        st.error(
                            f"No se pudo enviar el texto al grupo: {detalle_texto}"
                        )
                    else:
                        # 2. Generar y enviar UNA sola imagen del ranking.
                        imagen_grupo = (
                            generar_imagen_recuperacion_telegram(
                                tabla_general,
                                meta_individual,
                            )
                        )

                        ok_grupo, detalle_grupo = (
                            enviar_foto_telegram(
                                telegram_group_chat_id,
                                imagen_grupo,
                                "🏆 Ranking actualizado de recuperación",
                            )
                        )

                        if ok_grupo:
                            st.success(
                                "Texto + ranking en imagen enviados al grupo."
                            )
                        else:
                            st.error(
                                f"El texto se envió, pero la imagen falló: {detalle_grupo}"
                            )


            with tg_col2:
                estado_grupo = (
                    "Configurado"
                    if telegram_group_chat_id
                    else "Pendiente"
                )
                st.metric(
                    "Grupo Telegram",
                    estado_grupo,
                )

            if not telegram_group_chat_id:
                st.info(
                    "Agrega TELEGRAM_GROUP_CHAT_ID en Streamlit Secrets "
                    "para habilitar el envío al grupo."
                )



# =========================================================
# CARGAR REPORTES
# =========================================================

elif menu == "📥 Cargar reportes":

    st.markdown(
        """
        <div class="page-head-v22">
            <div class="page-head-kicker-v22">📥 DATOS · ACTUALIZACIÓN</div>
            <div class="page-head-title-v22">Cargar reportes</div>
            <div class="page-head-sub-v22">
                Ingresa Promesas de Pago y GEN CallCenter. El sistema detecta,
                valida y actualiza automáticamente los datos operativos.
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.markdown(
        """
        <div class="section-head-v22">
            <div>
                <div class="section-title-v22">Archivos del día</div>
                <div class="section-sub-v22">
                    Puedes cargar uno o ambos reportes; GEN Control identificará cada archivo.
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    periodo_carga_v27 = periodo_recuperacion_actual()
    if periodo_carga_v27["cierre_anterior"]:
        st.info(
            f"Del 1 al 5: Gestiones y Compromisos se registran para "
            f"{nombre_mes_es(fecha_local_actual().month)}; Recuperación se mantiene como "
            f"{periodo_carga_v27['etiqueta']}.",
            icon="📅",
        )

    archivos = st.file_uploader(
        "Seleccionar archivos",
        type=[
            "xls",
            "xlsx",
            "csv",
        ],
        accept_multiple_files=True,
    )

    if not archivos:
        st.info(
            "Carga Promesas de Pago y/o GEN CallCenter. "
            "GEN Control detectará automáticamente cada reporte."
        )

    else:
        resumen_cargas = []
        promesas_procesadas = False
        callcenter_procesado = False

        for archivo in archivos:
            try:
                df = leer_archivo(archivo)
                tipo = detectar_tipo_reporte(df)

                resumen_cargas.append(
                    {
                        "Archivo": archivo.name,
                        "Tipo detectado": tipo,
                        "Registros": len(df),
                        "Estado": "Correcto",
                    }
                )

                guardar_carga_supabase(
                    archivo.name,
                    tipo,
                    len(df),
                )

                if tipo == "PROMESAS":
                    (
                        resultado,
                        monto_sin_usuario,
                        distribucion,
                    ) = procesar_promesas(df)

                    # Regla de inicio de mes:
                    # G/C siempre son del archivo nuevo del mes actual.
                    # Recuperación puede seguir cerrando el mes anterior hasta el día 5.
                    periodo_upload_v28 = periodo_recuperacion_actual()
                    cierre_anterior_v28 = st.session_state.get(
                        "cierre_recuperacion_anterior_v28"
                    )

                    if (
                        periodo_upload_v28["cierre_anterior"]
                        and cierre_anterior_v28 is not None
                        and hasattr(cierre_anterior_v28, "empty")
                        and not cierre_anterior_v28.empty
                    ):
                        resultado = combinar_inicio_mes_v28(
                            resultado,
                            cierre_anterior_v28,
                        )

                    st.session_state.resultado_operadores = (
                        resultado
                    )
                    st.session_state.monto_sin_usuario = (
                        monto_sin_usuario
                    )
                    st.session_state.distribucion_sin_usuario = (
                        distribucion
                    )
                    st.session_state.promesas_cargado_en = datetime.now(
                        ZoneInfo("America/La_Paz")
                    )
                    st.session_state.promesas_nombre_archivo = archivo.name
                    st.session_state.promesas_mes_operativo_v32 = int(
                        fecha_local_actual().month
                    )
                    st.session_state.promesas_anio_operativo_v32 = int(
                        fecha_local_actual().year
                    )

                    if supabase_disponible():
                        guardar_snapshot_promesas_v93(
                            resultado,
                            monto_sin_usuario,
                            distribucion,
                            archivo.name,
                        )
                        guardar_resultados_supabase(
                            resultado,
                            fecha_local_actual(),
                            archivo.name,
                        )

                    promesas_procesadas = True

                elif tipo == "CALLCENTER":
                    st.session_state.callcenter_df = df.copy()
                    st.session_state.callcenter_cargado_en = datetime.now(
                        ZoneInfo("America/La_Paz")
                    )
                    st.session_state.callcenter_nombre_archivo = archivo.name

                    if supabase_disponible():
                        guardar_snapshot_callcenter_v93(
                            df,
                            archivo.name,
                        )

                    callcenter_procesado = True

            except Exception as e:
                resumen_cargas.append(
                    {
                        "Archivo": archivo.name,
                        "Tipo detectado": "ERROR",
                        "Registros": 0,
                        "Estado": str(e),
                    }
                )

        st.markdown("### Validación")
        st.dataframe(
            pd.DataFrame(resumen_cargas),
            use_container_width=True,
            hide_index=True,
        )

        if promesas_procesadas:
            st.success(
                "Promesas de Pago procesado correctamente. "
                "Los resultados ya están disponibles en Resumen y Mensajes diarios."
            )

        if callcenter_procesado:
            st.success(
                "GEN CallCenter procesado correctamente. "
                "Revisa Comportamiento diario para ver la evolución."
            )

        st.caption(
            "La pestaña Cargar reportes no repite rankings ni análisis "
            "para evitar duplicar información."
        )


# =========================================================
# HISTÓRICO
# =========================================================

elif menu == "🗂️ Histórico":

    st.markdown(
        """
        <div class="page-head-v22">
            <div class="page-head-kicker-v22">🗂️ HISTÓRICO · SUPABASE</div>
            <div class="page-head-title-v22">Histórico</div>
            <div class="page-head-sub-v22">
                Consulta cierres anteriores, filtra por operador y compara la
                evolución de gestiones, compromisos y recuperación.
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    if not supabase_disponible():
        st.warning(
            "Supabase debe estar conectado para consultar el histórico."
        )

    else:
        try:
            resp = (
                get_supabase()
                .table("resultados_diarios")
                .select("*")
                .order("fecha", desc=True)
                .execute()
            )

            historico = pd.DataFrame(
                resp.data or []
            )

            if historico.empty:
                st.info(
                    "Todavía no hay cierres guardados."
                )

            else:
                historico["fecha"] = pd.to_datetime(
                    historico["fecha"],
                    errors="coerce",
                ).dt.date

                operadores_hist = [
                    "Todos"
                ] + sorted(
                    historico["operador"]
                    .dropna()
                    .unique()
                    .tolist()
                )

                c1, c2 = st.columns(2)

                with c1:
                    operador_hist = st.selectbox(
                        "Operador",
                        operadores_hist,
                        key="hist_op_v14",
                    )

                with c2:
                    metrica_hist = st.selectbox(
                        "Indicador",
                        [
                            "porcentaje_recuperacion",
                            "recuperacion_acumulada",
                            "gestiones",
                            "compromisos",
                        ],
                        format_func=lambda x: {
                            "porcentaje_recuperacion": "% Recuperación",
                            "recuperacion_acumulada": "Recuperación Bs",
                            "gestiones": "Gestiones",
                            "compromisos": "Compromisos",
                        }[x],
                        key="hist_metrica_v14",
                    )

                vista_hist = historico.copy()

                if operador_hist != "Todos":
                    vista_hist = vista_hist[
                        vista_hist["operador"]
                        == operador_hist
                    ].copy()

                serie = (
                    vista_hist
                    .sort_values("fecha")
                    .groupby("fecha")[
                        metrica_hist
                    ]
                    .mean()
                )

                st.markdown("### Evolución histórica")
                st.line_chart(
                    serie,
                    use_container_width=True,
                )

                st.markdown("### Registros guardados")

                vista_hist = controles_ordenamiento(
                    vista_hist,
                    [
                        "fecha",
                        metrica_hist,
                        "operador",
                    ],
                    key_prefix="historico_v14",
                    columna_default="fecha",
                    descendente_default=True,
                    etiquetas={
                        "fecha": "Fecha",
                        metrica_hist: {
                            "porcentaje_recuperacion": "% Recuperación",
                            "recuperacion_acumulada": "Recuperación Bs",
                            "gestiones": "Gestiones",
                            "compromisos": "Compromisos",
                        }[metrica_hist],
                        "operador": "Operador (A-Z / Z-A)",
                    },
                )

                mostrar = vista_hist[
                    [
                        c for c in [
                            "fecha",
                            "operador",
                            "gestiones",
                            "compromisos",
                            "recuperacion_acumulada",
                            "porcentaje_recuperacion",
                            "archivo_origen",
                        ]
                        if c in vista_hist.columns
                    ]
                ].copy()

                if "recuperacion_acumulada" in mostrar:
                    mostrar[
                        "recuperacion_acumulada"
                    ] = mostrar[
                        "recuperacion_acumulada"
                    ].apply(formato_bs)

                if "porcentaje_recuperacion" in mostrar:
                    mostrar[
                        "porcentaje_recuperacion"
                    ] = mostrar[
                        "porcentaje_recuperacion"
                    ].apply(formato_porcentaje)

                st.dataframe(
                    mostrar,
                    use_container_width=True,
                    hide_index=True,
                )

        except Exception as e:
            st.error(
                f"No se pudo consultar el histórico: {e}"
            )


# =========================================================
# EQUIPO
# =========================================================


# =========================================================
# BONOS · PROTOTIPO DE PRUEBA
# =========================================================
elif menu == "💰 Bonos":

    st.markdown(
        """
        <style>
        .bonus-hero{
            background:linear-gradient(120deg,#102A43,#1B5B76);
            border-radius:18px;
            padding:20px 22px;
            margin-bottom:14px;
            color:white;
            box-shadow:0 14px 30px rgba(16,42,67,.12);
        }
        .bonus-hero h2{color:white!important;margin:0!important}
        .bonus-sub{color:#C5D7E5;font-size:10px;margin-top:4px}
        .bonus-card{
            border:1px solid #E4EBF3;
            border-radius:15px;
            background:#fff;
            padding:13px 14px;
            box-shadow:0 7px 20px rgba(16,42,67,.035);
        }
        .bonus-small{
            color:#71849A;font-size:9px;
        }
        .bonus-top-grid{
            display:grid;
            grid-template-columns:repeat(4,minmax(0,1fr));
            gap:10px;
            margin:10px 0 14px;
        }
        .bonus-top-card{
            border:1px solid #E4EBF3;
            background:#fff;
            border-radius:15px;
            padding:13px 14px;
            box-shadow:0 7px 20px rgba(16,42,67,.035);
            min-height:94px;
        }
        .bonus-top-label{
            font-size:8px;
            text-transform:uppercase;
            letter-spacing:.04em;
            color:#71849A;
            font-weight:800;
        }
        .bonus-top-value{
            font-size:23px;
            line-height:1.05;
            color:#102A43;
            font-weight:850;
            margin-top:7px;
        }
        .bonus-top-sub{
            font-size:9px;
            color:#7A8DA1;
            margin-top:5px;
        }
        .bonus-status{
            display:inline-block;
            padding:5px 8px;
            border-radius:999px;
            font-size:8px;
            font-weight:800;
        }
        .bonus-status-ok{
            background:#ECFDF3;
            color:#067647;
            border:1px solid #ABEFC6;
        }
        .bonus-status-warn{
            background:#FFF7ED;
            color:#B54708;
            border:1px solid #FED7AA;
        }
        .bonus-section-card{
            border:1px solid #E4EBF3;
            background:#fff;
            border-radius:16px;
            padding:14px 15px;
            box-shadow:0 7px 20px rgba(16,42,67,.03);
            margin-bottom:12px;
        }
        .bonus-flow{
            display:flex;
            align-items:center;
            justify-content:space-between;
            gap:8px;
            padding:10px 12px;
            border-radius:12px;
            background:#F7FAFC;
            border:1px solid #E7EDF4;
            margin:7px 0 11px;
            flex-wrap:wrap;
        }
        .bonus-flow-item{
            min-width:120px;
        }
        .bonus-flow-label{
            font-size:8px;
            color:#7A8DA1;
            text-transform:uppercase;
            font-weight:800;
        }
        .bonus-flow-value{
            font-size:16px;
            font-weight:850;
            color:#183B5B;
            margin-top:2px;
        }
        .bonus-arrow{
            color:#8AA0B5;
            font-weight:900;
        }
        @media(max-width:1000px){
            .bonus-top-grid{grid-template-columns:repeat(2,minmax(0,1fr));}
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    st.markdown(
        """
        <div class="bonus-hero">
            <h2>💰 Bonos</h2>
            <div class="bonus-sub">
                Simulación mensual con metas ajustables, prorrateo y cálculo ponderado del bono.
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.info(
        "Modo prueba: se cargaron los valores de JULIO 2026 del archivo de Bonos "
        "para validar la lógica antes de automatizarlo con los reportes de GEN Control."
    )

    # -------------------------------------------------
    # Datos de prueba tomados del archivo BONOS CC JULIO 2026
    # -------------------------------------------------
    bonos_julio = {
        "cvaca": {
            "nombre": "Carla Vaca",
            "productividad": 2474,
            "recuperacion": 195048.67,
            "promesas": 591,
            "satisfaccion": 1.00,
            "pecuf": 1.00,
            "pecn": 1.00,
            "meta_productividad": 2350,
        },
        "arodriguez": {
            "nombre": "Alisson Rodriguez",
            "productividad": 2239,
            "recuperacion": 231251.10,
            "promesas": 538,
            "satisfaccion": 1.00,
            "pecuf": 1.00,
            "pecn": 1.00,
            "meta_productividad": 2350,
        },
        "malvarez": {
            "nombre": "Anahir Alvarez",
            "productividad": 2404,
            "recuperacion": 236393.93,
            "promesas": 553,
            "satisfaccion": 1.00,
            "pecuf": 1.00,
            "pecn": 0.80,
            "meta_productividad": 2350,
        },
        "yrivas": {
            "nombre": "Yessica Rivas",
            "productividad": 1375,
            "recuperacion": 265245.54,
            "promesas": 478,
            "satisfaccion": 1.00,
            "pecuf": 0.95,
            "pecn": 1.00,
            "meta_productividad": 1375,
        },
        "yarinez": {
            "nombre": "Yanine Ariñez",
            "productividad": 2402,
            "recuperacion": 202458.85,
            "promesas": 561,
            "satisfaccion": 1.00,
            "pecuf": 1.00,
            "pecn": 0.80,
            "meta_productividad": 2350,
        },
        "projas": {
            "nombre": "Percy Rojas",
            "productividad": 2492,
            "recuperacion": 183702.80,
            "promesas": 553,
            "satisfaccion": 1.00,
            "pecuf": 1.00,
            "pecn": 1.00,
            "meta_productividad": 2350,
        },
        "jborja": {
            "nombre": "James Borja",
            "productividad": 2416,
            "recuperacion": 237951.05,
            "promesas": 555,
            "satisfaccion": 1.00,
            "pecuf": 1.00,
            "pecn": 1.00,
            "meta_productividad": 2350,
        },
        "avargas": {
            "nombre": "Aracely Peña",
            "productividad": 2405,
            "recuperacion": 243309.74,
            "promesas": 615,
            "satisfaccion": 1.00,
            "pecuf": 1.00,
            "pecn": 0.80,
            "meta_productividad": 2350,
        },
    }

    pesos_bono = {
        "Productividad": 0.20,
        "Recuperación": 0.30,
        "Promesas": 0.20,
        "Satisfacción": 0.10,
        "PECUF": 0.10,
        "PECN": 0.10,
    }

    metas_base = {
        "Recuperación": 170400.0,
        "Promesas": 550.0,
        "Satisfacción": 0.80,
        "PECUF": 0.95,
        "PECN": 0.90,
    }

    def cumplimiento_real_bono(alcance, objetivo):
        if objetivo is None or objetivo <= 0:
            return 0.0
        return max(
            float(alcance) / float(objetivo),
            0.0,
        )

    def cumplimiento_bono(alcance, objetivo):
        # Para el puntaje del bono el cumplimiento se topa en 100%,
        # tal como en el archivo original. El cumplimiento real se muestra aparte.
        return min(
            cumplimiento_real_bono(
                alcance,
                objetivo,
            ),
            1.0,
        )

    def monto_bono(puntaje):
        if puntaje >= 0.95:
            return 350
        if puntaje >= 0.90:
            return 200
        if puntaje >= 0.85:
            return 100
        return 0

    # -------------------------------------------------
    # Parámetros de prueba
    # -------------------------------------------------
    cperiodo, cmodo = st.columns([1.5, 1])

    resultado_bonos_actual = st.session_state.get(
        "resultado_operadores"
    )
    hay_datos_actuales_bonos = (
        resultado_bonos_actual is not None
        and not resultado_bonos_actual.empty
    )

    fuentes_bono = []
    if hay_datos_actuales_bonos:
        fuentes_bono.append(
            "Datos actuales de GEN Control"
        )
    fuentes_bono.append(
        "Julio 2026 · ejemplo del archivo"
    )

    with cperiodo:
        fuente_bono = st.selectbox(
            "Fuente de resultados",
            fuentes_bono,
            key="bonos_fuente_resultados_v4",
        )

    with cmodo:
        modo_ajuste_bono = st.selectbox(
            "Tipo de ajuste",
            [
                "Sin ajuste",
                "Prorrateo por horas",
                "Ajuste manual",
            ],
            key="bonos_modo_ajuste_v6",
        )

    st.markdown("### Configuración general del bono")

    cg1, cg2, cg3 = st.columns(3)
    with cg1:
        meta_prod_estandar = st.number_input(
            "Meta estándar · Productividad",
            min_value=0,
            value=2350,
            step=50,
            key="bono_meta_prod_estandar",
        )
    with cg2:
        meta_rec_estandar = st.number_input(
            "Meta estándar · Recuperación (USD)",
            min_value=0.0,
            value=170400.0,
            step=1000.0,
            key="bono_meta_rec_estandar",
        )
    with cg3:
        meta_prom_estandar = st.number_input(
            "Meta estándar · Promesas",
            min_value=0,
            value=550,
            step=10,
            key="bono_meta_prom_estandar",
        )

    st.caption(
        "Pesos del archivo: Productividad 20% · Recuperación 30% · Promesas 20% · "
        "Satisfacción 10% · PECUF 10% · PECN 10%."
    )

    st.divider()

    usuario_bonus = st.selectbox(
        "Operador para simular",
        list(bonos_julio.keys()),
        format_func=lambda u: bonos_julio[u]["nombre"],
        key="bono_operador_prueba",
    )

    base = dict(
        bonos_julio[usuario_bonus]
    )

    # Si se seleccionan datos actuales, Productividad, Recuperación y Promesas
    # provienen de la misma fuente que el Resumen de GEN Control.
    if (
        fuente_bono == "Datos actuales de GEN Control"
        and hay_datos_actuales_bonos
    ):
        fila_actual_bono = resultado_bonos_actual[
            resultado_bonos_actual["Usuario"].astype(str)
            == str(usuario_bonus)
        ]

        if not fila_actual_bono.empty:
            fila_actual_bono = fila_actual_bono.iloc[0]

            base["productividad"] = int(
                float(
                    fila_actual_bono.get(
                        "Gestiones",
                        base["productividad"],
                    )
                    or 0
                )
            )
            base["recuperacion"] = float(
                fila_actual_bono.get(
                    "Recuperación acumulada",
                    base["recuperacion"],
                )
                or 0
            )
            base["promesas"] = int(
                float(
                    fila_actual_bono.get(
                        "Compromisos",
                        base["promesas"],
                    )
                    or 0
                )
            )

            # Calidad proviene de otra fuente.
            # No reutilizar valores de Julio ni asumir 100%.
            base["satisfaccion"] = None
            base["pecuf"] = None
            base["pecn"] = None

    st.markdown(f"### {base['nombre']}")

    if fuente_bono == "Datos actuales de GEN Control":
        st.success(
            "Productividad, Recuperación y Promesas están tomando los mismos "
            "acumulados que el Resumen de GEN Control."
        )
    else:
        st.info(
            "Estás viendo los valores históricos de Julio 2026 usados solo como ejemplo."
        )

    # -------------------------------------------------
    # Ajuste/prorrateo
    # -------------------------------------------------
    st.markdown("#### 1. Meta válida para el bono")

    p1, p2, p3, p4 = st.columns(4)

    with p1:
        horas_planificadas = st.number_input(
            "Horas planificadas del mes",
            min_value=1.0,
            value=168.0,
            step=1.0,
            key=f"bono_horas_plan_{usuario_bonus}",
        )

    with p2:
        horas_fuera_cobranza = st.number_input(
            "Horas fuera de Cobranzas",
            min_value=0.0,
            value=0.0,
            step=1.0,
            key=f"bono_horas_fuera_{usuario_bonus}",
            help="Ej.: apoyo en Atención al Cliente, capacitación u otra función autorizada.",
        )

    disponibilidad = max(
        min(
            (horas_planificadas - horas_fuera_cobranza)
            / horas_planificadas,
            1.0,
        ),
        0.0,
    )

    meta_prod_prorrateada = round(
        meta_prod_estandar * disponibilidad
    )
    meta_rec_prorrateada = round(
        meta_rec_estandar * disponibilidad,
        2,
    )
    meta_prom_prorrateada = round(
        meta_prom_estandar * disponibilidad
    )

    with p3:
        aplicar_prorrateo = (
            modo_ajuste_bono == "Prorrateo por horas"
        )
        if aplicar_prorrateo:
            st.success("Prorrateo por horas activo")
        elif modo_ajuste_bono == "Ajuste manual":
            st.info("Ajuste manual activo")
        else:
            st.caption("Sin ajuste de metas")

    with p4:
        st.metric(
            "Disponibilidad en Cobranzas",
            f"{disponibilidad*100:.1f}%",
        )

    motivo = st.text_input(
        "Motivo del ajuste",
        value="",
        placeholder="Ej.: apoyo en Atención al Cliente · 2 jornadas",
        key=f"bono_motivo_{usuario_bonus}",
    )

    ma1, ma2, ma3 = st.columns(3)

    if modo_ajuste_bono == "Prorrateo por horas":
        meta_prod_default = meta_prod_prorrateada
        meta_rec_default = meta_rec_prorrateada
        meta_prom_default = meta_prom_prorrateada
    elif modo_ajuste_bono == "Ajuste manual":
        meta_prod_default = int(base["meta_productividad"])
        meta_rec_default = meta_rec_estandar
        meta_prom_default = meta_prom_estandar
    else:
        meta_prod_default = int(base["meta_productividad"])
        meta_rec_default = meta_rec_estandar
        meta_prom_default = meta_prom_estandar

    with ma1:
        meta_prod_final = st.number_input(
            "Meta definitiva · Productividad",
            min_value=0,
            value=int(meta_prod_default),
            step=1,
            key=f"bono_meta_prod_final_{usuario_bonus}_{int(aplicar_prorrateo)}",
        )

    with ma2:
        meta_rec_final = st.number_input(
            "Meta definitiva · Recuperación",
            min_value=0.0,
            value=float(meta_rec_default),
            step=100.0,
            key=f"bono_meta_rec_final_{usuario_bonus}_{int(aplicar_prorrateo)}",
        )

    with ma3:
        meta_prom_final = st.number_input(
            "Meta definitiva · Promesas",
            min_value=0,
            value=int(meta_prom_default),
            step=1,
            key=f"bono_meta_prom_final_{usuario_bonus}_{int(aplicar_prorrateo)}",
        )

    ajuste_prod_v6 = int(meta_prod_final) - int(base["meta_productividad"])
    ajuste_rec_v6 = float(meta_rec_final) - float(meta_rec_estandar)
    ajuste_prom_v6 = int(meta_prom_final) - int(meta_prom_estandar)

    st.markdown(
        f"""
        <div class="bonus-flow">
            <div class="bonus-flow-item">
                <div class="bonus-flow-label">Meta original · Productividad</div>
                <div class="bonus-flow-value">{formato_entero(base["meta_productividad"])}</div>
            </div>
            <div class="bonus-arrow">→</div>
            <div class="bonus-flow-item">
                <div class="bonus-flow-label">Ajuste</div>
                <div class="bonus-flow-value">{ajuste_prod_v6:+d}</div>
            </div>
            <div class="bonus-arrow">→</div>
            <div class="bonus-flow-item">
                <div class="bonus-flow-label">Meta válida</div>
                <div class="bonus-flow-value">{formato_entero(meta_prod_final)}</div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # -------------------------------------------------
    # Resultados / manuales
    # -------------------------------------------------
    st.markdown("#### 2. Resultados del mes")

    st.caption(
        "El resultado original se conserva. Si existe una regularización, registra "
        "el ajuste y el sistema calculará el resultado válido para el bono."
    )

    original_prod = int(base["productividad"])
    original_rec = float(base["recuperacion"])
    original_prom = int(base["promesas"])

    r1, r2, r3 = st.columns(3)

    with r1:
        st.metric("Original · Productividad", formato_entero(original_prod))
        ajuste_result_prod = st.number_input(
            "Ajuste · Productividad",
            value=0,
            step=1,
            key=f"bono_ajres_prod_{usuario_bonus}_{fuente_bono}",
        )
        alcance_prod = max(original_prod + int(ajuste_result_prod), 0)
        st.metric("Válido para bono", formato_entero(alcance_prod))

    with r2:
        st.metric("Original · Recuperación", formato_usd(original_rec))
        ajuste_result_rec = st.number_input(
            "Ajuste · Recuperación",
            value=0.0,
            step=100.0,
            key=f"bono_ajres_rec_{usuario_bonus}_{fuente_bono}",
        )
        alcance_rec = max(original_rec + float(ajuste_result_rec), 0.0)
        st.metric("Válido para bono", formato_usd(alcance_rec))

    with r3:
        st.metric("Original · Promesas", formato_entero(original_prom))
        ajuste_result_prom = st.number_input(
            "Ajuste · Promesas",
            value=0,
            step=1,
            key=f"bono_ajres_prom_{usuario_bonus}_{fuente_bono}",
        )
        alcance_prom = max(original_prom + int(ajuste_result_prom), 0)
        st.metric("Válido para bono", formato_entero(alcance_prom))

    hay_ajuste_resultado = (
        int(ajuste_result_prod) != 0
        or abs(float(ajuste_result_rec)) > 0.0001
        or int(ajuste_result_prom) != 0
    )

    motivo_ajuste_resultado = st.text_input(
        "Motivo del ajuste de resultados",
        value="",
        placeholder="Ej.: regularización aprobada / corrección de cierre",
        key=f"bono_motivo_ajres_{usuario_bonus}_{fuente_bono}",
    )

    ajuste_resultado_valido = (
        not hay_ajuste_resultado
        or bool(motivo_ajuste_resultado.strip())
    )

    if hay_ajuste_resultado and not motivo_ajuste_resultado.strip():
        st.warning(
            "Para aplicar el ajuste al bono debes registrar el motivo."
        )

    st.markdown("#### 2. Calidad")
    st.caption(
        "Satisfacción, PECUF y PECN se cargan manualmente desde la fuente de Calidad. "
        "Hasta completar los tres, el bono queda pendiente."
    )

    def parsear_porcentaje_calidad(valor):
        txt = str(valor or "").strip().replace("%", "").replace(",", ".")
        if not txt:
            return None
        try:
            numero = float(txt)
            # Permite escribir 95 o 0.95.
            if numero > 1:
                numero = numero / 100
            return min(max(numero, 0.0), 1.0)
        except Exception:
            return None

    q1, q2, q3 = st.columns(3)

    sat_default = (
        f"{float(base['satisfaccion']) * 100:.0f}"
        if base.get("satisfaccion") is not None
        else ""
    )
    pecuf_default = (
        f"{float(base['pecuf']) * 100:.0f}"
        if base.get("pecuf") is not None
        else ""
    )
    pecn_default = (
        f"{float(base['pecn']) * 100:.0f}"
        if base.get("pecn") is not None
        else ""
    )

    with q1:
        satisf_txt = st.text_input(
            "Satisfacción (%)",
            value=sat_default,
            placeholder="Ej.: 92",
            key=f"bono_sat_txt_{usuario_bonus}_{fuente_bono}",
        )
        satisf = parsear_porcentaje_calidad(satisf_txt)

    with q2:
        pecuf_txt = st.text_input(
            "PECUF (%)",
            value=pecuf_default,
            placeholder="Ej.: 96",
            key=f"bono_pecuf_txt_{usuario_bonus}_{fuente_bono}",
        )
        pecuf = parsear_porcentaje_calidad(pecuf_txt)

    with q3:
        pecn_txt = st.text_input(
            "PECN (%)",
            value=pecn_default,
            placeholder="Ej.: 91",
            key=f"bono_pecn_txt_{usuario_bonus}_{fuente_bono}",
        )
        pecn = parsear_porcentaje_calidad(pecn_txt)

    calidad_completa = all(
        valor is not None
        for valor in [satisf, pecuf, pecn]
    )

    calidad_cargados_v6 = sum(
        valor is not None
        for valor in [satisf, pecuf, pecn]
    )
    if calidad_completa:
        st.success(
            "✅ Calidad completa · 3 de 3 indicadores cargados."
        )
    else:
        st.warning(
            f"⚠️ Calidad pendiente · {calidad_cargados_v6} de 3 indicadores cargados."
        )

    indicadores = [
        ("Productividad", alcance_prod, meta_prod_final),
        ("Recuperación", alcance_rec, meta_rec_final),
        ("Promesas", alcance_prom, meta_prom_final),
    ]

    if calidad_completa:
        indicadores.extend(
            [
                ("Satisfacción", satisf, metas_base["Satisfacción"]),
                ("PECUF", pecuf, metas_base["PECUF"]),
                ("PECN", pecn, metas_base["PECN"]),
            ]
        )

    detalle_bono = []
    puntaje_total = 0.0

    for indicador, alcance, meta in indicadores:
        cumplimiento_real = cumplimiento_real_bono(
            alcance,
            meta,
        )
        cumplimiento = min(
            cumplimiento_real,
            1.0,
        )
        peso = pesos_bono[indicador]
        aporte = cumplimiento * peso
        puntaje_total += aporte

        detalle_bono.append(
            {
                "Indicador": indicador,
                "Meta válida": meta,
                "Alcance": alcance,
                "Cumplimiento real": cumplimiento_real,
                "Cumplimiento bono": cumplimiento,
                "Peso": peso,
                "Aporte": aporte,
            }
        )

    bono_bs = (
        monto_bono(puntaje_total)
        if calidad_completa
        else None
    )

    st.markdown("#### 3. Resultado proyectado")

    calculo_completo_v7 = calidad_completa and ajuste_resultado_valido

    if not calidad_completa:
        estado_bono_v6 = "Pendiente Calidad"
    elif not ajuste_resultado_valido:
        estado_bono_v6 = "Pendiente justificar ajuste"
    else:
        estado_bono_v6 = "Listo para enviar"
    estado_cls_v6 = (
        "bonus-status-ok"
        if calculo_completo_v7
        else "bonus-status-warn"
    )

    if bono_bs is None:
        bono_top_v6 = "Pendiente"
    else:
        bono_top_v6 = f"Bs {bono_bs}"

    if puntaje_total >= 0.95:
        siguiente_bono_txt_v6 = "Máximo alcanzado"
        puntos_faltan_v6 = 0.0
    elif puntaje_total >= 0.90:
        puntos_faltan_v6 = max(95 - puntaje_total * 100, 0)
        siguiente_bono_txt_v6 = (
            f"Faltan {puntos_faltan_v6:.1f} pts para Bs 350"
        )
    elif puntaje_total >= 0.85:
        puntos_faltan_v6 = max(90 - puntaje_total * 100, 0)
        siguiente_bono_txt_v6 = (
            f"Faltan {puntos_faltan_v6:.1f} pts para Bs 200"
        )
    else:
        puntos_faltan_v6 = max(85 - puntaje_total * 100, 0)
        siguiente_bono_txt_v6 = (
            f"Faltan {puntos_faltan_v6:.1f} pts para Bs 100"
        )

    st.markdown(
        f"""
        <div class="bonus-top-grid">
            <div class="bonus-top-card">
                <div class="bonus-top-label">Puntaje</div>
                <div class="bonus-top-value">{puntaje_total*100:.2f}%</div>
                <div class="bonus-top-sub">{siguiente_bono_txt_v6}</div>
            </div>
            <div class="bonus-top-card">
                <div class="bonus-top-label">Bono proyectado</div>
                <div class="bonus-top-value">{bono_top_v6}</div>
                <div class="bonus-top-sub">Según escala vigente</div>
            </div>
            <div class="bonus-top-card">
                <div class="bonus-top-label">Estado</div>
                <div style="margin-top:10px;">
                    <span class="bonus-status {estado_cls_v6}">{estado_bono_v6}</span>
                </div>
                <div class="bonus-top-sub">Calidad y cálculo</div>
            </div>
            <div class="bonus-top-card">
                <div class="bonus-top-label">Meta válida · Productividad</div>
                <div class="bonus-top-value">{formato_entero(meta_prod_final)}</div>
                <div class="bonus-top-sub">
                    Original {formato_entero(base["meta_productividad"])}
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    if calidad_completa:
        st.caption(
            "El cumplimiento real puede superar 100%. Para calcular el bono, "
            "cada indicador aporta como máximo el 100% de su peso."
        )
    else:
        st.warning(
            "Faltan indicadores de Calidad. El resultado todavía es parcial y "
            "no debe enviarse como bono definitivo."
        )


    # -------------------------------------------------
    # DESCARGA EXCEL · RÉPLICA OPERATIVA DEL ARCHIVO
    # -------------------------------------------------
    def generar_excel_bono_v7():
        """
        Genera el archivo a partir de la plantilla ORIGINAL de BONOS CC.
        No rediseña la hoja: conserva colores, combinaciones, bordes,
        anchos, alturas, fórmulas y las demás pestañas del archivo oficial.
        """
        from pathlib import Path
        from openpyxl import load_workbook

        ruta_template = Path(__file__).resolve().parent / "BONOS_TEMPLATE.xlsx"

        if not ruta_template.exists():
            raise FileNotFoundError(
                "Falta BONOS_TEMPLATE.xlsx en el repositorio. "
                "Sube la plantilla original junto a app.py."
            )

        wb_bonos = load_workbook(
            ruta_template,
            data_only=False,
        )

        if "BONOS CC" not in wb_bonos.sheetnames:
            raise ValueError(
                "La plantilla no contiene la hoja BONOS CC."
            )

        ws = wb_bonos["BONOS CC"]

        # Cada operador ocupa un bloque fijo en la plantilla original:
        # objetivo, alcance, porcentaje y aporte.
        bloques = {
            "cvaca": {"numero": 1, "num_col": "C", "nombre": "Carla Vaca", "obj": "G", "alc": "H", "pct": "I", "aporte": "J"},
            "arodriguez": {"numero": 2, "num_col": "M", "nombre": "Alisson Rodriguez", "obj": "Q", "alc": "R", "pct": "S", "aporte": "T"},
            "malvarez": {"numero": 3, "num_col": "V", "nombre": "Anahir Alvarez", "obj": "Z", "alc": "AA", "pct": "AB", "aporte": "AC"},
            "yrivas": {"numero": 4, "num_col": "AE", "nombre": "Yessica Rivas", "obj": "AI", "alc": "AJ", "pct": "AK", "aporte": "AL"},
            "yarinez": {"numero": 5, "num_col": "AO", "nombre": "Yanine Ariñez", "obj": "AS", "alc": "AT", "pct": "AU", "aporte": "AV"},
            "projas": {"numero": 6, "num_col": "AY", "nombre": "Percy Rojas", "obj": "BC", "alc": "BD", "pct": "BE", "aporte": "BF"},
            "jborja": {"numero": 7, "num_col": "BH", "nombre": "James Borja", "obj": "BL", "alc": "BM", "pct": "BN", "aporte": "BO"},
            "avargas": {"numero": 8, "num_col": "BQ", "nombre": "Aracely Peña", "obj": "BU", "alc": "BV", "pct": "BW", "aporte": "BX"},
        }

        filas_indicador = {
            "Productividad": 15,
            "Recuperación": 17,
            "Promesas": 19,
            "Satisfacción": 21,
            "PECUF": 23,
            "PECN": 25,
        }

        pesos_excel = {
            "Productividad": 0.20,
            "Recuperación": 0.30,
            "Promesas": 0.20,
            "Satisfacción": 0.10,
            "PECUF": 0.10,
            "PECN": 0.10,
        }

        metas_estandar_excel = {
            "Productividad": 2350,
            "Recuperación": 170400,
            "Promesas": 550,
            "Satisfacción": 0.80,
            "PECUF": 0.95,
            "PECN": 0.90,
        }

        # Fecha del archivo: se conserva el formato original y solo cambia el periodo.
        fecha_periodo = fecha_local_actual().replace(day=1)
        for celda_fecha in [
            "C2", "M2", "V2", "AE2",
            "AO2", "AY2", "BH2", "BQ2",
        ]:
            ws[celda_fecha] = fecha_periodo

        # Datos operativos actuales del sistema.
        resultado_actual_excel = st.session_state.get(
            "resultado_operadores"
        )

        def dato_operativo(usuario, columna, fallback):
            if (
                resultado_actual_excel is None
                or resultado_actual_excel.empty
                or "Usuario" not in resultado_actual_excel.columns
            ):
                return fallback

            fila = resultado_actual_excel[
                resultado_actual_excel["Usuario"].astype(str)
                == str(usuario)
            ]

            if fila.empty:
                return fallback

            valor = fila.iloc[0].get(columna, fallback)

            try:
                return float(valor)
            except Exception:
                return fallback

        # Cargar los 8 operadores en la misma estructura del archivo original.
        for usuario_x, bloque_x in bloques.items():
            obj_col = bloque_x["obj"]
            alc_col = bloque_x["alc"]
            pct_col = bloque_x["pct"]
            aporte_col = bloque_x["aporte"]
            num_col = bloque_x["num_col"]
            numero_operador = bloque_x["numero"]

            # Corregir numeración original en todos los bloques.
            for fila_num in (15, 17, 19, 21, 23, 25, 29):
                ws[f"{num_col}{fila_num}"] = numero_operador

            datos_ref = bonos_julio.get(
                usuario_x,
                {},
            )

            prod_x = dato_operativo(
                usuario_x,
                "Gestiones",
                datos_ref.get("productividad", 0),
            )
            rec_x = dato_operativo(
                usuario_x,
                "Recuperación acumulada",
                datos_ref.get("recuperacion", 0),
            )
            prom_x = dato_operativo(
                usuario_x,
                "Compromisos",
                datos_ref.get("promesas", 0),
            )

            # Para el operador actualmente editado se usa el resultado válido,
            # incluyendo cualquier regularización realizada en pantalla.
            if usuario_x == usuario_bonus:
                prod_x = alcance_prod
                rec_x = alcance_rec
                prom_x = alcance_prom

            # Calidad: toma lo que se haya cargado en la pestaña Bonos.
            # Si todavía no se cargó para un operador, queda en blanco.
            sat_key = f"bono_sat_txt_{usuario_x}_{fuente_bono}"
            pecuf_key = f"bono_pecuf_txt_{usuario_x}_{fuente_bono}"
            pecn_key = f"bono_pecn_txt_{usuario_x}_{fuente_bono}"

            sat_x = parsear_porcentaje_calidad(
                st.session_state.get(sat_key, "")
            )
            pecuf_x = parsear_porcentaje_calidad(
                st.session_state.get(pecuf_key, "")
            )
            pecn_x = parsear_porcentaje_calidad(
                st.session_state.get(pecn_key, "")
            )

            if usuario_x == usuario_bonus:
                sat_x = satisf
                pecuf_x = pecuf
                pecn_x = pecn

            resultados_x = {
                "Productividad": prod_x,
                "Recuperación": rec_x,
                "Promesas": prom_x,
                "Satisfacción": sat_x,
                "PECUF": pecuf_x,
                "PECN": pecn_x,
            }

            metas_x = dict(metas_estandar_excel)

            if usuario_x == usuario_bonus:
                metas_x["Productividad"] = meta_prod_final
                metas_x["Recuperación"] = meta_rec_final
                metas_x["Promesas"] = meta_prom_final

            for indicador_x, fila_x in filas_indicador.items():
                meta_x = metas_x[indicador_x]
                resultado_x = resultados_x[indicador_x]

                ws[f"{obj_col}{fila_x}"] = meta_x
                ws[f"{alc_col}{fila_x}"] = (
                    resultado_x
                    if resultado_x is not None
                    else None
                )

                # Mantener fórmulas equivalentes a las del archivo original.
                ws[f"{pct_col}{fila_x}"] = (
                    f'=IFERROR(MIN({alc_col}{fila_x}/'
                    f'{obj_col}{fila_x},1),0)'
                )
                ws[f"{aporte_col}{fila_x}"] = (
                    f'={pct_col}{fila_x}*'
                    f'{pesos_excel[indicador_x]}'
                )

        # Fórmula de puntaje y bono, conservando la posición original.
        totales = {
            "cvaca": ("I", "J"),
            "arodriguez": ("S", "T"),
            "malvarez": ("AB", "AC"),
            "yrivas": ("AK", "AL"),
            "yarinez": ("AU", "AV"),
            "projas": ("BE", "BF"),
            "jborja": ("BN", "BO"),
            "avargas": ("BW", "BX"),
        }

        for usuario_x, (col_score, col_bono) in totales.items():
            aporte_col = bloques[usuario_x]["aporte"]

            ws[f"{col_score}29"] = (
                f"={aporte_col}15+{aporte_col}17+"
                f"{aporte_col}19+{aporte_col}21+"
                f"{aporte_col}23+{aporte_col}25"
            )

            # IF anidado para máxima compatibilidad con Excel/LibreOffice.
            ws[f"{col_bono}29"] = (
                f'=IF({col_score}29>=95%,350,'
                f'IF({col_score}29>=90%,200,'
                f'IF({col_score}29>=85%,100,0)))'
            )
            ws[f"{col_bono}30"] = f"={col_bono}29"

        # Totales superiores de la hoja original.
        ws["A2"] = (
            "=SUM(H15,R15,AA15,AJ15,AT15,BD15,BM15,BV15)"
        )
        ws["A4"] = (
            "=SUM(H19,R19,AA19,AJ19,AT19,BD19,BM19,BV19)"
        )
        ws["A6"] = (
            "=SUM(H17,R17,AA17,AJ17,AT17,BD17,BM17,BV17)"
        )

        buffer_excel = BytesIO()
        wb_bonos.save(buffer_excel)
        buffer_excel.seek(0)
        return buffer_excel.getvalue()


    st.markdown("#### Descargar archivo oficial")
    st.caption(
        "Genera el mismo formato de BONOS CC del archivo original, conservando "
        "su diseño, estructura, fórmulas y las demás pestañas del libro."
    )

    try:
        excel_bono_v9 = generar_excel_bono_v7()

        st.download_button(
            "📥 Descargar BONOS CC · formato original",
            data=excel_bono_v9,
            file_name=(
                f"BONOS_{base['nombre'].replace(' ', '_')}_"
                f"{fecha_local_actual().strftime('%Y_%m')}.xlsx"
            ),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key=f"descargar_excel_bono_v9_{usuario_bonus}",
        )

    except Exception as error_excel_bono:
        st.error(
            "No se pudo preparar el Excel de Bonos. "
            f"Detalle técnico: {error_excel_bono}"
        )

    # -------------------------------------------------
    # 4. MENSAJE INDIVIDUAL DEL BONO
    # -------------------------------------------------
    st.markdown("#### 4. Envío individual")

    nombre_mensaje_bono = OPERADORES.get(
        usuario_bonus,
        {},
    ).get(
        "nombre_mensaje",
        base["nombre"].split()[0],
    )

    # Resumen breve de los indicadores principales.
    detalle_msg_bono = {
        fila["Indicador"]: fila
        for fila in detalle_bono
    }

    prod_pct_msg = (
        detalle_msg_bono["Productividad"]["Cumplimiento real"] * 100
    )
    rec_pct_msg = (
        detalle_msg_bono["Recuperación"]["Cumplimiento real"] * 100
    )
    prom_pct_msg = (
        detalle_msg_bono["Promesas"]["Cumplimiento real"] * 100
    )

    mensaje_bono = (
        f"💰 *Resultado de bono · {nombre_mensaje_bono}*\n\n"
        f"Tu evaluación mensual quedó de la siguiente manera:\n\n"
        f"📞 Productividad: {prod_pct_msg:.1f}%\n"
        f"💵 Recuperación: {rec_pct_msg:.1f}%\n"
        f"🎯 Promesas: {prom_pct_msg:.1f}%\n"
        + (
            f"⭐ Satisfacción: {satisf*100:.1f}%\n"
            f"✅ PECUF: {pecuf*100:.1f}%\n"
            f"✅ PECN: {pecn*100:.1f}%\n"
            if calidad_completa
            else "⭐ Calidad: pendiente de cargar\n"
        )
        + (
            f"📊 Puntaje final: *{puntaje_total*100:.2f}%*\n"
            f"🏅 Bono correspondiente: *Bs {bono_bs}*\n"
            if calidad_completa
            else f"📊 Puntaje parcial: *{puntaje_total*100:.2f}%*\n"
        )
    )

    if aplicar_prorrateo:
        mensaje_bono += (
            f"\nLa meta fue ajustada considerando "
            f"{disponibilidad*100:.1f}% de disponibilidad efectiva en Cobranzas."
        )

    if motivo.strip():
        mensaje_bono += (
            f"\nMotivo del ajuste: {motivo.strip()}."
        )
    if hay_ajuste_resultado:
        mensaje_bono += (
            f"\nRegularización de resultados: {motivo_ajuste_resultado.strip()}."
        )

    mensaje_bono += (
        "\n\nAnte cualquier consulta, puedes escribirme directamente."
        "\n👉 [José Carlos](https://t.me/josecarlos_27)"
    )

    with st.container(border=True):
        st.markdown("**Vista previa del mensaje**")
        st.code(
            mensaje_bono,
            language=None,
        )

        col_send_1, col_send_2 = st.columns([1, 1])

        with col_send_1:
            chat_id_bono = normalizar_telegram_chat_id(
                (
                    st.session_state.get(
                        "datos_contacto_operadores",
                        {}
                    ).get(
                        usuario_bonus,
                        {}
                    ).get(
                        "telegram_chat_id",
                        ""
                    )
                    if isinstance(
                        st.session_state.get(
                            "datos_contacto_operadores",
                            {}
                        ),
                        dict,
                    )
                    else ""
                )
            )

            # Fallback correcto: misma fuente que Mensajes diarios.
            if not chat_id_bono:
                try:
                    operadores_bono_db = cargar_operadores_supabase()
                    if (
                        operadores_bono_db is not None
                        and not operadores_bono_db.empty
                    ):
                        fila_bono_tg = operadores_bono_db[
                            operadores_bono_db["usuario"].astype(str)
                            == str(usuario_bonus)
                        ]
                        if not fila_bono_tg.empty:
                            chat_id_bono = normalizar_telegram_chat_id(
                                fila_bono_tg.iloc[0].get(
                                    "telegram_chat_id",
                                    "",
                                )
                            )
                except Exception:
                    chat_id_bono = ""

            if st.button(
                f"✈️ Enviar resultado a {nombre_mensaje_bono}",
                use_container_width=True,
                type="primary",
                disabled=(
                    not bool(chat_id_bono)
                    or not calidad_completa
                    or not ajuste_resultado_valido
                ),
                key=f"enviar_bono_{usuario_bonus}",
            ):
                ok_bono, detalle_envio_bono = enviar_mensaje_telegram(
                    chat_id_bono,
                    mensaje_bono,
                )

                if ok_bono:
                    st.success(
                        f"✅ Resultado del bono enviado a {nombre_mensaje_bono}."
                    )
                else:
                    st.error(
                        f"No se pudo enviar: {detalle_envio_bono}"
                    )

        with col_send_2:
            if not chat_id_bono:
                st.warning(
                    "Este operador todavía no tiene Telegram configurado."
                )
            elif not calidad_completa:
                st.warning(
                    "Telegram configurado · falta completar Calidad para habilitar el envío."
                )
            else:
                st.success(
                    "Telegram configurado · resultado completo listo para enviar."
                )


    df_detalle_bono = pd.DataFrame(
        detalle_bono
    )

    # Columnas visuales en porcentaje real (0–100).
    df_detalle_bono["Cumplimiento real %"] = (
        df_detalle_bono["Cumplimiento real"] * 100
    )
    df_detalle_bono["Cumplimiento bono %"] = (
        df_detalle_bono["Cumplimiento bono"] * 100
    )
    df_detalle_bono["Peso %"] = (
        df_detalle_bono["Peso"] * 100
    )
    df_detalle_bono["Aporte %"] = (
        df_detalle_bono["Aporte"] * 100
    )

    st.dataframe(
        df_detalle_bono[
            [
                "Indicador",
                "Meta válida",
                "Alcance",
                "Cumplimiento real %",
                "Peso %",
                "Aporte %",
            ]
        ],
        use_container_width=True,
        hide_index=True,
        column_config={
            "Meta válida": st.column_config.NumberColumn(
                "Meta válida",
                format="%.2f",
            ),
            "Alcance": st.column_config.NumberColumn(
                "Alcance",
                format="%.2f",
            ),
            "Cumplimiento real %": st.column_config.NumberColumn(
                "Cumplimiento real",
                format="%.1f%%",
            ),
            "Cumplimiento bono %": st.column_config.ProgressColumn(
                "Cumplimiento válido bono",
                min_value=0,
                max_value=100,
                format="%.1f%%",
            ),
            "Peso %": st.column_config.NumberColumn(
                "Peso",
                format="%.0f%%",
            ),
            "Aporte %": st.column_config.NumberColumn(
                "Aporte al puntaje",
                format="%.2f%%",
            ),
        },
    )

    if aplicar_prorrateo:
        st.success(
            f"Prorrateo aplicado: {disponibilidad*100:.1f}% de disponibilidad efectiva "
            f"en Cobranzas. Productividad pasa de {formato_entero(meta_prod_estandar)} "
            f"a {formato_entero(meta_prod_final)}."
        )

    if motivo.strip():
        st.caption(
            f"📝 Motivo registrado para la simulación: {motivo.strip()}"
        )

    st.divider()
    st.markdown("### Control general del mes")
    st.caption(
        "Seguimiento de los 8 operadores antes del cierre definitivo de BONOS CC."
    )

    filas_control_mes = []
    completos_mes = 0

    for idx_r, (usuario_r, datos_r) in enumerate(bonos_julio.items(), start=1):
        sat_r = parsear_porcentaje_calidad(
            st.session_state.get(f"bono_sat_txt_{usuario_r}_{fuente_bono}", "")
        )
        pecuf_r = parsear_porcentaje_calidad(
            st.session_state.get(f"bono_pecuf_txt_{usuario_r}_{fuente_bono}", "")
        )
        pecn_r = parsear_porcentaje_calidad(
            st.session_state.get(f"bono_pecn_txt_{usuario_r}_{fuente_bono}", "")
        )

        calidad_n = sum(v is not None for v in (sat_r, pecuf_r, pecn_r))
        completo_r = calidad_n == 3
        if completo_r:
            completos_mes += 1

        filas_control_mes.append({
            "Nº": idx_r,
            "Operador": datos_r["nombre"],
            "Calidad": f"{calidad_n}/3",
            "Estado": "✅ Completo" if completo_r else "⚠️ Pendiente Calidad",
        })

    c1_mes, c2_mes, c3_mes = st.columns(3)
    c1_mes.metric("Operadores completos", f"{completos_mes}/8")
    c2_mes.metric("Pendientes", 8 - completos_mes)
    c3_mes.metric("Cierre mensual", "✅ Listo" if completos_mes == 8 else "🟠 En revisión")

    if completos_mes == 8:
        st.success(
            "Los 8 operadores tienen Calidad completa. "
            "El BONOS CC consolidado ya está listo para cierre definitivo."
        )
    else:
        st.info(
            f"Faltan {8 - completos_mes} operador(es) por completar Satisfacción, PECUF y PECN. "
            "Puedes seguir descargando el consolidado para revisión."
        )

    st.dataframe(
        pd.DataFrame(filas_control_mes),
        use_container_width=True,
        hide_index=True,
    )


elif menu == "👥 Equipo":

    st.markdown(
        """
        <div class="page-head-v22">
            <div class="page-head-kicker-v22">👥 EQUIPO · OPERADORES</div>
            <div class="page-head-title-v22">Equipo</div>
            <div class="page-head-sub-v22">
                Administra operadores, correos, Telegram, estado y horarios de trabajo.
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    if st.session_state.operador_guardado_ok:
        st.success(
            "Datos del operador actualizados correctamente."
        )
        st.session_state.operador_guardado_ok = False

    operadores_db = cargar_operadores_supabase()

    if operadores_db is None or operadores_db.empty:
        sincronizar_operadores_base()
        operadores_db = cargar_operadores_supabase()

    if operadores_db is None or operadores_db.empty:
        st.warning(
            "No se pudieron cargar los operadores desde Supabase."
        )

    else:
        operadores_db = operadores_db.copy()

        operadores_db = controles_ordenamiento(
            operadores_db,
            [
                "nombre",
                "usuario",
                "correo",
                "telefono",
                "telegram_chat_id",
            ],
            key_prefix="equipo_v24",
            columna_default="nombre",
            descendente_default=False,
        )

        mostrar = operadores_db[
            [
                "nombre",
                "usuario",
                "correo",
                "telefono",
                "telegram_chat_id",
                "activo",
            ]
        ].copy()

        mostrar["telegram_chat_id"] = mostrar[
            "telegram_chat_id"
        ].apply(normalizar_telegram_chat_id)

        mostrar.columns = [
            "Operador",
            "Usuario CRM",
            "Correo",
            "Teléfono",
            "Telegram Chat ID",
            "Activo",
        ]

        st.dataframe(
            mostrar,
            use_container_width=True,
            hide_index=True,
        )

        st.markdown(
            """
            <div class="section-head-v22">
                <div>
                    <div class="section-title-v22">🕒 Horarios operativos</div>
                    <div class="section-sub-v22">Jornadas configuradas por operador y día.</div>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

        horarios_v70 = []
        nombres_dias_v94 = {
            0: "Lunes",
            1: "Martes",
            2: "Miércoles",
            3: "Jueves",
            4: "Viernes",
            5: "Sábado",
        }

        for usuario_h, horarios_dia_h in HORARIOS_OPERADORES.items():
            operador_h = OPERADORES.get(
                usuario_h,
                {},
            )

            for dia_h, datos_h in horarios_dia_h.items():
                break_h = (
                    f"{datos_h['break_inicio']}–{datos_h['break_fin']}"
                    if datos_h.get("break_inicio")
                    and datos_h.get("break_fin")
                    else "Sin break"
                )

                horarios_v70.append(
                    {
                        "Operador": operador_h.get(
                            "nombre",
                            usuario_h,
                        ),
                        "Día": nombres_dias_v94.get(
                            dia_h,
                            str(dia_h),
                        ),
                        "Entrada": datos_h["entrada"],
                        "Break": break_h,
                        "Salida": datos_h["salida"],
                    }
                )

        st.dataframe(
            pd.DataFrame(
                horarios_v70
            ),
            use_container_width=True,
            hide_index=True,
        )

        st.caption(
            "El esperado a la hora se calcula con el horario individual. "
            "Durante el break el esperado se congela."
        )

        st.caption(
            "Los cambios guardados se reflejan inmediatamente "
            "en esta tabla y en Mensajes diarios. Los datos editados "
            "ya no se sobrescriben al recargar la aplicación."
        )

        st.divider()

        st.markdown("### ✈️ Detectar IDs de Telegram")
        st.caption(
            "Cada operador debe abrir el bot y enviar /start una sola vez. "
            "Luego pulsa el botón para ver su Chat ID. Esto no envía mensajes."
        )

        if st.button(
            "🔎 Buscar operadores que escribieron al bot",
            use_container_width=True,
            key="buscar_ids_telegram",
        ):
            usuarios_tg, error_tg = obtener_usuarios_telegram_bot()

            if error_tg:
                st.info(error_tg)

            if usuarios_tg:
                st.success(
                    f"Se encontraron {len(usuarios_tg)} chats privados."
                )

                tabla_ids = pd.DataFrame(
                    [
                        {
                            "Nombre Telegram": u["nombre"],
                            "Usuario Telegram": u["username"],
                            "Chat ID": u["chat_id"],
                        }
                        for u in usuarios_tg
                    ]
                )

                st.dataframe(
                    tabla_ids,
                    use_container_width=True,
                    hide_index=True,
                )

                st.session_state[
                    "usuarios_telegram_detectados"
                ] = usuarios_tg

        st.divider()
        st.markdown("### Editar operador")

        seleccion = st.selectbox(
            "Selecciona un operador",
            operadores_db["usuario"].tolist(),
            format_func=lambda u: operadores_db.loc[
                operadores_db["usuario"] == u,
                "nombre",
            ].iloc[0],
        )

        fila_op = operadores_db[
            operadores_db["usuario"] == seleccion
        ].iloc[0]

        c1, c2, c3 = st.columns(3)

        with c1:
            nombre_op = st.text_input(
                "Nombre",
                value=str(
                    fila_op.get("nombre") or ""
                ),
            )

            nombre_mensaje_op = st.text_input(
                "Nombre corto para mensajes",
                value=str(
                    fila_op.get("nombre_mensaje") or ""
                ),
            )

        with c2:
            correo_op = st.text_input(
                "Correo corporativo",
                value=str(
                    fila_op.get("correo") or ""
                ),
            )

            telefono_op = st.text_input(
                "Teléfono",
                value=str(
                    fila_op.get("telefono") or ""
                ),
            )

        with c3:
            telegram_chat_id_op = st.text_input(
                "Telegram Chat ID",
                value=normalizar_telegram_chat_id(
                    fila_op.get("telegram_chat_id")
                ),
                help=(
                    "El operador debe iniciar primero el bot. "
                    "Después guarda aquí su Chat ID numérico."
                ),
            )

            ok_bot, detalle_bot = probar_conexion_telegram()

            if ok_bot:
                st.success(detalle_bot)
            else:
                st.caption(
                    f"Telegram: {detalle_bot}"
                )

        activo_op = st.checkbox(
            "Operador activo",
            value=bool(
                fila_op.get("activo", True)
            ),
        )

        if st.button(
            "💾 Guardar operador",
            type="primary",
        ):
            telefono_normalizado = (
                normalizar_telefono_whatsapp(
                    telefono_op
                )
            )

            payload = {
                "usuario": seleccion,
                "nombre": nombre_op.strip(),
                "nombre_mensaje": nombre_mensaje_op.strip(),
                "correo": correo_op.strip(),
                "telefono": telefono_normalizado,
                "telegram_chat_id": normalizar_telegram_chat_id(
                    telegram_chat_id_op
                ),
                "activo": bool(activo_op),
                "updated_at": datetime.now(
                    ZoneInfo("America/La_Paz")
                ).isoformat(),
            }

            ok, mensaje = guardar_operador_supabase(
                payload
            )

            if ok:
                st.session_state.operadores_supabase_cargados = False
                st.session_state.operador_guardado_ok = True
                st.rerun()
            else:
                st.error(
                    f"No se pudo guardar: {mensaje}"
                )


# =========================================================
# CONFIGURACIÓN
# =========================================================

elif menu == "⚙️ Configuración":

    st.markdown(
        """
        <div class="page-head-v22">
            <div class="page-head-kicker-v22">⚙️ ADMINISTRACIÓN</div>
            <div class="page-head-title-v22">Configuración y metas</div>
            <div class="page-head-sub-v22">
                Define metas, calendario, parámetros operativos y conexión de persistencia.
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    supa_ok, supa_msg = diagnostico_supabase()

    if supa_ok:
        st.success(
            "● Supabase conectado · cambios e históricos se guardan permanentemente."
        )
    else:
        st.error(
            "● Supabase no conectado."
        )
        st.caption(
            f"Diagnóstico: {supa_msg}"
        )
        st.info(
            "GEN Control seguirá funcionando en modo temporal "
            "hasta corregir la conexión."
        )

    if st.button(
        "🔄 Reintentar conexión Supabase",
        key="retry_supabase",
    ):
        get_supabase.clear()
        st.session_state["_supabase_error"] = ""
        st.rerun()

    # -----------------------------------------------------
    # ACCESO SOLO COORDINADOR
    # -----------------------------------------------------

    if not st.session_state.config_desbloqueada:

        st.info(
            "La configuración está protegida. "
            "Ingresa la clave de administrador para modificarla."
        )

        clave = st.text_input(
            "Clave de administrador",
            type="password",
            key="clave_admin_input",
        )

        if st.button(
            "Desbloquear configuración",
            type="primary",
        ):
            try:
                clave_correcta = st.secrets[
                    "ADMIN_PASSWORD"
                ]
            except Exception:
                clave_correcta = "GEN2026"

            if clave == clave_correcta:
                st.session_state.config_desbloqueada = True
                st.success(
                    "Configuración desbloqueada."
                )
                st.rerun()
            else:
                st.error(
                    "Clave incorrecta."
                )

    else:

        col_logout, _ = st.columns([1, 4])

        with col_logout:
            if st.button("🔒 Bloquear"):
                st.session_state.config_desbloqueada = False
                st.rerun()

        # -------------------------------------------------
        # METAS
        # -------------------------------------------------

        st.markdown("### Metas mensuales")

        c1, c2, c3 = st.columns(3)

        with c1:
            nueva_meta_g = st.number_input(
                "Meta mensual de gestiones",
                min_value=1,
                value=int(
                    st.session_state.meta_gestiones_cfg
                ),
                step=50,
            )

        with c2:
            nueva_meta_c = st.number_input(
                "Meta mensual de compromisos",
                min_value=1,
                value=int(
                    st.session_state.meta_compromisos_cfg
                ),
                step=10,
            )

        with c4:
            st.markdown(
                """
                <div style="border:1px solid #DDE6F0;border-radius:10px;padding:10px 12px;background:#F8FAFC;min-height:58px;">
                    <div style="font-size:9px;color:#71849A;font-weight:800;">META DIARIA DE COMPROMISOS</div>
                    <div style="font-size:12px;color:#183B5B;font-weight:850;margin-top:5px;">Automática según calendario</div>
                </div>
                """,
                unsafe_allow_html=True,
            )

        if st.button(
            "💾 Guardar metas",
            type="primary",
        ):
            st.session_state.meta_gestiones_cfg = int(
                nueva_meta_g
            )
            st.session_state.meta_compromisos_cfg = int(
                nueva_meta_c
            )
            st.session_state.meta_recuperacion_cfg = float(
                nueva_meta_r
            )
            st.session_state.meta_diaria_compromisos_cfg = int(
                nueva_meta_diaria_c
            )

            if supabase_disponible():
                ok, mensaje = guardar_configuracion_supabase()

                if ok:
                    st.success(
                        "Metas actualizadas y guardadas permanentemente."
                    )
                else:
                    st.warning(
                        f"Metas actualizadas en esta sesión. "
                        f"No se pudo guardar en Supabase: {mensaje}"
                    )
            else:
                st.success(
                    "Metas actualizadas para esta sesión."
                )

        st.divider()

        # -------------------------------------------------
        # CALENDARIO OPERATIVO
        # -------------------------------------------------

        st.markdown("### 📅 Calendario operativo")
        st.caption(
            "Marca los días que realmente cuentan como jornada laboral. "
            "Esto modifica automáticamente la meta diaria necesaria."
        )

        hoy = fecha_local_actual()

        meses = {
            1: "Enero",
            2: "Febrero",
            3: "Marzo",
            4: "Abril",
            5: "Mayo",
            6: "Junio",
            7: "Julio",
            8: "Agosto",
            9: "Septiembre",
            10: "Octubre",
            11: "Noviembre",
            12: "Diciembre",
        }

        col_mes, col_anio = st.columns(2)

        with col_mes:
            mes_sel = st.selectbox(
                "Mes",
                options=list(meses.keys()),
                format_func=lambda m: meses[m],
                index=hoy.month - 1,
            )

        with col_anio:
            anio_sel = st.number_input(
                "Año",
                min_value=2024,
                max_value=2035,
                value=hoy.year,
                step=1,
            )

        fecha_mes = date(
            int(anio_sel),
            int(mes_sel),
            1,
        )

        clave_mes = inicializar_calendario_mes(
            fecha_mes
        )

        calendario_mes = st.session_state.calendario_laboral[
            clave_mes
        ]

        if supabase_disponible():
            calendario_guardado = cargar_calendario_supabase(
                int(anio_sel),
                int(mes_sel),
            )

            if calendario_guardado:
                calendario_mes.update(
                    calendario_guardado
                )

        dias_mes = calendar.monthrange(
            int(anio_sel),
            int(mes_sel),
        )[1]

        nombres_dia = [
            "Lun",
            "Mar",
            "Mié",
            "Jue",
            "Vie",
            "Sáb",
            "Dom",
        ]

        st.markdown(
            "**Calendario laboral del mes**"
        )
        st.caption(
            "Activa únicamente los días que cuentan como jornada laboral. "
            "Los domingos vienen desmarcados por defecto."
        )

        encabezado = st.columns(7)
        for idx, nombre_dia in enumerate(
            ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]
        ):
            with encabezado[idx]:
                st.markdown(
                    f"<div style='text-align:center;font-weight:700;'>{nombre_dia}</div>",
                    unsafe_allow_html=True,
                )

        cal = calendar.Calendar(firstweekday=0)
        semanas = cal.monthdayscalendar(
            int(anio_sel),
            int(mes_sel),
        )

        for semana in semanas:
            columnas = st.columns(7)

            for idx, dia in enumerate(semana):
                with columnas[idx]:
                    if dia == 0:
                        st.markdown("&nbsp;", unsafe_allow_html=True)
                        continue

                    fecha_dia = date(
                        int(anio_sel),
                        int(mes_sel),
                        dia,
                    )

                    es_hoy = fecha_dia == fecha_local_actual()
                    etiqueta = (
                        f"⭐ {dia}"
                        if es_hoy
                        else str(dia)
                    )

                    valor = st.checkbox(
                        etiqueta,
                        value=calendario_mes.get(
                            dia,
                            fecha_dia.weekday() != 6,
                        ),
                        key=(
                            f"cal_{anio_sel}_{mes_sel}_{dia}"
                        ),
                        help=(
                            fecha_dia.strftime("%d/%m/%Y")
                        ),
                    )

                    calendario_mes[dia] = valor

        if st.button(
            "💾 Guardar calendario laboral",
            type="primary",
        ):
            if supabase_disponible():
                ok, mensaje = guardar_calendario_supabase(
                    int(anio_sel),
                    int(mes_sel),
                    calendario_mes,
                )

                if ok:
                    st.success(
                        "Calendario laboral guardado permanentemente."
                    )
                else:
                    st.error(
                        f"No se pudo guardar el calendario: {mensaje}"
                    )
            else:
                st.success(
                    "Calendario actualizado para esta sesión."
                )

        # -------------------------------------------------
        # CÁLCULO DE META DIARIA
        # -------------------------------------------------

        st.divider()
        st.markdown("### Meta diaria necesaria")

        fecha_calculo = (
            hoy
            if (
                int(anio_sel) == hoy.year
                and int(mes_sel) == hoy.month
            )
            else fecha_mes
        )

        jornadas = jornadas_configuradas(
            fecha_calculo
        )

        resultado = st.session_state.resultado_operadores

        hoy_es_laboral = (
            hoy in jornadas["dias"]
        )

        jornadas_antes_hoy = len(
            [d for d in jornadas["dias"] if d < hoy]
        )

        jornadas_despues_hoy = len(
            [d for d in jornadas["dias"] if d > hoy]
        )

        c1, c2, c3, c4 = st.columns(4)

        with c1:
            st.metric(
                "Jornadas del mes",
                jornadas["total"],
            )

        with c2:
            st.metric(
                "Completadas antes de hoy",
                jornadas_antes_hoy,
            )

        with c3:
            st.metric(
                "Hoy",
                "Laboral" if hoy_es_laboral else "No laboral",
            )

        with c4:
            st.metric(
                "Pendientes después de hoy",
                jornadas_despues_hoy,
            )

        st.info(
            f"Para calcular la meta de cierre de hoy se consideran "
            f"{jornadas['disponibles']} jornadas disponibles contando hoy. "
            f"Después de hoy quedan {jornadas_despues_hoy}."
        )

        if resultado is not None and not resultado.empty:

            promedio_g = float(
                resultado["Gestiones"].mean()
            )
            promedio_c = float(
                resultado["Compromisos"].mean()
            )
            promedio_r = float(
                resultado["Recuperación acumulada"].mean()
            )

            faltante_g = max(
                st.session_state.meta_gestiones_cfg
                - promedio_g,
                0,
            )

            faltante_c = max(
                st.session_state.meta_compromisos_cfg
                - promedio_c,
                0,
            )

            faltante_r = max(
                st.session_state.meta_recuperacion_cfg
                - promedio_r,
                0,
            )

            disponibles = max(
                jornadas["disponibles"],
                1,
            )

            meta_diaria_g = math.ceil(
                faltante_g / disponibles
            )

            calculo_comp_v35 = meta_diaria_compromisos_calendario_v35(
                fecha_calculo,
                promedio_c,
            )
            meta_diaria_c = int(calculo_comp_v35["meta_diaria"])

            meta_diaria_r = (
                faltante_r / disponibles
            )

            d1, d2, d3 = st.columns(3)

            with d1:
                st.metric(
                    "Gestiones necesarias por día",
                    formato_entero(
                        meta_diaria_g
                    ),
                )

            with d2:
                st.metric(
                    "Compromisos necesarios por día",
                    formato_entero(
                        meta_diaria_c
                    ),
                )

            with d3:
                st.metric(
                    "Recuperación necesaria por día",
                    formato_bs(
                        meta_diaria_r
                    ),
                )

            st.caption(
                f"Compromisos: base {calculo_comp_v35['base_mes']} por jornada · "
                f"faltan {formato_entero(calculo_comp_v35['faltante'])} · "
                f"{calculo_comp_v35['jornadas_disponibles']} jornadas disponibles. "
                "El calendario controla este cálculo automáticamente."
            )

        else:
            st.info(
                "Carga el reporte de Promesas de Pago para calcular "
                "la meta diaria real según el avance acumulado."
            )

        st.divider()

        # -------------------------------------------------
        # REGLA DE RECUPERACIÓN
        # -------------------------------------------------

        st.markdown("### Regla de recuperación")

        st.success(
            "Recuperación ajustada = recuperación individual "
            "+ (monto Sin usuario ÷ 8 operadores)."
        )

        st.info(
            "El porcentaje se calcula contra la meta mensual en USD "
            "de recuperación definida arriba."
        )

        st.warning(
            "La fórmula de distribución Sin usuario ÷ 8 "
            "se mantiene fija para evitar errores."
        )
